"""
Financial Statement Reconciliation Tool.

This tool is a rule-based prototype developed as the software artefact of a
Bachelor's thesis. It compares figures from two text-based PDF financial
statements and reconciles the older statement's current-period figures against
the comparative-period figures presented in the newer statement.

The tool is designed specifically for Finnish micro-enterprises, text-based PDF
financial statements and two-column financial statement layouts.

The purpose of the tool is not to replace human judgement. Instead, it reduces
manual comparison work, increases transparency and highlights rows that require
user review.

Main processing flow:
1. Extract all detectable rows from PDF financial statements.
2. Reconstruct logical financial statement structure (income statement and balance sheet).
3. Identify financial statement items using a vocabulary-supported canonical model.
4. Parse monetary values and align current and comparative periods.
5. Match rows based on label, canonical concept, section and structural context.
6. Evaluate matches and explicitly flag uncertain or unmatched cases.
7. Output results transparently in Excel and in the Streamlit user interface.

Transparency and reporting:
The tool produces an Excel workbook where financial statements are presented in a
PDF-like structure. The income statement and balance sheet are reconstructed with
hierarchical structure and indentation preserved.

Rows are color-coded to support quick review:
- Green: matched (values reconcile)
- Yellow: requires review (uncertain parsing or matching)
- Red: deviation or missing counterpart
- Grey: structural or header row (not subject to reconciliation)

This approach ensures that all extracted rows are visible and no data is hidden,
supporting a transparent and review-oriented workflow.

Limitations:
- Only text-based PDFs are supported. Scanned documents require OCR and are not
  included in this version.
- The tool does not perform a full statutory or compliance-level validation.
- Reliability is limited to the defined scope and supported layouts.
- The tool supports reconciliation work and does not replace human judgement.

Architecture:
The implementation follows a transparent, stage-based processing pipeline:
1. Configuration and vocabularies
2. Normalization and canonical taxonomy
3. PDF extraction and logical row reconstruction
4. Monetary row parsing and structuring
5. Period alignment and reconciliation matching
6. Transparent decision logic and flagging
7. Excel and Streamlit reporting

Each row-level result is designed to expose:
- the reconciliation decision
- the primary reason
- supporting factors
- and the underlying evidence

Author: Elina Malkki
Version: 1.6.0
"""

# =========================================================
# DESIGN PHILOSOPHY AND IMPLEMENTATION PRINCIPLES
# =========================================================
# This implementation follows a deliberately constrained and
# transparent design approach suitable for financial data processing.
#
# 1. Deterministic over probabilistic
#    The system relies on rule-based logic instead of machine learning.
#    This ensures traceability, reproducibility and explainability of the
#    reconciliation decisions.
#
# 2. Canonical normalization as a core mechanism
#    Financial statement labels are normalized into canonical forms using
#    a predefined vocabulary (terms.xlsx). This allows consistent matching
#    across documents despite differences in terminology, language or
#    presentation.
#
# 3. Structure-aware matching
#    Matching is not based on text similarity alone. Each row is evaluated
#    using several dimensions: canonical label, statement section,
#    hierarchical position and document context.
#
# 4. Controlled use of positional data
#    PDF X/Y coordinate data is used to support column detection and
#    validation. It is not used as the sole basis for reconciliation because
#    PDF layouts vary between source documents.
#
# 5. Risk-based reconciliation
#    The system prioritizes reliability over coverage. Clear matches are
#    accepted, while uncertain cases are explicitly flagged for manual review.
#
# 6. Explicit handling of critical financial items
#    Key totals and result figures, such as revenue, profit and balance sheet
#    totals, are handled with stricter rules and explicit mappings to reduce
#    the risk of incorrect matches.
#
# 7. Extended vocabulary for real-world variability
#    Additional canonical terms are included to support detailed financial
#    statement structures commonly found in practice. These terms improve
#    recognition coverage but do not extend the validated scope of the tool.
#
# 8. Strict scope definition
#    The implementation has been designed and validated using financial
#    statements of Finnish micro-enterprises. The vocabulary may include terms
#    from broader contexts, but reliability is only claimed within this scope.
#
# 9. Clear processing stages
#    The processing pipeline is structured as extraction, normalization,
#    row structuring, matching and reporting.
#
# 10. Prototype-level design with practical applicability
#     The system is a controlled prototype intended to support reconciliation
#     work within its scope. It is not a fully generalized production system.
#
# Overall principle:
# The system supports professional judgement; it does not replace it.


import io
import re
import traceback
from collections import Counter
from functools import lru_cache
from datetime import datetime
from difflib import SequenceMatcher
from typing import Optional, Iterable

import pandas as pd
import pdfplumber
try:
    import streamlit as st
except ModuleNotFoundError:  # Allows module-level testing without Streamlit installed.
    class _StreamlitFallback:
        """Lightweight fallback for CLI runs and automated tests."""
        def __getattr__(self, name):
            def _noop(*args, **kwargs):
                return None
            return _noop
    st = _StreamlitFallback()
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# =========================================================
# MODULE MAP AND READING GUIDE
# =========================================================
# The code broadly follows a processing pipeline, although certain components
# (such as vocabulary loading and UI integration) are intentionally placed in
# different locations for practical reasons.
#
# The main logical stages are:
#
# 1. Configuration and vocabularies
#    - define constants, thresholds and canonical vocabularies used throughout the workflow.
#
# 2. Normalization and taxonomy
#    - convert different PDF spellings into comparable canonical forms and metadata.
#
# 3. PDF extraction and logical row reconstruction
#    - read text-based PDF rows and use pdfplumber word coordinates to support
#      column interpretation.
#
# 4. Monetary row structuring
#    - create structured rows with label, values, section, hierarchy and
#      canonical identifier.
#
# 5. Reconciliation logic
#    - align periods and identify candidate matches based on section,
#      canonical concept and value evidence.
#
# 6. Transparent decision logic
#    - assign reconciliation status, primary reason and supporting evidence.
#
# 7. Excel and UI output
#    - present results in a PDF-like Excel format and Streamlit interface,
#      preserving structure and highlighting review-relevant rows.
#
# Note:
# The optional terms.xlsx vocabulary is loaded later in the file because it
# depends on normalization and validation helpers defined earlier. This keeps
# the implementation in a single-file prototype while maintaining logical readability.

# =========================================================
# BASE SETTINGS
# =========================================================
# The external vocabulary file (terms.xlsx) contains synonyms, translations
# and presentation variants used in real financial statements.
#
# Only deterministic rules and core thresholds are defined in code to ensure
# consistent and transparent behaviour.

TOLERANCE = 0.01
ROUNDING_WARNING_LIMIT = 1.00

MATCH_VERIFIED = "verified_match"
MATCH_SUGGESTED = "suggested_match"
MATCH_REJECTED = "rejected_match"

# Detects financial statement numbers in PDF rows, for example 1 234,56, -123,45 and (123,45).
NUMBER_PATTERN = re.compile(
    r"""
    (?<!\w)
    \(?
    -?
    \d+(?:[ .]\d{3})*
    (?:,\d{2})?
    \)?
    (?!\w)
    """,
    re.VERBOSE,
)

PAGE_NOISE_PATTERNS = [
    re.compile(r"^page\s+\d+$", re.I),
    re.compile(r"^\d{1,3}$"),
]


# =========================================================
# DOCUMENT STRUCTURE DETECTION
# =========================================================
# These terms help separate the actual income statement and balance sheet from
# tables of contents, notes, detailed schedules and signature pages.

TOC_KEYWORDS = ["sisällysluettelo", "sisällys", "sisältö"]

TOC_RELATED_TERMS = [
    "tuloslaskelma", "tase", "tase vastaavaa", "tase vastattavaa",
    "liitetiedot", "tilinpäätöksen allekirjoitus",
    "income statement", "balance sheet", "notes",
    "resultatrakning", "resultaträkning", "balansrakning", "balansräkning", "noter",
]

SKIP_SECTION_KEYWORDS = [
    "tase erittely", "tase-erittely", "erittely",
    "avoimet myyntilaskut", "avoimet ostolaskut", "avoimet matka ja kululaskut",
    "saldoluettelo", "paakirja", "pääkirja", "tililuettelo",
    "liitetiedot", "liite", "notes",
    "toimintakertomus", "rahavirtalaskelma", "cash flow",
]

POST_BALANCE_STOP_TITLES = [
    "liitetiedot", "rahoituslaskelma", "tilintarkastuskertomus", "allekirjoitukset",
    "toimintakertomus", "notes", "cash flow", "independent report", "signatures",
    "management report", "noter", "revisionsberattelse", "underskrifter",
    "verksamhetsberattelse",
]

DOCUMENT_STOP_TERMS = sorted(set(SKIP_SECTION_KEYWORDS + POST_BALANCE_STOP_TITLES))

SIGNATURE_AND_ROLE_TERMS = [
    "allekirjoitus", "allekirjoitukset",
    "hallituksen puheenjohtaja", "hallituksen jasen", "hallituksen jäsen",
    "hallituksen varajasen", "hallituksen varajäsen",
    "toimitusjohtaja", "ceo", "managing director",
]

LOCATION_SIGNATURE_TERMS = [
    "tampereella", "espoossa", "helsingissa", "helsingissä",
    "lahdessa", "heinolassa", "lappeenrannassa",
]

BANNED_LABEL_TERMS = [
    "tilinpaatos", "tuloslaskelma", "tase", "liitetiedot", "toimintakertomus",
    "annual report", "balance sheet", "income statement", "page",
]


# =========================================================
# CANONICAL TERM MAPPING
# =========================================================
# These canonical terms define the internal recognition layer used during
# reconciliation. Labels extracted from PDF documents are normalized into
# a standard internal representation, which allows the system to match
# equivalent items even when documents use different naming conventions
# or languages, for example "Liikevaihto", "Revenue" and "Net Sales".


PMA_INCOME_STATEMENT_CANONICALS = ['bruttotulos',
 'liikevaihto',
 'liiketoiminnan muut tuotot',
 'materiaalit ja palvelut',
 'aineet tarvikkeet ja tavarat',
 'aineet ja tarvikkeet ja tavarat',
 'ostot tilikauden aikana',
 'varastojen muutos',
 'varaston lisays',
 'ulkopuoliset palvelut',
 'henkilostokulut',
 'palkat ja palkkiot',
 'henkilosivukulut',
 'elakekulut',
 'muut henkilosivukulut',
 'poistot ja arvonalentumiset',
 'suunnitelman mukaiset poistot',
 'arvonalentumiset',
 'liiketoiminnan muut kulut',
 'liikevoitto',
 'liiketappio',
 'liiketulos',
 'rahoitustuotot ja kulut',
 'tuotot osuuksista saman konsernin yrityksissa',
 'tuotot osuuksista omistusyhteysyrityksissa',
 'tuotot muista pysyvien vastaavien sijoituksista',
 'muut korko ja rahoitustuotot',
 'arvonalentumiset pysyvien vastaavien sijoituksista',
 'arvonalentumiset vaihtuvien vastaavien rahoitusarvopapereista',
 'korkokulut ja muut rahoituskulut',
 'voitto tappio ennen tilinpaatossiirtoja ja veroja',
 'tulos ennen tilinpaatossiirtoja ja veroja',
 'tilinpaatossiirrot',
 'poistoeron muutos',
 'verotusperusteisten varausten muutos',
 'tuloverot',
 'muut valittomat verot',
 'tilikauden voitto',
 'tilikauden tappio',
 'tilikauden tulos',
 'hankinnan ja valmistuksen kulut',
 'bruttokate',
 'myynnin ja markkinoinnin kulut',
 'hallinnon kulut',
 'liiketoiminnan muut tuotot toimintokohtainen',
 'liiketoiminnan muut kulut toimintokohtainen',
 'vastikkeet',
 'hoitovastikkeet',
 'paomavastikkeet',
 'vuokratuotot',
 'kayttokorvaukset',
 'kiinteiston hoitokulut',
 'rahoitustuotot kiinteistokaava',
 'rahoituskulut kiinteistokaava',
 'varsinainen toiminta',
 'varainhankinta',
 'sijoitus ja rahoitustoiminta',
 'yleisavustukset']

PMA_BALANCE_ASSET_CANONICALS = ['pysyvat vastaavat',
 'aineettomat hyodykkeet',
 'kehittamismenot',
 'aineettomat oikeudet',
 'liikearvo',
 'muut pitkaaikaiset menot',
 'aineettomien hyodykkeiden ennakkomaksut',
 'aineelliset hyodykkeet',
 'maa ja vesialueet',
 'rakennukset ja rakennelmat',
 'koneet ja kalusto',
 'muut aineelliset hyodykkeet',
 'ennakkomaksut ja keskeneraiset hankinnat',
 'sijoitukset',
 'osuudet saman konsernin yrityksissa',
 'saamiset saman konsernin yrityksilta',
 'osuudet omistusyhteysyrityksissa',
 'saamiset omistusyhteysyrityksilta',
 'muut osakkeet ja osuudet',
 'muut saamiset',
 'vaihtuvat vastaavat',
 'vaihto omaisuus',
 'aineet ja tarvikkeet',
 'keskeneraiset tuotteet',
 'valmiit tuotteet',
 'tavarat',
 'muu vaihto omaisuus',
 'vaihto omaisuuden ennakkomaksut',
 'saamiset',
 'pitkaaikaiset saamiset',
 'lyhytaikaiset saamiset',
 'myyntisaamiset',
 'lainasaamiset',
 'maksamattomat osakkeet',
 'maksamattomat osuudet',
 'siirtosaamiset',
 'rahoitusarvopaperit',
 'rahoitusarvopaperit osuudet saman konsernin yrityksissa',
 'rahoitusarvopaperit muut osakkeet ja osuudet',
 'muut arvopaperit',
 'rahat ja pankkisaamiset',
 'vastaavaa yhteensa']

PMA_BALANCE_LIABILITY_CANONICALS = ['oma paaoma',
 'osake paoma',
 'osuus paoma',
 'muu vastaava paaoma',
 'osake osuus tai muu vastaava paaoma',
 'ylikurssirahasto',
 'arvonkorotusrahasto',
 'muut rahastot',
 'sijoitetun vapaan oman paaoman rahasto',
 'vararahasto',
 'yhtiojarjestyksen saantojen mukaiset rahastot',
 'muut rahastot yhteensa',
 'edellisten tilikausien voitto',
 'edellisten tilikausien tappio',
 'tilikauden voitto',
 'tilikauden tappio',
 'tilikauden tulos',
 'tilinpaatossiirtojen kertyma',
 'poistoero',
 'verotusperusteiset varaukset',
 'pakolliset varaukset',
 'elakevaraukset',
 'verovaraukset',
 'muut pakolliset varaukset',
 'vieras paaoma',
 'pitkaaikainen vieras paaoma',
 'lyhytaikainen vieras paaoma',
 'joukkovelkakirjalainat',
 'vaihtovelkakirjalainat',
 'lainat rahoituslaitoksilta',
 'takaisinlainat tyoelakevakuutuslaitoksilta',
 'saadut ennakot',
 'ostovelat',
 'rahoitusvekselit',
 'velat saman konsernin yrityksille',
 'velat omistusyhteysyrityksille',
 'muut velat',
 'siirtovelat',
 'vastattavaa yhteensa']

PMA_STRUCTURE_VARIATION_TERMS = ['bruttotulos',
 'bruttokate',
 'toimintokohtainen',
 'kiinteistokaava',
 'yhdistys ja saatiokaava',
 'varsinainen toiminta',
 'varainhankinta',
 'sijoitus ja rahoitustoiminta',
 'yleisavustukset']


# =========================================================
# SECTION-SPECIFIC CANONICAL VOCABULARIES
# =========================================================
# Canonical terms are grouped by financial statement section. This prevents
# visually similar labels from being matched across incompatible sections,
# such as asset-side and liability-side balance sheet rows.

GENERIC_CANONICAL_TERMS = [
    "tilikauden voitto",
    "vastaavaa yhteensa",
    "vastattavaa yhteensa",
]

SECTION_CANONICAL_TERMS = {
    "generic": GENERIC_CANONICAL_TERMS,
    "tuloslaskelma": PMA_INCOME_STATEMENT_CANONICALS,
    "tase_vastaavaa": PMA_BALANCE_ASSET_CANONICALS,
    "tase_vastattavaa": PMA_BALANCE_LIABILITY_CANONICALS,
}

SECTION_ALLOWED_CANONICALS = {
    "tuloslaskelma": set(GENERIC_CANONICAL_TERMS + PMA_INCOME_STATEMENT_CANONICALS),
    "tase_vastaavaa": set(GENERIC_CANONICAL_TERMS + PMA_BALANCE_ASSET_CANONICALS),
    "tase_vastattavaa": set(GENERIC_CANONICAL_TERMS + PMA_BALANCE_LIABILITY_CANONICALS),
}

# =========================================================
# DERIVED LABEL LISTS
# =========================================================
# These lists are derived from the canonical vocabulary and are used to
# support normalization, recognition and matching during reconciliation.
# They are internal helper structures rather than part of the external
# interface.
INCOME_STATEMENT_LABEL_TERMS = list(PMA_INCOME_STATEMENT_CANONICALS)
BALANCE_ASSET_LABEL_TERMS = list(PMA_BALANCE_ASSET_CANONICALS)
BALANCE_LIABILITY_LABEL_TERMS = list(PMA_BALANCE_LIABILITY_CANONICALS)
GENERIC_STATEMENT_LABEL_TERMS = list(GENERIC_CANONICAL_TERMS)

TERM_SYNONYM_MAP = {}

MAIN_ITEM_STARTERS = (
    "liikevaihto", "bruttotulos", "liikevoitto", "liiketulos", "liiketappio",
    "tilikauden voitto", "pysyvat vastaavat", "vaihtuvat vastaavat",
    "oma paaoma", "vieras paaoma", "vastaavaa yhteensa", "vastattavaa yhteensa",
)

SUM_KEYWORDS = ['yhteensa',
 'yhteensä',
 'summa',
 'total',
 'totalt',
 'subtotal',
 'vastaavaa yhteensa',
 'vastattavaa yhteensa',
 'oma paaoma yhteensa',
 'vieras paaoma yhteensa',
 'pysyvat vastaavat yhteensa',
 'vaihtuvat vastaavat yhteensa',
 'total assets',
 'total liabilities',
 'summa tillgangar',
 'summa tillgångar',
 'summa skulder',
 'summa eget kapital och skulder',
 'liikevoitto',
 'liiketappio',
 'tilikauden voitto',
 'tilikauden tappio',
 'tilikauden tulos',
 'profit for the financial year',
 'loss for the financial year',
 'arets resultat',
 'årets resultat',
 'arets vinst',
 'årets vinst',
 'arets forlust',
 'årets förlust']
VASTAAVAA_TOTAL_TERMS = ['vastaavaa yhteensa', 'varat yhteensa', 'total assets', 'assets total', 'summa tillgangar']
VASTATTAVAA_TOTAL_TERMS = ['vastattavaa yhteensa',
 'oma paaoma ja velat yhteensa',
 'total liabilities',
 'liabilities and equity total',
 'summa eget kapital och skulder',
 'summa skulder']


# =========================================================
# FORCED SYNONYM MAPPING FOR KEY FINANCIAL STATEMENT ITEMS
# =========================================================
# This dictionary defines explicit synonym mappings for critical financial
# statement line items. These mappings complement the general vocabulary and
# ensure that essential items are consistently recognized even when different
# naming conventions are used. This is especially important for profit-related
# rows and key subtotal figures.

FORCED_MAIN_ITEM_SYNONYMS = {
    "liikevaihto": ["liikevaihto", "myyntituotot"],
    "liikevoitto": ["liikevoitto", "liiketulos", "liiketappio", "liikevoitto tappio"],
    "voitto tappio ennen tilinpaatossiirtoja ja veroja": [
        "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "voitto ennen tilinpaatossiirtoja ja veroja",
        "tappio ennen tilinpaatossiirtoja ja veroja",
        "tulos ennen tilinpaatossiirtoja ja veroja",
        "voitto ennen tilinpäätössiirtoja ja veroja",
        "tappio ennen tilinpäätössiirtoja ja veroja",
        "tulos ennen tilinpäätössiirtoja ja veroja",
        "voitto tappio ennen veroja",
        "voitto ennen veroja",
        "tappio ennen veroja",
        "tulos ennen veroja",
        "tilikauden tulos ennen veroja",
        "tilikauden voitto ennen veroja",
        "tilikauden tappio ennen veroja",
        "tilikauden tulos ennen tilinpaatossiirtoja ja veroja",
        "tilikauden voitto ennen tilinpaatossiirtoja ja veroja",
        "tilikauden tappio ennen tilinpaatossiirtoja ja veroja",
        "tilikauden tulos ennen tilinpäätössiirtoja ja veroja",
        "tilikauden voitto ennen tilinpäätössiirtoja ja veroja",
        "tilikauden tappio ennen tilinpäätössiirtoja ja veroja",
        "tulos ennen veroja yhteensa",
        "voitto ennen veroja yhteensa",
        "tulos ennen tilinpaatossiirtoja",
        "voitto ennen tilinpaatossiirtoja",
    ],
    "tilikauden voitto": [
        "tilikauden voitto", "tilikauden tappio", "tilikauden tulos",
        "tilikauden voitto tappio", "tilikauden voitto (tappio)",
        "tilikauden tappio (voitto)", "tilikauden tulos tappio", "tilikauden tulos voitto",
    ],
    "vastaavaa yhteensa": ["vastaavaa yhteensa", "tase vastaavaa yhteensa", "varat yhteensa"],
    "vastattavaa yhteensa": [
        "vastattavaa yhteensa", "tase vastattavaa yhteensa", "oma paaoma ja velat yhteensa",
    ],
    "oma paaoma yhteensa": ["oma paaoma yhteensa", "oma pääoma yhteensä"],
    "vieras paaoma yhteensa": [
        "vieras paaoma yhteensa", "vieras pääoma yhteensä", "velat yhteensa", "velat yhteensä",
    ],
}

FORCED_MAIN_ITEMS_BY_SECTION = {
    "tuloslaskelma": {
        "liikevaihto", "liikevoitto", "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "tilikauden voitto",
    },
    "tase_vastaavaa": {"vastaavaa yhteensa"},
    "tase_vastattavaa": {
        "oma paaoma yhteensa", "vieras paaoma yhteensa", "vastattavaa yhteensa", "tilikauden voitto",
    },
}

FORCED_MAIN_ITEM_CANONICALS = set().union(*FORCED_MAIN_ITEMS_BY_SECTION.values())


def _normalize_dictionary_term(value) -> str:
    """Normalize input values for reliable comparison."""
    text = str(value).lower().replace("ä", "a").replace("ö", "o").replace("å", "a").replace("\xa0", " ")
    text = re.sub(r"\((.*?)\)", r"\1", text)
    return re.sub(r"\s+", " ", text).strip()


for _canonical, _synonyms in FORCED_MAIN_ITEM_SYNONYMS.items():
    TERM_SYNONYM_MAP.setdefault(_canonical, [])
    for _synonym in _synonyms:
        _normal = _normalize_dictionary_term(_synonym)
        if _normal not in TERM_SYNONYM_MAP[_canonical]:
            TERM_SYNONYM_MAP[_canonical].append(_normal)


# =========================================================
# FORCED ROOT-LEVEL FINANCIAL STATEMENT ITEMS
# =========================================================
# These key financial statement items are treated as root-level elements
# during reconciliation. They represent high-level totals or core figures,
# such as revenue, operating profit and balance sheet totals. They are handled
# with higher priority and stricter matching rules to reduce the risk of
# incorrect matches in critical summary-level figures.

FORCED_ROOT_ITEMS = {
    "liikevaihto", "liikevoitto", "voitto tappio ennen tilinpaatossiirtoja ja veroja",
    "tilikauden voitto", "vastaavaa yhteensa", "vastattavaa yhteensa",
    "oma paaoma yhteensa", "vieras paaoma yhteensa",
}

BALANCE_GRAND_TOTAL_ROOT_ITEMS = {"vastaavaa yhteensa", "vastattavaa yhteensa"}

BALANCE_GRAND_TOTAL_ITEMS = {
    "vastaavaa yhteensa", "tase vastaavaa yhteensa",
    "vastattavaa yhteensa", "tase vastattavaa yhteensa", "oma paaoma ja velat yhteensa",
}

MAJOR_GROUP_TOTAL_ITEMS = {
    "oma paaoma yhteensa", "vieras paaoma yhteensa",
    "pysyvat vastaavat yhteensa", "vaihtuvat vastaavat yhteensa",
}

SUBGROUP_TOTAL_ITEMS = {
    "aineettomat hyodykkeet yhteensa", "aineelliset hyodykkeet yhteensa",
    "sijoitukset yhteensa", "vaihto omaisuus yhteensa", "saamiset yhteensa",
    "pitkaaikaiset saamiset yhteensa", "lyhytaikaiset saamiset yhteensa",
    "pitkaaikainen yhteensa", "lyhytaikainen yhteensa",
    "pitkaaikainen vieras paaoma yhteensa", "lyhytaikainen vieras paaoma yhteensa",
    "materiaalit ja palvelut yhteensa", "aineet tarvikkeet ja tavarat yhteensa",
    "aineet ja tarvikkeet yhteensa", "henkilostokulut yhteensa",
    "henkilosivukulut yhteensa", "poistot ja arvonalentumiset yhteensa",
    "rahoitustuotot ja kulut yhteensa", "tuloverot yhteensa",
    "rahoitusarvopaperit yhteensa", "tilinpaatossiirtojen kertyma yhteensa",
    "pakolliset varaukset yhteensa",
}


# =========================================================
# ADDITIONAL CANONICAL BALANCE SHEET LABELS
# =========================================================
# These lists extend the base vocabulary with detailed asset-side and
# liability-side labels found in Finnish accounting data. The terms improve
# recognition coverage for detailed rows, such as tax receivables, bank
# accounts and accrued payroll liabilities. They do not extend the validated
# scope of the tool beyond Finnish micro-enterprises.

DETAILED_BALANCE_EXTRA_CANONICALS_ASSETS = ['verotilisaamiset',
 'kateisvarat',
 'pankkitili',
 'maksuliikenne ja tasmaytykset',
 'pankkikorttiostojen selvittelytili',
 'luottokorttiostojen selvittelytili']

DETAILED_BALANCE_EXTRA_CANONICALS_LIABILITIES = ['paaomapanokset',
 'peruspaaoma',
 'yhteiset oman paaoman tilit',
 'yksityistilit tilikaudella',
 'yksityissijoitukset rahana',
 'eronneiden yhtiomiesten kumulatiiviset tilit',
 'kumulatiiviset yksityistilit',
 'verotilivelka',
 'arvonlisaverovelka',
 'ennakonpidatysvelka',
 'sosiaaliturvamaksuvelka',
 'palkkamenot siirtovelat',
 'elakevakuutusmaksut siirtovelat',
 'menojaamat',
 'tyottomyysvakuutusmaksut siirtovelat']

SECTION_ALLOWED_CANONICALS["tase_vastaavaa"].update(DETAILED_BALANCE_EXTRA_CANONICALS_ASSETS)
SECTION_ALLOWED_CANONICALS["tase_vastattavaa"].update(DETAILED_BALANCE_EXTRA_CANONICALS_LIABILITIES)

# =========================================================
# EXTEND LABEL TERM LISTS WITH DETAILED CANONICAL ENTRIES
# =========================================================
# The base label lists are extended with additional canonical terms.
# dict.fromkeys() removes duplicate entries while preserving the original
# order, which expands recognition coverage without affecting matching
# consistency.

BALANCE_ASSET_LABEL_TERMS = list(dict.fromkeys(BALANCE_ASSET_LABEL_TERMS + DETAILED_BALANCE_EXTRA_CANONICALS_ASSETS))
BALANCE_LIABILITY_LABEL_TERMS = list(dict.fromkeys(BALANCE_LIABILITY_LABEL_TERMS + DETAILED_BALANCE_EXTRA_CANONICALS_LIABILITIES))
GENERIC_STATEMENT_LABEL_TERMS = list(dict.fromkeys(GENERIC_STATEMENT_LABEL_TERMS + list(FORCED_MAIN_ITEM_CANONICALS)))

# =========================================================
# CONSOLIDATED SET OF KNOWN FINANCIAL STATEMENT LABELS
# =========================================================
# This tuple aggregates all canonical labels used in the recognition process,
# including income statement items, balance sheet items, detailed balance sheet
# entries and forced key item mappings. It serves as the master reference set
# for identifying financial statement rows during parsing and reconciliation.

KNOWN_STATEMENT_LABELS = tuple(
    dict.fromkeys(
        PMA_INCOME_STATEMENT_CANONICALS
        + PMA_BALANCE_ASSET_CANONICALS
        + PMA_BALANCE_LIABILITY_CANONICALS
        + DETAILED_BALANCE_EXTRA_CANONICALS_ASSETS
        + DETAILED_BALANCE_EXTRA_CANONICALS_LIABILITIES
        + list(FORCED_MAIN_ITEM_CANONICALS)
    )
)


# =========================================================
# SECTION NORMALIZATION AND HIERARCHY RULES
# =========================================================
# Section names are normalized into internal identifiers. Hierarchy rules define
# parent-child relationships between financial statement rows so that the
# reconciliation engine can distinguish subtotals, groups and detailed items.

SECTION_NAME_MAP = {
    "tuloslaskelma": "tuloslaskelma",
    "tase_vastaavaa": "tase_vastaavaa",
    "tase_vastattavaa": "tase_vastattavaa",
    "tuloslaskelma ": "tuloslaskelma",
    "tase vastaavaa": "tase_vastaavaa",
    "tase vastattavaa": "tase_vastattavaa",
    "income_statement": "tuloslaskelma",
    "balance_assets": "tase_vastaavaa",
    "balance_liabilities": "tase_vastattavaa",
}

HIERARCHY_RULES = {'tuloslaskelma': {'materiaalit ja palvelut': {'aineet ja tarvikkeet ja tavarat',
                                               'ostot tilikauden aikana',
                                               'ulkopuoliset palvelut',
                                               'varastojen muutos',
                                               'varaston lisays'},
                   'henkilostokulut': {'elakekulut',
                                       'henkilosivukulut',
                                       'muut henkilosivukulut',
                                       'palkat ja palkkiot'},
                   'henkilosivukulut': {'muut henkilosivukulut', 'elakekulut'},
                   'poistot ja arvonalentumiset': {'suunnitelman mukaiset poistot'},
                   'rahoitustuotot ja kulut': {'korkokulut ja muut rahoituskulut',
                                               'muut korko ja rahoitustuotot',
                                               'rahoituskulut',
                                               'rahoitustuotot'},
                   'rahoitustuotot': {'muut korko ja rahoitustuotot'},
                   'rahoituskulut': {'korkokulut ja muut rahoituskulut'}},
 'tase_vastaavaa': {'pysyvat vastaavat': {'aineelliset hyodykkeet',
                                          'aineettomat hyodykkeet',
                                          'aineettomat oikeudet',
                                          'koneet ja kalusto',
                                          'liikearvo',
                                          'maa ja vesialueet',
                                          'muut pitkaaikaiset menot',
                                          'rakennukset ja rakennelmat'},
                    'aineettomat hyodykkeet': {'aineettomat oikeudet',
                                               'liikearvo',
                                               'muut pitkaaikaiset menot'},
                    'aineelliset hyodykkeet': {'koneet ja kalusto',
                                               'maa ja vesialueet',
                                               'rakennukset ja rakennelmat'},
                    'vaihtuvat vastaavat': {'aineet ja tarvikkeet',
                                            'ennakkomaksut',
                                            'kateisvarat',
                                            'luottokorttiostojen selvittelytili',
                                            'lyhytaikaiset',
                                            'maksuliikenne ja tasmaytykset',
                                            'muut saamiset',
                                            'myyntisaamiset',
                                            'pankkikorttiostojen selvittelytili',
                                            'pankkitili',
                                            'pitkaaikaiset',
                                            'rahat ja pankkisaamiset',
                                            'saamiset',
                                            'siirtosaamiset',
                                            'vaihto omaisuus',
                                            'verotilisaamiset'},
                    'saamiset': {'lyhytaikaiset',
                                 'muut saamiset',
                                 'myyntisaamiset',
                                 'pitkaaikaiset',
                                 'siirtosaamiset',
                                 'verotilisaamiset'},
                    'pitkaaikaiset': {'muut saamiset'},
                    'lyhytaikaiset': {'muut saamiset',
                                      'myyntisaamiset',
                                      'siirtosaamiset',
                                      'verotilisaamiset'},
                    'vaihto omaisuus': {'ennakkomaksut', 'aineet ja tarvikkeet'}},
 'tase_vastattavaa': {'oma paaoma': {'edellisten tilikausien tappio',
                                     'edellisten tilikausien voitto',
                                     'eronneiden yhtiomiesten kumulatiiviset tilit',
                                     'kumulatiiviset yksityistilit',
                                     'muut rahastot',
                                     'osake osuus tai muu vastaava paaoma',
                                     'osake paoma',
                                     'paaomapanokset',
                                     'peruspaaoma',
                                     'tilikauden tappio',
                                     'tilikauden tulos',
                                     'tilikauden voitto',
                                     'yhteiset oman paaoman tilit',
                                     'yhtiojarjestyksen saantojen mukaiset rahastot',
                                     'yksityissijoitukset rahana',
                                     'yksityistilit tilikaudella'},
                      'muut rahastot': {'yhtiojarjestyksen saantojen mukaiset rahastot'},
                      'muut velat': {'arvonlisaverovelka',
                                     'ennakonpidatysvelka',
                                     'sosiaaliturvamaksuvelka',
                                     'verotilivelka'},
                      'siirtovelat': {'elakevakuutusmaksut siirtovelat', 'palkkamenot siirtovelat'},
                      'vieras paaoma': {'lainat rahoituslaitoksilta',
                                        'lyhytaikainen',
                                        'muut velat',
                                        'ostovelat',
                                        'pitkaaikainen',
                                        'saadut ennakot',
                                        'siirtovelat'},
                      'pitkaaikainen': {'lainat rahoituslaitoksilta', 'muut velat'},
                      'lyhytaikainen': {'arvonlisaverovelka',
                                        'elakevakuutusmaksut siirtovelat',
                                        'ennakonpidatysvelka',
                                        'muut velat',
                                        'ostovelat',
                                        'palkkamenot siirtovelat',
                                        'saadut ennakot',
                                        'siirtovelat',
                                        'sosiaaliturvamaksuvelka',
                                        'verotilivelka'}}}

TOP_LEVEL_HIERARCHY_PARENTS = {'tuloslaskelma': {'henkilostokulut',
                   'liiketappio',
                   'liiketoiminnan muut kulut',
                   'liiketoiminnan muut tuotot',
                   'liiketulos',
                   'liikevaihto',
                   'liikevoitto',
                   'materiaalit ja palvelut',
                   'poistot ja arvonalentumiset',
                   'rahoituskulut',
                   'rahoitustuotot',
                   'rahoitustuotot ja kulut',
                   'tilikauden tappio',
                   'tilikauden tulos',
                   'tilikauden voitto',
                   'tuloverot'},
 'tase_vastaavaa': {'vastaavaa yhteensa', 'pysyvat vastaavat', 'vaihtuvat vastaavat'},
 'tase_vastattavaa': {'vastattavaa yhteensa', 'oma paaoma', 'vieras paaoma'}}

# Compatibility alias for older functions.
LEGAL_BASE_STRUCTURE = HIERARCHY_RULES

KEY_ITEMS_BY_SECTION = {
    "tuloslaskelma": ["liikevaihto", "liikevoitto", "liiketulos", "tilikauden voitto"],
    "tase_vastaavaa": ["vastaavaa yhteensa", "pysyvat vastaavat", "vaihtuvat vastaavat"],
    "tase_vastattavaa": ["vastattavaa yhteensa", "oma paaoma", "vieras paaoma"],
}


# =========================================================
# CENTRALIZED CANONICAL SOURCE MODEL
# =========================================================
#
# In previous development stages, the same information existed in several forms:
# - in the HIERARCHY_RULES structure
# - in the FORCED_MAIN_ITEMS_BY_SECTION lists
#
# CANONICAL_SOURCE_MODEL model from which technical indexes are built.
# CANONICAL_SOURCE_MODEL -> SECTION_ALLOWED_CANONICALS
# -> label lists
# -> metadata
#

CANONICAL_SOURCE_MODEL: dict[str, dict[str, dict]] = {}


def _source_add_item(
    model: dict,
    section: str,
    canonical: str,
    *,
    item_type: str = "allowed_item",
    parent: Optional[str] = None,
    children: Optional[Iterable[str]] = None,
    origin: str = "pma",
) -> None:
   
    section = normalize_section_name(section)
    canonical = normalize_label(canonical or "")
    parent = normalize_label(parent or "") or None
    if not section or not canonical:
        return

    section_model = model.setdefault(section, {})
    item = section_model.setdefault(canonical, {
        "canonical": canonical,
        "section": section,
        "types": set(),
        "parents": set(),
        "children": set(),
        "origins": set(),
    })
    item["types"].add(item_type or "allowed_item")
    item["origins"].add(origin)
    if parent:
        item["parents"].add(parent)
    for child in children or []:
        child_norm = normalize_label(child)
        if child_norm:
            item["children"].add(child_norm)


def build_canonical_source_model() -> dict[str, dict[str, dict]]:
    """Build a derived data structure used by the reconciliation workflow."""
    model: dict[str, dict[str, dict]] = {}

    base_by_section = {
        "tuloslaskelma": PMA_INCOME_STATEMENT_CANONICALS,
        "tase_vastaavaa": PMA_BALANCE_ASSET_CANONICALS + DETAILED_BALANCE_EXTRA_CANONICALS_ASSETS,
        "tase_vastattavaa": PMA_BALANCE_LIABILITY_CANONICALS + DETAILED_BALANCE_EXTRA_CANONICALS_LIABILITIES,
    }

    for section, items in base_by_section.items():
        for canonical in items:
            _source_add_item(model, section, canonical, item_type="allowed_item", origin="pma_list")

    for section, parents in HIERARCHY_RULES.items():
        for parent, children in parents.items():
            _source_add_item(
                model,
                section,
                parent,
                item_type="summary" if children else "main_item",
                children=children,
                origin="hierarchy",
            )
            for child in children:
                _source_add_item(
                    model,
                    section,
                    child,
                    item_type="detail",
                    parent=parent,
                    origin="hierarchy",
                )

    for section, items in FORCED_MAIN_ITEMS_BY_SECTION.items():
        for canonical in items:
            _source_add_item(model, section, canonical, item_type="key_item", origin="forced_key_item")

    for canonical in BALANCE_GRAND_TOTAL_ITEMS:
        norm = normalize_label(canonical)
        if norm in {"vastaavaa yhteensa", "tase vastaavaa yhteensa"}:
            _source_add_item(model, "tase_vastaavaa", norm, item_type="grand_total", origin="total_classification")
        elif norm in {"vastattavaa yhteensa", "tase vastattavaa yhteensa", "oma paaoma ja velat yhteensa"}:
            _source_add_item(model, "tase_vastattavaa", norm, item_type="grand_total", origin="total_classification")

    for canonical in MAJOR_GROUP_TOTAL_ITEMS:
        norm = normalize_label(canonical)
        if norm in {"pysyvat vastaavat yhteensa", "vaihtuvat vastaavat yhteensa"}:
            _source_add_item(model, "tase_vastaavaa", norm, item_type="section_total", origin="total_classification")
        elif norm in {"oma paaoma yhteensa", "vieras paaoma yhteensa"}:
            _source_add_item(model, "tase_vastattavaa", norm, item_type="section_total", origin="total_classification")

    return model


def _source_section_terms(section: str) -> set[str]:
    """Return a derived value used by the reconciliation workflow."""
    section = normalize_section_name(section)
    return set(CANONICAL_SOURCE_MODEL.get(section, {}).keys())


def _source_hierarchy_terms(section: str) -> set[str]:
    """Return a derived value used by the reconciliation workflow."""
    section = normalize_section_name(section)
    terms = set()
    for canonical, meta in CANONICAL_SOURCE_MODEL.get(section, {}).items():
        terms.add(canonical)
        terms.update(meta.get("children", set()))
    return terms


# =========================================================
# CANONICAL TAXONOMY ENRICHMENT AND SOURCE MODEL INTEGRATION
# =========================================================
# This section builds and maintains the canonical taxonomy used by the reconciliation engine.
#
# It integrates:
# - Internal canonical vocabularies (hardcoded base structure)
# - External vocabulary extensions (terms.xlsx)
# - Hierarchical relationships between financial statement items
#
# The purpose of this layer is to ensure that:
# 1. Financial statement items can be matched consistently across documents
# 2. Structural relationships (parent-child, totals, subtotals) are preserved
# 3. The system can adapt to real-world variation without changing core logic
#
# This design separates domain knowledge (taxonomy) from matching logic,
# which improves maintainability, transparency and explainability.


def _extend_unique_list(target: list, additions: Iterable[str]) -> list:
 
    existing = set(target)
    for item in additions:
        if item not in existing:
            target.append(item)
            existing.add(item)
    return target


def _source_item_type_from_terms(item_type: Optional[str], category: Optional[str]) -> str:
   
    raw = normalize_keyword_text(item_type or category or "")
    mapping = {
        "item": "allowed_item",
        "line_item": "allowed_item",
        "financial_statement_item": "allowed_item",
        "statement_item": "allowed_item",
        "detail": "detail",
        "summary": "summary",
        "heading": "summary",
        "total": "section_total",
        "subtotal": "section_total",
        "sum": "section_total",
        "section_total": "section_total",
        "grand_total": "grand_total",
        "key_item": "key_item",
    }
    return mapping.get(raw, raw or "detail")


def apply_external_terms_to_source_model(model: dict) -> None:
    """Update the active structural context."""
    allowed_categories = {
        "statement_item",
        "financial_statement_item",
        "line_item",
        "item",
        "summary",
        "total",
        "subtotal",
    }
    for row in EXTERNAL_TERM_ROWS:
        category = normalize_keyword_text(row.get("category") or "statement_item")
        if category not in allowed_categories:
            continue
        section = normalize_section_name(row.get("section"))
        canonical = normalize_label(row.get("canonical") or "")
        parent = normalize_label(row.get("parent") or "") or None
        if not section or not canonical:
            continue
        _source_add_item(
            model,
            section,
            canonical,
            item_type=_source_item_type_from_terms(row.get("type"), category),
            parent=parent,
            origin="terms_xlsx",
        )


def _flatten_hierarchy_terms(section: str) -> set[str]:
    """Return a derived value used by the reconciliation workflow."""
    section = normalize_section_name(section)
    if CANONICAL_SOURCE_MODEL:
        return _source_hierarchy_terms(section)

    terms: set[str] = set()
    for parent, children in HIERARCHY_RULES.get(section, {}).items():
        terms.add(parent)
        terms.update(children)
    return terms


def rebuild_pma_taxonomy_indexes() -> None:
    """Build a derived data structure used by the reconciliation workflow."""
    global CANONICAL_SOURCE_MODEL
    global SECTION_ALLOWED_CANONICALS
    global INCOME_STATEMENT_LABEL_TERMS, BALANCE_ASSET_LABEL_TERMS
    global BALANCE_LIABILITY_LABEL_TERMS, GENERIC_STATEMENT_LABEL_TERMS
    global KNOWN_STATEMENT_LABELS

    CANONICAL_SOURCE_MODEL = build_canonical_source_model()
    apply_external_terms_to_source_model(CANONICAL_SOURCE_MODEL)

    income_terms = _source_section_terms("tuloslaskelma")
    asset_terms = _source_section_terms("tase_vastaavaa")
    liability_terms = _source_section_terms("tase_vastattavaa")

    SECTION_ALLOWED_CANONICALS = {
        "tuloslaskelma": set(GENERIC_CANONICAL_TERMS) | income_terms,
        "tase_vastaavaa": set(GENERIC_CANONICAL_TERMS) | asset_terms,
        "tase_vastattavaa": set(GENERIC_CANONICAL_TERMS) | liability_terms,
    }

    SECTION_CANONICAL_TERMS["generic"] = list(dict.fromkeys(GENERIC_CANONICAL_TERMS))
    SECTION_CANONICAL_TERMS["tuloslaskelma"] = list(dict.fromkeys(sorted(income_terms)))
    SECTION_CANONICAL_TERMS["tase_vastaavaa"] = list(dict.fromkeys(sorted(asset_terms)))
    SECTION_CANONICAL_TERMS["tase_vastattavaa"] = list(dict.fromkeys(sorted(liability_terms)))

    INCOME_STATEMENT_LABEL_TERMS = list(dict.fromkeys(sorted(income_terms)))
    BALANCE_ASSET_LABEL_TERMS = list(dict.fromkeys(sorted(asset_terms)))
    BALANCE_LIABILITY_LABEL_TERMS = list(dict.fromkeys(sorted(liability_terms)))
    GENERIC_STATEMENT_LABEL_TERMS = list(dict.fromkeys(GENERIC_CANONICAL_TERMS + list(FORCED_MAIN_ITEM_CANONICALS)))

    KNOWN_STATEMENT_LABELS = tuple(
        dict.fromkeys(
            list(income_terms)
            + list(asset_terms)
            + list(liability_terms)
            + list(FORCED_MAIN_ITEM_CANONICALS)
        )
    )

    # Taxonomy changed: clear structural matching caches.
    try:
        _all_allowed_canonical_keys.cache_clear()
        _safe_structural_canonical_from_label.cache_clear()
        forced_main_item_canonical.cache_clear()
    except Exception:
        pass


# =========================================================
# PRACTICAL TAXONOMY EXTENSIONS FOR REAL-WORLD VARIABILITY
# =========================================================
# This section introduces practical extensions to the canonical taxonomy.
#
# It captures real-world variation in financial statement presentation,
# especially cases where:
# - items are renamed or aggregated
# - industry-specific structures are used (e.g. real estate, associations)
# - reporting formats deviate from standard layouts
#
# The purpose of this layer is to improve recognition coverage without
# modifying the core canonical model or introducing probabilistic logic.
#
# This reflects a key design decision:
# Instead of using machine learning to handle variability, the system
# explicitly models known variation patterns in a controlled way.
#
# This improves:
# - transparency (all mappings are visible)
# - explainability (no hidden model behaviour)
# - reliability within the defined scope

def apply_pma_practical_extension_terms() -> None:
    """Apply pma practical extension terms.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    income_terms = [
        "valmiiden ja keskeneraisten tuotteiden varastojen muutos",
        "valmistus omaan kayttoon",
        "arvonalentumiset pysyvien vastaavien hyodykkeista",
        "vaihtuvien vastaavien poikkeukselliset arvonalentumiset",
        "konserniavustus",
        "kiinteiston tuotot",
        "luottotappiot ja oikaisuerat",
        "hallinto",
        "kaytto ja huolto",
        "ulkoalueiden hoito",
        "siivous",
        "lammitys",
        "vesi ja jatevesi",
        "sahko ja kaasu",
        "jatehuolto",
        "vahinkovakuutukset",
        "kiinteistovero",
        "korjaukset",
        "muut hoitokulut",
        "hoitokate",
        "osinkotuotot",
        "korkotuotot",
        "muut rahoitustuotot",
        "korkokulut",
        "muut rahoituskulut",
        "valittomat verot",
        "tilikauden ylijaama",
        "tilikauden alijaama",
        "ylijaama",
        "alijaama",
        "tuotot",
        "kulut",
        "tuotto kulujaama",
        "tuottojaama",
        "kulujaama",
    ]
    asset_terms = [
        "muut aineettomat hyodykkeet",
        "kiinteistojen vuokraoikeuksien hankintamenot",
    ]
    liability_terms = [
        "kayvan arvon rahasto",
    ]
    variation_terms = [
        "lyhennetty tuloslaskelma",
        "bruttotulos kaava",
        "toimintokohtainen tuloslaskelma",
        "kiinteiston hallintaan perustuva tuloslaskelma",
        "yhdistys ja saatio",
        "nollarivit jatetty pois",
        "erat yhdistelty",
        "erat uudelleennimetty",
    ]

    _extend_unique_list(SECTION_CANONICAL_TERMS["tuloslaskelma"], income_terms)
    _extend_unique_list(SECTION_CANONICAL_TERMS["tase_vastaavaa"], asset_terms)
    _extend_unique_list(SECTION_CANONICAL_TERMS["tase_vastattavaa"], liability_terms)
    SECTION_ALLOWED_CANONICALS["tuloslaskelma"].update(income_terms)
    SECTION_ALLOWED_CANONICALS["tase_vastaavaa"].update(asset_terms)
    SECTION_ALLOWED_CANONICALS["tase_vastattavaa"].update(liability_terms)

    globals()["INCOME_STATEMENT_LABEL_TERMS"] = list(dict.fromkeys(INCOME_STATEMENT_LABEL_TERMS + income_terms))
    globals()["BALANCE_ASSET_LABEL_TERMS"] = list(dict.fromkeys(BALANCE_ASSET_LABEL_TERMS + asset_terms))
    globals()["BALANCE_LIABILITY_LABEL_TERMS"] = list(dict.fromkeys(BALANCE_LIABILITY_LABEL_TERMS + liability_terms))

    for term in variation_terms:
        if term not in PMA_STRUCTURE_VARIATION_TERMS:
            PMA_STRUCTURE_VARIATION_TERMS.append(term)


    SUBGROUP_TOTAL_ITEMS.update({
        "kiinteiston tuotot yhteensa",
        "kiinteiston hoitokulut yhteensa",
        "pakolliset varaukset yhteensa",
        "tilinpaatossiirtojen kertyma yhteensa",
    })

# =========================================================
# CANONICAL CONTEXT VALIDATION AND SEMANTIC CONTROL LAYER
# =========================================================
# This section implements the semantic validation layer of the reconciliation engine.
#
# It defines and evaluates the contextual correctness of financial statement items
# based on a canonical metadata model.
#
# The metadata model captures:
# - Allowed sections for each canonical item
# - Expected parent-child relationships (hierarchy)
# - Item types (detail, total, key item, result)
# - Expected financial statement side (assets/liabilities/income statement)
#
# The purpose of this layer is to:
# 1. Prevent structurally incorrect matches (e.g. asset matched with liability)
# 2. Validate whether detected rows follow expected financial statement logic
# 3. Provide explainable reasoning for reconciliation decisions
#
# This acts as a semantic control mechanism, similar to internal control checks
# in financial financial review, where relationships and structure are validated in
# addition to numeric consistency.
#
# Design principle:
# Instead of relying on statistical similarity, the system enforces structural
# and semantic correctness through explicit rules and metadata.

SECTION_SIDE_MAP = {
    "tuloslaskelma": "income_statement",
    "tase_vastaavaa": "vastaavaa",
    "tase_vastattavaa": "vastattavaa",
}

CANONICAL_ITEM_METADATA: dict[str, dict] = {}


def _ensure_metadata(canonical: Optional[str]) -> dict:
   
    canonical = normalize_label(canonical or "")
    if not canonical:
        return {}
    return CANONICAL_ITEM_METADATA.setdefault(canonical, {
        "canonical": canonical,
        "allowed_sections": set(),
        "parent_by_section": {},
        "allowed_parents": set(),
        "types": set(),
        "expected_sides": set(),
    })


def _add_canonical_metadata(canonical: Optional[str], section: Optional[str], parent: Optional[str] = None, item_type: str = "detail", expected_side: Optional[str] = None) -> None:
    """Add canonical metadata.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    canonical = normalize_label(canonical or "")
    section = normalize_section_name(section)
    parent = normalize_label(parent or "") or None
    expected_side = normalize_keyword_text(expected_side or "") or None
    if not canonical or not section:
        return
    meta = _ensure_metadata(canonical)
    meta["allowed_sections"].add(section)
    meta["types"].add(item_type or "detail")
    side = expected_side or SECTION_SIDE_MAP.get(section)
    if side:
        meta["expected_sides"].add(side)
    if parent:
        meta["parent_by_section"].setdefault(section, set()).add(parent)
        meta["allowed_parents"].add(parent)


def _build_canonical_item_metadata() -> None:
    """Build a derived data structure used by the reconciliation workflow."""
    CANONICAL_ITEM_METADATA.clear()

    for section, items in CANONICAL_SOURCE_MODEL.items():
        section = normalize_section_name(section)
        for canonical, meta in items.items():
            type_set = set(meta.get("types", set())) or {"allowed_item"}
            parents = set(meta.get("parents", set()))
            if parents:
                for parent in parents:
                    for item_type in type_set:
                        _add_canonical_metadata(canonical, section, parent=parent, item_type=item_type)
            else:
                for item_type in type_set:
                    _add_canonical_metadata(canonical, section, parent=None, item_type=item_type)

    # TOP_LEVEL_HIERARCHY_PARENTS
    for section, items in TOP_LEVEL_HIERARCHY_PARENTS.items():
        section = normalize_section_name(section)
        for item in items:
            _add_canonical_metadata(item, section, item_type="top_level")

    for item in ["tilikauden voitto", "tilikauden tappio", "tilikauden tulos"]:
        _add_canonical_metadata(item, "tuloslaskelma", item_type="result")
        _add_canonical_metadata(item, "tase_vastattavaa", parent="oma paaoma", item_type="equity_result")

    for item in ["poistoero", "verotusperusteiset varaukset"]:
        _add_canonical_metadata(item, "tuloslaskelma", parent="tilinpaatossiirrot", item_type="appropriation")
        _add_canonical_metadata(item, "tase_vastattavaa", parent="tilinpaatossiirtojen kertyma", item_type="balance_detail")


def get_canonical_metadata(canonical: Optional[str]) -> dict:
    """Get canonical metadata.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    canonical = normalize_label(canonical or "")
    if not canonical:
        return {}
    return CANONICAL_ITEM_METADATA.get(canonical, {})


def metadata_primary_value(values: set | list | tuple | None) -> Optional[str]:
    """Metadata primary value.
    
    Purpose: This function belongs to the number parsing and validation stage.
    Why: It reduces the risk that formatting differences or unreadable values are treated as reliable evidence.
    """
    if not values:
        return None
    return sorted(list(values))[0]


def evaluate_canonical_context(
    canonical: Optional[str],
    section_name: Optional[str],
    category_path: Optional[list[str] | str] = None,
    row_type: Optional[str] = None,
) -> dict:
    """Evaluate whether the item is valid in its detected context."""
    canonical = normalize_label(canonical or "")
    section_key = normalize_section_name(section_name)
    meta = get_canonical_metadata(canonical)

    if isinstance(category_path, str):
        path_items = [normalize_label(p) for p in category_path.split(">") if normalize_label(p)]
    else:
        path_items = [normalize_label(p) for p in (category_path or []) if normalize_label(p)]

    result = {
        "context_status": "Ei metadataa",
        "context_score": 0.55,
        "context_reason": "No PMA-based metadata has been defined for the item.",
        "allowed_sections": "",
        "expected_section": None,
        "expected_parent": None,
        "expected_side": None,
        "item_type": None,
        "section_ok": True,
        "parent_ok": True,
    }

    if not canonical or not section_key:
        result.update({
            "context_status": "Uncertain",
            "context_score": 0.45,
            "context_reason": "The item section could not be verified.",
            "section_ok": False,
        })
        return result

    if not meta:
        return result

    allowed_sections = set(meta.get("allowed_sections", set()))
    parent_by_section = meta.get("parent_by_section", {}) or {}
    expected_parents = set(parent_by_section.get(section_key, set()))
    type_set = set(meta.get("types", set()))

    result.update({
        "allowed_sections": ", ".join(sorted(allowed_sections)),
        "expected_section": metadata_primary_value(allowed_sections),
        "expected_parent": metadata_primary_value(expected_parents),
        "expected_side": metadata_primary_value(meta.get("expected_sides", set())),
        "item_type": metadata_primary_value(type_set),
    })

    if allowed_sections and section_key not in allowed_sections:
        result.update({
            "context_status": "Ristiriitainen",
            "context_score": 0.20,
            "context_reason": (
                "Based on the PMA structure, the item belongs to section "
                f"{', '.join(sorted(allowed_sections))}, but it was found in section {section_key}."
            ),
            "section_ok": False,
            "parent_ok": False,
        })
        return result

    if row_type in {"result", "grand_total", "section_total", "heading"} or type_set.intersection({"key_item", "top_level", "grand_total", "summary", "main_item"}):
        result.update({
            "context_status": "Looginen",
            "context_score": 1.0,
            "context_reason": "The item is in an allowed section and no mandatory parent path is required.",
            "section_ok": True,
            "parent_ok": True,
        })
        return result

    if expected_parents:
        if not path_items:
            result.update({
                "context_status": "Uncertain",
                "context_score": 0.65,
                "context_reason": (
                    "The item is in the correct section, but the expected parent path(s) "
                    f"{', '.join(sorted(expected_parents))} ei voitu varmistaa."
                ),
                "section_ok": True,
                "parent_ok": False,
            })
            return result

        if expected_parents.intersection(set(path_items)):
            result.update({
                "context_status": "Looginen",
                "context_score": 0.95,
                "context_reason": "The item is in the correct section and expected parent path.",
                "section_ok": True,
                "parent_ok": True,
            })
            return result

        result.update({
            "context_status": "Uncertain",
            "context_score": 0.55,
            "context_reason": (
                "The item is in the correct section, but the parent path differs from the expected "
                f"({', '.join(sorted(expected_parents))})."
            ),
            "section_ok": True,
            "parent_ok": False,
        })
        return result

    result.update({
        "context_status": "Looginen",
        "context_score": 0.85,
        "context_reason": "The item is in a section allowed by the PMA structure.",
        "section_ok": True,
        "parent_ok": True,
    })
    return result


# =========================================================
# DOCUMENT SECTION HEADINGS
# =========================================================
DOCUMENT_SECTION_HEADING_TERMS = {
    "tuloslaskelma": ["tuloslaskelma", "income statement", "profit and loss", "resultatrakning", "resultaträkning"],
    "tase_vastaavaa": ["tase vastaavaa", "vastaavaa", "balance assets", "assets", "tillgangar", "tillgångar"],
    "tase_vastattavaa": ["tase vastattavaa", "vastattavaa", "balance liabilities", "equity and liabilities", "liabilities and equity", "eget kapital och skulder"],
}

# Compatibility alias for document termination terms.
DOCUMENT_SECTION_STOP_TERMS = DOCUMENT_STOP_TERMS

# =========================================================
# TEXT NORMALIZATION AND PDF LINE CLEANING
# =========================================================
# This section prepares raw PDF text for reliable downstream processing.
#
# PDF-extracted text often contains inconsistent whitespace, non-breaking
# spaces, page numbers and layout artefacts. These functions normalize the
# text into a stable format before section detection, row extraction and
# reconciliation are performed.
#
# This stage is intentionally conservative: it removes technical noise but
# does not make reconciliation decisions.


def normalize_text(text: Optional[str]) -> str:
    """Convert an input value into a safe comparable representation."""
    if text is None:
        return ""
    try:
        if pd.isna(text):
            return ""
    except Exception:
        pass
    return str(text).replace("\xa0", " ").replace("\u202f", " ").strip()


def normalize_keyword_text(text: str) -> str:
    """Normalize input values for reliable comparison."""
    text = normalize_text(text).lower()
    replacements = {"ä": "a", "ö": "o", "å": "a", "\xa0": " ", "\u202f": " "}
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def clean_lines(text: str) -> list[str]:
    """Clean lines.
    
    Purpose: This function belongs to the document extraction stage.
    Why: It makes the PDF input usable before reconciliation decisions are made.
    """
    lines = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        norm = normalize_keyword_text(line)
        if any(p.fullmatch(norm) for p in PAGE_NOISE_PATTERNS):
            continue
        lines.append(line)
    return lines


# =========================================================
# LABEL NORMALIZATION, CANONICAL RECOGNITION AND HIERARCHY CONTEXT
# =========================================================
# This section converts noisy PDF labels into stable canonical financial
# statement concepts and validates their structural context.
#
# It handles:
# - PDF extraction artefacts, such as glued words.
# - Account code prefixes and formatting differences.
# - Finnish, English and Swedish terminology variants.
# - Canonical term recognition using exact, compact and fuzzy matching.
# - Financial statement hierarchy tracking.
#
# The purpose is to make labels comparable across documents before any
# reconciliation decision is made.
#
# This layer is deliberately rule-based so that recognition decisions remain
# explainable and explainable.

@lru_cache(maxsize=1)
def _known_label_phrases_for_degluing() -> tuple[tuple[str, str], ...]:
    """Normalize input values for reliable comparison."""
    phrases = set()
    try:
        for items in SECTION_CANONICAL_TERMS.values():
            phrases.update(str(x) for x in items)
    except Exception:
        pass
    try:
        phrases.update(str(x) for x in KNOWN_STATEMENT_LABELS)
    except Exception:
        pass
    try:
        phrases.update(str(x) for x in FORCED_MAIN_ITEM_CANONICALS)
    except Exception:
        pass
    try:
        for canonical, synonyms in TERM_SYNONYM_MAP.items():
            phrases.add(str(canonical))
            phrases.update(str(x) for x in synonyms)
    except Exception:
        pass

    normalized = []
    seen = set()
    for phrase in phrases:
        norm = normalize_keyword_text(phrase)
        norm = norm.replace("yhteensä", "yhteensa")
        norm = re.sub(r"[^a-z0-9\s]", " ", norm)
        norm = re.sub(r"\s+", " ", norm).strip()
        compact = re.sub(r"[^a-z0-9]", "", norm)
        if len(compact) >= 6 and compact not in seen:
            normalized.append((norm, compact))
            seen.add(compact)

    normalized.sort(key=lambda x: len(x[1]), reverse=True)
    return tuple(normalized)


@lru_cache(maxsize=20000)
def deglue_pdf_label_text(label: str) -> str:
    """Normalize input values for reliable comparison."""
    original = str(label or "")
    if not original:
        return ""

    base = normalize_keyword_text(original)
    base = base.replace("&", " ja ")
    base = base.replace("/", " ")
    base = base.replace("-", " ")
    base = base.replace("–", " ")
    base = base.replace("—", " ")
    base = base.replace("yhteensä", "yhteensa")
    base = re.sub(r"\s+", " ", base).strip()

    compact = re.sub(r"[^a-z0-9]", "", base)
    if not compact:
        return base

    phrases = _known_label_phrases_for_degluing()

    for phrase, phrase_compact in phrases:
        if compact == phrase_compact:
            return phrase

    account_match = re.match(r"^(?P<code>\d{3,6})(?P<body>[a-z].*)$", compact)
    if account_match:
        code = account_match.group("code")
        body = account_match.group("body")
        for phrase, phrase_compact in phrases:
            if body == phrase_compact or body.endswith(phrase_compact) or phrase_compact in body:
                readable = body.replace(phrase_compact, phrase, 1)
                readable = re.sub(r"(?<=\D)(\d+)(?=yhteensa$)", r" \1 ", readable)
                readable = readable.replace("yhteensa", " yhteensa")
                return re.sub(r"\s+", " ", f"{code} {readable}").strip()

    return base


def normalize_label(label: str) -> str:
    """Normalize label.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    label = deglue_pdf_label_text(label)
    label = normalize_keyword_text(label)
    replacements = {
        "&": " ja ", "/": " ", "-": " ", "–": " ", "—": " ",
        "(": " ", ")": " ", "[": " ", "]": " ", ",": " ", ".": " ",
        ":": " ", ";": " ", "*": " ", "€": " ", "yhteensä": "yhteensa",
        "lyhytaikaiset": "lyhytaikainen", "pitkäaikaiset": "pitkaaikainen",
        "pitkäaikainen": "pitkaaikainen", "pysyvät": "pysyvat", "pysyvä": "pysyva",
        "voitto (tappio)": "voitto tappio", "liikevoitto (-tappio)": "liikevoitto tappio",
        "tilikauden voitto (tappio)": "tilikauden voitto tappio",
    }
    for old, new in replacements.items():
        label = label.replace(old, new)
    label = re.sub(r"^\d+\s+", " ", label)
    label = re.sub(r"\s+\d+$", " ", label)
    label = re.sub(r"\bviite\s*\d+\b", " ", label)
    label = re.sub(r"\bnote\s*\d+\b", " ", label)
    label = re.sub(r"[^a-z0-9\s]", " ", label)
    label = re.sub(r"\s+", " ", label).strip()
    return label


def clean_display_label(label: str) -> str:
    """Clean display label.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    label = normalize_text(label)
    deglued = deglue_pdf_label_text(label)
    if deglued and re.sub(r"[^a-z0-9]", "", normalize_keyword_text(label)) == re.sub(r"[^a-z0-9]", "", normalize_keyword_text(deglued)):
        label = deglued
    label = re.sub(r"\s+", " ", label)
    return label.strip()


def strip_account_code_prefix(label: str) -> str:
   
    label = normalize_text(label)
    return re.sub(r"^\s*\d{3,6}\s+", "", label).strip()


def normalize_section_name(section_name: Optional[str]) -> Optional[str]:
    """Normalize section name.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    if section_name is None:
        return None
    norm = normalize_keyword_text(str(section_name)).replace(" ", "_")
    return SECTION_NAME_MAP.get(norm, SECTION_NAME_MAP.get(normalize_keyword_text(str(section_name)), norm))


def normalize_section_for_matching(section_name: Optional[str]) -> str:
    """Normalize section for matching.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    norm = normalize_section_name(section_name) or ""
    norm = norm.replace(" ", "_").strip().lower()
    aliases = {
        "tulos_laskelma": "tuloslaskelma",
        "income_statement": "tuloslaskelma",
        "resultatrakning": "tuloslaskelma",
        "tase_vastaavaa": "tase_vastaavaa",
        "balance_assets": "tase_vastaavaa",
        "assets": "tase_vastaavaa",
        "tase_vastattavaa": "tase_vastattavaa",
        "balance_liabilities": "tase_vastattavaa",
        "liabilities": "tase_vastattavaa",
        "equity_and_liabilities": "tase_vastattavaa",
    }
    return aliases.get(norm, norm)


def hierarchy_child_to_parent_map(section_key: Optional[str]) -> dict[str, set[str]]:
    """Hierarchy child to parent map.
    
    Purpose: This function belongs to the financial statement structure stage.
    Why: It prevents rows from being compared across incompatible statement sections or hierarchy levels.
    """
    section_key = normalize_section_name(section_key)
    rules = HIERARCHY_RULES.get(section_key, {})
    reverse: dict[str, set[str]] = {}
    for parent, children in rules.items():
        for child in children:
            reverse.setdefault(child, set()).add(parent)
    return reverse


def hierarchy_parent_to_children_map(section_key: Optional[str]) -> dict[str, set[str]]:
    """Hierarchy parent to children map.
    
    Purpose: This function belongs to the financial statement structure stage.
    Why: It prevents rows from being compared across incompatible statement sections or hierarchy levels.
    """
    section_key = normalize_section_name(section_key)
    return HIERARCHY_RULES.get(section_key, {})


def hierarchy_ancestors_for_label(label: str, section_key: Optional[str]) -> set[str]:
    """Return a derived value used by the reconciliation workflow."""
    section_key = normalize_section_name(section_key)
    reverse = hierarchy_child_to_parent_map(section_key)
    ancestors: set[str] = set()
    stack = list(reverse.get(label, set()))
    while stack:
        parent = stack.pop()
        if parent in ancestors:
            continue
        ancestors.add(parent)
        stack.extend(reverse.get(parent, set()))
    return ancestors


def is_legal_parent_label(canonical_label: Optional[str], section_name: Optional[str]) -> bool:
    """Is legal parent label.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    section_key = normalize_section_name(section_name)
    return bool(canonical_label and canonical_label in HIERARCHY_RULES.get(section_key, {}))


def is_legal_known_label(canonical_label: Optional[str], section_name: Optional[str]) -> bool:
    """Is legal known label.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    section_key = normalize_section_name(section_name)
    if not canonical_label or not section_key:
        return False
    if canonical_label in SECTION_ALLOWED_CANONICALS.get(section_key, set()):
        return True
    if canonical_label in HIERARCHY_RULES.get(section_key, {}):
        return True
    if canonical_label in TOP_LEVEL_HIERARCHY_PARENTS.get(section_key, set()):
        return True
    return False


def classify_legal_basis(canonical_label: Optional[str], section_name: Optional[str], row_type: Optional[str] = None) -> dict:
   
    section_key = normalize_section_name(section_name)
    if not canonical_label or not section_key:
        return {"legal_basis_match": False, "legal_basis_level": "unknown_item", "legal_basis_parent": None}

    if canonical_label in FORCED_MAIN_ITEMS_BY_SECTION.get(section_key, set()):
        return {"legal_basis_match": True, "legal_basis_level": "forced_key_item", "legal_basis_parent": None}

    structure = LEGAL_BASE_STRUCTURE.get(section_key, {})
    if canonical_label in structure or canonical_label in TOP_LEVEL_HIERARCHY_PARENTS.get(section_key, set()):
        return {"legal_basis_match": True, "legal_basis_level": "statutory_main_item", "legal_basis_parent": None}

    reverse = hierarchy_child_to_parent_map(section_key)
    parents = reverse.get(canonical_label, set())
    if parents:
        level = "permitted_detail_item" if row_type == "detail" else "statutory_sub_item"
        return {"legal_basis_match": True, "legal_basis_level": level, "legal_basis_parent": sorted(parents)[0]}

    if row_type == "detail":
        return {"legal_basis_match": False, "legal_basis_level": "permitted_detail_item", "legal_basis_parent": None}

    return {"legal_basis_match": False, "legal_basis_level": "unknown_item", "legal_basis_parent": None}


def update_category_path(category_path: list[str], canonical_label: Optional[str], section_name: Optional[str], row_type: Optional[str] = None) -> list[str]:
    """Update the active structural context."""
    section_key = normalize_section_name(section_name)
    if not canonical_label or not section_key:
        return category_path
    if canonical_label in FORCED_MAIN_ITEMS_BY_SECTION.get(section_key, set()):
        return category_path
    if row_type in {"grand_total", "result"} or is_total_row(canonical_label):
        return category_path

    rules = HIERARCHY_RULES.get(section_key, {})
    if canonical_label not in rules and canonical_label not in TOP_LEVEL_HIERARCHY_PARENTS.get(section_key, set()):
        return category_path

    ancestors = hierarchy_ancestors_for_label(canonical_label, section_key)
    if not ancestors:
        return [canonical_label]

    for i in range(len(category_path) - 1, -1, -1):
        if category_path[i] in ancestors:
            return category_path[: i + 1] + [canonical_label]

    #If parent is missing from the document, a logical minimum path is constructed through the expected parent.
    expected = sorted(ancestors)[0]
    if expected != canonical_label:
        return [expected, canonical_label]
    return [canonical_label]


def evaluate_hierarchy_position(canonical_label: Optional[str], category_path: Optional[list[str]], section_name: Optional[str], row_type: Optional[str] = None) -> dict:
    """Evaluate hierarchy position.
    
    Purpose: This function belongs to the financial statement structure stage.
    Why: It prevents rows from being compared across incompatible statement sections or hierarchy levels.
    """
    section_key = normalize_section_name(section_name)
    category_path = category_path or []
    immediate_parent = category_path[-1] if category_path else None
    top_parent = category_path[0] if category_path else None
    result = {
        "score": 0.0,
        "status": "Not used",
        "parent": immediate_parent,
        "top_parent": top_parent,
        "category_path": " > ".join(category_path),
        "expected_parent": None,
    }

    if not canonical_label or not section_key or section_key not in HIERARCHY_RULES:
        return result

    legal = classify_legal_basis(canonical_label, section_key, row_type=row_type)
    ancestors = hierarchy_ancestors_for_label(canonical_label, section_key)

    if canonical_label in FORCED_MAIN_ITEMS_BY_SECTION.get(section_key, set()):
        result.update({"score": 1.0, "status": "Forced key item", "parent": None, "top_parent": None, "category_path": canonical_label, "expected_parent": None})
        return result

    if row_type in {"grand_total", "result"} or is_total_row(canonical_label):
        result.update({"score": 1.0, "status": "Rakenteellinen"})
        return result

    if canonical_label in HIERARCHY_RULES.get(section_key, {}) or canonical_label in TOP_LEVEL_HIERARCHY_PARENTS.get(section_key, set()):
        result.update({"score": 1.0, "status": "Header row"})
        return result

    if ancestors:
        result["expected_parent"] = sorted(ancestors)[0]
        if immediate_parent in ancestors:
            result.update({"score": 1.0, "status": "Looginen"})
            return result
        if any(parent in category_path for parent in ancestors):
            result.update({"score": 0.90, "status": "Logical under main group"})
            return result
        if not category_path:
            result.update({"score": 0.65, "status": "Uncertain"})
            return result
        if legal.get("legal_basis_match"):
            result.update({"score": 0.55, "status": "Uncertain"})
            return result
        result.update({"score": 0.35, "status": "Ristiriitainen"})
        return result

    if legal.get("legal_basis_match"):
        result.update({"score": 0.70, "status": "Lakipohjainen"})
        return result

    result.update({"score": 0.50, "status": "Neutraali"})
    return result


def should_update_current_parent(canonical_label: Optional[str], section_name: Optional[str], row_type: Optional[str]) -> bool:
    """Should update current parent.
    
    Purpose: This function belongs to the financial statement structure stage.
    Why: It prevents rows from being compared across incompatible statement sections or hierarchy levels.
    """
    section_key = normalize_section_name(section_name)
    if not canonical_label or not section_key:
        return False
    if row_type in {"noise", "detail", "grand_total", "result"} or is_total_row(canonical_label):
        return False
    return is_legal_parent_label(canonical_label, section_key) or canonical_label in TOP_LEVEL_HIERARCHY_PARENTS.get(section_key, set())

def normalize_for_matching(text: str) -> str:
    """Normalize for matching.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    text = strip_account_code_prefix(text)
    text = normalize_label(text)
    replacements = {
        "financial year": "tilikausi",
        "for the financial year": "tilikausi",
        "for period": "tilikausi",
        "rakenskapsperiodens": "tilikauden",
        "arets": "tilikauden",
        "ovriga": "muut",
        "rorelse": "liiketoiminnan ",
        "rorelseintakter": "liiketoiminnan muut tuotot",
        "rorelsekostnader": "liiketoiminnan muut kulut",
        "vinst": "voitto",
        "forlust": "tappio",
        "resultat": "tulos",
        "skatter": "tuloverot",
        "inkomstskatter": "tuloverot",
        "kundfordringar": "myyntisaamiset",
        "leverantorsskulder": "ostovelat",
        "ovriga skulder": "muut velat",
        "ovriga fordringar": "muut saamiset",
        "kassa och bank": "rahat ja pankkisaamiset",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def token_sort_key(text: str) -> str:
   
    tokens = [t for t in normalize_for_matching(text).split() if t]
    return " ".join(sorted(tokens))


def fuzzy_ratio(a: str, b: str) -> float:
   
    a_norm = token_sort_key(a)
    b_norm = token_sort_key(b)
    if not a_norm or not b_norm:
        return 0.0
    return SequenceMatcher(None, a_norm, b_norm).ratio()


def get_term_aliases(canonical: str) -> list[str]:
   
    aliases = [canonical]
    aliases.extend(TERM_SYNONYM_MAP.get(canonical, []))
    return list(dict.fromkeys(normalize_for_matching(a) for a in aliases if a))


def find_best_canonical_term(label: str, allowed_terms: Iterable[str]) -> dict:
    """Find best canonical term.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    norm_label = normalize_for_matching(label)
    best = {"canonical": None, "alias": None, "score": 0.0, "match_type": "none"}
    if not norm_label:
        return best

    for canonical in allowed_terms:
        for alias in get_term_aliases(canonical):
            if not alias:
                continue
            if norm_label == alias:
                score = 1.0
                match_type = "exact"
            elif alias in norm_label or norm_label in alias:
                shorter = min(len(alias), len(norm_label))
                longer = max(len(alias), len(norm_label))
                score = 0.86 if longer == 0 else max(0.70, shorter / longer)
                match_type = "contains"
            else:
                score = fuzzy_ratio(norm_label, alias)
                match_type = "fuzzy"

            if score > best["score"]:
                best = {
                    "canonical": canonical,
                    "alias": alias,
                    "score": round(score, 4),
                    "match_type": match_type,
                }
    return best


def get_allowed_canonical_terms(section_name: Optional[str]) -> set[str]:
    """Get allowed canonical terms.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    section_key = normalize_section_name(section_name)
    if section_key in SECTION_ALLOWED_CANONICALS:
        return SECTION_ALLOWED_CANONICALS[section_key]
    return set(SECTION_CANONICAL_TERMS["generic"])


def evaluate_label_plausibility(label: str, section_name: Optional[str] = None, row_type: Optional[str] = None) -> dict:
    """Evaluate label plausibility.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    section_name = normalize_section_name(section_name)
    norm = normalize_for_matching(label)
    result = {
        "accepted": False,
        "score": 0.0,
        "canonical": None,
        "match_type": "none",
        "reason": None,
    }

    if not norm or len(norm) < 2:
        result["reason"] = "empty"
        return result

    if looks_like_non_data_line(label):
        result["reason"] = "noise"
        return result

    forced = forced_main_item_canonical(label, section_name)
    if forced:
        result.update({"accepted": True, "score": 1.0, "canonical": forced, "match_type": "forced_main_item"})
        return result

    compact_canonical = _compact_canonical_lookup(label, section_name)
    if compact_canonical:
        result.update({"accepted": True, "score": 1.0, "canonical": compact_canonical, "match_type": "compact_canonical"})
        return result

    if re.search(r"\b(oy|oyj|ab|abp|ky|ay|tmi|ltd|llc)\b", norm) and not is_total_row(label):
        result["reason"] = "company_form"
        return result

    if row_type in {"result", "grand_total", "section_total"}:
        result.update({"accepted": True, "score": 1.0, "canonical": canonicalize_common_labels(label), "match_type": "structural"})
        return result

    allowed = get_allowed_canonical_terms(section_name)
    best = find_best_canonical_term(label, allowed)
    generic_best = find_best_canonical_term(label, SECTION_CANONICAL_TERMS["generic"])

    chosen = best if best["score"] >= generic_best["score"] else generic_best

    if chosen["score"] >= 0.88:
        result.update({"accepted": True, "score": chosen["score"], "canonical": canonicalize_common_labels(chosen["canonical"]), "match_type": chosen["match_type"]})
        return result

    if chosen["score"] >= 0.76 and len(norm.split()) >= 2:
        result.update({"accepted": True, "score": chosen["score"], "canonical": canonicalize_common_labels(chosen["canonical"]), "match_type": f"soft_{chosen['match_type']}"})
        return result

    if looks_like_statement_label(label):
        result.update({
            "accepted": True,
            "score": max(chosen["score"], 0.82),
            "canonical": chosen["canonical"] or normalize_label(label),
            "match_type": "known_statement_label",
        })
        return result

    if is_total_row(label) and len(norm.split()) >= 2:
        result.update({"accepted": True, "score": max(chosen["score"], 0.74), "canonical": chosen["canonical"] or canonicalize_common_labels(label), "match_type": "total_fallback"})
        return result

    result.update({"reason": "low_plausibility", "score": chosen["score"], "canonical": chosen["canonical"], "match_type": chosen["match_type"]})
    return result


def _compact_canonical_lookup(label: str, section: Optional[str] = None) -> Optional[str]:
    """Return a derived value used by the reconciliation workflow."""
    norm = normalize_label(label)
    compact = re.sub(r"[^a-z0-9]", "", norm)
    if not compact:
        return None

    early_safe_variants = {
        "tulosennenveroja": "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "voittoennenveroja": "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "tappioennenveroja": "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "tulosennentilinpaatossiirtoja": "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "voittoennentilinpaatossiirtoja": "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "tappioennentilinpaatossiirtoja": "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "tulosennentilinpaatossiirtojajaveroja": "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "voittoennentilinpaatossiirtojajaveroja": "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "tappioennentilinpaatossiirtojajaveroja": "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "muutpitkavaikutteisetmenot": "muut pitkaaikaiset menot",
        "muutpitkaaikaisetmenot": "muut pitkaaikaiset menot",
    }
    if compact in early_safe_variants:
        return early_safe_variants[compact]

    candidates = set()
    try:
        candidates.update(SECTION_CANONICAL_TERMS.get("generic", []))
        if section:
            section_key = normalize_section_name(section)
            candidates.update(SECTION_CANONICAL_TERMS.get(section_key, []))
            candidates.update(SECTION_ALLOWED_CANONICALS.get(section_key, set()))
        else:
            for items in SECTION_CANONICAL_TERMS.values():
                candidates.update(items)
            for items in SECTION_ALLOWED_CANONICALS.values():
                candidates.update(items)
    except Exception:
        pass

    try:
        candidates.update(FORCED_MAIN_ITEM_CANONICALS)
        candidates.update(BALANCE_GRAND_TOTAL_ITEMS)
        candidates.update(MAJOR_GROUP_TOTAL_ITEMS)
        candidates.update(SUBGROUP_TOTAL_ITEMS)
    except Exception:
        pass

    candidates.update({
        "lyhytaikainen yhteensa",
        "lyhytaikaiset yhteensa",
        "pitkaaikainen yhteensa",
        "pitkaaikaiset yhteensa",
        "tuloverot yhteensa",
        "aineet tarvikkeet ja tavarat yhteensa",
        "rahoitustuotot ja kulut yhteensa",
        "varaston lisays tai vahennys",
        "voitto tappio ennen satunnaisia eria",
        "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "tulos ennen tilinpaatossiirtoja ja veroja",
        "tulos ennen veroja",
        "voitto ennen veroja",
        "tappio ennen veroja",
        "laskennalliset verot",
        "muut pitkavaikutteiset menot",
        "muut pitkaaikaiset menot",
    })

    compact_map = {}
    for candidate in candidates:
        cand_norm = normalize_label(candidate)
        cand_compact = re.sub(r"[^a-z0-9]", "", cand_norm)
        if cand_compact:
            compact_map.setdefault(cand_compact, cand_norm)

    if compact in compact_map:
        return compact_map[compact]

    safe_variants = {
        "lyhytaikaisetyhteensa": "lyhytaikainen yhteensa",
        "lyhytaikainenyhteensa": "lyhytaikainen yhteensa",
        "pitkaaikaisetyhteensa": "pitkaaikainen yhteensa",
        "pitkaaikainenyhteensa": "pitkaaikainen yhteensa",
        "tuloverotyhteensa": "tuloverot yhteensa",
        "tulosennenveroja": "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "voittoennenveroja": "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "tappioennenveroja": "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "tulosennentilinpaatossiirtoja": "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "voittoennentilinpaatossiirtoja": "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "tulosennentilinpaatossiirtojajaveroja": "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "voittoennentilinpaatossiirtojajaveroja": "voitto tappio ennen tilinpaatossiirtoja ja veroja",
        "muutpitkavaikutteisetmenot": "muut pitkaaikaiset menot",
        "muutpitkaaikaisetmenot": "muut pitkaaikaiset menot",
    }
    return safe_variants.get(compact)


def canonicalize_common_labels(label: str) -> str:
    """Normalize input values for reliable comparison."""
    label = normalize_label(label)
    compact_lookup = _compact_canonical_lookup(label)
    if compact_lookup:
        return compact_lookup

    compact = label.replace(" ", "")
    compact_synonyms = {
        "aineettomathyodykkeetyhteensa": "aineettomat hyodykkeet yhteensa",
        "aineellisethyodykkeetyhteensa": "aineelliset hyodykkeet yhteensa",
        "pysyvatvastaavatyhteensa": "pysyvat vastaavat yhteensa",
        "vaihtuvatvastaavatyhteensa": "vaihtuvat vastaavat yhteensa",
        "saamisetyhteensa": "saamiset yhteensa",
        "rahatjapankkisaamisetyhteensa": "rahat ja pankkisaamiset yhteensa",
        "tasevastaavaayhteensa": "vastaavaa yhteensa",
        "tasevastattavaayhteensa": "vastattavaa yhteensa",
        "liiketulos": "liikevoitto tappio",
        "tilikauden tulos": "tilikauden voitto tappio",
    }
    if compact in compact_synonyms:
        return compact_synonyms[compact]

    if "tulos ennen veroja" in label or "voitto ennen veroja" in label or "tappio ennen veroja" in label:
        return "voitto tappio ennen tilinpaatossiirtoja ja veroja"
    if "tulos ennen tilinpaatossiirtoja" in label or "voitto ennen tilinpaatossiirtoja" in label or "tappio ennen tilinpaatossiirtoja" in label:
        return "voitto tappio ennen tilinpaatossiirtoja ja veroja"

    if "muut pitkavaikutteiset menot" in label or "muut pitkaaikaiset menot" in label:
        return "muut pitkaaikaiset menot"

    synonyms = {
        "liikevaihto": "liikevaihto",
        "materiaalit ja palvelut": "materiaalit ja palvelut",
        "henkilostokulut": "henkilostokulut",
        "poistot ja arvonalentumiset": "poistot ja arvonalentumiset",
        "liikevoitto tappio": "liikevoitto tappio",
        "tilikauden voitto tappio": "tilikauden voitto tappio",
        "vastaavaa yhteensa": "vastaavaa yhteensa",
        "vastattavaa yhteensa": "vastattavaa yhteensa",
    }
    for key, value in synonyms.items():
        if key in label:
            return value
    return label


# =========================================================
# NUMERIC PARSING AND PRESENTATION-INTEGRITY CONTROLS
# =========================================================
# This section converts Finnish-formatted monetary values into numeric values
# and checks whether the original PDF presentation is reliable enough for
# reconciliation.
#
# It handles:
# - Finnish decimal commas.
# - Space or dot thousand separators.
# - Negative values shown with parentheses.
# - Document-level dominant number formatting.
# - Decimal and separator inconsistencies.
# - Cases where a value was technically repaired before comparison.
#
# The purpose is not only to parse numbers, but also to detect when the
# presentation of a number is too uncertain to be accepted silently.
#
# This supports explainability by keeping a distinction between:
# - The interpreted numeric value.
# - The original value as presented in the PDF.
# - Any formatting or repair issue affecting reliability.


def parse_finnish_number(value: str) -> float:
    """Parse finnish number.
    
    Purpose: This function belongs to the number parsing and validation stage.
    Why: It reduces the risk that formatting differences or unreadable values are treated as reliable evidence.
    """
    value = str(value).strip().replace("\xa0", "").replace("\u202f", "").replace(" ", "")
    negative = False
    if value.startswith("(") and value.endswith(")"):
        negative = True
        value = value[1:-1]
    value = value.replace(".", "")
    value = value.replace(",", ".")
    if value in {"", "-"}:
        raise ValueError("Arvo ei ole numero")
    number = float(value)
    return -number if negative else number


def parse_number_style(raw: str) -> dict:
    """Parse number style.
    
    Purpose: This function belongs to the number parsing and validation stage.
    Why: It reduces the risk that formatting differences or unreadable values are treated as reliable evidence.
    """
    raw = normalize_text(raw)
    cleaned = raw
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = cleaned[1:-1]
    cleaned = cleaned.lstrip("-")

    decimal_separator = None
    decimal_count = 0
    if "," in cleaned:
        decimal_separator = ","
        decimal_count = len(cleaned.split(",")[-1])
    elif "." in cleaned:
        last_dot = cleaned.rfind(".")
        right = cleaned[last_dot + 1:]
        left = cleaned[:last_dot]
        if right.isdigit() and 1 <= len(right) <= 2 and re.search(r"\d", left):
            decimal_separator = "."
            decimal_count = len(right)

    thousands_separator = None
    if re.search(r"\d \d{3}(?:\D|$)", cleaned):
        thousands_separator = "space"
    elif decimal_separator == "," and re.search(r"\d\.\d{3}(?:,\d+)?$", cleaned):
        thousands_separator = "dot"
    elif decimal_separator is None and re.search(r"\d\.\d{3}(?:\D|$)", cleaned):
        thousands_separator = "dot"

    numeric_value = None
    try:
        numeric_value = parse_finnish_number(raw)
    except Exception:
        pass

    return {
        "raw": raw,
        "decimal_separator": decimal_separator,
        "decimal_count": decimal_count,
        "thousands_separator": thousands_separator,
        "numeric_value": numeric_value,
    }


def classify_number_format(raw: str) -> str:
    """Classify number format.
    
    Purpose: This function belongs to the number parsing and validation stage.
    Why: It reduces the risk that formatting differences or unreadable values are treated as reliable evidence.
    """
    style = parse_number_style(raw)
    if style["thousands_separator"] == "space" and style["decimal_separator"] == "," and style["decimal_count"] == 2:
        return "space_thousands_comma_decimal"
    if style["thousands_separator"] is None and style["decimal_separator"] == "," and style["decimal_count"] == 2:
        return "no_thousands_comma_decimal"
    if style["decimal_separator"] == ".":
        return "dot_decimal"
    if style["thousands_separator"] == "dot" and style["decimal_separator"] == ",":
        return "dot_thousands_comma_decimal"
    if style["thousands_separator"] is None and style["decimal_separator"] is None:
        return "integer_only"
    return "other"


def only_thousands_separator_diff(raw1: Optional[str], raw2: Optional[str]) -> bool:
    
    if not raw1 or not raw2:
        return False
    s1 = parse_number_style(raw1)
    s2 = parse_number_style(raw2)
    if s1["numeric_value"] is None or s2["numeric_value"] is None:
        return False
    if abs(s1["numeric_value"] - s2["numeric_value"]) >= TOLERANCE:
        return False
    if s1["decimal_separator"] != s2["decimal_separator"]:
        return False
    if s1["decimal_count"] != s2["decimal_count"]:
        return False
    return s1["thousands_separator"] != s2["thousands_separator"]


def choose_document_dominant_format(df: pd.DataFrame) -> dict:
    """Identify the dominant numeric presentation style used in a document.

    The detected format is later used as a benchmark for identifying unusual
    number presentations, such as missing decimals or inconsistent separators."""
    decimal_separators = []
    decimal_counts = []
    thousands_separators = []
    for col in ["current_value_raw_original", "comparison_value_raw_original"]:
        if col not in df.columns:
            continue
        for raw in df[col].dropna():
            style = parse_number_style(raw)
            if style["decimal_separator"] is not None:
                decimal_separators.append(style["decimal_separator"])
            if style["decimal_count"] > 0:
                decimal_counts.append(style["decimal_count"])
            if style["numeric_value"] is not None and abs(style["numeric_value"]) >= 1000 and style["thousands_separator"] is not None:
                thousands_separators.append(style["thousands_separator"])
    return {
        "decimal_separator": Counter(decimal_separators).most_common(1)[0][0] if decimal_separators else ",",
        "decimal_count": Counter(decimal_counts).most_common(1)[0][0] if decimal_counts else 2,
        "thousands_separator": Counter(thousands_separators).most_common(1)[0][0] if thousands_separators else "space",
    }


def document_format_issue(raw: Optional[str], dominant_format: Optional[dict]) -> bool:
    """Document format issue.
    
    Purpose: This function belongs to the risk and quality-control stage.
    Why: It makes uncertainty explicit instead of silently accepting weak evidence.
    """
    if raw is None or dominant_format is None:
        return False
    style = parse_number_style(raw)
    if style["numeric_value"] is None:
        return True
    if dominant_format["decimal_separator"] is not None and style["decimal_separator"] != dominant_format["decimal_separator"]:
        return True
    if dominant_format["decimal_count"] is not None and style["decimal_count"] != dominant_format["decimal_count"]:
        return True
    if abs(style["numeric_value"]) >= 1000 and style["thousands_separator"] != dominant_format["thousands_separator"]:
        return True
    return False


def has_decimal_presentation_defect(raw: Optional[str], dominant_format: Optional[dict]) -> bool:
  
    if raw is None or dominant_format is None:
        return False
    style = parse_number_style(raw)
    if dominant_format.get("decimal_count") == 2:
        if style["decimal_separator"] is None and style["decimal_count"] == 0:
            return True
        if style["decimal_count"] not in (0, 2):
            return True
    return False


def parse_original_display_number(raw: Optional[str]) -> Optional[float]:
    """Parse and interpret the numeric value shown in the PDF."""
    if raw is None or str(raw).strip() == "":
        return None
    try:
        return parse_finnish_number(str(raw))
    except Exception:
        return None


def has_review_unsafe_numeric_repair(raw: Optional[str], was_repaired: bool, dominant_format: Optional[dict]) -> bool:
    """Determine whether a numeric value is unsafe to accept without manual review.

    A value is considered unsafe for automatic acceptance if it was technically repaired, if it
    contains suspicious digit grouping, or if its decimal presentation differs
    from the document's dominant numeric format."""
    if was_repaired:
        return True
    raw_text = str(raw or "").strip()
    if re.search(r"\d+\s+\d{4,}$", raw_text) and not re.search(r"[,\.]\d{2}\b", raw_text):
        return True
    if has_decimal_presentation_defect(raw, dominant_format):
        return True
    return False


def describe_numeric_integrity_issue(raw_old, raw_new, old_raw_numeric, new_raw_numeric, repaired_issue, decimal_presentation_issue) -> str:
    """Describe numeric integrity issue.
    
    Purpose: This function belongs to the risk and quality-control stage.
    Why: It makes uncertainty explicit instead of silently accepting weak evidence.
    """
    parts = []
    if repaired_issue:
        parts.append("the value was produced with a technical repair")
    if decimal_presentation_issue:
        parts.append("the decimal separator is missing or differs from the expected format")
    if old_raw_numeric is not None and new_raw_numeric is not None:
        parts.append(f"the original presented values are {old_raw_numeric:.2f} and {new_raw_numeric:.2f}")
    elif raw_old is not None or raw_new is not None:
        parts.append(f"original presentations: {raw_old} / {raw_new}")
    return "; ".join(parts) or "the value could not be verified without interpretation"


# =========================================================
# ROW FILTERING, KEY ITEM DETECTION AND ROW TYPE CLASSIFICATION
# =========================================================
# This section separates usable financial statement rows from document noise
# and classifies the role of each extracted label.
#
# It identifies:
# - Company metadata and page markers.
# - Non-data lines, signatures and document headings.
# - Forced key financial statement items.
# - Detail account rows, totals, grand totals and result rows.
#
# The purpose is to prevent non-financial text from entering the reconciliation
# logic and to assign each valid row a structural role before matching.
#
# This improves reliability because totals, result rows and detail rows are not
# treated as interchangeable evidence.


def is_business_id(line: str) -> bool:
   
    return bool(re.search(r"\b\d{7}-\d\b", normalize_keyword_text(line)))


def is_page_marker(line: str) -> bool:
  
    norm = normalize_keyword_text(line)
    if any(p.fullmatch(norm) for p in PAGE_NOISE_PATTERNS):
        return True
    return bool(re.match(r"^(sivu|page)\s*\d+", norm))


def is_company_name_line(line: str) -> bool:
   
    raw = normalize_text(line)
    norm = normalize_keyword_text(line)

    if re.match(r"^\s*\d{3,6}\s+", raw):
        return False

    company_forms = [" oy", " oyj", " ab", " abp", " ky", " ay", " tmi", " ltd", " llc"]
    if any(form in f" {norm} " for form in company_forms):
        return len(norm.split()) <= 6
    return False


def looks_like_company_metadata_line(line: str) -> bool:
    
    norm = normalize_keyword_text(line)
    raw = normalize_text(line)

    if not norm:
        return True

    if re.match(r"^\s*\d{3,6}\s+", raw):
        return False

    if is_business_id(line):
        return True

    if is_company_name_line(line):
        return True

    metadata_terms = [
        "y tunnus",
        "y-tunnus",
        "business id",
        "organisationsnummer",
        "fo nummer",
        "fo-nummer",
        "kotipaikka",
        "domicile",
        "hemort",
        "lahiosoite",
        "postiosoite",
        "address",
    ]
    if any(term in norm for term in metadata_terms):
        return True

    return False


def is_metadata_line(line: str) -> bool:
    """Is metadata line.
    
    Purpose: This function belongs to the document extraction stage.
    Why: It makes the PDF input usable before reconciliation decisions are made.
    """
    raw = normalize_text(line)

    if re.match(r"^\s*\d{3,6}\s+", raw):
        return False

    return (
        is_business_id(raw)
        or is_company_name_line(raw)
        or is_page_marker(raw)
        or looks_like_company_metadata_line(raw)
    )


def looks_like_non_data_line(line: str) -> bool:
    """Looks like non data line.
    
    Purpose: This function belongs to the document extraction stage.
    Why: It makes the PDF input usable before reconciliation decisions are made.
    """
    norm = normalize_keyword_text(line)

    if not norm:
        return True

    if is_metadata_line(line):
        return True

    if len(norm) < 2:
        return True

    if any(p.fullmatch(norm) for p in PAGE_NOISE_PATTERNS):
        return True

    if norm in {"page", "€"}:
        return True

    if re.fullmatch(r"[\d\s\-–—]+", norm):
        return True

    if re.search(r"\b\d{1,2}\.\d{1,2}\.\d{4}\b", norm):
        return True

    signature_terms = [
        "allekirjoitus",
        "allekirjoitukset",
        "hallituksen puheenjohtaja",
        "hallituksen varajasen",
        "hallituksen varajäsen",
        "toimitusjohtaja",
        "tampereella",
        "espoossa",
        "lahdessa",
        "heinolassa",
        "lappeenrannassa",
    ]
    if any(term in norm for term in signature_terms):
        return True

    banned_terms = [
        "tilinpaatos",
        "tuloslaskelma",
        "tase",
        "liitetiedot",
        "toimintakertomus",
        "annual report",
        "balance sheet",
        "income statement",
        "resultatrakning",
        "balansrakning",
    ]
    if any(term == norm for term in banned_terms):
        return True

    return False


def looks_like_detail_account_row(label: str) -> bool:
  
    raw = normalize_text(label)
    norm = normalize_keyword_text(label)
    return bool(re.match(r"^\d{3,6}\s+\S+", raw) or re.match(r"^\d{3,6}\s+[a-z]", norm))


def is_total_row(label: str) -> bool:
    
    norm = normalize_label(label)
    total_terms = ["yhteensa", "summa", "total", "vastaavaa yhteensa", "vastattavaa yhteensa"]
    return any(term in norm for term in total_terms)


@lru_cache(maxsize=20000)
def forced_main_item_canonical(label: str, section_name: Optional[str] = None) -> Optional[str]:
    
    norm = normalize_label(label)
    compact = norm.replace(" ", "")
    section_key = normalize_section_name(section_name)
    allowed = FORCED_MAIN_ITEMS_BY_SECTION.get(section_key, FORCED_MAIN_ITEM_CANONICALS)
    for canonical in allowed:
        canonical_norm = normalize_label(canonical)
        candidates = [canonical_norm] + [normalize_label(s) for s in FORCED_MAIN_ITEM_SYNONYMS.get(canonical, [])]
        for cand in candidates:
            if not cand:
                continue
            # Previous substring logic was too risky; exact or compact equality is required here.
            if norm == cand or compact == cand.replace(" ", ""):
                return canonical
    return None

def is_forced_main_item(label: str, section_name: Optional[str] = None) -> bool:
  
    return forced_main_item_canonical(label, section_name) is not None

def is_balance_grand_total(label: str) -> bool:
   
    canonical = forced_main_item_canonical(label, None)
    return canonical in {"vastaavaa yhteensa", "vastattavaa yhteensa"}


def is_forced_root_item(label: str, section_name: Optional[str] = None) -> bool:
    
    canonical = forced_main_item_canonical(label, section_name)
    return canonical in FORCED_ROOT_ITEMS


def is_balance_total_root_item(label: str, section_name: Optional[str] = None) -> bool:
  
    canonical = forced_main_item_canonical(label, section_name)
    section_key = normalize_section_name(section_name)
    return (
        (section_key == "tase_vastaavaa" and canonical == "vastaavaa yhteensa")
        or (section_key == "tase_vastattavaa" and canonical == "vastattavaa yhteensa")
    )


def classify_total_level(label: str, section_name: Optional[str] = None) -> str:
    """Classify the structural level of a total row.

    The function separates balance sheet grand totals, major group totals,
    subgroup totals and generic totals so that totals are matched only against
    structurally comparable totals."""

    norm = normalize_label(label)
    compact = norm.replace(" ", "")
    section_key = normalize_section_name(section_name)
    canonical = forced_main_item_canonical(label, section_key) or canonicalize_common_labels(label)
    canonical_norm = normalize_label(canonical)

    def _matches_any(terms: set[str]) -> bool:
        """Matches any.
        
        Purpose: This function belongs to the reconciliation and candidate evaluation stage.
        Why: It supports controlled matching while keeping uncertain cases visible for manual review.
        """
        for term in terms:
            term_norm = normalize_label(term)
            if not term_norm:
                continue
            if norm == term_norm or compact == term_norm.replace(" ", ""):
                return True
        return False

    if section_key == "tase_vastaavaa" and (
        canonical_norm == "vastaavaa yhteensa"
        or norm in {"vastaavaa yhteensa", "tase vastaavaa yhteensa"}
    ):
        return "balance_grand_total"

    if section_key == "tase_vastattavaa" and (
        canonical_norm == "vastattavaa yhteensa"
        or norm in {"vastattavaa yhteensa", "tase vastattavaa yhteensa", "oma paaoma ja velat yhteensa"}
    ):
        return "balance_grand_total"

    if _matches_any(BALANCE_GRAND_TOTAL_ITEMS):
        if any(x in norm for x in ["vastaavaa", "vastattavaa", "total assets", "liabilities", "eget kapital och skulder"]):
            return "balance_grand_total"

    if _matches_any(MAJOR_GROUP_TOTAL_ITEMS):
        return "major_group_total"

    if _matches_any(SUBGROUP_TOTAL_ITEMS):
        return "subgroup_total"

    if is_named_total_label(label):
        return "subgroup_total"

    if is_generic_total_label(label):
        return "generic_total"

    return "not_total"

def forced_root_category_path(canonical_label: Optional[str], section_name: Optional[str]) -> str:
   
    section_key = normalize_section_name(section_name)
    section_display = {
        "tuloslaskelma": "Tuloslaskelma",
        "tase_vastaavaa": "Tase vastaavaa",
        "tase_vastattavaa": "Tase vastattavaa",
    }.get(section_key, str(section_name or ""))
    if canonical_label:
        return f"{section_display} > {canonical_label}" if section_display else canonical_label
    return section_display


def classify_row_type(label: str) -> str:
  
    label = strip_account_code_prefix(label)
    norm = normalize_label(label)
    forced = forced_main_item_canonical(label, None)
    if forced in {"vastaavaa yhteensa", "vastattavaa yhteensa", "oma paaoma yhteensa", "vieras paaoma yhteensa"}:
        return "grand_total"
    if forced in {"liikevoitto", "tilikauden voitto", "voitto tappio ennen tilinpaatossiirtoja ja veroja"}:
        return "result"
    if looks_like_detail_account_row(label):
        return "detail"
    if looks_like_non_data_line(label):
        return "noise"
    if any(term in norm for term in ["tilikauden voitto", "tilikauden tappio", "tilikauden tulos", "liikevoitto", "liiketulos"]):
        return "result"
    if any(term in norm for term in [
        "vastaavaa yhteensa", "vastattavaa yhteensa", "oma paaoma yhteensa", "vieras paaoma yhteensa",
        "pysyvat vastaavat yhteensa", "vaihtuvat vastaavat yhteensa",
    ]):
        return "grand_total"
    if "yhteensa" in norm:
        return "section_total"
    return "main"


def label_contains_total_term(label: str) -> bool:
    """Label contains total term.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    return "yhteensa" in normalize_label(label) or "total" in normalize_label(label)


def label_contains_result_term(label: str) -> bool:
    """Label contains result term.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    norm = normalize_label(label)
    return any(t in norm for t in ["tilikauden voitto", "tilikauden tappio", "tilikauden tulos", "liiketulos", "liikevoitto", "liiketappio"])


def _label_term_match_count(norm_label: str, candidates: list[str]) -> int:
    """Label term match count.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    return sum(1 for term in candidates if term in norm_label)


def is_plausible_statement_label(label: str, section_name: Optional[str] = None, row_type: Optional[str] = None) -> bool:
    """Is plausible statement label.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    norm = canonicalize_common_labels(label)

    if not norm or len(norm) < 2:
        return False

    if looks_like_non_data_line(label) or looks_like_detail_account_row(label):
        return False

    if len(norm.split()) == 1 and norm not in {
        "saamiset", "pitkaaikaiset", "lyhytaikaiset", "oma paaoma", "vieras paaoma",
        "liikevaihto", "henkilostokulut", "tuloverot", "liiketulos", "liikevoitto", "liiketappio"
    }:
        return False

    if re.search(r"\b(oy|oyj|ab|abp|ky|ay|tmi|ltd|llc)\b", norm) and not is_total_row(label):
        return False

    if row_type in {"result", "grand_total", "section_total"}:
        return True

    generic_hits = _label_term_match_count(norm, GENERIC_STATEMENT_LABEL_TERMS)

    if section_name == "tuloslaskelma":
        section_hits = _label_term_match_count(norm, INCOME_STATEMENT_LABEL_TERMS)
    elif section_name == "tase_vastaavaa":
        section_hits = _label_term_match_count(norm, BALANCE_ASSET_LABEL_TERMS)
    elif section_name == "tase_vastattavaa":
        section_hits = _label_term_match_count(norm, BALANCE_LIABILITY_LABEL_TERMS)
    else:
        section_hits = 0

    if section_hits > 0 or generic_hits > 0:
        return True

    if len(norm.split()) >= 2 and is_total_row(label):
        return True

    return False


# =========================================================
# PDF SECTION DETECTION AND STATEMENT BLOCK SELECTION
# =========================================================
# This section identifies the main financial statement sections from a text-based PDF.
#
# It detects:
# - Tables of contents and pages that should be skipped.
# - Income statement, balance sheet asset side and balance sheet liability side.
# - Section boundaries and terminal rows.
# - The most plausible section block when several candidates exist.
#
# The purpose is to extract only the actual financial statement blocks before
# monetary rows are parsed and reconciled.
#
# Candidate blocks are scored to reduce the risk of selecting notes, summaries,
# table of contents entries or detailed schedules instead of the statutory
# statement sections.


def looks_like_table_of_contents(lines: list[str]) -> bool:
  
    joined = " ".join(normalize_keyword_text(x) for x in lines[:30])
    return any(k in joined for k in TOC_KEYWORDS)


def detect_scanned_pdf(pdf_bytes: bytes, max_pages_to_check: int = 5) -> bool:
    
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        if not pdf.pages:
            return True
        lengths = []
        for page in pdf.pages[:max_pages_to_check]:
            lengths.append(len(normalize_text(page.extract_text() or "")))
    return sum(lengths) < 80 or (sum(lengths) / max(1, len(lengths))) < 20


def extract_document_logical_lines(pdf_bytes: bytes) -> tuple[list[dict], list[int], list[int]]:
    """Extract document logical lines.
    
    Purpose: This function belongs to the document extraction stage.
    Why: It makes the PDF input usable before reconciliation decisions are made.
    """
    rows: list[dict] = []
    skipped_toc_pages: list[int] = []
    skipped_report_pages: list[int] = []

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            lines = build_logical_lines_from_page(page)
            if not lines:
                continue

            if looks_like_table_of_contents(lines):
                skipped_toc_pages.append(page_no)
                continue

            header_norm = " ".join(normalize_keyword_text(x) for x in lines[:25])
            if any(term in header_norm for term in SKIP_SECTION_KEYWORDS):
                skipped_report_pages.append(page_no)

            for line_idx, line in enumerate(lines):
                rows.append({
                    "page": page_no,
                    "line_idx": line_idx,
                    "text": line,
                    "norm": normalize_keyword_text(line),
                })

    return rows, skipped_toc_pages, skipped_report_pages


def build_source_line_lookup(pdf_bytes: bytes) -> dict:
   
    lookup: dict[str, list[dict]] = {}
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page_no, page in enumerate(pdf.pages, start=1):
                for line_idx, line in enumerate(build_logical_lines_from_page(page)):
                    key = normalize_keyword_text(line)
                    if not key:
                        continue
                    lookup.setdefault(key, []).append({
                        "source_page": page_no,
                        "source_line_idx": line_idx,
                        "source_line_text": line,
                    })
    except Exception:
        return {}
    return lookup


def attach_source_locations(df: pd.DataFrame, source_lookup: dict) -> pd.DataFrame:
   
    df = df.copy()
    if df.empty or "source_line" not in df.columns:
        df["source_page"] = None
        df["source_line_idx"] = None
        return df

    used_counter: dict[str, int] = {}
    pages = []
    line_idxs = []
    for line in df["source_line"].fillna("").astype(str):
        key = normalize_keyword_text(line)
        matches = source_lookup.get(key, [])
        use_idx = used_counter.get(key, 0)
        if matches:
            chosen = matches[min(use_idx, len(matches) - 1)]
            pages.append(chosen.get("source_page"))
            line_idxs.append(chosen.get("source_line_idx"))
            used_counter[key] = use_idx + 1
        else:
            pages.append(None)
            line_idxs.append(None)
    df["source_page"] = pages
    df["source_line_idx"] = line_idxs
    return df


def classify_document_section_heading(norm_line: str) -> Optional[str]:
    """Classify document section heading.
    
    Purpose: This function belongs to the financial statement structure stage.
    Why: It prevents rows from being compared across incompatible statement sections or hierarchy levels.
    """
    if not norm_line:
        return None
    if "yhteensa" in norm_line or "total" in norm_line:
        return None
    if any(term == norm_line or term in norm_line for term in DOCUMENT_SECTION_HEADING_TERMS["tase_vastaavaa"]):
        if "vastattavaa" not in norm_line:
            return "tase_vastaavaa"
    if any(term == norm_line or term in norm_line for term in DOCUMENT_SECTION_HEADING_TERMS["tase_vastattavaa"]):
        return "tase_vastattavaa"
    if any(term == norm_line or term in norm_line for term in DOCUMENT_SECTION_HEADING_TERMS["tuloslaskelma"]):
        return "tuloslaskelma"
    return None


def is_document_stop_heading(norm_line: str) -> bool:
  
    if not norm_line:
        return False
    return any(term in norm_line for term in DOCUMENT_SECTION_STOP_TERMS)


def is_section_terminal_line(line: str, section_name: str) -> bool:
   
    norm = normalize_label(line)
    compact = norm.replace(" ", "")
    section_key = normalize_section_name(section_name)

    if section_key == "tase_vastaavaa":
        return (
            norm in {"vastaavaa yhteensa", "tase vastaavaa yhteensa"}
            or "vastaavaayhteensa" in compact
            or "tasevastaavaayhteensa" in compact
        )

    if section_key == "tase_vastattavaa":
        return (
            norm in {"vastattavaa yhteensa", "tase vastattavaa yhteensa", "oma paaoma ja velat yhteensa"}
            or "vastattavaayhteensa" in compact
            or "tasevastattavaayhteensa" in compact
            or "omapaaomajavelatyhteensa" in compact
        )

    if section_key == "tuloslaskelma":
        return any(term in norm for term in [
            "tilikauden voitto tappio",
            "tilikauden voitto",
            "tilikauden tappio",
            "tilikauden tulos",
            "arets resultat",
            "profit for the financial year",
            "loss for the financial year",
        ])

    return False


def trim_section_lines_to_terminal(lines: list[str], section_name: str) -> list[str]:
   
    if not lines:
        return lines
    trimmed = []
    for line in lines:
        trimmed.append(line)
        if is_section_terminal_line(line, section_name):
            break
    return trimmed


def score_section_block(lines: list[str], section_name: str) -> float:
    """Score section block.
    
    Purpose: This function belongs to the financial statement structure stage.
    Why: It prevents rows from being compared across incompatible statement sections or hierarchy levels.
    """
    if not lines:
        return -9999.0
    parsed, diag = parse_section_lines_to_df(lines, section_name)
    if parsed.empty:
        return -999.0

    score = 0.0
    score += len(parsed) * 2.0
    score += float(parsed.get("summary_row", pd.Series(False, index=parsed.index)).fillna(False).sum()) * 6.0
    if "row_type" in parsed.columns:
        score += float((parsed["row_type"] == "grand_total").sum()) * 10.0
        score += float((parsed["row_type"] == "section_total").sum()) * 4.0
        score -= float((parsed["row_type"] == "detail").sum()) * 0.5
    if "block_type" in parsed.columns:
        score += float((parsed["block_type"] == "main_statement").sum()) * 1.5
        score -= float((parsed["block_type"] == "statement_detail").sum()) * 0.25
    if {"current_value", "comparison_value"}.issubset(parsed.columns):
        score += float(parsed[["current_value", "comparison_value"]].notna().sum(axis=1).eq(2).sum()) * 3.0
    score -= diag.get("rejected_by_whitelist", 0) * 1.0
    score -= diag.get("noise_rows", 0) * 0.2
    return score


def extract_all_sections_row_level(pdf_bytes: bytes) -> dict:
    """Extract all sections row level.
    
    Purpose: This function belongs to the document extraction stage.
    Why: It makes the PDF input usable before reconciliation decisions are made.
    """
    rows, skipped_toc_pages, skipped_report_pages = extract_document_logical_lines(pdf_bytes)

    boundaries: list[dict] = []
    for pos, row in enumerate(rows):
        heading = classify_document_section_heading(row["norm"])
        if heading:
            boundaries.append({"kind": "section", "section": heading, "pos": pos, "page": row["page"]})
        elif is_document_stop_heading(row["norm"]):
            boundaries.append({"kind": "stop", "section": None, "pos": pos, "page": row["page"]})

    if not boundaries:
        raise ValueError("Not all main sections were detected in the document.")

    candidates = {"tuloslaskelma": [], "tase_vastaavaa": [], "tase_vastattavaa": []}

    for i, b in enumerate(boundaries):
        if b["kind"] != "section":
            continue
        section = b["section"]
        start_pos = b["pos"] + 1
        end_pos = len(rows)
        for next_b in boundaries[i + 1:]:
            if next_b["kind"] in {"section", "stop"}:
                end_pos = next_b["pos"]
                break
        block_rows = rows[start_pos:end_pos]
        block_lines = [r["text"] for r in block_rows]
        block_lines = trim_section_lines_to_terminal(block_lines, section)
        candidates[section].append({
            "lines": block_lines,
            "start_page": b["page"],
            "score": score_section_block(block_lines, section),
        })

    sections = {}
    start_pages = {}
    for section_name, blocks in candidates.items():
        usable = [b for b in blocks if b["lines"]]
        if not usable:
            sections[section_name] = []
            start_pages[section_name] = None
            continue
        usable.sort(key=lambda b: (b["score"], -b["start_page"]), reverse=True)
        best = usable[0]
        sections[section_name] = best["lines"]
        start_pages[section_name] = best["start_page"]

    if not sections["tuloslaskelma"] or not sections["tase_vastaavaa"] or not sections["tase_vastattavaa"]:
        raise ValueError("Not all main sections were detected in the document.")

    return {
        "tuloslaskelma": sections["tuloslaskelma"],
        "tase_vastaavaa": sections["tase_vastaavaa"],
        "tase_vastattavaa": sections["tase_vastattavaa"],
        "pages": {
            "tuloslaskelma": sorted({b["start_page"] for b in candidates["tuloslaskelma"] if b["lines"]}),
            "tase_vastaavaa": sorted({b["start_page"] for b in candidates["tase_vastaavaa"] if b["lines"]}),
            "tase_vastattavaa": sorted({b["start_page"] for b in candidates["tase_vastattavaa"] if b["lines"]}),
            "skipped_toc_pages": skipped_toc_pages,
            "skipped_report_pages": skipped_report_pages,
        },
        "start_pages": start_pages,
        "candidate_blocks": candidates,
    }


def find_sme_statement_pages_from_bytes(pdf_bytes: bytes):
    """Find sme statement pages from bytes.
    
    Purpose: This function belongs to the document extraction stage.
    Why: It makes the PDF input usable before reconciliation decisions are made.
    """
    result = {
        "tuloslaskelma": [],
        "tase_vastaavaa": [],
        "tase_vastattavaa": [],
        "skipped_toc_pages": [],
    }
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for i, page in enumerate(pdf.pages):
            text = normalize_text(page.extract_text() or "")
            if not text:
                continue
            lines = clean_lines(text)
            if not lines:
                continue
            if looks_like_table_of_contents(lines):
                result["skipped_toc_pages"].append(i + 1)
                continue
            header = " ".join(normalize_keyword_text(x) for x in lines[:25])
            if "tuloslaskelma" in header:
                result["tuloslaskelma"].append(i + 1)
            if "tase vastaavaa" in header or ("tase" in header and "vastaavaa" in header):
                result["tase_vastaavaa"].append(i + 1)
            if "tase vastattavaa" in header or ("tase" in header and "vastattavaa" in header):
                result["tase_vastattavaa"].append(i + 1)
    for k in ["tuloslaskelma", "tase_vastaavaa", "tase_vastattavaa"]:
        result[k] = list(dict.fromkeys(result[k]))
    return result


def extract_section_text_across_pages_from_bytes(pdf_bytes: bytes, start_page_1_based: int, stop_titles: list[str]) -> list[str]:
    """Extract section text across pages from bytes.
    
    Purpose: This function belongs to the document extraction stage.
    Why: It makes the PDF input usable before reconciliation decisions are made.
    """
    collected = []
    normalized_stops = [normalize_keyword_text(s) for s in stop_titles]
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for i in range(start_page_1_based - 1, len(pdf.pages)):
            text = normalize_text(pdf.pages[i].extract_text() or "")
            if not text:
                continue
            lines = clean_lines(text)
            if not lines:
                continue
            header = " ".join(normalize_keyword_text(x) for x in lines[:25])
            if i > start_page_1_based - 1 and any(stop in header for stop in normalized_stops):
                break
            collected.extend(lines)
    return collected


def extract_all_sections_from_bytes(pdf_bytes: bytes):
    """Extract all sections from bytes.
    
    Purpose: This function belongs to the document extraction stage.
    Why: It makes the PDF input usable before reconciliation decisions are made.
    """
    pages = find_sme_statement_pages_from_bytes(pdf_bytes)
    if not pages["tuloslaskelma"] or not pages["tase_vastaavaa"] or not pages["tase_vastattavaa"]:
        raise ValueError("Not all main sections were detected in the document.")
    s_income = pages["tuloslaskelma"][0]
    s_assets = pages["tase_vastaavaa"][0]
    s_liab = pages["tase_vastattavaa"][0]
    income_lines = extract_section_text_across_pages_from_bytes(pdf_bytes, s_income, ["tase vastaavaa", "tase vastattavaa"])
    assets_lines = extract_section_text_across_pages_from_bytes(pdf_bytes, s_assets, ["tase vastattavaa"])
    liab_lines = extract_section_text_across_pages_from_bytes(pdf_bytes, s_liab, POST_BALANCE_STOP_TITLES)
    return {
        "tuloslaskelma": income_lines,
        "tase_vastaavaa": assets_lines,
        "tase_vastattavaa": liab_lines,
        "pages": pages,
        "start_pages": {
            "tuloslaskelma": s_income,
            "tase_vastaavaa": s_assets,
            "tase_vastattavaa": s_liab,
        },
    }


# ---------------------------------------------------------
# PDF LAYOUT RECONSTRUCTION (WORD-TO-ROW GROUPING)
# ---------------------------------------------------------
# This function reconstructs logical text rows from PDF word-level coordinates.
#
# PDF extraction provides individual words with positional metadata, but not
# reliable line structures. This function groups words into rows based on their
# vertical alignment (y-coordinate proximity).
#
# The purpose is to rebuild human-readable rows before label parsing and
# numeric extraction are performed.
#
# This is a critical preprocessing step, because incorrect row grouping would
# lead to incorrect label-value associations in later stages.


def _group_words_into_rows(words: list[dict], y_tolerance: float = 3.0) -> list[list[dict]]:
    """Group PDF-extracted words into logical rows using vertical alignment.

    Each word contains positional coordinates. This function clusters words
    into rows based on their y-position, allowing for small vertical variation
    caused by PDF rendering inconsistencies.
    
      Parameters:
        words (list[dict]): List of word objects with positional metadata.
        y_tolerance (float): Maximum vertical distance for grouping words into the same row.

    Returns:
        list[list[dict]]: A list of rows, where each row is a list of word objects
        sorted from left to right.

    Why this matters:
    PDF files do not store text as structured rows. Reconstructing rows is
    necessary before extracting financial statement labels and values."""
    rows: list[list[dict]] = []
    sorted_words = sorted(words, key=lambda w: (round(w.get("top", 0), 1), w.get("x0", 0)))
    for word in sorted_words:
        top = float(word.get("top", 0))
        placed = False
        for row in rows:
            row_top = sum(float(w.get("top", 0)) for w in row) / len(row)
            if abs(top - row_top) <= y_tolerance:
                row.append(word)
                placed = True
                break
        if not placed:
            rows.append([word])
    for row in rows:
        row.sort(key=lambda w: w.get("x0", 0))
    return rows


# =========================================================
# X/Y COORDINATE-ASSISTED COLUMN EXTRACTION AND ROW SAFETY CHECKS
# =========================================================
# This section uses PDF word coordinates to support amount-column extraction.
#
# PDF text extraction may place labels, account numbers and monetary values on
# the same logical line without preserving table structure. These functions use
# x/y positions to identify the right-side amount columns and rebuild a cleaner
# label-value line.
#
# Coordinate data is used only as supporting evidence. It does not override the
# later semantic, numeric or reconciliation checks.
#
# The section also includes row-safety checks that prevent invalid merged labels
# and detailed account rows from being treated as main statement rows.

def _normalize_amount_token_text(text: str) -> str:
    """Normalize input values for reliable comparison."""
    text = normalize_text(text or "")
    text = text.replace("−", "-").replace("–", "-").replace("—", "-")
    text = text.replace("\u202f", " ").replace("\xa0", " ")
    return text.strip()


def _is_amount_candidate_word(text: str) -> bool:
   
    t = _normalize_amount_token_text(text)
    if not t:
        return False
    if t in {"-", "–", "—", "(", ")"}:
        return True
    return bool(re.search(r"\d", t))


def _looks_like_amount_group(parts: list[str]) -> bool:
  
    if not parts:
        return False
    joined = " ".join(_normalize_amount_token_text(p) for p in parts).strip()
    compact = joined.replace(" ", "")
    if compact in {"-", "–", "—"}:
        return True
    return bool(re.search(r"\d+[,.]\d{2}\)?-?$", compact))


def _clean_amount_group(parts: list[str]) -> str:
    """Convert an input value into a safe comparable representation."""
    cleaned = [_normalize_amount_token_text(p) for p in parts if _normalize_amount_token_text(p)]
    if not cleaned:
        return ""
    compact = "".join(cleaned).replace(" ", "")
    if compact in {"-", "–", "—"}:
        return "0,00"

    suffix_negative = cleaned[-1] in {"-", "–", "—"}
    prefix_negative = cleaned[0] in {"-", "–", "—"}
    cleaned = [p for p in cleaned if p not in {"-", "–", "—"}]
    if not cleaned:
        return "0,00"

    amount = " ".join(cleaned)
    amount = re.sub(r"\s+", " ", amount).strip()
    if suffix_negative or prefix_negative:
        amount = "-" + amount.lstrip("-")
    return amount



# ---------------------------------------------------------
# COORDINATE-AWARE COLUMN VALUE METADATA
# ---------------------------------------------------------
# These helpers keep the PDF column evidence attached to the logical row.
# The marker is internal only: user-facing source lines are cleaned before
# reporting. This allows the old text-based parser to remain in place while
# preventing a single comparative-period value from being treated as a
# current-period value.

COLUMN_VALUE_MARKER_PREFIX = "__FSRT_COLUMN_VALUES__"
COLUMN_VALUE_MARKER_RE = re.compile(r"\s+__FSRT_COLUMN_VALUES__\{([^}]*)\}\s*$")


def _escape_marker_value(value) -> str:
    """Escape marker values without introducing external dependencies."""
    text = "" if value is None else str(value)
    return text.replace("\\", "\\\\").replace("|", "\\p").replace("=", "\\e").replace("}", "\\c")


def _unescape_marker_value(value: str) -> str:
    """Unescape values stored in an internal coordinate marker."""
    text = "" if value is None else str(value)
    return text.replace("\\c", "}").replace("\\e", "=").replace("\\p", "|").replace("\\\\", "\\")


def attach_column_value_marker(line: str, metadata: dict) -> str:
    """Attach internal column metadata to a logical text line."""
    if not metadata:
        return line
    parts = []
    for key in (
        "current_value_raw_parsed",
        "comparison_value_raw_parsed",
        "column_parse_status",
        "column_parse_basis",
    ):
        value = metadata.get(key)
        if value is not None:
            parts.append(f"{key}={_escape_marker_value(value)}")
    if not parts:
        return line
    return f"{line} {COLUMN_VALUE_MARKER_PREFIX}{{{'|'.join(parts)}}}"


def extract_column_value_marker(line: str) -> tuple[str, dict]:
    """Separate an internal column marker from a logical line."""
    text = str(line or "")
    match = COLUMN_VALUE_MARKER_RE.search(text)
    if not match:
        return text, {}
    clean_line = text[:match.start()].rstrip()
    metadata = {}
    for part in match.group(1).split("|"):
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        metadata[key] = _unescape_marker_value(value)
    return clean_line, metadata


def strip_column_value_marker(line: str) -> str:
    """Remove internal coordinate metadata from a line before display."""
    clean_line, _metadata = extract_column_value_marker(line)
    return clean_line


def assign_value_groups_to_statement_columns(value_groups: list[dict], page_width: float) -> dict:
    """Assign coordinate-detected value groups to current and comparative columns.

    Finnish two-column financial statements usually show the current period in
    the left numeric column and the comparative period in the right numeric
    column. The function is intentionally conservative for single-value rows:
    a lone value on the far-right side is treated as a comparative-period value,
    while a lone value in the ambiguous middle-right area is not forced into the
    current period.
    """
    result = {
        "current_value_raw_parsed": None,
        "comparison_value_raw_parsed": None,
        "column_parse_status": "no_value_found",
        "column_parse_basis": "coordinate_groups",
    }
    groups = sorted([g for g in value_groups if g.get("text")], key=lambda g: float(g.get("x0", 0)))
    if not groups:
        return result

    if len(groups) >= 2:
        selected = groups[-2:]
        result["current_value_raw_parsed"] = selected[0].get("text")
        result["comparison_value_raw_parsed"] = selected[1].get("text")
        result["column_parse_status"] = "two_values_by_x_order"
        return result

    only = groups[0]
    x0 = float(only.get("x0", 0))
    text = only.get("text")
    far_right_limit = page_width * 0.72 if page_width else 0
    likely_current_limit = page_width * 0.58 if page_width else 0

    if page_width and x0 >= far_right_limit:
        result["comparison_value_raw_parsed"] = text
        result["column_parse_status"] = "single_value_in_comparative_column"
    elif page_width and x0 <= likely_current_limit:
        result["current_value_raw_parsed"] = text
        result["column_parse_status"] = "single_value_in_current_column"
    else:
        result["column_parse_status"] = "single_value_in_ambiguous_column"
    return result

def build_coordinate_aware_line_from_row(row: list[dict], page_width: float) -> str:
    """Rebuild a financial statement line using PDF word coordinates.

    The function identifies monetary value groups from the right side of the
    page and separates them from the label area. This helps recover rows where
    normal PDF text extraction does not preserve table columns.

    Coordinate logic is used conservatively: if the detected value groups are
    unclear, the original full line is returned instead of forcing a repair."""
    if not row:
        return ""

    row = sorted(row, key=lambda w: float(w.get("x0", 0)))
    full_line = normalize_text(" ".join(normalize_text(w.get("text", "")) for w in row if normalize_text(w.get("text", ""))))
    if not full_line:
        return ""

    # account numbers, page numbers and years getting mixed up in values.
    min_amount_x = page_width * 0.32
    groups: list[dict] = []
    current: list[dict] = []

    def flush_current():
      
        nonlocal current
        if not current:
            return
        parts = [w.get("text", "") for w in current]
        if _looks_like_amount_group(parts):
            groups.append({
                "x0": min(float(w.get("x0", 0)) for w in current),
                "x1": max(float(w.get("x1", 0)) for w in current),
                "text": _clean_amount_group(parts),
                "parts": parts,
            })
        current = []

    previous_x1 = None
    for w in row:
        text = normalize_text(w.get("text", ""))
        x0 = float(w.get("x0", 0))
        x1 = float(w.get("x1", 0))
        is_candidate = x0 >= min_amount_x and _is_amount_candidate_word(text)
        if not is_candidate:
            flush_current()
            previous_x1 = None
            continue

        gap = None if previous_x1 is None else x0 - previous_x1

        # Dash-aware column selector:
        current_texts = [_normalize_amount_token_text(x.get("text", "")) for x in current]
        current_is_single_dash = len(current_texts) == 1 and current_texts[0] in {"-", "–", "—"}
        next_is_numeric_amount = bool(re.search(r"\d+[,.]\d{2}", _normalize_amount_token_text(text)))
        dash_belongs_to_next_number = current_is_single_dash and next_is_numeric_amount and gap is not None and gap <= 75

        if current and gap is not None and gap > 20 and not dash_belongs_to_next_number:
            flush_current()
        current.append(w)
        previous_x1 = x1
    flush_current()

    if not groups:
        return full_line

    groups = sorted(groups, key=lambda g: g["x0"])

    # rightmost, because the leftmost digits are often account numbers or
    # other itemized information.
    value_groups = groups[-2:] if len(groups) >= 2 else groups[-1:]

    if len(value_groups) == 2 and abs(float(value_groups[1]["x0"]) - float(value_groups[0]["x0"])) < 18:
        return full_line

    first_value_x0 = min(g["x0"] for g in value_groups)

    label_tokens = []
    for w in row:
        x1 = float(w.get("x1", 0))
        txt = normalize_text(w.get("text", ""))
        if not txt:
            continue
        if x1 < first_value_x0 - 2:
            label_tokens.append(txt)

    label = normalize_text(" ".join(label_tokens)).strip(" -–,;:")
    if not label:
        return full_line

    value_text = " ".join(g["text"] for g in value_groups if g.get("text"))
    coordinate_line = normalize_text(f"{label} {value_text}")

    column_metadata = assign_value_groups_to_statement_columns(value_groups, page_width)
    if coordinate_line and column_metadata.get("column_parse_status") != "no_value_found":
        return attach_column_value_marker(coordinate_line, column_metadata)
    return coordinate_line or full_line


def looks_like_statement_label(label: str) -> bool:
   
    norm = normalize_label(label)
    if not norm:
        return False

    known = {normalize_label(x) for x in KNOWN_STATEMENT_LABELS}
    starters = tuple(normalize_keyword_text(x) for x in MAIN_ITEM_STARTERS)
    forced = {normalize_label(x) for x in FORCED_MAIN_ITEM_CANONICALS}

    return (
        norm in known
        or norm in forced
        or normalize_keyword_text(norm).startswith(starters)
    )


def line_starts_new_main_item(line: str) -> bool:
   
    return looks_like_statement_label(line)


def merged_label_looks_invalid(label: str) -> bool:
 
    norm = normalize_keyword_text(label)
    conflicting_pairs = [
        ("henkilostokulut", "liiketoiminnan muut kulut"),
        ("myyntisaamiset", "muut saamiset"),
        ("oma paaoma", "vieras paaoma"),
        ("vastaavaa", "vastattavaa"),
    ]
    return any(a in norm and b in norm for a, b in conflicting_pairs)


def has_account_code_prefix(text: Optional[str]) -> bool:
 
    return bool(re.match(r"^\s*\d{3,6}\s+", normalize_text(text or "")))


def detect_row_block_type(source_line: Optional[str], row_type: str, structure: str, label: str) -> str:
    """Classify whether a row belongs to the main financial statement or a detail schedule.

    The function uses account-code prefixes, row type and detected layout
    structure to prevent detailed account rows from being mixed with statutory
    statement rows."""

    norm_label = normalize_label(label or "")

    if has_account_code_prefix(source_line):
        return "statement_detail"

    if any(term in norm_label for term in ("yhteensa", "total assets", "total liabilities")) and structure != "single-column":
        return "main_statement"

    if row_type in {"grand_total", "section_total"} and structure != "single-column":
        return "main_statement"

    if structure != "single-column" and row_type in {"main", "section_total", "grand_total"}:
        return "main_statement"

    if structure == "single-column" and row_type in {"detail", "section_total", "grand_total"}:
        return "statement_detail"

    if row_type == "detail":
        return "statement_detail"

    return "main_statement" if structure != "single-column" else "statement_detail"


# =========================================================
# LOGICAL LINE RECONSTRUCTION AND BROKEN ROW REPAIR
# =========================================================
# This section reconstructs usable logical lines from PDF pages.
#
# It handles cases where PDF extraction splits one financial statement row
# across multiple physical lines, for example when:
# - an account-code row is separated from its amount,
# - an amount appears alone on the next line,
# - a label continues after a line break,
# - PDF word extraction fails and raw text extraction must be used instead.
#
# The purpose is to repair obvious layout fragmentation before section
# detection and monetary row parsing.
#
# Repairs are conservative: lines are merged only when the continuation pattern
# is structurally plausible and does not create conflicting financial labels.

def _has_amount_token(line: str) -> bool:
  
    if not line:
        return False
    return bool(re.search(r"-?\d+(?:[ .\u00a0\u202f]\d{3})*,\d{2}\b", normalize_text(line)))

def _is_amounts_only_line(line: str) -> bool:
    norm = normalize_text(line)
    if not norm:
        return False
    tmp = re.sub(r"-?\d+(?:[ .\u00a0\u202f]\d{3})*,\d{2}", "", norm)
    tmp = re.sub(r"[\s\-–—€]+", "", tmp)
    return tmp == ""

def _starts_with_account_code(line: str) -> bool:
  
    return bool(re.match(r"^\s*\d{3,6}\b", normalize_text(line)))

def _looks_like_label_continuation(line: str) -> bool:
    
    if not line:
        return False
    if line_starts_new_main_item(line):
        return False
    if _starts_with_account_code(line):
        return False
    if _has_amount_token(line):
        return False
    if looks_like_non_data_line(line):
        return False
    return len(normalize_keyword_text(line).split()) <= 8

def _merge_broken_account_detail_lines(lines: list[str]) -> list[str]:
    """Merge account-detail rows that were split across multiple PDF lines.

    Some PDFs separate the account label and monetary value onto different
    extracted lines. This function joins only structurally plausible fragments
    so that account-code rows can be parsed as complete label-value rows.
    """
    out: list[str] = []
    i = 0
    while i < len(lines):
        cur = lines[i]
        if _starts_with_account_code(cur) and not _has_amount_token(cur):
            parts = [cur]
            j = i + 1
            merged = False
            while j < len(lines) and len(parts) < 4:
                nxt = lines[j]
                if _is_amounts_only_line(nxt):
                    parts.append(nxt)
                    out.append(" ".join(parts))
                    i = j + 1
                    merged = True
                    break
                if _looks_like_label_continuation(nxt):
                    parts.append(nxt)
                    j += 1
                    continue
                break
            if merged:
                continue
        if (not _has_amount_token(cur)) and i + 1 < len(lines) and _is_amounts_only_line(lines[i + 1]):
            if not line_starts_new_main_item(cur) and not looks_like_non_data_line(cur):
                out.append(f"{cur} {lines[i + 1]}")
                i += 2
                continue
        out.append(cur)
        i += 1
    return out

def build_logical_lines_from_page(page) -> list[str]:
    """ Build cleaned logical text lines from a single PDF page.

    The function first tries coordinate-based word extraction. If that fails,
    it falls back to plain text extraction. It then merges safe continuation
    lines and repairs broken account-detail rows."""
    try:
        words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
    except Exception:
        text = normalize_text(page.extract_text() or "")
        return clean_lines(text)

    if not words:
        text = normalize_text(page.extract_text() or "")
        return clean_lines(text)

    rows = _group_words_into_rows(words)
    lines: list[str] = []
    page_width = float(getattr(page, "width", 0) or 0)
    for row in rows:
        line = build_coordinate_aware_line_from_row(row, page_width) if page_width else " ".join(
            normalize_text(w.get("text", "")) for w in row if normalize_text(w.get("text", ""))
        )
        line = normalize_text(line)
        norm = normalize_keyword_text(line)
        if not line or any(p.fullmatch(norm) for p in PAGE_NOISE_PATTERNS) or norm == "page":
            continue
        lines.append(line)

    combined: list[str] = []
    i = 0
    while i < len(lines):
        current = lines[i]
        if i == len(lines) - 1:
            combined.append(current)
            break
        next_line = lines[i + 1]
        current_norm = normalize_keyword_text(current)
        current_incomplete = current.endswith(("JA", "SEKÄ", "-", "–", "(", "/")) or current_norm.endswith(("ja", "seka"))
        if current_incomplete:
            if line_starts_new_main_item(next_line):
                combined.append(current)
                i += 1
                continue
            merged = f"{current} {next_line}".strip()
            if merged_label_looks_invalid(merged):
                combined.append(current)
                i += 1
                continue
            combined.append(merged)
            i += 2
            continue
        combined.append(current)
        i += 1
    return _merge_broken_account_detail_lines(combined)


def build_logical_lines_from_pdf_section(pdf_bytes: bytes, start_page_1_based: int, stop_titles: list[str]) -> list[str]:
    """Build logical lines from pdf section.
    
    Purpose: This function belongs to the document extraction stage.
    Why: It makes the PDF input usable before reconciliation decisions are made.
    """
    out: list[str] = []
    normalized_stops = [normalize_keyword_text(s) for s in stop_titles]
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for i in range(start_page_1_based - 1, len(pdf.pages)):
            page = pdf.pages[i]
            lines = build_logical_lines_from_page(page)
            if not lines:
                continue
            header = " ".join(normalize_keyword_text(x) for x in lines[:25])
            if i > start_page_1_based - 1 and any(stop in header for stop in normalized_stops):
                break
            out.extend(lines)
    return out


# =========================================================
# MONETARY ROW PARSING AND DOCUMENT RELIABILITY ASSESSMENT
# =========================================================
# This section parses financial statement lines into structured monetary rows
# and evaluates whether the extracted document is reliable enough for automated
# reconciliation.
#
# It handles:
# - Broken numeric tokens and uncertain decimal repairs.
# - Separation of labels from current-period and comparative-period values.
# - Row-level canonical recognition and hierarchy validation.
# - Value parse status and row parse quality.
# - Document-level reliability scoring.
#
# The purpose is to keep extraction transparent: rows are not only parsed into
# values, but also given quality indicators that show whether they are safe,
# uncertain or require manual review.

def repair_broken_number_tokens(tokens: list[str]) -> list[dict]:

    result = []
    i = 0
    while i < len(tokens):
        if i < len(tokens) - 1:
            a = tokens[i].strip()
            b = tokens[i + 1].strip()
            if re.fullmatch(r"-?\d{1,3}", a) and re.fullmatch(r"-?\d{4,}", b):
                sign = "-" if b.startswith("-") else ""
                b_digits = b.lstrip("-")
                repaired = f"{sign}{a} {b_digits[:-2]},{b_digits[-2:]}"
                original = f"{a} {b}"
                try:
                    parse_finnish_number(repaired)
                    result.append({
                        "original": original,
                        "parsed": repaired,
                        "repaired": True,
                        "repair_type": "uncertain_decimal_insert",
                    })
                    i += 2
                    continue
                except Exception:
                    pass
        result.append({
            "original": tokens[i],
            "parsed": tokens[i],
            "repaired": False,
            "repair_type": None,
        })
        i += 1
    return result


def split_label_and_numbers(line: str, section_name: Optional[str] = None):
    """Split label and numbers.
    
    Purpose: This function belongs to the number parsing and validation stage.
    Why: It reduces the risk that formatting differences or unreadable values are treated as reliable evidence.
    """
    original_line_with_metadata = line
    line, column_metadata = extract_column_value_marker(line)
    raw_match_objs = list(NUMBER_PATTERN.finditer(line))
    raw_matches = [m.group(0).strip() for m in raw_match_objs]

    if raw_match_objs:
        first_raw = raw_match_objs[0].group(0).strip()
        if raw_match_objs[0].start() <= 1 and re.fullmatch(r"\d{3,6}", first_raw) and len(raw_match_objs) >= 2:
            raw_matches_for_values = [m.group(0).strip() for m in raw_match_objs[1:]]
            first_value_index = raw_match_objs[1].start()
        else:
            raw_matches_for_values = raw_matches
            first_anchor = raw_matches_for_values[0].split()[0] if raw_matches_for_values else ""
            first_value_index = line.find(first_anchor)
    else:
        raw_matches_for_values = []
        first_value_index = -1

    token_infos = repair_broken_number_tokens(raw_matches_for_values)
    if len(token_infos) == 0:
        return None
    if first_value_index < 0:
        first_anchor = token_infos[0]["original"].split()[0]
        first_value_index = line.find(first_anchor)
    if first_value_index < 0:
        return None
    original_label = line[:first_value_index].strip(" -–,;:")
    label = original_label
    if not label or len(normalize_label(label)) < 2 or looks_like_non_data_line(label):
        return None

    if looks_like_non_data_line(label):
        return None

    norm_label = normalize_label(label)
    company_form_patterns = [
        r"\boy\b",
        r"\boyj\b",
        r"\bab\b",
        r"\babp\b",
        r"\bky\b",
        r"\bay\b",
        r"\btmi\b",
        r"\bltd\b",
        r"\bllc\b",
    ]
    if any(re.search(pattern, norm_label) for pattern in company_form_patterns):
        if len(norm_label.split()) <= 6 and not is_total_row(label):
            return None

    # Banned document-title terms are checked as whole normalized terms, not as substrings.
    # This prevents valid standard rows such as "tulos ennen tilinpaatossiirtoja ja veroja"
    # from being rejected only because the word stem "tilinpaatos" appears inside
    # "tilinpaatossiirtoja". Forced standard items are always allowed to continue
    # to numeric parsing, but their values still need to pass the normal reconciliation checks.
    banned_label_hit = any(
        re.search(rf"(?<![a-z0-9]){re.escape(term)}(?![a-z0-9])", norm_label)
        for term in BANNED_LABEL_TERMS
    )
    if banned_label_hit and not is_total_row(label) and not forced_main_item_canonical(label, section_name):
        return None

    preliminary_row_type = classify_row_type(label)
    plausibility = evaluate_label_plausibility(label, section_name=section_name, row_type=preliminary_row_type)

    # Extraction-first principle:
    extraction_status = "taxonomy_recognized" if plausibility.get("accepted") else "extracted_unmapped"

    parsed_numbers = []
    for item in token_infos:
        try:
            parsed_numbers.append({
                "original": item["original"],
                "parsed": item["parsed"],
                "value": parse_finnish_number(item["parsed"]),
                "repaired": item["repaired"],
                "repair_type": item.get("repair_type"),
            })
        except Exception:
            continue
    if len(parsed_numbers) == 0:
        return None

    # Coordinate-aware column override: when the PDF extractor preserved x-coordinate
    # evidence, use it instead of assuming that a single number belongs to the
    # current period. This prevents comparative-period values from being used as
    # older current-period values.
    if column_metadata:
        current_raw = column_metadata.get("current_value_raw_parsed") or None
        comparison_raw = column_metadata.get("comparison_value_raw_parsed") or None
        current_item = None
        comparison_item = None

        if current_raw:
            try:
                current_item = {
                    "original": current_raw,
                    "parsed": current_raw,
                    "value": parse_finnish_number(current_raw),
                    "repaired": False,
                    "repair_type": None,
                }
            except Exception:
                current_item = None
        if comparison_raw:
            try:
                comparison_item = {
                    "original": comparison_raw,
                    "parsed": comparison_raw,
                    "value": parse_finnish_number(comparison_raw),
                    "repaired": False,
                    "repair_type": None,
                }
            except Exception:
                comparison_item = None

        if current_item is not None or comparison_item is not None:
            return {
                "label": label,
                "current_value_raw_original": current_item["original"] if current_item else None,
                "comparison_value_raw_original": comparison_item["original"] if comparison_item else None,
                "current_value_raw_parsed": current_item["parsed"] if current_item else None,
                "comparison_value_raw_parsed": comparison_item["parsed"] if comparison_item else None,
                "current_value_was_repaired": False,
                "comparison_value_was_repaired": False,
                "current_value_repair_type": None,
                "comparison_value_repair_type": None,
                "all_numbers_raw": " | ".join(x["original"] for x in parsed_numbers),
                "structure": "2-columns-coordinate" if current_item and comparison_item else "single-value-coordinate",
                "source_line": strip_column_value_marker(original_line_with_metadata),
                "label_match_score": plausibility.get("score"),
                "label_match_canonical": plausibility.get("canonical"),
                "label_match_rule": plausibility.get("match_type"),
                "label_match_accepted": bool(plausibility.get("accepted")),
                "label_match_reason": plausibility.get("reason"),
                "extraction_status": extraction_status,
                "column_parse_status": column_metadata.get("column_parse_status"),
                "column_parse_basis": column_metadata.get("column_parse_basis"),
            }

        if column_metadata.get("column_parse_status") == "single_value_in_ambiguous_column":
            return None

    meaningful = []
    for item in parsed_numbers:
        raw = item["parsed"]
        val = item["value"]
        raw_clean = raw.replace(" ", "").replace(".", "").replace(",", "")
        if re.fullmatch(r"(19|20)\d{2}", raw_clean):
            continue
        if "," in raw or abs(val) >= 10:
            meaningful.append(item)

    if len(meaningful) >= 2:
        current_item = meaningful[-2]
        comparison_item = meaningful[-1]
        return {
            "label": label,
            "current_value_raw_original": current_item["original"],
            "comparison_value_raw_original": comparison_item["original"],
            "current_value_raw_parsed": current_item["parsed"],
            "comparison_value_raw_parsed": comparison_item["parsed"],
            "current_value_was_repaired": current_item["repaired"],
            "comparison_value_was_repaired": comparison_item["repaired"],
            "current_value_repair_type": current_item.get("repair_type"),
            "comparison_value_repair_type": comparison_item.get("repair_type"),
            "all_numbers_raw": " | ".join(x["original"] for x in parsed_numbers),
            "structure": "2-columns-filtered",
            "source_line": line,
            "label_match_score": plausibility.get("score"),
            "label_match_canonical": plausibility.get("canonical"),
            "label_match_rule": plausibility.get("match_type"),
            "label_match_accepted": bool(plausibility.get("accepted")),
            "label_match_reason": plausibility.get("reason"),
            "extraction_status": extraction_status,
            "column_parse_status": column_metadata.get("column_parse_status", "text_order_fallback"),
            "column_parse_basis": column_metadata.get("column_parse_basis", "text_order"),
        }

    single_item = meaningful[-1] if meaningful else parsed_numbers[-1]
    return {
        "label": label,
        "current_value_raw_original": single_item["original"],
        "comparison_value_raw_original": None,
        "current_value_raw_parsed": single_item["parsed"],
        "comparison_value_raw_parsed": None,
        "current_value_was_repaired": single_item["repaired"],
        "comparison_value_was_repaired": False,
        "current_value_repair_type": single_item.get("repair_type"),
        "comparison_value_repair_type": None,
        "all_numbers_raw": " | ".join(x["original"] for x in parsed_numbers),
        "structure": "single-column",
        "source_line": line,
        "label_match_score": plausibility.get("score"),
        "label_match_canonical": plausibility.get("canonical"),
        "label_match_rule": plausibility.get("match_type"),
        "label_match_accepted": bool(plausibility.get("accepted")),
        "label_match_reason": plausibility.get("reason"),
        "extraction_status": extraction_status,
        "column_parse_status": column_metadata.get("column_parse_status", "single_value_text_order_fallback"),
        "column_parse_basis": column_metadata.get("column_parse_basis", "text_order"),
    }


def _heading_candidate_from_line(line: str, section_key: Optional[str]) -> Optional[str]:
    
    if not line or NUMBER_PATTERN.search(line):
        return None
    if looks_like_non_data_line(line):
        return None
    canonical = canonicalize_common_labels(line)
    if is_legal_parent_label(canonical, section_key) or canonical in TOP_LEVEL_HIERARCHY_PARENTS.get(normalize_section_name(section_key), set()):
        return canonical
    return None


def classify_value_parse_status(raw_value, was_repaired: bool = False, repair_type: Optional[str] = None) -> str:
    
    if raw_value is None:
        return "missing"
    try:
        if pd.isna(raw_value):
            return "missing"
    except Exception:
        pass
    raw_text = str(raw_value).strip()
    if not raw_text:
        return "missing"
    if repair_type == "uncertain_decimal_insert":
        return "uncertain"
    if was_repaired:
        return "repaired"
    digits_only = re.sub(r"\D", "", raw_text)
    if len(digits_only) >= 5 and not re.search(r"[,\.]\d{2}\b", raw_text) and not re.search(r"\s", raw_text):
        return "uncertain"
    return "found"


def value_status_is_safe_for_verified_match(status: str) -> bool:
    """Value status is safe for verified match.
    
    Purpose: This function belongs to the number parsing and validation stage.
    Why: It reduces the risk that formatting differences or unreadable values are treated as reliable evidence.
    """
    return str(status or "").lower() == "found"


def parse_section_lines_to_df(section_lines: list[str], section_name: Optional[str] = None):
    """Parse section lines to df.
    
    Purpose: This function belongs to the document extraction stage.
    Why: It makes the PDF input usable before reconciliation decisions are made.
    """
    records = []
    section_key = normalize_section_name(section_name)
    diag = {
        "all_lines": len(section_lines),
        "parsed_rows": 0,
        "detail_rows": 0,
        "noise_rows": 0,
        "heading_rows": 0,
        "rejected_by_whitelist": 0,
        "hierarchy_conflicts": 0,
        "legal_basis_unknown_rows": 0,
    }
    category_path: list[str] = []

    for row in section_lines:
        row_display = strip_column_value_marker(row)
        row_norm = normalize_keyword_text(row_display)
        if any(p.fullmatch(row_norm) for p in PAGE_NOISE_PATTERNS) or row_norm == "page":
            diag["noise_rows"] += 1
            continue
        if looks_like_non_data_line(row_display):
            diag["noise_rows"] += 1
            continue

        heading_canonical = _heading_candidate_from_line(row_display, section_key)
        if heading_canonical:
            category_path = update_category_path(category_path, heading_canonical, section_key, row_type="heading")
            diag["heading_rows"] += 1
            continue

        result = split_label_and_numbers(row, section_name=section_key)
        if not result:
            continue

        raw_label = result["label"]
        row_type = classify_row_type(raw_label)
        if row_type == "noise":
            diag["noise_rows"] += 1
            continue
        if row_type == "detail":
            diag["detail_rows"] += 1

        plausibility = evaluate_label_plausibility(raw_label, section_name=section_key, row_type=row_type)
        if not plausibility["accepted"]:
            # Keep the row visible in diagnostics instead of silently accepting it.
            diag["rejected_by_whitelist"] += 1

        forced_canonical = forced_main_item_canonical(raw_label, section_key)
        canonical_label = (
            forced_canonical
            or (result.get("label_match_canonical") if result.get("label_match_accepted") else None)
            or (plausibility.get("canonical") if plausibility.get("accepted") else None)
            or normalize_label(raw_label)
        )
        forced_main_item = bool(forced_canonical)
        if forced_main_item:
            if forced_canonical in {"vastaavaa yhteensa", "vastattavaa yhteensa", "oma paaoma yhteensa", "vieras paaoma yhteensa"}:
                row_type = "grand_total"
            elif forced_canonical in {"liikevoitto", "tilikauden voitto", "voitto tappio ennen tilinpaatossiirtoja ja veroja"}:
                row_type = "result"
        forced_root_item = bool(forced_main_item and canonical_label in FORCED_ROOT_ITEMS)
        balance_total_root_item = bool(forced_root_item and canonical_label in BALANCE_GRAND_TOTAL_ROOT_ITEMS)
        total_level = classify_total_level(raw_label, section_key)

        legal_basis = classify_legal_basis(canonical_label, section_key, row_type=row_type)
        if forced_root_item:
            legal_basis.update({
                "legal_basis_match": True,
                "legal_basis_level": "statutory_grand_total" if total_level == "balance_grand_total" else ("statutory_major_group_total" if total_level == "major_group_total" else "statutory_root_item"),
                "legal_basis_parent": None,
            })
        if legal_basis.get("legal_basis_level") == "unknown_item":
            diag["legal_basis_unknown_rows"] += 1

        if forced_root_item:
            hierarchy = {
                "score": 1.0,
                "status": "Forced main-level item",
                "parent": None,
                "top_parent": None,
                "category_path": forced_root_category_path(canonical_label, section_key),
                "expected_parent": None,
            }
        else:
            hierarchy = evaluate_hierarchy_position(canonical_label, category_path, section_key, row_type=row_type)
        if hierarchy["status"] == "Ristiriitainen":
            diag["hierarchy_conflicts"] += 1

        try:
            current_val = None
            if result.get("current_value_raw_parsed") is not None:
                current_val = parse_finnish_number(result.get("current_value_raw_parsed"))
            comparison_val = None
            if result.get("comparison_value_raw_parsed") is not None:
                comparison_val = parse_finnish_number(result.get("comparison_value_raw_parsed"))
            if current_val is None and comparison_val is None:
                continue
            block_type = detect_row_block_type(result.get("source_line"), row_type, result.get("structure", ""), raw_label)

            immediate_parent = hierarchy.get("parent")
            top_parent = hierarchy.get("top_parent")
            full_category_path = [p for p in category_path if p]
            if forced_root_item:
                immediate_parent = None
                top_parent = None
                full_category_path_for_row = [forced_root_category_path(canonical_label, section_key)]
            elif forced_main_item:
                immediate_parent = None
                top_parent = None
                full_category_path_for_row = [canonical_label]
            elif canonical_label and (not full_category_path or full_category_path[-1] != canonical_label):
                full_category_path_for_row = full_category_path + [canonical_label]
            else:
                full_category_path_for_row = full_category_path

            canonical_context = evaluate_canonical_context(canonical_label, section_key, full_category_path_for_row, row_type=row_type)

            records.append({
                "label": clean_display_label(raw_label),
                "normalized_label": normalize_label(raw_label),
                "summary_row": is_total_row(raw_label),
                "row_type": row_type,
                "current_value_raw_original": result["current_value_raw_original"],
                "comparison_value_raw_original": result["comparison_value_raw_original"],
                "current_value_raw_parsed": result["current_value_raw_parsed"],
                "comparison_value_raw_parsed": result["comparison_value_raw_parsed"],
                "current_value_was_repaired": result["current_value_was_repaired"],
                "comparison_value_was_repaired": result["comparison_value_was_repaired"],
                "current_value_repair_type": result.get("current_value_repair_type"),
                "comparison_value_repair_type": result.get("comparison_value_repair_type"),
                "current_value_status": classify_value_parse_status(
                    result.get("current_value_raw_original"),
                    bool(result.get("current_value_was_repaired")),
                    result.get("current_value_repair_type"),
                ),
                "comparison_value_status": classify_value_parse_status(
                    result.get("comparison_value_raw_original"),
                    bool(result.get("comparison_value_was_repaired")),
                    result.get("comparison_value_repair_type"),
                ),
                "row_parse_quality": "high" if classify_value_parse_status(
                    result.get("current_value_raw_original"),
                    bool(result.get("current_value_was_repaired")),
                    result.get("current_value_repair_type"),
                ) == "found" and classify_value_parse_status(
                    result.get("comparison_value_raw_original"),
                    bool(result.get("comparison_value_was_repaired")),
                    result.get("comparison_value_repair_type"),
                ) in {"found", "missing"} else "medium",
                "all_numbers_raw": result["all_numbers_raw"],
                "structure": result["structure"],
                "current_value": current_val,
                "comparison_value": comparison_val,
                "source_line": strip_column_value_marker(result["source_line"]),
                "label_match_score": result.get("label_match_score", plausibility.get("score")),
                "label_match_canonical": canonical_label,
                "label_match_rule": result.get("label_match_rule", plausibility.get("match_type")),
                "label_match_accepted": bool(result.get("label_match_accepted", plausibility.get("accepted"))),
                "label_match_reason": result.get("label_match_reason", plausibility.get("reason")),
                "extraction_status": result.get("extraction_status", "taxonomy_recognized" if plausibility.get("accepted") else "extracted_unmapped"),
                "column_parse_status": result.get("column_parse_status", "text_order_fallback"),
                "column_parse_basis": result.get("column_parse_basis", "text_order"),
                "parent_label": immediate_parent,
                "top_parent_label": top_parent,
                "category_path": " > ".join(full_category_path_for_row),
                "total_context_key": total_context_key_from_parts(section_key, clean_display_label(raw_label), " > ".join(full_category_path_for_row), immediate_parent, top_parent),
                "expected_parent_label": hierarchy.get("expected_parent"),
                "hierarchy_score": hierarchy.get("score", 0.0),
                "hierarchy_status": hierarchy.get("status"),
                "legal_basis_match": bool(legal_basis.get("legal_basis_match")),
                "legal_basis_level": legal_basis.get("legal_basis_level"),
                "legal_basis_parent": legal_basis.get("legal_basis_parent"),
                "canonical_expected_section": canonical_context.get("expected_section"),
                "canonical_allowed_sections": canonical_context.get("allowed_sections"),
                "canonical_expected_parent": canonical_context.get("expected_parent"),
                "canonical_expected_side": canonical_context.get("expected_side"),
                "canonical_item_type": canonical_context.get("item_type"),
                "canonical_context_status": canonical_context.get("context_status"),
                "canonical_context_reason": canonical_context.get("context_reason"),
                "canonical_context_score": canonical_context.get("context_score"),
                "forced_main_item": forced_main_item,
                "forced_main_item_canonical": forced_canonical,
                "forced_root_item": forced_root_item,
                "balance_total_root_item": balance_total_root_item,
                "total_level": total_level,
                "hierarchy_level": 0 if forced_root_item else (len(full_category_path_for_row) - 1 if full_category_path_for_row else None),
                "match_priority": 110 if total_level == "balance_grand_total" else (100 if forced_root_item else (90 if total_level == "major_group_total" else (75 if total_level == "subgroup_total" else (80 if forced_main_item else 50)))),
                "block_type": block_type,
            })
        except Exception:
            continue

        if forced_root_item:
            category_path = []
        else:
            category_path = update_category_path(category_path, canonical_label, section_key, row_type=row_type)

    diag["parsed_rows"] = len(records)
    df = pd.DataFrame(records)
    return df, diag


def add_section_name(df: pd.DataFrame, section_name: str) -> pd.DataFrame:
    """Add section name.
    
    Purpose: This function belongs to the financial statement structure stage.
    Why: It prevents rows from being compared across incompatible statement sections or hierarchy levels.
    """
    df = df.copy()
    df["section"] = section_name
    return df


def _safe_ratio(part: int, whole: int) -> float:
    """Safe ratio.
    
    Purpose: This function belongs to the risk and quality-control stage.
    Why: It makes uncertainty explicit instead of silently accepting weak evidence.
    """
    if whole <= 0:
        return 0.0
    return part / whole


def classify_document_usability(score: int) -> str:
   
    if score >= 85:
        return "Usable"
    if score >= 60:
        return "Vaatii tarkistusta"
    return "Not recommended without manual review"


def find_missing_key_items(df: pd.DataFrame) -> list[str]:
    
    if df is None or df.empty:
        return [f"{sec}:{item}" for sec, items in KEY_ITEMS_BY_SECTION.items() for item in items]
    missing = []
    for section, items in KEY_ITEMS_BY_SECTION.items():
        sec_df = df[df.get("section", pd.Series(dtype=str)).astype(str).apply(normalize_section_for_matching) == section].copy()
        labels = set()
        if not sec_df.empty:
            if "label_match_canonical" in sec_df.columns:
                labels.update(sec_df["label_match_canonical"].fillna("").astype(str).map(canonicalize_common_labels))
            if "label" in sec_df.columns:
                labels.update(sec_df["label"].fillna("").astype(str).map(canonicalize_common_labels))
        for item in items:
            item_can = canonicalize_common_labels(item)
            # Profit/loss label variants are alternatives, so not every wording variant is mandatory.
            if item_can in {"liikevoitto", "liiketulos"}:
                if any(x in labels for x in {"liikevoitto", "liiketulos", "liiketappio"}):
                    continue
            if item_can in {"tilikauden voitto", "tilikauden tappio", "tilikauden tulos"}:
                if any(x in labels for x in {"tilikauden voitto", "tilikauden tappio", "tilikauden tulos", "tilikauden voitto tappio"}):
                    continue
            if item_can not in labels:
                missing.append(f"{section}:{item_can}")
    return missing

def build_document_reliability(parsed_result: dict) -> dict:
   
    df = parsed_result.get("df_all", pd.DataFrame()).copy()
    diagnostics = parsed_result.get("diagnostics", {}) or {}
    pages = parsed_result.get("pages", {}) or {}
    start_pages = parsed_result.get("start_pages", {}) or {}
    balance_info = parsed_result.get("balance_validation", {}) or {}

    row_count = len(df)

    repaired_rows = 0
    if not df.empty and {"current_value_was_repaired", "comparison_value_was_repaired"}.issubset(df.columns):
        repaired_rows = int((df["current_value_was_repaired"].fillna(False) | df["comparison_value_was_repaired"].fillna(False)).sum())

    fallback_rows = 0
    if not df.empty and "structure" in df.columns:
        fallback_rows = int((df["structure"].fillna("") == "fallback").sum())

    summary_rows = 0
    if not df.empty and "summary_row" in df.columns:
        summary_rows = int(df["summary_row"].fillna(False).sum())

    missing_summary_rows = int(summary_rows == 0)

    unbalanced_current = int(balance_info.get("current_match") is False)
    comparison_status = balance_info.get("comparison_status")
    unbalanced_comparison = int(balance_info.get("comparison_match") is False and comparison_status != "Ei vertailusarjaa")

    missing_sections = 0
    if len(parsed_result.get("df_income", pd.DataFrame())) == 0:
        missing_sections += 1
    if len(parsed_result.get("df_vastaavaa", pd.DataFrame())) == 0:
        missing_sections += 1
    if len(parsed_result.get("df_vastattavaa", pd.DataFrame())) == 0:
        missing_sections += 1

    uncertain_section_boundaries = 0
    if start_pages.get("tuloslaskelma") is None:
        uncertain_section_boundaries += 1
    if start_pages.get("tase_vastaavaa") is None:
        uncertain_section_boundaries += 1
    if start_pages.get("tase_vastattavaa") is None:
        uncertain_section_boundaries += 1

    skipped_toc_pages = len(pages.get("skipped_toc_pages", []))
    skipped_report_pages = len(pages.get("skipped_report_pages", []))

    hierarchy_conflicts = 0
    if not df.empty and "hierarchy_status" in df.columns:
        hierarchy_conflicts = int((df["hierarchy_status"].fillna("") == "Ristiriitainen").sum())
    elif isinstance(diagnostics, dict):
        hierarchy_conflicts = int(diagnostics.get("hierarchy_conflicts", 0) or 0)

    missing_key_items = find_missing_key_items(df)

    score = 100
    reasons = []

    repaired_ratio = _safe_ratio(repaired_rows, row_count)
    fallback_ratio = _safe_ratio(fallback_rows, row_count)

    if repaired_rows > 0:
        penalty = min(20, max(4, round(repaired_ratio * 100)))
        score -= penalty
        reasons.append(f"Repaired numeric values were detected on {repaired_rows} rows")

    if fallback_rows > 0:
        penalty = min(25, max(4, round(fallback_ratio * 120)))
        score -= penalty
        reasons.append(f"Fallback interpretation was used on {fallback_rows} rows")

    if missing_sections > 0:
        penalty = missing_sections * 12
        score -= penalty
        reasons.append(f"Mandatory main sections were missing or not detected ({missing_sections})")

    if uncertain_section_boundaries > 0:
        penalty = uncertain_section_boundaries * 6
        score -= penalty
        reasons.append(f"Section boundaries remained uncertain in {uncertain_section_boundaries} locations")

    if missing_summary_rows > 0:
        score -= 8
        reasons.append("No total rows were detected")

    if hierarchy_conflicts > 0:
        score -= min(8, hierarchy_conflicts * 2)
        reasons.append(f"Rakenteellisia hierarkiakonflikteja havaittiin {hierarchy_conflicts}")

    if missing_key_items:
        score -= min(15, len(missing_key_items) * 3)
        reasons.append(f"Key items were not detected ({len(missing_key_items)})")

    if unbalanced_current:
        score -= 18
        reasons.append("The balance sheet does not balance for the current period")

    if unbalanced_comparison:
        score -= 18
        reasons.append("The balance sheet does not balance for the comparative period")

    if skipped_toc_pages > 0:
        reasons.append(f"Table-of-contents pages were skipped ({skipped_toc_pages})")

    if skipped_report_pages > 0:
        reasons.append(f"Toimintakertomus- tai raporttisivuja ohitettiin {skipped_report_pages}")

    if isinstance(diagnostics, dict) and diagnostics.get("document_quality") == "Erittelypainotteinen":
        score -= 4
        reasons.append("The document contains many detailed schedules, which may make main total interpretation more difficult")

    if row_count == 0:
        score = 0
        reasons.append("No rows could be extracted from the document")

    score = max(0, min(100, score))
    classification = classify_document_usability(score)

    if not reasons:
        reasons.append("No material reliability issues were detected")

    return {
        "score": score,
        "classification": classification,
        "reasons": reasons,
        "repaired_rows": repaired_rows,
        "fallback_rows": fallback_rows,
        "summary_rows": summary_rows,
        "missing_sections": missing_sections,
        "uncertain_section_boundaries": uncertain_section_boundaries,
        "row_count": row_count,
        "skipped_toc_pages": skipped_toc_pages,
        "skipped_report_pages": skipped_report_pages,
        "unbalanced_current": bool(unbalanced_current),
        "unbalanced_comparison": bool(unbalanced_comparison),
        "hierarchy_conflicts": hierarchy_conflicts,
        "missing_key_items": missing_key_items,
    }


# =========================================================
# PERIOD DETECTION AND RECONCILIATION VALUE SELECTION
# =========================================================
# This section detects the reporting periods shown in financial statement
# columns and selects the value that should be used in reconciliation.
#
# The tool compares:
# - the older statement's current-period value
# - against the newer statement's comparative-period value
#
# Period keys are detected from full date ranges, balance dates or year-only
# fallbacks. The selected value source is stored explicitly so that the
# reconciliation result remains traceable.
#
# A key control is that the newer file never falls back to the current-period
# value when the comparative-period value is missing. This prevents silent
# false-positive reconciliations.


def _unique_preserve_order(items: list[str]) -> list[str]:
  
    out: list[str] = []
    for item in items:
        item = str(item).strip()
        if item and item not in out:
            out.append(item)
    return out


def normalize_period_key(text: str) -> str:
    """Normalize input values for reliable comparison."""
    text = normalize_text(str(text or ""))
    text = text.replace("–", "-").replace("—", "-")
    text = re.sub(r"\s+", "", text)
    return text.strip()


def _period_end_date(period_key: Optional[str]) -> Optional[str]:
    
    key = normalize_period_key(period_key or "")
    if not key:
        return None
    if "-" in key:
        return key.split("-")[-1]
    if re.fullmatch(r"\d{1,2}\.\d{1,2}\.(?:19|20)\d{2}", key):
        return key
    return None


def _period_year(period_key: Optional[str]) -> Optional[str]:
   
    end_date = _period_end_date(period_key)
    if end_date:
        m = re.search(r"((?:19|20)\d{2})$", end_date)
        if m:
            return m.group(1)
    key = normalize_period_key(period_key or "")
    m = re.search(r"\b((?:19|20)\d{2})\b", key)
    return m.group(1) if m else None


def extract_column_period_keys(section_lines: list[str]) -> dict:
    
    header = " ".join(str(x) for x in section_lines[:12])
    header_norm = normalize_period_key(header)

    ranges = re.findall(
        r"\d{1,2}\.\d{1,2}\.(?:19|20)\d{2}-\d{1,2}\.\d{1,2}\.(?:19|20)\d{2}",
        header_norm,
    )
    ranges = _unique_preserve_order([normalize_period_key(x) for x in ranges])
    if len(ranges) >= 2:
        return {
            "current_period_key": ranges[0],
            "comparison_period_key": ranges[1],
            "period_detection_basis": "full_period_range",
        }

    dates = re.findall(r"\d{1,2}\.\d{1,2}\.(?:19|20)\d{2}", header_norm)
    dates = _unique_preserve_order([normalize_period_key(x) for x in dates])
    if len(dates) >= 2:
        return {
            "current_period_key": dates[0],
            "comparison_period_key": dates[1],
            "period_detection_basis": "balance_date",
        }

    years = _unique_preserve_order(re.findall(r"\b((?:19|20)\d{2})\b", header_norm))
    return {
        "current_period_key": years[0] if len(years) >= 1 else None,
        "comparison_period_key": years[1] if len(years) >= 2 else None,
        "period_detection_basis": "year_fallback" if years else "not_detected",
    }


def annotate_df_with_period_keys(df: pd.DataFrame, period_info: dict) -> pd.DataFrame:
 
    df = df.copy()
    df["current_period_key"] = period_info.get("current_period_key")
    df["comparison_period_key"] = period_info.get("comparison_period_key")
    df["period_detection_basis"] = period_info.get("period_detection_basis")
    return df


def get_row_value_for_period(
    row,
    target_period_key: Optional[str],
    file_role: str = "older",
) -> tuple[Optional[float], Optional[str], bool, str]:
   
    target = normalize_period_key(target_period_key) if target_period_key is not None else None
    cur_key = row.get("current_period_key")
    cmp_key = row.get("comparison_period_key")
    cur_key = normalize_period_key(cur_key) if cur_key is not None and pd.notna(cur_key) else None
    cmp_key = normalize_period_key(cmp_key) if cmp_key is not None and pd.notna(cmp_key) else None

    if file_role == "older":
        value = row.get("current_value")
        if value is not None and pd.notna(value):
            if target and cur_key == target:
                source = "older_left_current_period_match"
            elif target and cur_key and cur_key != target:
                source = "older_left_current_period_mismatch_but_role_rule"
            elif target and not cur_key:
                source = "older_left_current_no_detected_period_but_role_rule"
            else:
                source = "older_left_current_role_rule"
            return (
                value,
                row.get("current_value_raw_original"),
                bool(row.get("current_value_was_repaired")),
                source,
                row.get("current_value_status", classify_value_parse_status(row.get("current_value_raw_original"), bool(row.get("current_value_was_repaired")), row.get("current_value_repair_type"))),
            )
        return None, None, False, "older_left_current_missing", "missing"

    if file_role == "newer":
        value = row.get("comparison_value")
        raw = row.get("comparison_value_raw_original")
        if value is not None and pd.notna(value):
            if target and cmp_key == target:
                source = "newer_right_comparison_period_match"
            elif target and cmp_key and cmp_key != target:
                source = "newer_right_comparison_period_mismatch_but_role_rule"
            elif target and not cmp_key:
                source = "newer_right_comparison_no_detected_period_but_role_rule"
            else:
                source = "newer_right_comparison_role_rule"
            return (
                value,
                raw,
                bool(row.get("comparison_value_was_repaired")),
                source,
                row.get("comparison_value_status", classify_value_parse_status(raw, bool(row.get("comparison_value_was_repaired")), row.get("comparison_value_repair_type"))),
            )

        # Intentional hard stop: no fallback to the newer current value.
        return None, None, False, "newer_right_comparison_missing_no_comparison_period_present", "missing"

    return None, None, False, "unknown_file_role", "missing"

def _period_keys_from_info(info: dict) -> list[str]:
    
    if not isinstance(info, dict):
        return []
    return _unique_preserve_order([
        normalize_period_key(info.get("current_period_key")) if info.get("current_period_key") else "",
        normalize_period_key(info.get("comparison_period_key")) if info.get("comparison_period_key") else "",
    ])


def determine_common_period_keys_by_section(parsed_older: dict, parsed_newer: dict) -> dict:
    
    result: dict = {}
    older_infos = parsed_older.get("period_keys", {}) or {}
    newer_infos = parsed_newer.get("period_keys", {}) or {}

    for section_key in ("tuloslaskelma", "tase_vastaavaa", "tase_vastattavaa"):
        old_info = older_infos.get(section_key, {}) or {}
        new_info = newer_infos.get(section_key, {}) or {}
        old_keys = _period_keys_from_info(old_info)
        new_keys = _period_keys_from_info(new_info)
        common = [k for k in old_keys if k and k in new_keys]

        preferred = normalize_period_key(old_info.get("current_period_key")) if old_info.get("current_period_key") else None
        if preferred and preferred in common:
            result[section_key] = preferred
            continue

        preferred_end = _period_end_date(preferred)
        if preferred_end:
            for nk in new_keys:
                if _period_end_date(nk) == preferred_end:
                    result[section_key] = nk
                    break
            if section_key in result:
                continue

        result[section_key] = common[0] if common else None

    return result


def _target_period_for_row(row, target_periods_by_section: dict) -> Optional[str]:
    section_key = normalize_section_for_matching(row.get("section"))
    return target_periods_by_section.get(section_key)


def add_match_values_for_common_period(
    df: pd.DataFrame,
    target_periods_by_section: dict,
    file_role: str,
) -> pd.DataFrame:
    """Add match values for common period.
    
    Purpose: This function belongs to the number parsing and validation stage.
    Why: It reduces the risk that formatting differences or unreadable values are treated as reliable evidence.
    """
    df = df.copy()
    if df.empty:
        df["match_value"] = pd.Series(dtype="float64")
        df["match_value_raw"] = pd.Series(dtype="object")
        df["match_value_was_repaired"] = pd.Series(dtype="bool")
        df["match_value_source"] = pd.Series(dtype="object")
        df["match_value_status"] = pd.Series(dtype="object")
        df["comparison_target_period_key"] = pd.Series(dtype="object")
        return df

    targets = df.apply(lambda row: _target_period_for_row(row, target_periods_by_section), axis=1)
    extracted = df.apply(
        lambda row: get_row_value_for_period(
            row,
            _target_period_for_row(row, target_periods_by_section),
            file_role=file_role,
        ),
        axis=1,
    )
    df["match_value"] = extracted.apply(lambda x: x[0])
    df["match_value_raw"] = extracted.apply(lambda x: x[1])
    df["match_value_was_repaired"] = extracted.apply(lambda x: x[2])
    df["match_value_source"] = extracted.apply(lambda x: x[3])
    df["match_value_status"] = extracted.apply(lambda x: x[4] if len(x) > 4 else classify_value_parse_status(x[1], bool(x[2])))
    df["comparison_target_period_key"] = targets
    df["file_role_for_value_selection"] = file_role
    return df

# =========================================================
# BALANCE SHEET SECTION RESCUE AND MISCLASSIFICATION CORRECTION
# =========================================================
# This section corrects obvious balance sheet section misclassifications.
#
# PDF extraction may occasionally place asset-side or liability-side rows under
# the wrong section because the source layout does not preserve table boundaries
# reliably.
#
# These functions infer the correct balance sheet side from known canonical
# asset and liability labels. If a row is moved, the related canonical metadata,
# hierarchy path and total classification are refreshed to keep later matching
# consistent.
#
# This is a conservative correction layer: it only rescues rows when the label
# clearly indicates either the asset side or the liability side.

def infer_balance_section_from_label(label: str) -> Optional[str]:
    """Infer balance section from label.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    norm = normalize_label(label)
    compact = norm.replace(" ", "")
    liability_terms = [
        "oma paaoma", "vieras paaoma", "vastattavaa yhteensa",
        "ostovelat", "muut velat", "siirtovelat", "menojaamat",
        "palkkamenot siirtovelat", "elakevakuutusmaksut siirtovelat",
        "tyottomyysvakuutusmaksut siirtovelat", "arvonlisaverovelka",
        "verotilivelka", "ennakonpidatysvelka", "sosiaaliturvamaksuvelka",
        "lainat rahoituslaitoksilta", "paaomapanos", "peruspaaoma",
        "yksityistilit", "tilikauden voitto", "edellisten tilikausien voitto",
    ]
    asset_terms = [
        "pysyvat vastaavat", "vaihtuvat vastaavat", "vastaavaa yhteensa",
        "aineettomat hyodykkeet", "aineelliset hyodykkeet", "vaihto omaisuus",
        "saamiset", "myyntisaamiset", "siirtosaamiset", "muut saamiset",
        "rahat ja pankkisaamiset", "pankkisaamiset", "kateisvarat", "verotilisaamiset",
    ]
    if any(t in norm or t.replace(" ", "") in compact for t in liability_terms):
        return "Tase vastattavaa"
    if any(t in norm or t.replace(" ", "") in compact for t in asset_terms):
        return "Tase vastaavaa"
    return None

def rescue_misclassified_balance_sections(df: pd.DataFrame) -> pd.DataFrame:
   
    if df is None or df.empty or "label" not in df.columns or "section" not in df.columns:
        return df
    df = df.copy()
    rescued = []
    for idx, row in df.iterrows():
        current_section = str(row.get("section", ""))
        current_key = normalize_section_name(current_section)
        inferred = infer_balance_section_from_label(str(row.get("label", "")))

        if inferred and current_key in {"tuloslaskelma", "tase_vastaavaa", "tase_vastattavaa"}:
            inferred_key = normalize_section_name(inferred)
            if inferred_key != current_key:
                label = str(row.get("label", ""))
                canonical = forced_main_item_canonical(label, inferred_key) or row.get("label_match_canonical") or canonicalize_common_labels(label)
                total_level = classify_total_level(label, inferred_key)

                df.at[idx, "section"] = inferred
                df.at[idx, "label_match_canonical"] = canonical
                df.at[idx, "total_level"] = total_level
                df.at[idx, "forced_main_item"] = bool(forced_main_item_canonical(label, inferred_key))
                df.at[idx, "forced_main_item_canonical"] = forced_main_item_canonical(label, inferred_key)
                df.at[idx, "forced_root_item"] = bool(canonical in FORCED_ROOT_ITEMS)
                df.at[idx, "balance_total_root_item"] = bool(
                    (inferred_key == "tase_vastaavaa" and canonical == "vastaavaa yhteensa")
                    or (inferred_key == "tase_vastattavaa" and canonical == "vastattavaa yhteensa")
                )
                if df.at[idx, "forced_root_item"]:
                    df.at[idx, "parent_label"] = None
                    df.at[idx, "top_parent_label"] = None
                    df.at[idx, "category_path"] = forced_root_category_path(canonical, inferred_key)
                    df.at[idx, "hierarchy_status"] = "Forced main-level item"
                df.at[idx, "total_context_key"] = total_context_key_from_parts(
                    inferred_key,
                    label,
                    str(df.at[idx, "category_path"]) if "category_path" in df.columns else "",
                    df.at[idx, "parent_label"] if "parent_label" in df.columns else None,
                    df.at[idx, "top_parent_label"] if "top_parent_label" in df.columns else None,
                )
                rescued.append(True)
            else:
                rescued.append(False)
        else:
            rescued.append(False)
    df["section_rescued"] = rescued
    return df

# =========================================================
# PDF FINANCIAL STATEMENT PARSING ORCHESTRATOR
# =========================================================
# This section orchestrates the full PDF parsing workflow for one financial statement.
#
# It performs:
# - Scanned-PDF detection.
# - Section extraction with row-level detection and page-level fallback.
# - Period-key detection.
# - Row parsing for income statement, asset side and liability side.
# - Source-location attachment.
# - Balance-side rescue.
# - Document-level number-format and reliability assessment.
#
# The purpose is to convert one text-based PDF financial statement into a
# structured, validated dataset that can be used by the reconciliation engine.


def parse_sme_financial_statement_from_bytes(pdf_bytes: bytes, filename: str):
    """Parse sme financial statement from bytes.
    
    Purpose: This function belongs to the number parsing and validation stage.
    Why: It reduces the risk that formatting differences or unreadable values are treated as reliable evidence.
    """
    if detect_scanned_pdf(pdf_bytes):
        raise ValueError("The document appears to be a scanned image-based PDF. OCR is required before reconciliation.")

    section_extraction_method = "row_level"
    section_extraction_warning = None

    try:
        section_bundle = extract_all_sections_row_level(pdf_bytes)
    except Exception as exc:
        section_extraction_method = "page_level_fallback"
        section_extraction_warning = f"Row-level section detection failed: {exc}"
        section_bundle = extract_all_sections_from_bytes(pdf_bytes)

    pages = section_bundle["pages"]
    start_pages = section_bundle["start_pages"]

    income_lines = section_bundle["tuloslaskelma"]
    assets_lines = section_bundle["tase_vastaavaa"]
    liab_lines = section_bundle["tase_vastattavaa"]

    income_periods = extract_column_period_keys(income_lines)
    assets_periods = extract_column_period_keys(assets_lines)
    liab_periods = extract_column_period_keys(liab_lines)

    df_income_raw, income_diag = parse_section_lines_to_df(income_lines, "Tuloslaskelma")
    df_assets_raw, assets_diag = parse_section_lines_to_df(assets_lines, "Tase vastaavaa")
    df_liab_raw, liab_diag = parse_section_lines_to_df(liab_lines, "Tase vastattavaa")

    df_income = annotate_df_with_period_keys(add_section_name(df_income_raw, "Tuloslaskelma"), income_periods)
    df_assets = annotate_df_with_period_keys(add_section_name(df_assets_raw, "Tase vastaavaa"), assets_periods)
    df_liab = annotate_df_with_period_keys(add_section_name(df_liab_raw, "Tase vastattavaa"), liab_periods)

    source_lookup = build_source_line_lookup(pdf_bytes)
    df_income = attach_source_locations(df_income, source_lookup)
    df_assets = attach_source_locations(df_assets, source_lookup)
    df_liab = attach_source_locations(df_liab, source_lookup)

    df_all = pd.concat([df_income, df_assets, df_liab], ignore_index=True)
    df_all = rescue_misclassified_balance_sections(df_all)
    if df_all.empty:
        raise ValueError("No rows could be extracted from the document for reconciliation.")

    dominant_format = choose_document_dominant_format(df_all)
    diagnostics = {
        "tuloslaskelma": income_diag,
        "tase_vastaavaa": assets_diag,
        "tase_vastattavaa": liab_diag,
        "document_quality": "Medium",
        "section_extraction_method": section_extraction_method,
        "section_extraction_warning": section_extraction_warning,
    }
    detail_share = (income_diag["detail_rows"] + assets_diag["detail_rows"] + liab_diag["detail_rows"]) / max(1, len(df_all))
    if detail_share > 0.4:
        diagnostics["document_quality"] = "Erittelypainotteinen"
    elif len(df_all) >= 20:
        diagnostics["document_quality"] = "Good"

    parsed_result = {
        "document": filename,
        "pages": pages,
        "start_pages": start_pages,
        "df_income": df_income,
        "df_vastaavaa": df_assets,
        "df_vastattavaa": df_liab,
        "df_all": df_all,
        "section_lines": {
            "tuloslaskelma": income_lines,
            "tase_vastaavaa": assets_lines,
            "tase_vastattavaa": liab_lines,
        },
        "source_line_lookup": source_lookup,
        "dominant_number_format": dominant_format,
        "diagnostics": diagnostics,
        "period_keys": {
            "tuloslaskelma": income_periods,
            "tase_vastaavaa": assets_periods,
            "tase_vastattavaa": liab_periods,
        },
    }
    parsed_result["document_reliability"] = build_document_reliability(parsed_result)
    return parsed_result


# =========================================================
# RECONCILIATION MATCHING CONTROLS AND CONFIDENCE CLASSIFICATION
# =========================================================
# This section defines the control logic used when deciding whether two rows
# can be treated as a verified reconciliation match.
#
# It checks:
# - Canonical label equivalence.
# - Section compatibility.
# - Total-row context compatibility.
# - Semantic level compatibility.
# - Numeric equality within tolerance.
# - Safe value extraction status.
# - Absence of repair, fallback and context conflicts.
#
# The purpose is to avoid false-positive reconciliations. A row is accepted as
# a clear verified match only when both the label evidence and value evidence
# support the decision.
#
# Uncertain, repaired or structurally weak matches are deliberately downgraded
# to manual review instead of being accepted automatically.


def canonical_values_equivalent(a, b, section: Optional[str] = None) -> bool:
    """Compare canonical values while tolerating compact PDF formatting variants."""
    a_norm = canonicalize_common_labels(a)
    b_norm = canonicalize_common_labels(b)
    if a_norm and b_norm and a_norm == b_norm:
        return True
    a_compact = re.sub(r"[^a-z0-9]", "", normalize_label(a_norm))
    b_compact = re.sub(r"[^a-z0-9]", "", normalize_label(b_norm))
    return bool(a_compact and b_compact and a_compact == b_compact)


# Perustelu:


def _match_reason_has_value_evidence(reason: str) -> bool:
    """Return True only when the selected candidate is supported by value evidence.

    This is intentionally stricter than accepting a row because it has a unique
    canonical label. The thesis artefact must avoid silent false positives: if a
    row is only structurally plausible but not value-supported, it is left for
    manual review.
    """
    reason = str(reason or "").lower()
    value_tokens = (
        "value_checked",
        "value_supported",
        "text_and_value",
        "label_and_value",
        "same_canonical_value_checked",
        "exact_label_value_checked",
    )
    return any(token in reason for token in value_tokens)


def _values_match_within_tolerance(row, tolerance: float = TOLERANCE) -> bool:
    """Verify the actual numeric equality instead of trusting the status text."""
    diff = _row_get(row, "difference", None)
    if diff is not None:
        try:
            return abs(float(diff)) <= tolerance
        except Exception:
            return False

    old_value = _row_get(row, "value_older_current", None)
    new_value = _row_get(row, "value_newer_comparison", None)
    if old_value is None or new_value is None:
        return False
    try:
        return abs(float(new_value) - float(old_value)) <= tolerance
    except Exception:
        return False


def _row_has_safe_value_statuses(row) -> bool:
    """Row has safe value statuses.
    
    Purpose: This function belongs to the number parsing and validation stage.
    Why: It reduces the risk that formatting differences or unreadable values are treated as reliable evidence.
    """
    old_value_status = str(_row_get(row, "value_status_older", "") or "")
    new_value_status = str(_row_get(row, "value_status_newer", "") or "")
    return value_status_is_safe_for_verified_match(old_value_status) and value_status_is_safe_for_verified_match(new_value_status)


def is_clear_verified_match_decision(row_or_status=None, **kwargs) -> bool:
    """Return True only for clearly verified, automatically acceptable matches.

    Earlier versions accepted some matches too easily when the canonical label
    looked plausible. That creates the highest reconciliation risk: a wrong row
    can pass as reconciled. This gate now requires all of the following:
    same section, compatible canonical labels, safe value extraction, numeric
    equality within tolerance, no structural/context conflict, no numeric repair,
    and explicit value-supported match evidence.
    """
    row = row_or_status if hasattr(row_or_status, "get") else None
    if row is None:
        return False

    status = str(_row_get(row, "status", "") or "")
    if status != "Match":
        return False

    if not _values_match_within_tolerance(row):
        return False

    old_section = str(_row_get(row, "section_match_norm_older", "") or "")
    new_section = str(_row_get(row, "section_match_norm_newer", "") or "")
    if not old_section or not new_section or old_section != new_section:
        return False

    old_can = str(_row_get(row, "label_match_canonical_older", "") or "")
    new_can = str(_row_get(row, "label_match_canonical_newer", "") or "")
    if not old_can or not new_can or not canonical_values_equivalent(old_can, new_can):
        return False

    if bool(_row_get(row, "context_conflict", False)):
        return False
    if bool(_row_get(row, "context_uncertain", False)):
        return False

    hierarchy_status = str(_row_get(row, "hierarchy_status", "") or "")
    if hierarchy_status in {"Ristiriitainen", "Uncertain"}:
        return False

    context_status_old = str(_row_get(row, "canonical_context_status_older", "") or "")
    context_status_new = str(_row_get(row, "canonical_context_status_newer", "") or "")
    if context_status_old in {"Ristiriitainen", "Uncertain"} or context_status_new in {"Ristiriitainen", "Uncertain"}:
        return False

    if bool(_row_get(row, "numeric_integrity_issue", False)) or bool(_row_get(row, "repaired_issue", False)):
        return False
    if bool(_row_get(row, "fallback_used", False)):
        return False
    if not _row_has_safe_value_statuses(row):
        return False

    method = str(_row_get(row, "match_method", "") or "")
    reason = str(_row_get(row, "match_reason", "") or "")

    # Review/failure reasons are intentionally never upgraded to verified
    # matches, even if another field still says status == "Match".
    if reason.startswith("review:") or reason.startswith("fail:"):
        return False

    strong_methods = {"exact_value_match", "compact_value_match", "exact", "compact", "canonical"}
    if method not in strong_methods:
        return False

    return _match_reason_has_value_evidence(reason)


def determine_row_confidence(
    status: str,
    uncertain_reason: Optional[str],
    repaired_issue: bool,
    separator_issue: bool,
    consistency_issue: bool,
    fallback_used: bool,
    label_match_score: float = 0.0,
    match_method: Optional[str] = None,
    hierarchy_score: float = 0.0,
    hierarchy_conflict: bool = False,
    context_uncertain: bool = False,
    value_status_issue: bool = False,
    match_reason: Optional[str] = None,
) -> str:
    """Classify confidence conservatively.

    The confidence label is not only a display value. It controls whether a row
    can be accepted without manual review. Any repaired value, uncertain parse,
    fallback, context uncertainty or soft match lowers the confidence.
    """
    if status == "Match":
        if hierarchy_conflict or uncertain_reason or value_status_issue:
            return "Low"
        if repaired_issue or separator_issue or consistency_issue or fallback_used or context_uncertain:
            return "Medium"

        strong_method = match_method in {
            "canonical",
            "exact",
            "exact_value_match",
            "compact",
            "compact_value_match",
        }
        value_evidence = _match_reason_has_value_evidence(match_reason or "")

        if strong_method and value_evidence and label_match_score >= 0.98 and hierarchy_score >= 0.80:
            return "High"
        if strong_method and value_evidence and label_match_score >= 0.95:
            return "High"
        if strong_method and value_evidence and hierarchy_score >= 0.90:
            return "High"

        return "Medium"

    if status == "Value differs":
        if match_method in {"canonical", "exact", "compact", "exact_value_match", "compact_value_match"}:
            return "High"
        return "Medium"

    if status == "Missing row":
        return "High"

    if status == "Could not verify":
        return "Low"

    if uncertain_reason:
        return "Low"

    return "Medium"


def determine_row_severity(status: str, confidence_level: str, repaired_issue: bool, separator_issue: bool, consistency_issue: bool, decimal_presentation_issue: bool) -> str:
   
    if status in {"Value differs", "Missing row", "Could not verify"}:
        return "Tarkistettava"
    if repaired_issue or separator_issue or consistency_issue or decimal_presentation_issue or confidence_level != "High":
        return "Technical note"
    return "Ei huomioita"


def build_row_explanation(status: str, uncertain_reason: Optional[str], repaired_issue: bool, separator_issue: bool, consistency_issue: bool, decimal_presentation_issue: bool, difference):
   
    if status == "Match":
        return "Values match"
    if status == "Value differs":
        return "Value differs"
    if status == "Missing row":
        return "Rivi puuttuu toisesta tiedostosta"
    if status == "Could not verify":
        return "The row could not be matched with sufficient certainty"
    return "Tarkista havainto"


def make_keys(label: str, section: str) -> dict:
   
    canonical = canonicalize_common_labels(label)
    compact = canonical.replace(" ", "")
    return {"canonical": f"{section.lower()}||{canonical}", "compact": f"{section.lower()}||{compact}"}


def _safe_norm(value) -> str:
   
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return normalize_label(str(value))


def is_generic_total_label(label: str) -> bool:
    """Is generic total label.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    norm = normalize_label(label)
    return norm in {"yhteensa", "summa", "total", "totalt"}


def is_named_total_label(label: str) -> bool:
    
    norm = normalize_label(label)
    if is_generic_total_label(norm):
        return False
    return "yhteensa" in norm or "total" in norm or "totalt" in norm


def is_context_sensitive_total_row(row_or_label) -> bool:

    if isinstance(row_or_label, (pd.Series, dict)):
        label = str(row_or_label.get("label", ""))
    else:
        label = str(row_or_label or "")
    return is_generic_total_label(label)


def total_context_key_from_parts(section, label, category_path=None, parent_label=None, top_parent_label=None) -> str:
   
    section_norm = normalize_section_for_matching(section)
    label_norm = normalize_label(label)
    forced = forced_main_item_canonical(label, section_norm)
    if forced in {"vastaavaa yhteensa", "vastattavaa yhteensa", "oma paaoma yhteensa", "vieras paaoma yhteensa"}:
        return f"{section_norm}||ROOT||{forced}"
    path_norm = _safe_norm(category_path)
    parent_norm = _safe_norm(parent_label)
    top_norm = _safe_norm(top_parent_label)
    if is_generic_total_label(label) or label_norm == "yhteensa":
        context = path_norm or " > ".join(x for x in [top_norm, parent_norm] if x) or parent_norm or top_norm
        return f"{section_norm}||CTX_TOTAL||{context}"
    if is_named_total_label(label):
        return f"{section_norm}||NAMED_TOTAL||{label_norm}"
    return f"{section_norm}||ROW||{label_norm}"


def total_context_key(row) -> str:
   
    return total_context_key_from_parts(
        row.get("section"),
        row.get("label"),
        row.get("category_path"),
        row.get("parent_label"),
        row.get("top_parent_label"),
    )



def _row_total_flag(row) -> bool:
    """Return whether the row is an aggregate or total-like row."""
    if isinstance(row, (pd.Series, dict)):
        if bool(row.get("summary_row", False)):
            return True
        row_type = str(row.get("row_type", "") or "")
        total_level = str(row.get("total_level", "") or "")
        label = str(row.get("label", "") or "")
        return row_type in {"grand_total", "section_total", "result"} or bool(total_level and total_level != "nan") or is_total_row(label)
    return is_total_row(str(row or ""))


def _row_total_level_value(row) -> str:
    """Return a stable total level value for structural comparison."""
    if not isinstance(row, (pd.Series, dict)):
        return "total" if is_total_row(str(row or "")) else "detail"
    total_level = str(row.get("total_level", "") or "").strip().lower()
    if total_level and total_level != "nan":
        return total_level
    row_type = str(row.get("row_type", "") or "").strip().lower()
    if row_type in {"grand_total", "section_total", "result"}:
        return row_type
    return "total" if _row_total_flag(row) else "detail"


def rows_are_structurally_compatible_for_matching(r_old, r_new) -> tuple[bool, str]:
    """Validate row-level structure before any label or value match is accepted.

    This gate prevents aggregate rows from matching detail rows and prevents
    totals from different hierarchy contexts from being paired only because
    their labels or values look similar.
    """
    old_total = _row_total_flag(r_old)
    new_total = _row_total_flag(r_new)
    if old_total != new_total:
        return False, "fail:total_vs_non_total_mismatch"

    old_level = _row_total_level_value(r_old)
    new_level = _row_total_level_value(r_new)
    if old_total and new_total and old_level != new_level:
        # Result rows are allowed to match result rows even when terminology differs.
        if not (old_level == "result" and new_level == "result"):
            return False, "fail:total_level_mismatch"

    if old_total and new_total and not total_context_compatible(r_old, r_new):
        return False, "fail:total_context_mismatch"

    old_parent = _safe_norm(r_old.get("parent_label") if isinstance(r_old, (pd.Series, dict)) else None)
    new_parent = _safe_norm(r_new.get("parent_label") if isinstance(r_new, (pd.Series, dict)) else None)
    if old_parent and new_parent and old_parent != new_parent:
        return False, "fail:parent_context_mismatch"

    return True, "ok:structurally_compatible"

def total_context_compatible(r_old, r_new) -> bool:

    if not (_row_total_flag(r_old) or _row_total_flag(r_new) or is_context_sensitive_total_row(r_old) or is_context_sensitive_total_row(r_new)):
        return True
    old_key = total_context_key(r_old)
    new_key = total_context_key(r_new)
    return bool(old_key and new_key and old_key == new_key)


def map_user_status(status: str) -> str:
    """Map internal reconciliation statuses to Finnish user-facing labels."""
    return {
        "Match": "Täsmää",
        "Value differs": "Luku poikkeaa",
        "Missing row": "Rivi puuttuu",
        "Could not verify": "Ei voitu varmistaa",
        "Manual review": "Tarkista",
        "Virhe": "Luku poikkeaa",
        "Puuttuu": "Rivi puuttuu",
    }.get(status, status)


def map_user_confidence(confidence: str) -> str:
    """Map internal confidence levels to Finnish user-facing labels."""
    return {"High": "Vahva", "Medium": "Keskitaso", "Low": "Matala", "Uncertain": "Epävarma"}.get(confidence, confidence)


def normalize_label_strict(label: str) -> str:
    """Normalize label strict.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    norm = normalize_label(label)
    replacements = {
        "liikevoitto tappio": "liikevoitto",
        "liiketappio": "liikevoitto",
        "tilikauden voitto tappio": "tilikauden voitto",
        "tilikauden tappio": "tilikauden voitto",
        "voitto tappio": "voitto",
    }
    for old_part, new_part in replacements.items():
        norm = norm.replace(old_part, new_part)
    return re.sub(r"\s+", " ", norm).strip()


def normalize_for_exact_match(label: str) -> str:
    """Normalize for exact match.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    return normalize_label_strict(label)


def normalize_for_compact_match(label: str) -> str:
    """Normalize for compact match.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    return normalize_label_strict(label).replace(" ", "")


def get_canonical_group(label: str) -> str:
    """Get canonical group.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    return canonicalize_common_labels(label)


def strong_exact_match(r_old, r_new, tolerance: float = TOLERANCE) -> bool:
    """Strong exact match.
    
    Purpose: This function belongs to the reconciliation and candidate evaluation stage.
    Why: It supports controlled matching while keeping uncertain cases visible for manual review.
    """
    label_a = normalize_for_exact_match(r_old.get("label", ""))
    label_b = normalize_for_exact_match(r_new.get("label", ""))
    if not label_a or not label_b or label_a != label_b:
        return False
    if not total_context_compatible(r_old, r_new):
        return False

    value_a = r_old.get("match_value") if "match_value" in getattr(r_old, "index", []) else r_old.get("current_value")
    value_b = _candidate_comparison_value(r_new)
    if value_a is None or value_b is None:
        return False

    try:
        return abs(float(value_a) - float(value_b)) <= tolerance
    except Exception:
        return False


def label_similarity_score(label_a: str, label_b: str) -> tuple[float, str]:
    """Label similarity score.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    exact_a = normalize_for_exact_match(label_a)
    exact_b = normalize_for_exact_match(label_b)

    compact_a = normalize_for_compact_match(label_a)
    compact_b = normalize_for_compact_match(label_b)

    canon_a = get_canonical_group(label_a)
    canon_b = get_canonical_group(label_b)

    if exact_a and exact_b and exact_a == exact_b:
        return 1.00, "exact"

    if compact_a and compact_b and compact_a == compact_b:
        return 0.95, "compact"

    if canon_a and canon_b and canon_a == canon_b and exact_a != exact_b:
        return 0.70, "synonym"

    ratio = fuzzy_ratio(label_a, label_b)
    if ratio >= 0.92:
        return float(ratio), "fuzzy_strong"
    if ratio >= 0.85:
        return float(ratio), "fuzzy"
    return 0.0, "none"


def value_support_score(value_old, value_new, tolerance=TOLERANCE) -> float:
    """Value support score.
    
    Purpose: This function belongs to the number parsing and validation stage.
    Why: It reduces the risk that formatting differences or unreadable values are treated as reliable evidence.
    """
    if value_old is None or value_new is None:
        return 0.0
    try:
        diff = abs(float(value_old) - float(value_new))
    except Exception:
        return 0.0
    if diff <= tolerance:
        return 0.20
    if diff <= ROUNDING_WARNING_LIMIT:
        return 0.10
    return 0.0


def _candidate_comparison_value(r_new) -> Optional[float]:
    # Do not silently fall back when the selected comparison-period value is missing.
    """Candidate comparison value.
    
    Purpose: This function belongs to the number parsing and validation stage.
    Why: It reduces the risk that formatting differences or unreadable values are treated as reliable evidence.
    """
    if "match_value" in getattr(r_new, "index", []):
        mv = r_new.get("match_value")
        role = r_new.get("file_role_for_value_selection") if "file_role_for_value_selection" in getattr(r_new, "index", []) else None
        if mv is not None and pd.notna(mv):
            return mv
        if role == "newer":
            return None
    if r_new.get("structure") == "single-column":
        return r_new.get("current_value")
    if pd.notna(r_new.get("comparison_value")):
        return r_new.get("comparison_value")
    return r_new.get("current_value")


def _candidate_comparison_raw(r_new) -> Optional[str]:
    """Candidate comparison raw.
    
    Purpose: This function belongs to the reconciliation and candidate evaluation stage.
    Why: It supports controlled matching while keeping uncertain cases visible for manual review.
    """
    if "match_value_raw" in getattr(r_new, "index", []):
        raw = r_new.get("match_value_raw")
        role = r_new.get("file_role_for_value_selection") if "file_role_for_value_selection" in getattr(r_new, "index", []) else None
        if raw is not None and pd.notna(raw):
            return raw
        if role == "newer":
            return None
    if r_new.get("structure") == "single-column":
        return r_new.get("current_value_raw_original")
    if pd.notna(r_new.get("comparison_value")) and r_new.get("comparison_value_raw_original") is not None:
        return r_new.get("comparison_value_raw_original")
    return r_new.get("current_value_raw_original")


def _candidate_comparison_repaired(r_new) -> bool:
    """Candidate comparison repaired.
    
    Purpose: This function belongs to the reconciliation and candidate evaluation stage.
    Why: It supports controlled matching while keeping uncertain cases visible for manual review.
    """
    if "match_value_was_repaired" in getattr(r_new, "index", []):
        role = r_new.get("file_role_for_value_selection") if "file_role_for_value_selection" in getattr(r_new, "index", []) else None
        mv = r_new.get("match_value") if "match_value" in getattr(r_new, "index", []) else None
        if mv is not None and pd.notna(mv):
            return bool(r_new.get("match_value_was_repaired"))
        if role == "newer":
            return False
    if r_new.get("structure") == "single-column":
        return bool(r_new.get("current_value_was_repaired"))
    if pd.notna(r_new.get("comparison_value")):
        return bool(r_new.get("comparison_value_was_repaired"))
    return bool(r_new.get("current_value_was_repaired"))


def _candidate_comparison_status(r_new) -> str:
    """Return a derived value used by the reconciliation workflow."""
    idx = getattr(r_new, "index", [])

    if "match_value_status" in idx:
        status = r_new.get("match_value_status")
        role = r_new.get("file_role_for_value_selection") if "file_role_for_value_selection" in idx else None
        mv = r_new.get("match_value") if "match_value" in idx else None
        if status is not None and pd.notna(status) and mv is not None and pd.notna(mv):
            return str(status)
        if role == "newer":
            return "missing"

    if r_new.get("structure") == "single-column":
        return str(r_new.get(
            "current_value_status",
            classify_value_parse_status(
                r_new.get("current_value_raw_original"),
                bool(r_new.get("current_value_was_repaired", False)),
            ),
        ))

    if pd.notna(r_new.get("comparison_value")):
        return str(r_new.get(
            "comparison_value_status",
            classify_value_parse_status(
                r_new.get("comparison_value_raw_original"),
                bool(r_new.get("comparison_value_was_repaired", False)),
            ),
        ))

    return str(r_new.get(
        "current_value_status",
        classify_value_parse_status(
            r_new.get("current_value_raw_original"),
            bool(r_new.get("current_value_was_repaired", False)),
        ),
    ))
def _numeric_diff(a, b) -> Optional[float]:
   
    if a is None or b is None:
        return None
    try:
        return abs(float(a) - float(b))
    except Exception:
        return None


def _row_semantic_level(row) -> str:
    """Return a derived value used by the reconciliation workflow."""
    block_type = str(row.get("block_type", "") or "")
    row_type = str(row.get("row_type", "") or "")
    label = str(row.get("label", "") or "")
    source_line = str(row.get("source_line", "") or "")

    if has_account_code_prefix(source_line) or has_account_code_prefix(label) or block_type == "statement_detail":
        return "detail"
    if row_type in {"grand_total", "section_total"} or bool(row.get("summary_row", False)):
        return "total"
    return "main"


def _match_levels_compatible(r_old, r_new, *, exact_label: bool = False, label_score: float = 0.0) -> bool:
    """Evaluate whether the item is valid in its detected context."""
    old_level = _row_semantic_level(r_old)
    new_level = _row_semantic_level(r_new)

    if old_level == new_level:
        return True

    if old_level == "total" or new_level == "total":
        return False

    old_label = str(r_old.get("label", "") or "")
    new_label = str(r_new.get("label", "") or "")

    if exact_label:
        if normalize_for_exact_match(old_label) == normalize_for_exact_match(new_label):
            return True
        if normalize_for_compact_match(old_label) == normalize_for_compact_match(new_label):
            return True

    if label_score >= 0.92:
        return True

    old_forced = (
        r_old.get("forced_main_item_canonical")
        if "forced_main_item_canonical" in getattr(r_old, "index", [])
        else None
    ) or forced_main_item_canonical(old_label, r_old.get("section"))

    new_forced = (
        r_new.get("forced_main_item_canonical")
        if "forced_main_item_canonical" in getattr(r_new, "index", [])
        else None
    ) or forced_main_item_canonical(new_label, r_new.get("section"))

    if old_forced and new_forced and old_forced == new_forced:
        return True

    return False


def _row_is_taxonomy_recognized(row) -> bool:
    """Return a derived value used by the reconciliation workflow."""
    return bool(row.get("label_match_accepted", False)) or str(row.get("extraction_status", "")) == "taxonomy_recognized"


def _fuzzy_matching_allowed_for_pair(r_old, r_new) -> bool:
   
    if _row_is_taxonomy_recognized(r_old) and _row_is_taxonomy_recognized(r_new):
        return True

    old_forced = forced_main_item_canonical(r_old.get("label", ""), r_old.get("section"))
    new_forced = forced_main_item_canonical(r_new.get("label", ""), r_new.get("section"))
    return bool(old_forced and new_forced and old_forced == new_forced)


# =========================================================
# CANDIDATE SELECTION, MATCH FAILURE DIAGNOSTICS AND FINAL RECONCILIATION
# =========================================================
# This section performs the final reconciliation between the older and newer
# financial statement datasets.
#
# It handles:
# - Canonical match-key construction.
# - Candidate selection within the same statement section.
# - Exact, compact, canonical and fuzzy value-supported matching.
# - Diagnosis of why a row could not be matched.
# - Manual-review gating for uncertain matches.
# - Repair of reciprocal missing-row findings.
# - Final result row construction and sorting.
#
# The purpose is to compare the older statement's current-period rows against
# the newer statement's comparative-period rows while avoiding false-positive
# matches.
#
# Matching is intentionally conservative: label similarity alone is not enough.
# A verified match must also be supported by value evidence, section
# compatibility, semantic level compatibility and context checks.

# ---------------------------------------------------------
# CANONICAL MATCH KEY CONSTRUCTION
# ---------------------------------------------------------

@lru_cache(maxsize=1)
def _all_allowed_canonical_keys() -> frozenset[str]:
    """Return a derived value used by the reconciliation workflow."""
    keys = set()
    try:
        for values in SECTION_ALLOWED_CANONICALS.values():
            keys.update(normalize_label(v) for v in values if normalize_label(v))
    except Exception:
        pass
    try:
        keys.update(normalize_label(v) for v in KNOWN_STATEMENT_LABELS if normalize_label(v))
    except Exception:
        pass
    try:
        keys.update(normalize_label(v) for v in FORCED_MAIN_ITEM_CANONICALS if normalize_label(v))
    except Exception:
        pass
    return frozenset(keys)


def _strip_safe_total_suffix(canonical: str) -> str:

    canonical = normalize_label(canonical)
    explicit_aliases = {
        "tase vastaavaa yhteensa": "vastaavaa yhteensa",
        "tase vastattavaa yhteensa": "vastattavaa yhteensa",
        "oma paaoma ja velat yhteensa": "vastattavaa yhteensa",
    }
    canonical = explicit_aliases.get(canonical, canonical)

    protected_totals = {
        "vastaavaa yhteensa",
        "vastattavaa yhteensa",
        "oma paaoma yhteensa",
        "vieras paaoma yhteensa",
        "pysyvat vastaavat yhteensa",
        "vaihtuvat vastaavat yhteensa",
    }
    if canonical in protected_totals:
        return canonical

    if canonical.endswith(" yhteensa"):
        base = canonical[: -len(" yhteensa")].strip()
        if base and base in _all_allowed_canonical_keys():
            return base
    return canonical


@lru_cache(maxsize=50000)
def _safe_structural_canonical_from_label(label: str, section=None) -> str:
    """Normalize input values for reliable comparison."""
    label = str(label or "")
    forced = forced_main_item_canonical(label, section)
    if forced:
        return _strip_safe_total_suffix(forced)

    compact_canonical = _compact_canonical_lookup(label, section)
    if compact_canonical:
        return _strip_safe_total_suffix(compact_canonical)

    norm = normalize_label(label)
    return _strip_safe_total_suffix(norm)


def _canonical_match_key(row) -> str:
    
    label = str(row.get("label", "") or "")
    section = row.get("section")

    if has_account_code_prefix(label):
        cleaned = normalize_label(re.sub(r"^\s*\d{3,6}\s+", "", label))
        cleaned = re.sub(r"\s+yhteensa$", "", cleaned).strip()
        return cleaned or normalize_label(label)

    return _safe_structural_canonical_from_label(label, section)


def _same_canonical_match_allowed(r_old, r_new) -> bool:
    """Return True only when same-canonical candidates are structurally comparable.

    Earlier development versions allowed a same-canonical pair even when the
    semantic row level differed. That made the tool feel helpful, but it also
    hid an important uncertainty: a main statement row and a detail row may have
    the same normalized name while still representing different levels of
    presentation.

    This transparent version keeps the candidate available only when the rows
    are on the same semantic level. Cross-level relationships are handled by
    the presentation-change and one-to-many review layers instead of being
    silently accepted as a one-to-one match.
    """
    old_key = _canonical_match_key(r_old)
    new_key = _canonical_match_key(r_new)
    if not old_key or old_key != new_key:
        return False

    return _row_semantic_level(r_old) == _row_semantic_level(r_new)


def _candidate_signature_for_disambiguation(row) -> tuple:
    """Return a structural signature used to distinguish rows with identical labels."""
    return (
        normalize_section_for_matching(row.get("section")),
        str(row.get("block_type", "") or ""),
        str(row.get("row_type", "") or ""),
        _row_semantic_level(row),
        bool(row.get("summary_row", False)),
        normalize_label(row.get("parent_label") or ""),
        normalize_label(row.get("top_parent_label") or ""),
        total_context_key(row),
    )

# ---------------------------------------------------------
# MATCH FAILURE DIAGNOSTICS AND EXCEPTION CLASSIFICATION
# ---------------------------------------------------------

def choose_unique_supported_candidate(candidates_df: pd.DataFrame, r_old, old_value):
    """Choose a candidate only when the choice is explainable.

    The function intentionally separates three situations:
    1. One candidate with value support: safe candidate for an automatic result.
    2. One candidate without value support but with an actual numeric difference:
       usable for reporting a value difference, because the same item appears to
       exist but the amounts do not reconcile.
    3. One candidate without comparable values or only structural uniqueness:
       not selected automatically; the row remains missing or moves to manual
       review through the normal result pipeline.

    This prevents the tool from silently accepting a row only because the label
    or canonical key was unique.
    """
    if candidates_df is None or candidates_df.empty:
        return None
    frame = candidates_df.copy()
    if "candidate_value" not in frame.columns:
        frame["candidate_value"] = frame.apply(_candidate_comparison_value, axis=1)
    if "value_support" not in frame.columns:
        frame["value_support"] = frame["candidate_value"].apply(lambda v: value_support_score(old_value, v))
    if "value_diff" not in frame.columns:
        frame["value_diff"] = frame["candidate_value"].apply(lambda v: _numeric_diff(old_value, v))

    supported = frame[frame["value_support"] > 0].copy()
    if len(supported) == 1:
        return supported.sort_values(by=["value_support", "value_diff"], ascending=[False, True], na_position="last").iloc[0]
    if len(supported) > 1:
        old_sig = _candidate_signature_for_disambiguation(r_old)
        supported["signature"] = supported.apply(_candidate_signature_for_disambiguation, axis=1)
        same_sig = supported[supported["signature"] == old_sig]
        if len(same_sig) == 1:
            return same_sig.iloc[0]
        return None

    # No equal-value support exists. Select only if there is exactly one
    # candidate and both sides contain numeric values, so the row can be shown as
    # a transparent value difference rather than as a hidden forced match.
    comparable = frame[frame["value_diff"].notna()].copy()
    if len(comparable) == 1 and len(frame) == 1:
        return comparable.iloc[0]

    return None


def find_best_row_match(r_old, df_new: pd.DataFrame, used_new: set):
    """Normalize input values for reliable comparison."""
    available = df_new[~df_new.index.isin(used_new)].copy()
    if available.empty:
        return None, None, 0.0, "fail:no_unused_candidates"

    old_section_norm = normalize_section_for_matching(r_old.get("section"))
    available["section_match_norm"] = available["section"].apply(normalize_section_for_matching)
    candidates = available[available["section_match_norm"] == old_section_norm].copy()
    if candidates.empty:
        return None, None, 0.0, "fail:section_mismatch"

    compatibility = candidates.apply(lambda r: rows_are_structurally_compatible_for_matching(r_old, r), axis=1)
    candidates["structural_match_allowed"] = compatibility.apply(lambda x: bool(x[0]))
    candidates["structural_match_reason"] = compatibility.apply(lambda x: x[1])
    candidates = candidates[candidates["structural_match_allowed"]].copy()
    if candidates.empty:
        return None, None, 0.0, "fail:structural_context_mismatch"

    if is_context_sensitive_total_row(r_old) or _row_total_flag(r_old):
        candidates["total_context_key"] = candidates.apply(total_context_key, axis=1)
        old_total_context_key = total_context_key(r_old)
        context_candidates = candidates[candidates["total_context_key"] == old_total_context_key].copy()
        if not context_candidates.empty:
            candidates = context_candidates
        else:
            return None, None, 0.0, "fail:total_context_mismatch"

    old_label = r_old["label"]
    old_value = r_old.get("match_value") if "match_value" in getattr(r_old, "index", []) else r_old.get("current_value")
    old_exact = normalize_for_exact_match(old_label)
    old_compact = normalize_for_compact_match(old_label)
    old_block_type = str(r_old.get("block_type", "main_statement") or "main_statement")

    candidates["exact_norm"] = candidates["label"].apply(normalize_for_exact_match)
    candidates["compact_norm"] = candidates["label"].apply(normalize_for_compact_match)
    candidates["block_match_bonus"] = candidates.get("block_type", pd.Series("main_statement", index=candidates.index)).fillna("main_statement").astype(str).apply(
        lambda x: 0.20 if x == old_block_type else (-0.20 if old_block_type == "main_statement" and x == "statement_detail" else 0.0)
    )
    old_forced_canonical = r_old.get("forced_main_item_canonical") if "forced_main_item_canonical" in getattr(r_old, "index", []) else None
    if not old_forced_canonical:
        old_forced_canonical = forced_main_item_canonical(old_label, r_old.get("section"))
    candidates["forced_match_bonus"] = candidates.apply(
        lambda r: 0.35 if old_forced_canonical and (r.get("forced_main_item_canonical") == old_forced_canonical or forced_main_item_canonical(r.get("label", ""), r.get("section")) == old_forced_canonical) else 0.0,
        axis=1,
    )

    candidates["semantic_level"] = candidates.apply(_row_semantic_level, axis=1)
    old_semantic_level = _row_semantic_level(r_old)

    # 0) Exact label / canonical match even when values differ and row levels differ:
    broad_exact_candidates = candidates[
        (candidates["exact_norm"] == old_exact)
    ].copy()
    if not broad_exact_candidates.empty and old_exact:
        broad_exact_candidates["candidate_value"] = broad_exact_candidates.apply(_candidate_comparison_value, axis=1)
        broad_exact_candidates["value_diff"] = broad_exact_candidates["candidate_value"].apply(lambda v: _numeric_diff(old_value, v))
        broad_exact_candidates["value_support"] = broad_exact_candidates["candidate_value"].apply(lambda v: value_support_score(old_value, v))
        same_level = broad_exact_candidates[broad_exact_candidates["semantic_level"] == old_semantic_level].copy()
        if len(same_level) == 1:
            best = same_level.iloc[0]
            reason = "matched:exact_label_same_level_value_checked" if float(best.get("value_support", 0.0) or 0.0) > 0 else "matched:exact_label_same_level_value_diff"
            return best, "exact", 1.10, reason
        if len(broad_exact_candidates) == 1:
            best = broad_exact_candidates.iloc[0]
            reason = "matched:exact_label_unique_cross_level_value_checked" if float(best.get("value_support", 0.0) or 0.0) > 0 else "review:exact_label_cross_level_value_diff"
            return best, "exact", 1.00, reason

    # 0) Strict exact-label match even when values differ:
    exact_candidates = candidates[
        (candidates["exact_norm"] == old_exact)
        & (candidates["semantic_level"] == old_semantic_level)
    ].copy()
    if not exact_candidates.empty and old_exact:
        exact_candidates["candidate_value"] = exact_candidates.apply(_candidate_comparison_value, axis=1)
        exact_candidates["value_diff"] = exact_candidates["candidate_value"].apply(lambda v: _numeric_diff(old_value, v))
        exact_candidates["value_support"] = exact_candidates["candidate_value"].apply(lambda v: value_support_score(old_value, v))
        best = choose_unique_supported_candidate(exact_candidates, r_old, old_value)
        if best is not None:
            reason = "matched:exact_label_value_checked" if float(best.get("value_support", 0.0) or 0.0) > 0 else "matched:exact_label_value_diff"
            return best, "exact", 1.10, reason

    # 0) Safe identical/canonical match:
    old_safe_key = _canonical_match_key(r_old) if "_canonical_match_key" in globals() else normalize_for_exact_match(old_label)
    candidates["safe_match_key"] = candidates.apply(lambda r: _canonical_match_key(r) if "_canonical_match_key" in globals() else normalize_for_exact_match(r.get("label", "")), axis=1)
    safe_candidates = candidates[
        (candidates["safe_match_key"] == old_safe_key)
        & (candidates["semantic_level"] == old_semantic_level)
    ].copy()
    if not safe_candidates.empty and old_safe_key:
        # If there are multiple keys of the same level, only select if the value
        safe_candidates["candidate_value"] = safe_candidates.apply(_candidate_comparison_value, axis=1)
        safe_candidates["value_diff"] = safe_candidates["candidate_value"].apply(lambda v: _numeric_diff(old_value, v))
        safe_candidates["value_support"] = safe_candidates["candidate_value"].apply(lambda v: value_support_score(old_value, v))
        best = choose_unique_supported_candidate(safe_candidates, r_old, old_value)
        if best is not None:
            reason = "matched:same_canonical_or_label_value_checked" if float(best.get("value_support", 0.0) or 0.0) > 0 else "review:same_canonical_without_value_support"
            return best, "canonical", 1.05, reason

    # 0B) Broad canonical match if unique:
    broad_canonical_candidates = candidates[
        (candidates["safe_match_key"] == old_safe_key)
    ].copy()
    if not broad_canonical_candidates.empty and old_safe_key:
        broad_canonical_candidates["candidate_value"] = broad_canonical_candidates.apply(_candidate_comparison_value, axis=1)
        broad_canonical_candidates["value_diff"] = broad_canonical_candidates["candidate_value"].apply(lambda v: _numeric_diff(old_value, v))
        broad_canonical_candidates["value_support"] = broad_canonical_candidates["candidate_value"].apply(lambda v: value_support_score(old_value, v))
        if len(broad_canonical_candidates) == 1:
            best = broad_canonical_candidates.iloc[0]
            reason = "matched:same_canonical_unique_cross_level_value_checked" if float(best.get("value_support", 0.0) or 0.0) > 0 else "review:same_canonical_cross_level_value_diff"
            return best, "canonical", 1.00, reason

    # 0) Canonical-first match:
    old_canonical_match_key = _canonical_match_key(r_old)
    candidates["canonical_match_key"] = candidates.apply(_canonical_match_key, axis=1)
    canonical_candidates = candidates[
        (candidates["canonical_match_key"] == old_canonical_match_key)
        & (candidates.apply(lambda r: _same_canonical_match_allowed(r_old, r), axis=1))
    ].copy()
    if not canonical_candidates.empty and old_canonical_match_key:
        canonical_candidates["candidate_value"] = canonical_candidates.apply(_candidate_comparison_value, axis=1)
        canonical_candidates["value_diff"] = canonical_candidates["candidate_value"].apply(lambda v: _numeric_diff(old_value, v))
        canonical_candidates["value_support"] = canonical_candidates["candidate_value"].apply(lambda v: value_support_score(old_value, v))
        canonical_candidates["same_block_bonus"] = canonical_candidates.get("block_type", pd.Series("", index=canonical_candidates.index)).fillna("").astype(str).apply(lambda x: 0.10 if x == old_block_type else 0.0)
        canonical_candidates["same_semantic_bonus"] = canonical_candidates["semantic_level"].apply(lambda x: 0.10 if x == old_semantic_level else 0.0)

        best = choose_unique_supported_candidate(canonical_candidates, r_old, old_value)
        if best is not None:
            reason = "matched:same_canonical_value_checked" if float(best.get("value_support", 0.0) or 0.0) > 0 else "review:same_canonical_without_value_support"
            return best, "canonical", 1.05, reason

    # 0A) Strict compact-label and value-supported match:
    compact_value_candidates = candidates[
        (candidates["compact_norm"] == old_compact)
    ].copy()
    if not compact_value_candidates.empty and old_compact:
        compact_value_candidates["candidate_value"] = compact_value_candidates.apply(_candidate_comparison_value, axis=1)
        compact_value_candidates["value_diff"] = compact_value_candidates["candidate_value"].apply(lambda v: _numeric_diff(old_value, v))
        compact_value_candidates["value_support"] = compact_value_candidates["candidate_value"].apply(lambda v: value_support_score(old_value, v))
        compact_value_candidates = compact_value_candidates[compact_value_candidates["value_support"] > 0].copy()
        if not compact_value_candidates.empty:
            compact_value_candidates["summary_bonus"] = compact_value_candidates["summary_row"].fillna(False).apply(lambda x: 0.10 if bool(x) == bool(r_old.get("summary_row")) else 0.0)
            compact_value_candidates["rowtype_bonus"] = compact_value_candidates["row_type"].fillna("").apply(lambda x: 0.05 if str(x) == str(r_old.get("row_type", "")) else 0.0)
            compact_value_candidates["structure_bonus"] = compact_value_candidates.get("structure", pd.Series("", index=compact_value_candidates.index)).fillna("").astype(str).apply(lambda x: 0.10 if x != "single-column" else 0.0)
            compact_value_candidates["candidate_total"] = (
                1.10
                + compact_value_candidates["value_support"]
                + compact_value_candidates["summary_bonus"]
                + compact_value_candidates["rowtype_bonus"]
                + compact_value_candidates["structure_bonus"]
                + compact_value_candidates["block_match_bonus"]
                + compact_value_candidates.get("forced_match_bonus", pd.Series(0.0, index=compact_value_candidates.index))
            )
            compact_value_candidates = compact_value_candidates.sort_values(
                by=["value_support", "candidate_total", "value_diff"],
                ascending=[False, False, True],
                na_position="last",
            )
            best = compact_value_candidates.iloc[0]
            return best, "compact_value_match", 0.98, "matched:compact_label_and_value"

    # 0) Strong exact match: same normalized text + same value wins immediately.
    strong_exact_candidates = candidates[
        candidates.apply(lambda r: strong_exact_match(r_old, r), axis=1)
    ].copy()
    if not strong_exact_candidates.empty:
        if old_block_type == "main_statement" and (strong_exact_candidates.get("block_type", pd.Series("", index=strong_exact_candidates.index)) == "main_statement").any():
            strong_exact_candidates = strong_exact_candidates[strong_exact_candidates.get("block_type", pd.Series("", index=strong_exact_candidates.index)) == "main_statement"].copy()
        strong_exact_candidates["summary_bonus"] = strong_exact_candidates["summary_row"].fillna(False).apply(lambda x: 0.10 if bool(x) == bool(r_old.get("summary_row")) else 0.0)
        strong_exact_candidates["rowtype_bonus"] = strong_exact_candidates["row_type"].fillna("").apply(lambda x: 0.05 if str(x) == str(r_old.get("row_type", "")) else 0.0)
        strong_exact_candidates["structure_bonus"] = strong_exact_candidates.get("structure", pd.Series("", index=strong_exact_candidates.index)).fillna("").astype(str).apply(lambda x: 0.20 if x != "single-column" else 0.0)
        strong_exact_candidates["candidate_value"] = strong_exact_candidates.apply(_candidate_comparison_value, axis=1)
        strong_exact_candidates["value_support"] = strong_exact_candidates["candidate_value"].apply(lambda v: value_support_score(old_value, v))
        strong_exact_candidates["value_diff"] = strong_exact_candidates["candidate_value"].apply(lambda v: _numeric_diff(old_value, v))
        strong_exact_candidates["candidate_total"] = 1.20 + strong_exact_candidates["value_support"] + strong_exact_candidates["summary_bonus"] + strong_exact_candidates["rowtype_bonus"] + strong_exact_candidates["structure_bonus"] + strong_exact_candidates["block_match_bonus"] + strong_exact_candidates.get("forced_match_bonus", pd.Series(0.0, index=strong_exact_candidates.index))
        best = choose_unique_supported_candidate(strong_exact_candidates, r_old, old_value)
        if best is not None:
            return best, "exact_value_match", 1.20, "matched:exact_text_and_value"

    # 1) Exact same normalized label always gets first priority.
    exact_candidates = candidates[
        (candidates["exact_norm"] == old_exact)
    ].copy()
    if not exact_candidates.empty:
        if old_block_type == "main_statement" and (exact_candidates.get("block_type", pd.Series("", index=exact_candidates.index)) == "main_statement").any():
            exact_candidates = exact_candidates[exact_candidates.get("block_type", pd.Series("", index=exact_candidates.index)) == "main_statement"].copy()
        exact_candidates["candidate_value"] = exact_candidates.apply(_candidate_comparison_value, axis=1)
        exact_candidates["value_diff"] = exact_candidates["candidate_value"].apply(lambda v: _numeric_diff(old_value, v))
        exact_candidates["value_support"] = exact_candidates["candidate_value"].apply(lambda v: value_support_score(old_value, v))
        exact_candidates["summary_bonus"] = exact_candidates["summary_row"].fillna(False).apply(lambda x: 0.10 if bool(x) == bool(r_old.get("summary_row")) else 0.0)
        exact_candidates["rowtype_bonus"] = exact_candidates["row_type"].fillna("").apply(lambda x: 0.05 if str(x) == str(r_old.get("row_type", "")) else 0.0)
        exact_candidates["candidate_total"] = 1.00 + exact_candidates["value_support"] + exact_candidates["summary_bonus"] + exact_candidates["rowtype_bonus"] + exact_candidates["block_match_bonus"] + exact_candidates.get("forced_match_bonus", pd.Series(0.0, index=exact_candidates.index))
        best = choose_unique_supported_candidate(exact_candidates, r_old, old_value)
        if best is not None:
            return best, "exact", 1.00, ("matched:exact_value_supported" if float(best.get("value_support", 0.0) or 0.0) > 0 else "review:exact_label_without_value_support")

    # 2) Compact same label can be accepted, but only if values support the pair.
    compact_candidates = candidates[
        (candidates["compact_norm"] == old_compact)
    ].copy()
    if not compact_candidates.empty:
        if old_block_type == "main_statement" and (compact_candidates.get("block_type", pd.Series("", index=compact_candidates.index)) == "main_statement").any():
            compact_candidates = compact_candidates[compact_candidates.get("block_type", pd.Series("", index=compact_candidates.index)) == "main_statement"].copy()
        compact_candidates["candidate_value"] = compact_candidates.apply(_candidate_comparison_value, axis=1)
        compact_candidates["value_support"] = compact_candidates["candidate_value"].apply(lambda v: value_support_score(old_value, v))
        compact_candidates = compact_candidates[compact_candidates["value_support"] > 0].copy()
        if not compact_candidates.empty:
            compact_candidates["summary_bonus"] = compact_candidates["summary_row"].fillna(False).apply(lambda x: 0.10 if bool(x) == bool(r_old.get("summary_row")) else 0.0)
            compact_candidates["rowtype_bonus"] = compact_candidates["row_type"].fillna("").apply(lambda x: 0.05 if str(x) == str(r_old.get("row_type", "")) else 0.0)
            compact_candidates["candidate_total"] = 0.95 + compact_candidates["value_support"] + compact_candidates["summary_bonus"] + compact_candidates["rowtype_bonus"] + compact_candidates["block_match_bonus"] + compact_candidates.get("forced_match_bonus", pd.Series(0.0, index=compact_candidates.index))
            compact_candidates = compact_candidates.sort_values(by=["block_match_bonus", "candidate_total"], ascending=False)
            best = compact_candidates.iloc[0]
            return best, "compact", 0.95, "matched:compact_value_supported"

    # 3) Synonym/fuzzy candidates are allowed only when values also support the match.
    best_idx = None
    best_total = -1.0
    best_method = None
    best_label_score = 0.0

    for idx, r_new in candidates.iterrows():
        candidate_block_type = str(r_new.get("block_type", "main_statement") or "main_statement")

        label_score, method = label_similarity_score(old_label, r_new["label"])
        if method in {"none", "exact", "compact"}:
            continue

        value_support = value_support_score(old_value, _candidate_comparison_value(r_new))
        if value_support <= 0:
            continue

        if not _fuzzy_matching_allowed_for_pair(r_old, r_new):
            continue

        if not _match_levels_compatible(r_old, r_new, exact_label=False, label_score=label_score):
            continue

        if old_block_type == "main_statement" and candidate_block_type == "statement_detail" and label_score < 0.92:
            continue

        if not total_context_compatible(r_old, r_new):
            continue

        total_score = label_score + value_support + float(r_new.get("block_match_bonus", 0.0) or 0.0)
        if bool(r_old.get("summary_row")) == bool(r_new.get("summary_row")):
            total_score += 0.10
        if str(r_old.get("row_type", "")) == str(r_new.get("row_type", "")):
            total_score += 0.05
        if bool(r_old.get("hierarchy_status") in {"Looginen", "Header row", "Rakenteellinen"}) == bool(r_new.get("hierarchy_status") in {"Looginen", "Header row", "Rakenteellinen"}):
            total_score += 0.03
        total_score += float(r_new.get("label_match_score", 0.0) or 0.0) * 0.02

        if total_score > best_total:
            best_total = total_score
            best_idx = idx
            best_method = method
            best_label_score = label_score

    if best_idx is None:
        return None, None, 0.0, "fail:no_value_supported_candidate"

    min_threshold = 0.95 if best_method == "synonym" else 0.90
    if best_total < min_threshold:
        return None, None, 0.0, f"fail:below_threshold:{best_method}:{best_total:.2f}"

    return candidates.loc[best_idx], best_method, float(best_label_score), f"matched:{best_method}_value_supported"

# ---------------------------------------------------------
# MANUAL REVIEW GATE AND VERIFIED-MATCH UPGRADE
# ---------------------------------------------------------

def diagnose_match_failure(r_old, df_new: pd.DataFrame, used_new: set) -> str:
    """Diagnose match failure.
    
    Purpose: This function belongs to the reconciliation and candidate evaluation stage.
    Why: It supports controlled matching while keeping uncertain cases visible for manual review.
    """
    available = df_new.copy()
    old_section_norm = normalize_section_for_matching(r_old.get("section"))
    available["section_match_norm"] = available["section"].apply(normalize_section_for_matching)

    if available.empty:
        return "fail:no_new_rows"

    same_section_all = available[available["section_match_norm"] == old_section_norm].copy()
    if same_section_all.empty:
        return "fail:section_mismatch"

    same_section_unused = same_section_all[~same_section_all.index.isin(used_new)].copy()
    if same_section_unused.empty:
        return "fail:all_section_candidates_used"

    old_exact = normalize_for_exact_match(r_old.get("label"))
    old_compact = normalize_for_compact_match(r_old.get("label"))
    old_value = r_old.get("match_value") if "match_value" in getattr(r_old, "index", []) else r_old.get("current_value")

    for frame in (same_section_all, same_section_unused, available):
        frame["exact_norm"] = frame["label"].apply(normalize_for_exact_match)
        frame["compact_norm"] = frame["label"].apply(normalize_for_compact_match)
        frame["candidate_value"] = frame.apply(_candidate_comparison_value, axis=1)
        frame["value_support"] = frame["candidate_value"].apply(lambda v: value_support_score(old_value, v))

    exact_all = same_section_all[same_section_all["exact_norm"] == old_exact]
    exact_unused = same_section_unused[same_section_unused["exact_norm"] == old_exact]
    if not exact_all.empty:
        if exact_unused.empty:
            return "fail:exact_candidate_already_used"
        if (exact_unused["value_support"] > 0).any():
            return "fail:exact_candidate_not_selected_even_with_value_support"
        return "fail:exact_label_found_but_value_not_supporting"

    compact_all = same_section_all[same_section_all["compact_norm"] == old_compact]
    compact_unused = same_section_unused[same_section_unused["compact_norm"] == old_compact]
    if not compact_all.empty:
        if compact_unused.empty:
            return "fail:compact_candidate_already_used"
        if (compact_unused["value_support"] > 0).any():
            return "fail:compact_candidate_not_selected_even_with_value_support"
        return "fail:compact_label_found_but_value_not_supporting"

    same_section_unused["label_ratio"] = same_section_unused["label"].apply(lambda x: fuzzy_ratio(r_old.get("label"), x))
    best_ratio = float(same_section_unused["label_ratio"].max()) if not same_section_unused.empty else 0.0
    best_value_support = float(same_section_unused["value_support"].max()) if not same_section_unused.empty else 0.0
    if best_ratio >= 0.85 and best_value_support <= 0:
        return f"fail:similar_label_but_value_not_supporting:{best_ratio:.2f}"
    if best_ratio >= 0.85:
        return f"fail:fuzzy_candidate_not_selected:{best_ratio:.2f}"

    if "block_type" in same_section_unused.columns:
        same_block = same_section_unused[same_section_unused["block_type"].fillna("main_statement").astype(str) == str(r_old.get("block_type", "main_statement"))]
        if same_block.empty and not same_section_unused.empty:
            return "fail:candidate_found_only_in_other_block_type"

    global_exact = available[available["exact_norm"] == old_exact]
    if not global_exact.empty:
        return "fail:exact_label_found_in_other_section"

    return "fail:no_plausible_candidate"


def normalize_exception_category(value: Optional[str]) -> str:
    
    raw = str(value or "").strip()
    key = normalize_keyword_text(raw)
    mapping = {
        "": "No exception",
        "no exception": "No exception",
        "ei luokiteltu": "No exception",
        "rivi puuttuu": "Missing row",
        "puuttuva rivi": "Missing row",
        "value differs": "Value differs",
        "ei voitu varmistaa": "Could not verify",
        "presentation change": "Esitystapaa muutettu",
        "mahdollinen esitystavan muutos": "Esitystapaa muutettu",
        "fallback tulkinta": "Could not verify",
        "fallback-tulkinta": "Could not verify",
        "osiorajauksen tarkistus": "Could not verify",
    }
    return mapping.get(key, raw[:1].upper() + raw[1:] if raw else "No exception")


def classify_exception_category(status: str, match_reason: Optional[str], repaired_issue: bool, fallback_used: bool, difference, match_method: Optional[str] = None) -> str:
    
    method = str(match_method or "")
    reason = str(match_reason or "")

    if status == "Value differs":
        return "Value differs"
    if status == "Missing row":
        if method == "unmatched_newer" or "missing_from_older" in reason:
            return "Missing from comparative-period financial statement"
        return "Missing from current-period financial statement"
    if status == "Could not verify":
        return "Could not verify"
    if fallback_used:
        return "Could not verify"
    if "section" in reason:
        return "Could not verify"
    return "No exception"


def _row_get(row, key: str, default=None):
    """Safe helper function for reading values ​​from a pandas row."""
    try:
        value = row.get(key, default)
        if value is None:
            return default
        if pd.isna(value):
            return default
        return value
    except Exception:
        return default


def is_strong_verified_match_row(row) -> bool:
   
    return is_clear_verified_match_decision(row)

# ---------------------------------------------------------
# FINAL ROW-BY-ROW RECONCILIATION
# ---------------------------------------------------------


def manual_review_reason_for_row(row) -> Optional[str]:
    """Return a manual-review reason for any non-verified automated decision."""
    status = str(_row_get(row, "status", "") or "")

    if status not in {"Match", "Could not verify"}:
        return None

    if is_strong_verified_match_row(row):
        return None

    reasons = []

    old_section = str(_row_get(row, "section_match_norm_older", "") or "")
    new_section = str(_row_get(row, "section_match_norm_newer", "") or "")
    if old_section and new_section and old_section != new_section:
        reasons.append("the item was found in a different section")

    old_can = str(_row_get(row, "label_match_canonical_older", "") or "")
    new_can = str(_row_get(row, "label_match_canonical_newer", "") or "")
    if not old_can or not new_can:
        reasons.append("canonical support is missing for at least one side")
    elif not canonical_values_equivalent(old_can, new_can):
        reasons.append("canonical-nimikkeet ovat ristiriidassa")

    if bool(_row_get(row, "context_conflict", False)):
        reasons.append("kontekstissa on ristiriita")
    if bool(_row_get(row, "context_uncertain", False)):
        reasons.append("PMA context is uncertain")

    hierarchy_status = str(_row_get(row, "hierarchy_status", "") or "")
    if hierarchy_status == "Ristiriitainen":
        reasons.append("rakennepolku on ristiriitainen")
    elif hierarchy_status == "Uncertain":
        reasons.append("rakennepolku on epävarma")

    context_status_old = str(_row_get(row, "canonical_context_status_older", "") or "")
    context_status_new = str(_row_get(row, "canonical_context_status_newer", "") or "")
    if context_status_old == "Ristiriitainen" or context_status_new == "Ristiriitainen":
        reasons.append("the item section or parent metadata is inconsistent")
    elif context_status_old == "Uncertain" or context_status_new == "Uncertain":
        reasons.append("the item section or parent metadata is uncertain")

    old_status = str(_row_get(row, "value_status_older", "") or "")
    new_status = str(_row_get(row, "value_status_newer", "") or "")
    if old_status == "missing" or new_status == "missing":
        reasons.append("vertailuarvo puuttuu")
    elif old_status == "uncertain" or new_status == "uncertain":
        reasons.append("value extraction is genuinely uncertain")

    if bool(_row_get(row, "numeric_integrity_issue", False)) or bool(_row_get(row, "repaired_issue", False)):
        reasons.append("numeric extraction or presentation required repair or interpretation")

    if bool(_row_get(row, "fallback_used", False)):
        reasons.append("fallback interpretation was used")

    method = str(_row_get(row, "match_method", "") or "")
    reason = str(_row_get(row, "match_reason", "") or "")
    if reason.startswith("review:"):
        reasons.append("the candidate was found, but the evidence is not strong enough for automatic acceptance")
    if method in {"fuzzy", "fuzzy_strong", "synonym"}:
        reasons.append("the match is based on a soft label match")

    if status == "Match" and not _match_reason_has_value_evidence(reason):
        reasons.append("accepted match lacks explicit value-supported evidence")

    if status == "Match" and not _values_match_within_tolerance(row):
        reasons.append("numeric equality could not be independently verified")

    confidence = str(_row_get(row, "confidence_level", "") or "")
    if confidence in {"Low", "Medium", "Uncertain"}:
        reasons.append(f"confidence level is {confidence}")

    if status == "Could not verify" and not reasons:
        reasons.append("the automated decision remained uncertain")

    if status == "Match" and not reasons:
        reasons.append("match was not strong enough for automatic acceptance")

    return "; ".join(dict.fromkeys(reasons)) if reasons else None


def upgrade_clear_verified_rows(df: pd.DataFrame) -> pd.DataFrame:
  
    if df is None or df.empty:
        return df

    out = df.copy()
    strong_mask = out.apply(is_strong_verified_match_row, axis=1)

    if "confidence_level" in out.columns:
        out.loc[strong_mask, "confidence_level"] = "High"

    if "manual_review_required" in out.columns:
        out.loc[strong_mask, "manual_review_required"] = False

    if "manual_review_reason" in out.columns:
        out.loc[strong_mask, "manual_review_reason"] = None

    if "status" in out.columns:
        out.loc[strong_mask, "status"] = "Match"

    if "severity" in out.columns:
        out.loc[strong_mask, "severity"] = "No exception"

    if "review_priority" in out.columns:
        out.loc[strong_mask, "review_priority"] = 4

    if "match_level" in out.columns:
        out.loc[strong_mask, "match_level"] = MATCH_VERIFIED

    return out


def apply_manual_review_gate(df: pd.DataFrame) -> pd.DataFrame:
    
    if df is None or df.empty:
        return df

    out = df.copy()

    if "original_status" not in out.columns:
        out["original_status"] = out.get("status")

    out = upgrade_clear_verified_rows(out)

    out["manual_review_reason"] = out.apply(manual_review_reason_for_row, axis=1)

    strong_mask = out.apply(is_strong_verified_match_row, axis=1)
    out.loc[strong_mask, "manual_review_reason"] = None

    out["manual_review_required"] = out["manual_review_reason"].fillna("").astype(str).str.len() > 0

    mask = out["manual_review_required"] & out["status"].isin(["Match", "Could not verify"])
    out.loc[mask, "status"] = "Manual review"
    out.loc[mask, "severity"] = "Tarkistettava"
    out.loc[mask, "review_priority"] = 3
    out.loc[mask, "selite"] = out.loc[mask].apply(
        lambda r: f"Manual review required: {r.get('manual_review_reason')}. Original automated decision: {r.get('original_status')}.",
        axis=1,
    )
    out.loc[mask, "exception_category"] = out.loc[mask, "exception_category"].fillna("uncertain automated interpretation")
    out.loc[mask & (out["exception_category"].astype(str).str.strip() == ""), "exception_category"] = "uncertain automated interpretation"

    out = upgrade_clear_verified_rows(out)

    return out


def repair_reciprocal_missing_exact_rows(df: pd.DataFrame, tolerance: float = TOLERANCE) -> pd.DataFrame:
    
    if df is None or df.empty:
        return df

    required = {"status", "section", "label_older", "label_newer", "value_older_current", "value_newer_comparison"}
    if not required.issubset(df.columns):
        return df

    out = df.copy()

    def _missing(v) -> bool:
       
        if v is None:
            return True
        try:
            if pd.isna(v):
                return True
        except Exception:
            pass
        return str(v).strip().lower() in {"", "nan", "none", "nat"}

    def _present(v) -> bool:
        
        return not _missing(v)

    def _first_present(*values):
        
        for value in values:
            if _present(value):
                return value
        return None

    def _present_label(row) -> str:
        """Present label.
        
        Purpose: This function belongs to the normalization and canonical recognition stage.
        Why: It converts inconsistent financial statement wording into comparable internal concepts.
        """
        return str(_first_present(
            row.get("label_older"), row.get("label_newer"),
            row.get("label_match_canonical_older"), row.get("label_match_canonical_newer"),
            row.get("forced_main_item_canonical_older"), row.get("forced_main_item_canonical_newer"),
            row.get("label_norm_older"), row.get("label_norm_newer"),
        ) or "")

    def _section_key(row) -> str:
        """Section key.
        
        Purpose: This function belongs to the financial statement structure stage.
        Why: It prevents rows from being compared across incompatible statement sections or hierarchy levels.
        """
        old_sec = row.get("section_match_norm_older") if "section_match_norm_older" in row.index else None
        new_sec = row.get("section_match_norm_newer") if "section_match_norm_newer" in row.index else None
        return normalize_section_for_matching(_first_present(old_sec, new_sec, row.get("section")))

    def _row_keys(row) -> set[tuple[str, str, str]]:
        
        section = _section_key(row)
        values = [_present_label(row)]
        for col in [
            "label_match_canonical_older", "label_match_canonical_newer",
            "forced_main_item_canonical_older", "forced_main_item_canonical_newer",
            "label_norm_older", "label_norm_newer",
        ]:
            if col in row.index and _present(row.get(col)):
                values.append(str(row.get(col)))

        keys: set[tuple[str, str, str]] = set()
        for value in values:
            if not _present(value):
                continue
            exact = normalize_for_exact_match(str(value))
            compact = normalize_for_compact_match(str(value))
            canonical = normalize_for_exact_match(_safe_structural_canonical_from_label(str(value), section))
            if exact:
                keys.add((section, "exact", exact))
            if compact:
                keys.add((section, "compact", compact))
            if canonical:
                keys.add((section, "canonical", canonical))
        return keys

    def _keys_overlap(row_a, row_b) -> bool:
        
        keys_a = _row_keys(row_a)
        keys_b = _row_keys(row_b)
        return bool(keys_a and keys_b and keys_a.intersection(keys_b))

    def _context_compatible(row_a, row_b) -> bool:
       
        try:
            label_a = _present_label(row_a)
            label_b = _present_label(row_b)
            if is_generic_total_label(label_a) or is_generic_total_label(label_b):
                key_a = _first_present(row_a.get("total_context_key_older"), row_a.get("total_context_key_newer"))
                key_b = _first_present(row_b.get("total_context_key_older"), row_b.get("total_context_key_newer"))
                if _present(key_a) and _present(key_b) and key_a != key_b:
                    return False
        except Exception:
            pass
        return True

    def _value_diff_for_pair(row_newer_only, row_older_only):
        """Value diff for pair.
        
        Purpose: This function belongs to the number parsing and validation stage.
        Why: It reduces the risk that formatting differences or unreadable values are treated as reliable evidence.
        """
        old_value = row_older_only.get("value_older_current")
        new_value = row_newer_only.get("value_newer_comparison")
        try:
            if old_value is not None and new_value is not None and pd.notna(old_value) and pd.notna(new_value):
                return abs(float(new_value) - float(old_value))
        except Exception:
            pass
        return None

    missing = out[out["status"].astype(str).eq("Missing row")].copy()
    newer_only = missing[missing["label_older"].apply(_missing) & missing["label_newer"].apply(_present)].copy()
    older_only = missing[missing["label_older"].apply(_present) & missing["label_newer"].apply(_missing)].copy()

    if newer_only.empty or older_only.empty:
        return out

    used_old_only = set()
    rows_to_drop = set()

    for idx_newer, row_newer in newer_only.iterrows():
        candidates = []
        for idx_older, row_older in older_only.iterrows():
            if idx_older in used_old_only:
                continue
            if _section_key(row_newer) != _section_key(row_older):
                continue
            if not _keys_overlap(row_newer, row_older):
                continue
            if not _context_compatible(row_newer, row_older):
                continue
            candidates.append((idx_older, row_older, _value_diff_for_pair(row_newer, row_older)))

        if not candidates:
            continue
        if len(candidates) > 1:
            numeric_candidates = [c for c in candidates if c[2] is not None]
            if numeric_candidates:
                numeric_candidates.sort(key=lambda c: c[2])
                if len(numeric_candidates) > 1 and abs(numeric_candidates[0][2] - numeric_candidates[1][2]) <= tolerance:
                    continue
                idx_older, row_older, _ = numeric_candidates[0]
            else:
                continue
        else:
            idx_older, row_older, _ = candidates[0]

        used_old_only.add(idx_older)
        rows_to_drop.add(idx_older)

        old_value = row_older.get("value_older_current")
        new_value = row_newer.get("value_newer_comparison")
        difference = None
        status = "Could not verify"
        try:
            if old_value is not None and new_value is not None and pd.notna(old_value) and pd.notna(new_value):
                difference = float(new_value) - float(old_value)
                status = "Match" if abs(difference) <= tolerance else "Value differs"
        except Exception:
            status = "Could not verify"

        label = _first_present(row_newer.get("label_newer"), row_older.get("label_older"), _present_label(row_newer), _present_label(row_older))
        label = clean_display_label(str(label or ""))

        out.at[idx_newer, "label_older"] = label
        out.at[idx_newer, "label_newer"] = label
        out.at[idx_newer, "value_older_current"] = old_value
        out.at[idx_newer, "value_newer_comparison"] = new_value
        out.at[idx_newer, "display_older_value"] = row_older.get("display_older_value")
        out.at[idx_newer, "display_newer_value"] = row_newer.get("display_newer_value")
        out.at[idx_newer, "difference"] = difference
        out.at[idx_newer, "status"] = status
        out.at[idx_newer, "match_method"] = "reciprocal_missing_label_repair"
        out.at[idx_newer, "match_reason"] = "matched:reciprocal_missing_rows_same_label_or_canonical"
        out.at[idx_newer, "match_level"] = MATCH_VERIFIED if status in {"Match", "Value differs"} else MATCH_SUGGESTED
        out.at[idx_newer, "label_match_score"] = max(float(row_newer.get("label_match_score", 0) or 0), float(row_older.get("label_match_score", 0) or 0), 1.0)
        out.at[idx_newer, "confidence_level"] = "High"
        out.at[idx_newer, "severity"] = "Tarkistettava" if status == "Value differs" else ("No exception" if status == "Match" else "Tarkistettava")
        out.at[idx_newer, "review_priority"] = 1 if status == "Value differs" else (4 if status == "Match" else 2)
        out.at[idx_newer, "selite"] = "The same item was found in both files. The previous two missing-row findings were merged into one reconciliation row."
        out.at[idx_newer, "exception_category"] = "Value differs" if status == "Value differs" else status

        for col in list(out.columns):
            if col.endswith("_older") and _missing(out.at[idx_newer, col]) and col in row_older.index:
                out.at[idx_newer, col] = row_older.get(col)

        if "label_norm_older" in out.columns:
            out.at[idx_newer, "label_norm_older"] = normalize_for_exact_match(label)
        if "label_norm_newer" in out.columns:
            out.at[idx_newer, "label_norm_newer"] = normalize_for_exact_match(label)
        if "section_match_norm_older" in out.columns and _missing(out.at[idx_newer, "section_match_norm_older"]):
            out.at[idx_newer, "section_match_norm_older"] = _section_key(row_older)
        if "section_match_norm_newer" in out.columns and _missing(out.at[idx_newer, "section_match_norm_newer"]):
            out.at[idx_newer, "section_match_norm_newer"] = _section_key(row_newer)

    if rows_to_drop:
        out = out.drop(index=list(rows_to_drop))
    return out

def compare_all_rows(parsed_older: dict, parsed_newer: dict, tolerance: float = TOLERANCE):
    """
    Reconcile all extracted rows between the older and newer financial statements.

    The function aligns both datasets to the common comparison period, searches
    for the best candidate match for each older row, classifies the result as a
    match, value difference, missing row or manual-review case, and preserves a
    detailed decision trace for each decision.
    """
    df_old = parsed_older["df_all"].copy()
    df_new = parsed_newer["df_all"].copy()
    older_dominant = parsed_older.get("dominant_number_format")
    newer_dominant = parsed_newer.get("dominant_number_format")

    comparison_target_periods_by_section = determine_common_period_keys_by_section(parsed_older, parsed_newer)
    df_old = add_match_values_for_common_period(df_old, comparison_target_periods_by_section, file_role="older")
    df_new = add_match_values_for_common_period(df_new, comparison_target_periods_by_section, file_role="newer")

    df_old["canonical_key"] = df_old.apply(lambda r: make_keys(r["label"], r["section"])["canonical"], axis=1)
    df_new["canonical_key"] = df_new.apply(lambda r: make_keys(r["label"], r["section"])["canonical"], axis=1)
    df_old["compact_key"] = df_old.apply(lambda r: make_keys(r["label"], r["section"])["compact"], axis=1)
    df_new["compact_key"] = df_new.apply(lambda r: make_keys(r["label"], r["section"])["compact"], axis=1)

    used_new = set()
    rows = []

    for _, r_old in df_old.iterrows():
        if r_old.get("row_type") == "noise":
            continue

        r_new, match_method, label_match_score, match_reason = find_best_row_match(r_old, df_new, used_new)
        if r_new is not None:
            used_new.add(r_new.name)
        else:
            match_reason = diagnose_match_failure(r_old, df_new, used_new)

        label_old = r_old["label"]
        label_new = r_new["label"] if r_new is not None else None
        value_old = r_old.get("match_value") if "match_value" in getattr(r_old, "index", []) else r_old["current_value"]
        raw_old = r_old.get("match_value_raw") if "match_value_raw" in getattr(r_old, "index", []) else r_old["current_value_raw_original"]
        value_status_old = r_old.get("match_value_status") if "match_value_status" in getattr(r_old, "index", []) else r_old.get("current_value_status", classify_value_parse_status(raw_old, bool(r_old.get("current_value_was_repaired", False)), r_old.get("current_value_repair_type")))

        value_new = None
        raw_new = None
        new_repaired = False
        if r_new is not None:
            value_new = _candidate_comparison_value(r_new)
            raw_new = _candidate_comparison_raw(r_new)
            new_repaired = _candidate_comparison_repaired(r_new)
            value_status_new = _candidate_comparison_status(r_new)
        else:
            value_status_new = "missing"

        repaired_issue = bool(r_old["current_value_was_repaired"]) or new_repaired
        separator_issue = False
        consistency_issue = False
        decimal_presentation_issue = False
        difference = None
        uncertain_reason = None

        if r_new is None:
            status = "Missing row"
            label_match_score = float(r_old.get("label_match_score", 0.0) or 0.0)
        else:
            if value_old is None or value_new is None:
                status = "Could not verify"
                uncertain_reason = "Vertailuarvo puuttuu toisesta sarakkeesta"
            else:
                difference = value_new - value_old
                separator_issue = repaired_issue or (classify_number_format(str(raw_old)) != classify_number_format(str(raw_new)))
                if only_thousands_separator_diff(raw_old, raw_new):
                    separator_issue = False
                consistency_issue = document_format_issue(raw_old, older_dominant) or document_format_issue(raw_new, newer_dominant)
                if only_thousands_separator_diff(raw_old, raw_new):
                    consistency_issue = False
                old_raw_numeric = parse_original_display_number(raw_old)
                new_raw_numeric = parse_original_display_number(raw_new)
                decimal_presentation_issue = has_decimal_presentation_defect(raw_old, older_dominant) or has_decimal_presentation_defect(raw_new, newer_dominant)
                review_unsafe_numeric_issue = (
                    has_review_unsafe_numeric_repair(raw_old, bool(r_old.get("match_value_was_repaired", r_old.get("current_value_was_repaired", False))), older_dominant)
                    or has_review_unsafe_numeric_repair(raw_new, new_repaired, newer_dominant)
                )
                value_status_issue = not (value_status_is_safe_for_verified_match(value_status_old) and value_status_is_safe_for_verified_match(value_status_new))
                if abs(difference) <= tolerance:
                    strong_value_match_method = match_method in {
                        "canonical", "exact", "exact_value_match", "compact", "compact_value_match"
                    }
                    if value_status_old in {"missing", "uncertain"} or value_status_new in {"missing", "uncertain"}:
                        status = "Could not verify"
                        uncertain_reason = "The comparative value is missing or genuinely uncertain"
                    elif review_unsafe_numeric_issue and not strong_value_match_method:
                        status = "Could not verify"
                        uncertain_reason = "Numeropoiminta vaatii tarkistuksen"
                    else:
                        status = "Match"
                else:
                    status = "Value differs"

        fallback_used_row = bool(str(r_old.get("match_value_source", "")).startswith("fallback")) or bool(r_new is not None and str(r_new.get("match_value_source", "")).startswith("fallback"))

        combined_label_score = max(
            float(r_old.get("label_match_score", 0.0) or 0.0),
            float(r_new.get("label_match_score", 0.0) or 0.0) if r_new is not None else 0.0,
            float(label_match_score or 0.0),
        )
        combined_hierarchy_score = max(
            float(r_old.get("hierarchy_score", 0.0) or 0.0),
            float(r_new.get("hierarchy_score", 0.0) or 0.0) if r_new is not None else 0.0,
        )
        hierarchy_conflict = bool(r_old.get("hierarchy_status") == "Ristiriitainen") or bool(r_new is not None and r_new.get("hierarchy_status") == "Ristiriitainen")
        context_conflict = bool(r_old.get("canonical_context_status") == "Ristiriitainen") or bool(r_new is not None and r_new.get("canonical_context_status") == "Ristiriitainen")
        context_uncertain = bool(r_old.get("canonical_context_status") == "Uncertain") or bool(r_new is not None and r_new.get("canonical_context_status") == "Uncertain")
        if context_conflict:
            hierarchy_conflict = True
        if hierarchy_conflict:
            combined_hierarchy_status = "Ristiriitainen"
        elif bool(r_old.get("hierarchy_status") in {"Looginen", "Header row", "Rakenteellinen"}) or bool(r_new is not None and r_new.get("hierarchy_status") in {"Looginen", "Header row", "Rakenteellinen"}):
            combined_hierarchy_status = "Looginen"
        elif bool(r_old.get("hierarchy_status") == "Uncertain") or bool(r_new is not None and r_new.get("hierarchy_status") == "Uncertain"):
            combined_hierarchy_status = "Uncertain"
        else:
            combined_hierarchy_status = r_old.get("hierarchy_status") or (r_new.get("hierarchy_status") if r_new is not None else None)
        confidence_level = determine_row_confidence(
            status, uncertain_reason, repaired_issue, separator_issue, consistency_issue, fallback_used_row,
            label_match_score=combined_label_score, match_method=match_method,
            hierarchy_score=combined_hierarchy_score, hierarchy_conflict=hierarchy_conflict,
            context_uncertain=context_uncertain, value_status_issue=value_status_issue,
            match_reason=match_reason,
        )
        severity = determine_row_severity(status, confidence_level, repaired_issue, separator_issue, consistency_issue, decimal_presentation_issue)
        selite = build_row_explanation(status, uncertain_reason, repaired_issue, separator_issue, consistency_issue, decimal_presentation_issue, difference)
        exception_category = classify_exception_category(status, match_reason, repaired_issue, fallback_used_row, difference, match_method)

        rows.append({
            "section": r_old["section"],
            "label_older": label_old,
            "label_newer": label_new,
            "value_older_current": value_old,
            "value_newer_comparison": value_new,
            "display_older_value": raw_old,
            "display_newer_value": raw_new,
            "difference": difference,
            "status": status,
            "summary_row": bool(r_old.get("summary_row", False)) or bool(r_new.get("summary_row", False)) if r_new is not None else bool(r_old.get("summary_row", False)),
            "severity": severity,
            "confidence_level": confidence_level,
            "review_priority": 2 if status != "Match" else 4,
            "selite": selite,
            "source_line_older": r_old.get("source_line"),
            "source_page_older": r_old.get("source_page"),
            "source_line_idx_older": r_old.get("source_line_idx"),
            "source_line_newer": r_new.get("source_line") if r_new is not None else None,
            "source_page_newer": r_new.get("source_page") if r_new is not None else None,
            "source_line_idx_newer": r_new.get("source_line_idx") if r_new is not None else None,
            "raw_older_original": raw_old,
            "raw_newer_original": raw_new,
            "raw_older_numeric_as_presented": parse_original_display_number(raw_old),
            "raw_newer_numeric_as_presented": parse_original_display_number(raw_new),
            "numeric_integrity_issue": bool(repaired_issue or decimal_presentation_issue),
            "value_status_older": value_status_old,
            "value_status_newer": value_status_new,
            "current_value_status_older": r_old.get("current_value_status"),
            "comparison_value_status_older": r_old.get("comparison_value_status"),
            "current_value_status_newer": r_new.get("current_value_status") if r_new is not None else None,
            "comparison_value_status_newer": r_new.get("comparison_value_status") if r_new is not None else None,
            "row_parse_quality_older": r_old.get("row_parse_quality"),
            "row_parse_quality_newer": r_new.get("row_parse_quality") if r_new is not None else None,
            "value_source_older": r_old.get("match_value_source"),
            "value_source_newer": r_new.get("match_value_source") if r_new is not None else None,
            "target_period_older": r_old.get("comparison_target_period_key"),
            "target_period_newer": r_new.get("comparison_target_period_key") if r_new is not None else None,
            "older_current_period_key": r_old.get("current_period_key"),
            "older_comparison_period_key": r_old.get("comparison_period_key"),
            "newer_current_period_key": r_new.get("current_period_key") if r_new is not None else None,
            "newer_comparison_period_key": r_new.get("comparison_period_key") if r_new is not None else None,
            "fallback_used": fallback_used_row,
            "exception_category": exception_category,
            "repaired_issue": repaired_issue,
            "match_method": match_method,
            "match_reason": match_reason,
            "match_level": classify_match_level_from_row({"status": status, "match_method": match_method, "match_reason": match_reason, "confidence_level": confidence_level}),
            "label_match_score": combined_label_score,
            "section_match_norm_older": normalize_section_for_matching(r_old.get("section")),
            "section_match_norm_newer": normalize_section_for_matching(r_new.get("section")) if r_new is not None else None,
            "block_type_older": r_old.get("block_type"),
            "block_type_newer": r_new.get("block_type") if r_new is not None else None,
            "label_norm_older": normalize_for_exact_match(label_old),
            "label_norm_newer": normalize_for_exact_match(label_new) if label_new is not None else None,
            "label_match_canonical_older": r_old.get("label_match_canonical"),
            "label_match_canonical_newer": r_new.get("label_match_canonical") if r_new is not None else None,
            "extraction_status_older": r_old.get("extraction_status"),
            "extraction_status_newer": r_new.get("extraction_status") if r_new is not None else None,
            "label_match_accepted_older": r_old.get("label_match_accepted"),
            "label_match_accepted_newer": r_new.get("label_match_accepted") if r_new is not None else None,
            "parent_label_older": r_old.get("parent_label"),
            "parent_label_newer": r_new.get("parent_label") if r_new is not None else None,
            "top_parent_label_older": r_old.get("top_parent_label"),
            "top_parent_label_newer": r_new.get("top_parent_label") if r_new is not None else None,
            "category_path_older": r_old.get("category_path"),
            "category_path_newer": r_new.get("category_path") if r_new is not None else None,
            "total_context_key_older": r_old.get("total_context_key") or total_context_key(r_old),
            "total_context_key_newer": (r_new.get("total_context_key") or total_context_key(r_new)) if r_new is not None else None,
            "legal_basis_level_older": r_old.get("legal_basis_level"),
            "legal_basis_level_newer": r_new.get("legal_basis_level") if r_new is not None else None,
            "forced_main_item_older": r_old.get("forced_main_item"),
            "forced_main_item_newer": r_new.get("forced_main_item") if r_new is not None else None,
            "forced_main_item_canonical_older": r_old.get("forced_main_item_canonical"),
            "forced_main_item_canonical_newer": r_new.get("forced_main_item_canonical") if r_new is not None else None,
            "forced_root_item_older": r_old.get("forced_root_item"),
            "forced_root_item_newer": r_new.get("forced_root_item") if r_new is not None else None,
            "balance_total_root_item_older": r_old.get("balance_total_root_item"),
            "balance_total_root_item_newer": r_new.get("balance_total_root_item") if r_new is not None else None,
            "total_level_older": r_old.get("total_level"),
            "total_level_newer": r_new.get("total_level") if r_new is not None else None,
            "legal_basis_match_older": r_old.get("legal_basis_match"),
            "legal_basis_match_newer": r_new.get("legal_basis_match") if r_new is not None else None,
            "canonical_expected_section_older": r_old.get("canonical_expected_section"),
            "canonical_expected_section_newer": r_new.get("canonical_expected_section") if r_new is not None else None,
            "canonical_allowed_sections_older": r_old.get("canonical_allowed_sections"),
            "canonical_allowed_sections_newer": r_new.get("canonical_allowed_sections") if r_new is not None else None,
            "canonical_expected_parent_older": r_old.get("canonical_expected_parent"),
            "canonical_expected_parent_newer": r_new.get("canonical_expected_parent") if r_new is not None else None,
            "canonical_expected_side_older": r_old.get("canonical_expected_side"),
            "canonical_expected_side_newer": r_new.get("canonical_expected_side") if r_new is not None else None,
            "canonical_item_type_older": r_old.get("canonical_item_type"),
            "canonical_item_type_newer": r_new.get("canonical_item_type") if r_new is not None else None,
            "canonical_context_status_older": r_old.get("canonical_context_status"),
            "canonical_context_status_newer": r_new.get("canonical_context_status") if r_new is not None else None,
            "canonical_context_reason_older": r_old.get("canonical_context_reason"),
            "canonical_context_reason_newer": r_new.get("canonical_context_reason") if r_new is not None else None,
            "canonical_context_score_older": r_old.get("canonical_context_score"),
            "canonical_context_score_newer": r_new.get("canonical_context_score") if r_new is not None else None,
            "context_conflict": context_conflict,
            "context_uncertain": context_uncertain,
            "hierarchy_score": combined_hierarchy_score,
            "hierarchy_status": combined_hierarchy_status,
        })

    unused_newer = df_new[~df_new.index.isin(used_new)].copy()
    for _, r_new in unused_newer.iterrows():
        if r_new.get("row_type") == "noise":
            continue
        value_new = _candidate_comparison_value(r_new)
        raw_new = _candidate_comparison_raw(r_new)
        rows.append({
            "section": r_new["section"],
            "label_older": None,
            "label_newer": r_new["label"],
            "value_older_current": None,
            "value_newer_comparison": value_new,
            "display_older_value": None,
            "display_newer_value": raw_new,
            "difference": None,
            "status": "Missing row",
            "summary_row": bool(r_new.get("summary_row", False)),
            "severity": "Tarkistettava",
            "confidence_level": "Low",
            "review_priority": 2,
            "selite": "The row exists only in the newer file",
            "source_line_older": None,
            "source_page_older": None,
            "source_line_idx_older": None,
            "source_line_newer": r_new.get("source_line"),
            "source_page_newer": r_new.get("source_page"),
            "source_line_idx_newer": r_new.get("source_line_idx"),
            "raw_older_original": None,
            "raw_newer_original": raw_new,
            "value_source_older": None,
            "value_source_newer": r_new.get("match_value_source"),
            "target_period_older": None,
            "target_period_newer": r_new.get("comparison_target_period_key"),
            "older_current_period_key": None,
            "older_comparison_period_key": None,
            "newer_current_period_key": r_new.get("current_period_key"),
            "newer_comparison_period_key": r_new.get("comparison_period_key"),
            "fallback_used": bool(str(r_new.get("match_value_source", "")).startswith("fallback")),
            "exception_category": "Missing from comparative-period financial statement",
            "repaired_issue": False,
            "match_method": "unmatched_newer",
            "match_reason": "fail:missing_from_older",
            "match_level": MATCH_REJECTED,
            "label_match_score": float(r_new.get("label_match_score", 0.0) or 0.0),
            "section_match_norm_older": None,
            "section_match_norm_newer": normalize_section_for_matching(r_new.get("section")),
            "label_norm_older": None,
            "label_norm_newer": normalize_for_exact_match(r_new.get("label")),
            "label_match_canonical_older": None,
            "label_match_canonical_newer": r_new.get("label_match_canonical"),
            "parent_label_older": None,
            "parent_label_newer": r_new.get("parent_label"),
            "top_parent_label_older": None,
            "top_parent_label_newer": r_new.get("top_parent_label"),
            "category_path_older": None,
            "category_path_newer": r_new.get("category_path"),
            "total_context_key_older": None,
            "total_context_key_newer": r_new.get("total_context_key") or total_context_key(r_new),
            "legal_basis_level_older": None,
            "legal_basis_level_newer": r_new.get("legal_basis_level"),
            "forced_main_item_older": None,
            "forced_main_item_newer": r_new.get("forced_main_item"),
            "forced_main_item_canonical_older": None,
            "forced_main_item_canonical_newer": r_new.get("forced_main_item_canonical"),
            "forced_root_item_older": None,
            "forced_root_item_newer": r_new.get("forced_root_item"),
            "balance_total_root_item_older": None,
            "balance_total_root_item_newer": r_new.get("balance_total_root_item"),
            "total_level_older": None,
            "total_level_newer": r_new.get("total_level"),
            "legal_basis_match_older": None,
            "legal_basis_match_newer": r_new.get("legal_basis_match"),
            "hierarchy_score": float(r_new.get("hierarchy_score", 0.0) or 0.0),
            "hierarchy_status": r_new.get("hierarchy_status"),
        })

    result_df = pd.DataFrame(rows)
    if not result_df.empty:
        result_df = apply_presentation_change_detection(result_df)
        result_df = repair_reciprocal_missing_exact_rows(result_df, tolerance=tolerance)
        result_df = enrich_missing_row_findings(result_df)
        result_df = apply_manual_review_gate(result_df)
        result_df = upgrade_clear_verified_rows(result_df)
        result_df = enrich_missing_row_findings(result_df)
        result_df = result_df.sort_values(by=["review_priority", "summary_row", "section", "label_newer", "label_older"], ascending=[True, False, True, True, True], na_position="last").reset_index(drop=True)
    return result_df


# =========================================================
# REPORTING, EXPLANATIONS AND REVIEW OUTPUT DATASETS
# =========================================================
# This section converts technical reconciliation results into user-facing
# review tables, summaries and decision-trace datasets.
#
# It builds:
# - Document and test summaries.
# - Human-readable match explanations.
# - Manual-review reasons.
# - Main display tables.
# - Decision trace and technical debug views.
# - Management dashboard metrics.
# - Key-item control summaries.
#
# The purpose is to make reconciliation results understandable, traceable and
# suitable for both user review and thesis testing documentation.


def build_test_summary(parsed_result) -> dict:
    
    df = parsed_result["df_all"]
    repaired_rows = int((df["current_value_was_repaired"].fillna(False) | df["comparison_value_was_repaired"].fillna(False)).sum()) if not df.empty and {"current_value_was_repaired", "comparison_value_was_repaired"}.issubset(df.columns) else 0
    fallback_rows = int((df["structure"].fillna("") == "fallback").sum()) if not df.empty and "structure" in df.columns else 0
    fallback_ratio = round((fallback_rows / len(df)) * 100, 1) if len(df) > 0 else 0
    reliability = parsed_result.get("document_reliability", {}) or {}
    return {
        "Asiakirja": parsed_result["document"],
        "Kokonaisluotettavuuspisteet": reliability.get("score"),
        "Use classification": reliability.get("classification"),
        "Extracted rows": len(df),
        "Income statement rows": len(parsed_result["df_income"]),
        "Balance sheet asset rows": len(parsed_result["df_vastaavaa"]),
        "Balance sheet liability rows": len(parsed_result["df_vastattavaa"]),
        "Detected total rows": int(df["summary_row"].fillna(False).sum()) if "summary_row" in df.columns else 0,
        "Korjattuja lukuja": repaired_rows,
        "Fallback rows": fallback_rows,
        "Fallback-rivien osuus %": fallback_ratio,
        "Tuloslaskelman aloitussivu": parsed_result.get("start_pages", {}).get("tuloslaskelma"),
        "Tase vastaavaa -aloitussivu": parsed_result.get("start_pages", {}).get("tase_vastaavaa"),
        "Tase vastattavaa -aloitussivu": parsed_result.get("start_pages", {}).get("tase_vastattavaa"),
        "Dokumentin hallitseva numeromuoto": str(parsed_result.get("dominant_number_format")),
        "Skipped table-of-contents pages": len(parsed_result.get("pages", {}).get("skipped_toc_pages", [])),
        "Ohitettuja toimintakertomussivuja": len(parsed_result.get("pages", {}).get("skipped_report_pages", [])),
        "Rows rejected by whitelist": int(sum((parsed_result.get("diagnostics", {}).get(s, {}) or {}).get("rejected_by_whitelist", 0) for s in ["tuloslaskelma", "tase_vastaavaa", "tase_vastattavaa"])),
        "Hierarkiakonflikteja": reliability.get("hierarchy_conflicts"),
        "Luottamuksen laskun syyt": " | ".join(reliability.get("reasons", [])),
    }


def _value_present(value) -> bool:
    """Return True when a display or numeric value is meaningfully present."""
    return value is not None and str(value).strip() not in {"", "nan", "None"}


def _yes_no_mark(value: bool) -> str:
    """Format a boolean evidence point for user-facing review tables."""
    return "Yes" if bool(value) else "No"


def _is_blank_display_value(value) -> bool:
    """Return True when a value should be treated as empty in explanations."""
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    return str(value).strip().lower() in {"", "nan", "none", "nat"}


def _safe_display_text(value, fallback: str = "-") -> str:
    """Convert technical values into compact display text."""
    if _is_blank_display_value(value):
        return fallback
    return str(value)


def _values_equal_for_display(row) -> bool:
    """Return True when the compared values are numerically equal within tolerance."""
    diff = _row_get(row, "difference", None) if "_row_get" in globals() else row.get("difference")
    if diff is not None:
        try:
            return abs(float(diff)) <= TOLERANCE
        except Exception:
            return False
    old_value = row.get("value_older_current")
    new_value = row.get("value_newer_comparison")
    if old_value is None or new_value is None:
        return False
    try:
        return abs(float(new_value) - float(old_value)) <= TOLERANCE
    except Exception:
        return False


def _sections_equal_for_display(row) -> bool:
    """Return True when both sides belong to the same normalized section."""
    old_section = str(row.get("section_match_norm_older", "") or "")
    new_section = str(row.get("section_match_norm_newer", "") or "")
    return bool(old_section and new_section and old_section == new_section)


def _canonicals_equal_for_display(row) -> bool:
    """Return True when both sides have compatible canonical labels."""
    old_can = str(row.get("label_match_canonical_older", "") or "")
    new_can = str(row.get("label_match_canonical_newer", "") or "")
    if not old_can or not new_can:
        return False
    try:
        return canonical_values_equivalent(old_can, new_can)
    except Exception:
        return False


def build_user_decision(row) -> str:
    """Build a concise user-facing decision label for one reconciliation row."""
    status = str(row.get("status", "") or "")
    manual = bool(row.get("manual_review_required", False))

    if status == "Match" and not manual:
        return "Täsmää"
    if status == "Value differs":
        return "Luku poikkeaa"
    if status == "Missing row":
        return "Rivi puuttuu"
    if status in {"Could not verify", "Manual review"} or manual:
        return "Tarkista manuaalisesti"
    return status or "Tarkista"


def build_plain_language_reason(row) -> str:
    """Explain the decision in plain language without requiring technical columns."""
    status = str(row.get("status", "") or "")
    manual = bool(row.get("manual_review_required", False))
    old_label = _safe_display_text(row.get("label_older"))
    new_label = _safe_display_text(row.get("label_newer"))
    difference = row.get("difference")

    if status == "Missing row":
        if old_label == "-":
            return "Erä löytyi uudemmasta tilinpäätöksestä, mutta vastaavaa erää ei löytynyt vanhemmasta vertailukauden aineistosta."
        if new_label == "-":
            return "Erä löytyi vanhemmasta tilinpäätöksestä, mutta vastaavaa vertailukauden erää ei löytynyt uudemmasta tilinpäätöksestä."
        return "Vastaavaa riviä ei löytynyt luotettavasti toisesta tilinpäätöksestä."

    if status == "Value differs":
        if difference is not None and not _is_blank_display_value(difference):
            return f"Sama tai riittävän vastaava erä löytyi, mutta luvut poikkeavat toisistaan erotuksella {difference}."
        return "Sama tai riittävän vastaava erä löytyi, mutta arvoja ei voitu todeta samoiksi."

    if status == "Match" and not manual:
        return "Rivi hyväksyttiin täsmääväksi, koska osio, nimike/käsite ja numeerinen arvo tukevat samaa päätelmää."

    if manual:
        return "Ohjelma löysi mahdollisen yhteyden, mutta näyttö ei riitä automaattiseen hyväksyntään. Rivi on jätetty käyttäjän tarkistettavaksi."

    return "Rivin tulkinta jäi epävarmaksi ja vaatii tarkistusta."


def build_evidence_summary(row) -> str:
    """Build a compact Finnish evidence checklist for the main result table."""
    evidence = []
    evidence.append(f"Sama osio: {_yes_no_mark(_sections_equal_for_display(row))}")
    evidence.append(f"Sama sanastokäsite: {_yes_no_mark(_canonicals_equal_for_display(row))}")
    evidence.append(f"Arvot samat: {_yes_no_mark(_values_equal_for_display(row))}")

    method = _safe_display_text(row.get("match_method"))
    level = _safe_display_text(row.get("match_level"))
    confidence = map_user_confidence(_safe_display_text(row.get("confidence_level")))
    evidence.append(f"Menetelmä: {method}")
    evidence.append(f"Täsmäytystaso: {level}")
    evidence.append(f"Luottamustaso: {confidence}")

    if bool(row.get("numeric_integrity_issue", False)) or bool(row.get("repaired_issue", False)):
        evidence.append("Huomio: lukuarvo vaati teknistä tulkintaa")
    if bool(row.get("context_conflict", False)):
        evidence.append("Huomio: rakenteellinen konteksti ristiriitainen")
    elif bool(row.get("context_uncertain", False)):
        evidence.append("Huomio: rakenteellinen konteksti epävarma")
    if bool(row.get("manual_review_required", False)):
        reason = _safe_display_text(row.get("manual_review_reason"), "tarkistus vaaditaan")
        evidence.append(f"Tarkistuksen syy: {reason}")

    return " | ".join(evidence)


def build_confidence_explanation(row) -> str:
    """Explain why the row confidence is high, medium or low."""
    confidence = str(row.get("confidence_level", "") or "")
    if confidence in {"High", "Vahva"} and not bool(row.get("manual_review_required", False)):
        return "Vahva: osio, nimike/käsite ja arvo tukevat täsmäytystä ilman merkittävää epävarmuutta."
    if confidence in {"Medium", "Keskitaso"}:
        return "Keskitaso: täsmäytykselle on tukea, mutta jokin tekijä, kuten esitystapa, konteksti tai tekninen tulkinta, heikentää varmuutta."
    if confidence in {"Low", "Matala", "Uncertain"}:
        return "Matala: ohjelma ei saanut riittävän vahvaa näyttöä automaattiseen päätelmään."
    if bool(row.get("manual_review_required", False)):
        return "Tarkistettava: päätös jätettiin käyttäjälle, koska automaattinen näyttö ei ollut riittävä."
    return "Luottamustasoa ei voitu määrittää yksiselitteisesti."

# ---------------------------------------------------------
# HUMAN-READABLE DECISION EXPLANATIONS
# ---------------------------------------------------------

def explain_match_decision(row) -> str:
    """Build a human-readable explanation for a reconciliation decision.

    The explanation describes the matching method, canonical support, section
    compatibility, hierarchy/context evidence and numeric outcome. This makes
    the result understandable without reading the technical technical detail columns."""
    status = str(row.get("status", "") or "")
    method = str(row.get("match_method", "") or "")
    reason = str(row.get("match_reason", "") or "")
    section_old = str(row.get("section_match_norm_older", "") or "")
    section_new = str(row.get("section_match_norm_newer", "") or "")
    level_old = str(row.get("block_type_older", "") or "")
    level_new = str(row.get("block_type_newer", "") or "")
    canonical_old = str(row.get("label_match_canonical_older", "") or "")
    canonical_new = str(row.get("label_match_canonical_newer", "") or "")
    context_conflict = bool(row.get("context_conflict", False))
    context_uncertain = bool(row.get("context_uncertain", False))
    hierarchy_status = str(row.get("hierarchy_status", "") or "")
    extraction_old = str(row.get("extraction_status_older", "") or "")
    extraction_new = str(row.get("extraction_status_newer", "") or "")
    difference = row.get("difference")

    parts = []

    if status == "Missing row":
        if not _value_present(row.get("label_older")):
            return "The row was found only in the newer financial statement; no corresponding row was found in the older data."
        return "The row was found in the older financial statement, but no corresponding row was found in the comparative period of the newer statement."

    if method in {"exact_value_match", "exact"}:
        parts.append("The same normalized row label was found in the same section.")
    elif method in {"compact_value_match", "compact"}:
        parts.append("The row label matched using compact comparison that tolerates whitespace and PDF formatting differences.")
    elif method == "canonical":
        parts.append("Rows were matched by the same canonical item before value comparison.")
    elif method == "synonym":
        parts.append("Rivit yhdistettiin sanaston tai canonical-nimikkeen perusteella.")
    elif method:
        parts.append(f"Reconciliation was performed using method {method}.")
    else:
        parts.append("No reliable matching method was formed for the row.")

    if canonical_old and canonical_new and canonical_values_equivalent(canonical_old, canonical_new):
        parts.append(f"Canonical-nimike on sama: {canonicalize_common_labels(canonical_old)}.")
    elif canonical_old or canonical_new:
        parts.append(f"Canonical-nimikkeet poikkeavat tai puuttuvat ({canonical_old or '-'} / {canonical_new or '-'}).")

    if section_old and section_new:
        if section_old == section_new:
            parts.append(f"The section matches ({section_old}).")
        else:
            parts.append(f"Osio poikkeaa ({section_old} / {section_new}).")

    if level_old and level_new and level_old != level_new:
        parts.append(f"The row level differs ({level_old} / {level_new}), so the match is treated cautiously.")

    if context_conflict:
        parts.append("PMA-konteksti on ristiriitainen.")
    elif context_uncertain or hierarchy_status == "Uncertain":
        parts.append("The PMA context or parent structure is uncertain.")
    elif hierarchy_status in {"Looginen", "Rakenteellinen", "Header row"}:
        parts.append("PMA-/parent-konteksti tukee osumaa.")

    if extraction_old == "extracted_unmapped" or extraction_new == "extracted_unmapped":
        parts.append("Ainakin toinen rivi poimittiin dataksi ilman varmaa PMA-sanastotunnistusta.")

    if status == "Match":
        parts.append("Values match within tolerance.")
    elif status == "Value differs":
        if difference is not None:
            parts.append(f"The same item was found, but the values differ by {difference}.")
        else:
            parts.append("The same item was found, but the value could not be compared numerically.")
    elif status == "Could not verify":
        parts.append("The match or values could not be verified with sufficient certainty.")

    return " ".join(p for p in parts if p)


def explain_review_reason(row) -> str:
    """Return a derived value used by the reconciliation workflow."""
    status = str(row.get("status", "") or "")
    manual_reason = str(row.get("manual_review_reason", "") or "")
    if manual_reason and manual_reason.lower() != "nan":
        return manual_reason
    if status == "Match":
        return ""

    if status == "Value differs":
        return "The same item was matched, but the amount differs."
    if status == "Missing row":
        return "No counterpart was found for the row in the other financial statement."
    if status == "Could not verify":
        return "The reliability of the match or value was not sufficient for a verified match."

    if bool(row.get("context_conflict", False)):
        return "PMA-konteksti on ristiriitainen."
    if bool(row.get("context_uncertain", False)):
        return "The PMA context is uncertain."
    if str(row.get("extraction_status_older", "")) == "extracted_unmapped" or str(row.get("extraction_status_newer", "")) == "extracted_unmapped":
        return "The row was extracted as data, but it was not identified reliably with the PMA vocabulary."

    return str(row.get("selite", "") or "Rivi vaatii tarkistusta.")


def classify_transparency_reasons(row) -> tuple[str, str]:
    """Return the primary reason and secondary factors behind the row decision.

    This function is the core transparency layer of the thesis prototype. It
    turns technical matching signals into one clear primary reason and a short
    list of supporting factors, so the user can see why a complete comparison
    was or was not accepted.
    """
    status = str(row.get("status", "") or "")
    manual = bool(row.get("manual_review_required", False))

    if status == "Match" and not manual:
        return (
            "Täsmäytys hyväksyttiin",
            "Sama osio, sama tai vastaava sanastokäsite ja sama arvo tukevat päätöstä.",
        )

    primary = None
    factors: list[str] = []

    old_label = _safe_display_text(row.get("label_older"))
    new_label = _safe_display_text(row.get("label_newer"))

    if status == "Missing row" or old_label == "-" or new_label == "-":
        primary = "Vastinrivi puuttuu"
        factors.append("Erä näkyy vain toisessa tilinpäätöksessä tai vertailukauden sarakkeessa.")

    if status == "Value differs":
        primary = primary or "Luvut poikkeavat"
        diff = row.get("difference")
        if not _is_blank_display_value(diff):
            factors.append(f"Vastaava erä löytyi, mutta erotus on {diff}.")
        else:
            factors.append("Vastaava erä löytyi, mutta arvoja ei voitu todeta samoiksi.")

    if status in {"Could not verify", "Manual review"} or manual:
        primary = primary or "Ei riittävää näyttöä automaattiseen täsmäytykseen"

    old_section = str(row.get("section_match_norm_older", "") or "")
    new_section = str(row.get("section_match_norm_newer", "") or "")
    if old_section and new_section:
        if old_section == new_section:
            factors.append("Osio täsmää.")
        else:
            primary = primary or "Osio ei täsmää"
            factors.append(f"Rivit ovat eri osioissa ({old_section} / {new_section}).")

    old_can = str(row.get("label_match_canonical_older", "") or "")
    new_can = str(row.get("label_match_canonical_newer", "") or "")
    if old_can and new_can:
        try:
            if canonical_values_equivalent(old_can, new_can):
                factors.append(f"Sanastokäsite täsmää ({canonicalize_common_labels(old_can)}).")
            else:
                primary = primary or "Sanastokäsitteet poikkeavat"
                factors.append(f"Sanastokäsitteet poikkeavat ({old_can} / {new_can}).")
        except Exception:
            factors.append("Sanastokäsitettä ei voitu varmistaa teknisesti.")
    elif status not in {"Missing row"}:
        factors.append("Sanastotuki puuttuu vähintään toiselta riviltä.")

    if bool(row.get("context_conflict", False)):
        primary = primary or "Rakenteellinen konteksti on ristiriitainen"
        factors.append("Parent-/rakennepolku viittaa ristiriitaan.")
    elif bool(row.get("context_uncertain", False)):
        factors.append("Rakenteellinen konteksti on epävarma.")

    if bool(row.get("numeric_integrity_issue", False)) or bool(row.get("repaired_issue", False)):
        factors.append("Lukuarvo vaati teknistä tulkintaa tai korjausta.")

    if bool(row.get("fallback_used", False)):
        factors.append("Käytössä oli varatulkinta, joten tulos vaatii varovaisuutta.")

    method = str(row.get("match_method", "") or "")
    if method in {"fuzzy", "fuzzy_strong", "synonym"}:
        factors.append("Nimikeosuma perustuu pehmeään vastaavuuteen.")
    elif method:
        factors.append(f"Täsmäytystapa: {method}.")

    manual_reason = str(row.get("manual_review_reason", "") or "")
    if manual and manual_reason and manual_reason.lower() != "nan":
        factors.append(manual_reason)

    if primary is None:
        primary = "Täydellistä täsmäytystä ei voitu perustella"

    cleaned_factors = [f for f in dict.fromkeys(factors) if f and f != primary]
    return primary, " ".join(cleaned_factors)


def explain_why_not_perfect_match(row) -> str:
    """Explain in one sentence why a row was or was not accepted."""
    primary, additional = classify_transparency_reasons(row)
    if additional:
        return f"{primary}: {additional}"
    return primary


def transparency_primary_reason(row) -> str:
    """Return the main reason shown in the compact Excel report."""
    return classify_transparency_reasons(row)[0]


def transparency_additional_factors(row) -> str:
    """Return secondary factors shown in the compact Excel report."""
    return classify_transparency_reasons(row)[1]

def build_user_action(row) -> str:
    """Return one concise next action for the user."""
    status = str(row.get("status", "") or "")
    manual = bool(row.get("manual_review_required", False))
    if status == "Match" and not manual:
        return "Ei toimenpidettä."
    if status == "Value differs":
        return "Tarkista, johtuuko ero oikeasta lukupoikkeamasta vai esitystavasta."
    if status == "Missing row":
        return "Tarkista, onko erä aidosti puuttuva, yhdistetty toiseen riviin tai jätetty nollarivinä pois."
    return "Tarkista rivi manuaalisesti ennen johtopäätöstä."

# ---------------------------------------------------------
# USER-FACING DISPLAY TABLES
# ---------------------------------------------------------

def make_display_df(df: pd.DataFrame) -> pd.DataFrame:
    """Build the compact user-facing comparison table for the thesis prototype.

    The table is intentionally simple: one decision, one primary reason, short
    supporting evidence and one recommended action. Technical details remain in
    the separate Decision details sheet.
    """
    columns = [
        "Osio",
        "Päätös",
        "Erä vanhemmassa tilinpäätöksessä",
        "Erä uudemman tilinpäätöksen vertailukaudessa",
        "Vanhemman tilikauden arvo",
        "Uudemman vertailukauden arvo",
        "Erotus",
        "Pääperuste",
        "Lisäperusteet",
        "Päätöksen näyttö",
        "Käyttäjän toimenpide",
        "Luottamustaso",
    ]
    if df is None or df.empty:
        return pd.DataFrame(columns=columns)

    out = pd.DataFrame(index=df.index)
    out["Osio"] = df.get("section")
    out["Päätös"] = df.apply(build_user_decision, axis=1)
    out["Erä vanhemmassa tilinpäätöksessä"] = df.get("label_older").apply(_clean_display_label) if "label_older" in df.columns else ""
    out["Erä uudemman tilinpäätöksen vertailukaudessa"] = df.get("label_newer").apply(_clean_display_label) if "label_newer" in df.columns else ""
    out["Vanhemman tilikauden arvo"] = df.get("display_older_value")
    out["Uudemman vertailukauden arvo"] = df.get("display_newer_value")
    out["Erotus"] = df.get("difference")
    out["Pääperuste"] = df.apply(transparency_primary_reason, axis=1)
    out["Lisäperusteet"] = df.apply(transparency_additional_factors, axis=1)
    out["Päätöksen näyttö"] = df.apply(build_evidence_summary, axis=1)
    out["Käyttäjän toimenpide"] = df.apply(build_user_action, axis=1)
    out["Luottamustaso"] = df.get("confidence_level", pd.Series(index=df.index, dtype=str)).apply(map_user_confidence)
    return out.reset_index(drop=True)

# ---------------------------------------------------------
# DECISION TRACE AND TECHNICAL DEBUG OUTPUTS
# ---------------------------------------------------------


def build_decision_trace_display_df(df: pd.DataFrame) -> pd.DataFrame:
    """Build a concise Finnish decision-detail table for transparent review."""
    if df is None or df.empty:
        return pd.DataFrame()
    out = pd.DataFrame()
    out["Osio"] = df.get("section")
    out["Päätös"] = df.apply(build_user_decision, axis=1)
    out["Pääperuste"] = df.apply(transparency_primary_reason, axis=1)
    out["Lisäperusteet"] = df.apply(transparency_additional_factors, axis=1)
    out["Erä vanhemmassa"] = df.get("label_older")
    out["Erä uudemmassa"] = df.get("label_newer")
    out["Sama osio"] = df.apply(_sections_equal_for_display, axis=1)
    out["Sanastokäsite vanhemmassa"] = df.get("label_match_canonical_older")
    out["Sanastokäsite uudemmassa"] = df.get("label_match_canonical_newer")
    out["Sama sanastokäsite"] = df.apply(_canonicals_equal_for_display, axis=1)
    out["Vanhemman arvo"] = df.get("display_older_value")
    out["Uudemman arvo"] = df.get("display_newer_value")
    out["Arvot samat"] = df.apply(_values_equal_for_display, axis=1)
    out["Erotus"] = df.get("difference")
    out["Täsmäytystapa"] = df.get("match_method")
    out["Tekninen peruste"] = df.get("match_reason")
    out["Luottamustaso"] = df.get("confidence_level")
    out["Vaatii tarkistuksen"] = df.get("manual_review_required")
    out["Tarkistuksen syy"] = df.get("manual_review_reason")
    out["Vanhemman sivu"] = df.get("source_page_older")
    out["Uudemman sivu"] = df.get("source_page_newer")
    out["Vanhemman lähderivi"] = df.get("source_line_older")
    out["Uudemman lähderivi"] = df.get("source_line_newer")
    return out


def build_tech_debug_df(df: pd.DataFrame) -> pd.DataFrame:
   
    if df is None or df.empty:
        return pd.DataFrame()
    out = pd.DataFrame()
    technical_cols = {
        "Section": "section",
        "Older label raw": "label_older",
        "Newer label raw": "label_newer",
        "Older normalized label": "label_norm_older",
        "Newer normalized label": "label_norm_newer",
        "Section norm vanha": "section_match_norm_older",
        "Section norm uusi": "section_match_norm_newer",
        "Older parent": "parent_label_older",
        "Newer parent": "parent_label_newer",
        "Older top parent": "top_parent_label_older",
        "Newer top parent": "top_parent_label_newer",
        "Older structure path": "category_path_older",
        "Newer structure path": "category_path_newer",
        "Lakipohjainen taso vanha": "legal_basis_level_older",
        "Lakipohjainen taso uusi": "legal_basis_level_newer",
        "Forced key item vanha": "forced_main_item_canonical_older",
        "Forced key item uusi": "forced_main_item_canonical_newer",
        "Older root item": "forced_root_item_older",
        "Newer root item": "forced_root_item_newer",
        "Taseen loppusumma vanha": "balance_total_root_item_older",
        "Taseen loppusumma uusi": "balance_total_root_item_newer",
        "Total-taso vanha": "total_level_older",
        "Total-taso uusi": "total_level_newer",
        "Hierarkiapisteet": "hierarchy_score",
        "Hierarkiastatus": "hierarchy_status",
        "Label-osuman pisteet": "label_match_score",
        "Match reason": "match_reason",
    }
    for display, source_col in technical_cols.items():
        out[display] = df.get(source_col)
    return out


def _clean_display_label(value) -> str:
    """Return a derived value used by the reconciliation workflow."""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text_value = str(value).strip()
    if text_value.lower() in {"nan", "none", "nat"}:
        return ""
    return text_value


def enrich_missing_row_findings(df: pd.DataFrame) -> pd.DataFrame:

    if df is None or df.empty or "status" not in df.columns:
        return df

    out = df.copy()
    missing_mask = out["status"].fillna("").eq("Missing row")
    if not missing_mask.any():
        return out

    def _missing_side(row) -> str:
        
        old_label = _clean_display_label(row.get("label_older"))
        new_label = _clean_display_label(row.get("label_newer"))
        old_value = row.get("value_older_current")
        new_value = row.get("value_newer_comparison")

        if old_label and not new_label:
            return "from the comparative period of the newer financial statement"
        if new_label and not old_label:
            return "from the older financial statement for the comparative period"

        try:
            if pd.notna(old_value) and (new_value is None or pd.isna(new_value)):
                return "from the comparative period of the newer financial statement"
        except Exception:
            pass
        try:
            if pd.notna(new_value) and (old_value is None or pd.isna(old_value)):
                return "from the older financial statement for the comparative period"
        except Exception:
            pass
        return "toisesta tiedostosta"

    def _present_side(row) -> str:
       
        old_label = _clean_display_label(row.get("label_older"))
        new_label = _clean_display_label(row.get("label_newer"))
        if old_label and not new_label:
            return "from the older financial statement"
        if new_label and not old_label:
            return "from the newer financial statement"
        return "from only one financial statement"

    def _item_name(row) -> str:
        
        return _clean_display_label(row.get("label_older")) or _clean_display_label(row.get("label_newer")) or "Unnamed item"

    for idx, row in out.loc[missing_mask].iterrows():
        missing_side = _missing_side(row)
        present_side = _present_side(row)
        item = _item_name(row)

        out.at[idx, "exception_category"] = "Missing row"
        out.at[idx, "severity"] = "Tarkistettava"
        out.at[idx, "review_priority"] = 2
        out.at[idx, "selite"] = f"Item '{item}' was found {present_side}, but is missing {missing_side}."
        if "manual_review_reason" in out.columns:
            out.at[idx, "manual_review_reason"] = "Counterpart row was not found in the other financial statement."
        if "manual_review_required" in out.columns:
            out.at[idx, "manual_review_required"] = True

    return out


def filter_review_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Return a derived value used by the reconciliation workflow."""
    if df is None or df.empty:
        return pd.DataFrame()

    status = df.get("status", pd.Series(index=df.index, dtype=str)).fillna("")
    pres = df.get("presentation_change_flag", pd.Series(index=df.index, dtype=str)).fillna("")

    true_review_statuses = [
        "Value differs",
        "Missing row",
        "Could not verify",
        "Manual review",
    ]

    return df[
        status.isin(true_review_statuses)
        | pres.isin([PRESENTATION_CHANGE_CLEAR])
    ].copy()


def build_top_findings_df(df: pd.DataFrame, limit: int = 10) -> pd.DataFrame:
    """Build a Finnish table of the most important findings for the workbook and UI."""
    review = filter_review_rows(df)
    columns = ["Tärkeys", "Osio", "Erä", "Päätös", "Pääperuste", "Erotus", "Käyttäjän toimenpide"]
    if review.empty:
        return pd.DataFrame(columns=columns)
    priority = {
        "Value differs": 1,
        "Missing row": 2,
        "Could not verify": 3,
        "Manual review": 3,
        "Match": 9,
    }
    review = review.copy()
    review["_priority"] = review.get("status", "").map(priority).fillna(4)
    review = review.sort_values(["_priority", "section", "label_older"], ascending=True).head(limit)
    out = pd.DataFrame(index=review.index)
    out["Tärkeys"] = review["_priority"].map(lambda x: int(x) if pd.notna(x) else "")
    out["Osio"] = review.get("section")
    out["Erä"] = review.get("label_older").fillna(review.get("label_newer"))
    out["Päätös"] = review.apply(build_user_decision, axis=1)
    out["Pääperuste"] = review.apply(transparency_primary_reason, axis=1)
    out["Erotus"] = review.get("difference")
    out["Käyttäjän toimenpide"] = review.apply(build_user_action, axis=1)
    return out.reset_index(drop=True)


def status_row_style(row):
    """Return UI row colours that match the Excel workbook status colours."""
    status = str(
        row.get("Päätös", "")
        or row.get("Status", "")
        or row.get("Paatos", "")
        or row.get("Decision", "")
        or row.get("Finding", "")
        or ""
    )
    if status in {"Täsmää", "Match", "Usable", "Ei toimenpidettä."}:
        return ["background-color: #d9f2d9"] * len(row)
    if status in {"Tarkista manuaalisesti", "Could not verify", "Manual review", "Vaatii tarkistusta", "Tarkista"}:
        return ["background-color: #fff3cd"] * len(row)
    if status in {"Rivi puuttuu", "Missing row", "Puuttuu", "Row missing from one document"}:
        return ["background-color: #ffe5b4"] * len(row)
    if status in {"Luku poikkeaa", "Value differs", "Not recommended without manual review"}:
        return ["background-color: #f8d7da"] * len(row)
    return [""] * len(row)


def build_priority_summary(df: pd.DataFrame) -> pd.DataFrame:
   
    if df.empty:
        return pd.DataFrame(columns=["Prioriteetti", "Count", "Kuvaus"])
    priorities = [
        ("1. Poikkeavat luvut", int((df["status"] == "Value differs").sum()), "Rivit, joissa luku poikkeaa."),
        ("2. Tarkistettavat", int(df["status"].isin(["Missing row", "Could not verify", "Manual review"]).sum()), "Rivit, joita ei voitu varmistaa tai joista puuttuu vastinrivi."),
        ("3. Technical notes", int((df["severity"] == "Technical note").sum()), "Technical notes are shown in the decision trace, not in the main user worklist."),
    ]
    return pd.DataFrame(priorities, columns=["Prioriteetti", "Count", "Kuvaus"])


def build_match_debug_df(df: pd.DataFrame) -> pd.DataFrame:
    """Build match debug df.
    
    Purpose: This function belongs to the reconciliation and candidate evaluation stage.
    Why: It supports controlled matching while keeping uncertain cases visible for manual review.
    """
    if df is None or df.empty:
        return pd.DataFrame(columns=[
            "Older section", "Newer section", "Older normalized section", "Newer normalized section",
            "Older label raw", "Newer label raw", "Older normalized label", "Newer normalized label",
            "Matching method", "Matching reason", "Label-osuman pisteet",
            "Old value", "New value", "Difference", "Status"
        ])
    out = pd.DataFrame()
    out["Older section"] = df.get("section")
    out["Newer section"] = df.get("section")
    out["Older normalized section"] = df.get("section_match_norm_older")
    out["Newer normalized section"] = df.get("section_match_norm_newer")
    out["Older label raw"] = df.get("label_older")
    out["Newer label raw"] = df.get("label_newer")
    out["Older normalized label"] = df.get("label_norm_older")
    out["Newer normalized label"] = df.get("label_norm_newer")
    out["Exception category"] = df.get("exception_category")
    out["Presentation change"] = df.get("presentation_change_flag")
    out["Presentation-change basis"] = df.get("presentation_change_reason")
    out["Matching method"] = df.get("match_method")
    out["Matching reason"] = df.get("match_reason")
    out["Label-osuman pisteet"] = df.get("label_match_score")
    out["Old value"] = df.get("display_older_value")
    out["New value"] = df.get("display_newer_value")
    out["Difference"] = df.get("difference")
    out["Status"] = df.get("status")
    return out

# ---------------------------------------------------------
# DASHBOARD, TEST METRICS AND KEY ITEM CONTROLS
# ---------------------------------------------------------



def _count_present_reconciliation_rows(df: pd.DataFrame, label_column: str, value_column: str) -> int:
    """Count rows that are present on one side of the comparison result."""
    if df is None or df.empty:
        return 0
    if label_column not in df.columns and value_column not in df.columns:
        return 0
    labels = df.get(label_column, pd.Series(index=df.index, dtype=object)).apply(_clean_display_label)
    values = df.get(value_column, pd.Series(index=df.index, dtype=object))
    value_present = values.apply(lambda value: not _is_blank_display_value(value))
    return int((labels.astype(str).str.len().gt(0) | value_present).sum())

def build_management_dashboard_df(df: pd.DataFrame) -> pd.DataFrame:
    """Build a compact Finnish summary for UI, Excel and thesis test reporting."""
    if df is None or df.empty:
        base = {
            "Vertailtuja rivejä yhteensä": 0,
            "Löydettyjä rivejä kuluvan kauden tilinpäätöksestä": 0,
            "Löydettyjä rivejä vertailukauden tilinpäätöksestä": 0,
            "Täsmäävät rivit": 0,
            "Luvut poikkeavat": 0,
            "Puuttuvat tai tarkistettavat rivit": 0,
            "Onnistumisaste %": 0.0,
            "Matalan luottamuksen rivit": 0,
            "Varatulkintaa käyttäneet rivit": 0,
            "Korjattuja lukuja": 0,
            "Esitystavan muutos -havaintoja": 0,
        }
    else:
        total = len(df)
        ok = int((df["status"] == "Match").sum()) if "status" in df.columns else 0
        errors = int((df["status"] == "Value differs").sum()) if "status" in df.columns else 0
        uncertain = int(df["status"].isin(["Missing row", "Could not verify", "Manual review"]).sum()) if "status" in df.columns else 0
        low_conf = int((df.get("confidence_level", pd.Series(index=df.index, dtype=str)).fillna("") == "Low").sum())
        fallback = int(df.get("fallback_used", pd.Series(False, index=df.index)).fillna(False).astype(bool).sum())
        repaired = int(df.get("repaired_issue", pd.Series(False, index=df.index)).fillna(False).astype(bool).sum())
        presentation_changes = int(df.get("presentation_change_flag", pd.Series("", index=df.index)).fillna("").isin([PRESENTATION_CHANGE_POSSIBLE, PRESENTATION_CHANGE_CLEAR]).sum())
        older_found = _count_present_reconciliation_rows(df, "label_older", "value_older_current")
        newer_found = _count_present_reconciliation_rows(df, "label_newer", "value_newer_comparison")
        base = {
            "Vertailtuja rivejä yhteensä": total,
            "Löydettyjä rivejä kuluvan kauden tilinpäätöksestä": older_found,
            "Löydettyjä rivejä vertailukauden tilinpäätöksestä": newer_found,
            "Täsmäävät rivit": ok,
            "Luvut poikkeavat": errors,
            "Puuttuvat tai tarkistettavat rivit": uncertain,
            "Onnistumisaste %": round(ok / max(1, total) * 100, 1),
            "Matalan luottamuksen rivit": low_conf,
            "Varatulkintaa käyttäneet rivit": fallback,
            "Korjattuja lukuja": repaired,
            "Esitystavan muutos -havaintoja": presentation_changes,
        }

    return pd.DataFrame([{"Mittari": k, "Arvo": v} for k, v in base.items()])


def build_exception_category_summary(df: pd.DataFrame) -> pd.DataFrame:
   
    if df is None or df.empty or "exception_category" not in df.columns:
        return pd.DataFrame(columns=["Exception category", "Count"])
    out = df["exception_category"].fillna("ei luokiteltu").value_counts().reset_index()
    out.columns = ["Exception category", "Count"]
    return out


def build_test_case_metrics(parsed_older: dict, parsed_newer: dict, comparison_df: pd.DataFrame) -> pd.DataFrame:
    """Build one metrics row for a single test case."""
    rel_old = parsed_older.get("document_reliability", {}) or {}
    rel_new = parsed_newer.get("document_reliability", {}) or {}
    total = len(comparison_df) if comparison_df is not None else 0
    ok = int((comparison_df["status"] == "Match").sum()) if comparison_df is not None and not comparison_df.empty else 0
    errors = int((comparison_df["status"] == "Value differs").sum()) if comparison_df is not None and not comparison_df.empty else 0
    uncertain = int(comparison_df["status"].isin(["Missing row", "Could not verify", "Manual review"]).sum()) if comparison_df is not None and not comparison_df.empty else 0
    return pd.DataFrame([{
        "Older document": parsed_older.get("document"),
        "Newer document": parsed_newer.get("document"),
        "Vanhemman luotettavuuspisteet": rel_old.get("score"),
        "Uudemman luotettavuuspisteet": rel_new.get("score"),
        "Compared rows": total,
        "Matched rows": ok,
        "Differing values": errors,
        "Uncertain/missing rows": uncertain,
        "Onnistumisaste %": round(ok / max(1, total) * 100, 1),
        "Fallback rows": int(comparison_df.get("fallback_used", pd.Series(False, index=comparison_df.index)).fillna(False).astype(bool).sum()) if comparison_df is not None and not comparison_df.empty else 0,
        "Korjattuja lukuja": int(comparison_df.get("repaired_issue", pd.Series(False, index=comparison_df.index)).fillna(False).astype(bool).sum()) if comparison_df is not None and not comparison_df.empty else 0,
    }])


def _to_float_or_none(value):
  
    try:
        if value is None or pd.isna(value):
            return None
        return float(value)
    except Exception:
        return None


def build_decision_trace_df(df: pd.DataFrame) -> pd.DataFrame:
    """Build a traceable decision trace for each reconciliation decision.

    The decision trace preserves labels, values, source rows, source pages,
    matching methods, confidence levels and structural context for later review."""
    if df is None or df.empty:
        return pd.DataFrame(columns=[
            "Section", "Era vanhemmassa", "Era uudemmassa", "Paatos", "Exception category",
            "Tasmaystapa", "Tasmayksen peruste", "Old value", "New value", "Difference",
            "Older page", "Older source row", "Newer page", "Newer source row", "Confidence level",
            "Older value source", "Uuden arvolahde", "Older structure path", "Newer structure path"
        ])
    out = pd.DataFrame()
    out["Osio"] = df.get("section")
    out["Era vanhemmassa"] = df.get("label_older")
    out["Era uudemmassa"] = df.get("label_newer")
    out["Paatos"] = df.get("status")
    out["Exception category"] = df.get("exception_category")
    out["Presentation change"] = df.get("presentation_change_flag")
    out["Presentation-change basis"] = df.get("presentation_change_reason")
    out["Tasmaystapa"] = df.get("match_method")
    out["Tasmayksen peruste"] = df.get("match_reason")
    out["Old value"] = df.get("display_older_value")
    out["New value"] = df.get("display_newer_value")
    out["Difference"] = df.get("difference")
    out["Older page"] = df.get("source_page_older")
    out["Older source row"] = df.get("source_line_older")
    out["Newer page"] = df.get("source_page_newer")
    out["Newer source row"] = df.get("source_line_newer")
    out["Confidence level"] = df.get("confidence_level")
    out["Older value source"] = df.get("value_source_older")
    out["Uuden arvolahde"] = df.get("value_source_newer")
    out["Older structure path"] = df.get("category_path_older")
    out["Newer structure path"] = df.get("category_path_newer")
    return out


def build_key_item_control_df(df: pd.DataFrame) -> pd.DataFrame:
    """Summarize whether key financial statement items were found and reconciled."""
    rows = []
    if df is None or df.empty:
        return pd.DataFrame(columns=["Section", "Avainera", "Loytyy vanhemmasta", "Loytyy uudemmasta", "Status", "Huomio"])

    def _has_canon(row, side: str, canonical: str) -> bool:
     
        val = row.get(f"forced_main_item_canonical_{side}")
        if pd.notna(val) and str(val) == canonical:
            return True
        lab = row.get(f"label_{'older' if side == 'older' else 'newer'}")
        sec = row.get("section")
        try:
            return forced_main_item_canonical(lab, sec) == canonical
        except Exception:
            return False

    for section, items in FORCED_MAIN_ITEMS_BY_SECTION.items():
        for canonical in sorted(items):
            mask_old = df.apply(lambda r: _has_canon(r, "older", canonical), axis=1)
            mask_new = df.apply(lambda r: _has_canon(r, "newer", canonical), axis=1)
            found_old = bool(mask_old.any())
            found_new = bool(mask_new.any())
            related = df[mask_old | mask_new]
            if related.empty:
                status = "Rivi puuttuu molemmista"
                note = "Avaineraa ei tunnistettu vertailutuloksessa."
            else:
                statuses = set(related.get("status", pd.Series(dtype=str)).dropna().astype(str))
                if "Match" in statuses:
                    status = "Match"
                    note = "A matching pair was found for the key item."
                elif found_old and found_new:
                    status = "Tarkista"
                    note = "The key item was found in both documents, but it could not be verified as a match."
                else:
                    status = "Row missing from one document"
                    note = "Avainera loytyi vain toisesta dokumentista."
            rows.append({
                "Section": section,
                "Avainera": canonical,
                "Loytyy vanhemmasta": found_old,
                "Loytyy uudemmasta": found_new,
                "Status": status,
                "Huomio": note,
            })
    return pd.DataFrame(rows)


def _canonical_for_side(row, side: str) -> str:
    """Canonical for side.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    val = row.get(f"forced_main_item_canonical_{side}")
    if pd.notna(val) and str(val).strip():
        return str(val)
    lab = row.get(f"label_{'older' if side == 'older' else 'newer'}")
    return canonicalize_common_labels(lab or "")


def _row_value_for_side(row, side: str):
    """Row value for side.
    
    Purpose: This function belongs to the number parsing and validation stage.
    Why: It reduces the risk that formatting differences or unreadable values are treated as reliable evidence.
    """
    if side == "older":
        return _to_float_or_none(row.get("value_older_current"))
    return _to_float_or_none(row.get("value_newer_comparison"))


def _text_contains_canonical(row, side: str, canonical: str) -> bool:
    """Text contains canonical.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    texts = []
    for col in [
        f"label_{'older' if side == 'older' else 'newer'}",
        f"parent_label_{side}",
        f"top_parent_label_{side}",
        f"category_path_{side}",
        f"forced_main_item_canonical_{side}",
        f"label_match_canonical_{side}",
    ]:
        v = row.get(col)
        if pd.notna(v):
            texts.append(normalize_label(str(v)))
    c = normalize_label(canonical)
    return any(c and c in t for t in texts)


# =========================================================
# HIERARCHICAL ONE-TO-MANY PRESENTATION CHANGE ANALYSIS
# =========================================================
# This section detects possible presentation changes where one main row in one
# financial statement corresponds to several detail rows or a subtotal in the
# other financial statement.
#
# It handles:
# - Main item versus subtotal comparisons.
# - Main item versus account-detail sum comparisons.
# - Context-based grouping using parent, top-parent and category-path metadata.
# - Suggestions where values support a possible one-to-many relationship.
#
# The purpose is not to automatically reconcile these cases, but to generate
# review suggestions for presentation changes that require professional judgement.

BLOCK_SUM_TOLERANCE = max(TOLERANCE, 0.01)


def _as_float(value):
   
    try:
        if value is None or pd.isna(value):
            return None
        return float(value)
    except Exception:
        return None


def _norm_for_block(value) -> str:
    
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return normalize_label(str(value))


def _compact_for_block(value) -> str:
    
    return _norm_for_block(value).replace(" ", "")


def _row_order_value(row, side: str):
    """Row order value.
    
    Purpose: This function belongs to the number parsing and validation stage.
    Why: It reduces the risk that formatting differences or unreadable values are treated as reliable evidence.
    """
    page = row.get(f"source_page_{side}")
    idx = row.get(f"source_line_idx_{side}")
    try:
        page = int(page) if page is not None and not pd.isna(page) else 10**9
    except Exception:
        page = 10**9
    try:
        idx = int(idx) if idx is not None and not pd.isna(idx) else 10**9
    except Exception:
        idx = 10**9
    return (page, idx)


def _side_label(row, side: str):
    """Side label.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    return row.get(f"label_{side}")


def _side_value(row, side: str):
    """Side value.
    
    Purpose: This function belongs to the number parsing and validation stage.
    Why: It reduces the risk that formatting differences or unreadable values are treated as reliable evidence.
    """
    col = "value_older_current" if side == "older" else "value_newer_comparison"
    return _as_float(row.get(col))


def _side_source_line(row, side: str):
    """Side source line.
    
    Purpose: This function belongs to the document extraction stage.
    Why: It makes the PDF input usable before reconciliation decisions are made.
    """
    return row.get(f"source_line_{side}")


def _side_parent(row, side: str):
    """Side parent.
    
    Purpose: This function belongs to the financial statement structure stage.
    Why: It prevents rows from being compared across incompatible statement sections or hierarchy levels.
    """
    return row.get(f"parent_label_{side}")


def _side_top_parent(row, side: str):
    """Side top parent.
    
    Purpose: This function belongs to the financial statement structure stage.
    Why: It prevents rows from being compared across incompatible statement sections or hierarchy levels.
    """
    return row.get(f"top_parent_label_{side}")


def _side_path(row, side: str):
   
    return row.get(f"category_path_{side}")


def _is_account_detail_label(label) -> bool:
    """Is account detail label.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    raw = str(label or "").strip()
    compact = raw.replace(" ", "")
    return bool(re.match(r"^\d{3,6}", raw) or re.match(r"^\d{3,6}", compact))


def _is_probable_group_total_or_subtotal(label) -> bool:
    
    norm = _norm_for_block(label)
    if not norm or _is_account_detail_label(label):
        return False
    subtotal_terms = [
        "yhteensa", "summa", "menojaamat", "menojäämät", "saamiset", "velat",
        "pitkaaikainen", "lyhytaikainen", "pysyvat vastaavat", "vaihtuvat vastaavat",
        "oma paaoma", "vieras paaoma", "aineelliset hyodykkeet", "aineettomat hyodykkeet",
        "henkilosivukulut", "henkilostokulut", "materiaalit ja palvelut",
    ]
    return any(term in norm for term in subtotal_terms)


def _block_context_tokens_from_row(row, side: str) -> set[str]:
   
    tokens = set()
    for value in [_side_parent(row, side), _side_top_parent(row, side), _side_path(row, side), _side_label(row, side)]:
        norm = _norm_for_block(value)
        compact = norm.replace(" ", "")
        if norm:
            tokens.add(norm)
        if compact:
            tokens.add(compact)
        canon = canonicalize_common_labels(norm) if norm else ""
        if canon:
            tokens.add(canon)
            tokens.add(canon.replace(" ", ""))
    return {t for t in tokens if t}


def _main_label_context_tokens(label: str) -> set[str]:
    """Main label context tokens.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    norm = _norm_for_block(label)
    canon = canonicalize_common_labels(norm)
    parts = {norm, norm.replace(" ", "")}
    if canon:
        parts.add(canon)
        parts.add(canon.replace(" ", ""))
    synonym_map = {
        "siirtovelat": {"menojaamat", "menojäämät", "palkkamenot siirtovelat", "elakevakuutusmaksut siirtovelat", "tyottomyysvakuutusmaksut siirtovelat"},
        "saamiset": {"myyntisaamiset", "muut saamiset", "siirtosaamiset", "pitkaaikaiset saamiset", "lyhytaikaiset saamiset"},
        "muut velat": {"arvonlisaverovelka", "verotilivelka", "ennakonpidatysvelka", "sosiaaliturvamaksuvelka"},
        "rahat ja pankkisaamiset": {"pankkisaamiset", "pankkitili", "kateisvarat"},
    }
    for key, vals in synonym_map.items():
        if key in parts or key.replace(" ", "") in parts:
            parts.update(vals)
            parts.update(v.replace(" ", "") for v in vals)
    return {p for p in parts if p}


def _context_supports_main(row, side: str, main_label: str) -> bool:
    
    main_tokens = _main_label_context_tokens(main_label)
    row_tokens = _block_context_tokens_from_row(row, side)
    if main_tokens & row_tokens:
        return True
    raw_line = _norm_for_block(_side_source_line(row, side))
    return any(t and t in raw_line for t in main_tokens if len(t) >= 6)


def _build_side_rows(df: pd.DataFrame, side: str, section: str) -> pd.DataFrame:
   
    label_col = f"label_{side}"
    if df is None or df.empty or label_col not in df.columns:
        return pd.DataFrame()
    rows = df[df[label_col].notna()].copy()
    rows = rows[rows.get("section", pd.Series(index=rows.index)).fillna("").astype(str) == str(section)].copy()
    if rows.empty:
        return rows
    rows["_side_label"] = rows[label_col].astype(str)
    rows["_side_value"] = rows.apply(lambda r: _side_value(r, side), axis=1)
    rows["_is_account_detail"] = rows["_side_label"].apply(_is_account_detail_label)
    rows["_is_subtotal_candidate"] = rows["_side_label"].apply(_is_probable_group_total_or_subtotal)
    rows["_order_key"] = rows.apply(lambda r: _row_order_value(r, side), axis=1)
    return rows.sort_values("_order_key").copy()


def _extract_block_candidates_for_main(df: pd.DataFrame, main_row, main_side: str, detail_side: str) -> list[dict]:
    """Extract block candidates for main.
    
    Purpose: This function belongs to the document extraction stage.
    Why: It makes the PDF input usable before reconciliation decisions are made.
    """
    main_label = _side_label(main_row, main_side)
    main_value = _side_value(main_row, main_side)
    section = main_row.get("section")
    if not main_label or main_value is None or section is None:
        return []

    pool = _build_side_rows(df, detail_side, section)
    if pool.empty:
        return []

    context_pool = pool[pool.apply(lambda r: _context_supports_main(r, detail_side, str(main_label)), axis=1)].copy()
    if context_pool.empty:
        return []

    suggestions = []
    main_norm = normalize_for_exact_match(str(main_label))

    subtotal_pool = context_pool[
        (context_pool["_side_value"].notna())
        & (~context_pool["_is_account_detail"])
        & (context_pool["_is_subtotal_candidate"])
        & (context_pool["_side_label"].apply(lambda x: normalize_for_exact_match(x) != main_norm))
    ].copy()
    if not subtotal_pool.empty:
        subtotal_pool["_diff"] = subtotal_pool["_side_value"].apply(lambda v: abs(float(v) - float(main_value)))
        matching_subtotals = subtotal_pool[subtotal_pool["_diff"] <= BLOCK_SUM_TOLERANCE].copy()
        for _, subtotal in matching_subtotals.head(5).iterrows():
            block_details = context_pool[(context_pool["_is_account_detail"]) & (context_pool["_side_value"].notna())].copy()
            detail_labels = block_details["_side_label"].astype(str).head(12).tolist()
            detail_sum = float(block_details["_side_value"].sum()) if not block_details.empty else None
            suggestions.append({
                "suggestion_type": "subtotal_equals_main",
                "main_label": main_label,
                "main_value": float(main_value),
                "subtotal_label": subtotal["_side_label"],
                "subtotal_value": float(subtotal["_side_value"]),
                "detail_sum": detail_sum,
                "detail_labels": detail_labels,
                "difference": float(subtotal["_side_value"] - float(main_value)),
                "confidence": "High",
                "reason": "The main item value matches a subtotal row in the same heading context in the other file.",
            })

    details = context_pool[(context_pool["_is_account_detail"]) & (context_pool["_side_value"].notna())].copy()
    if not details.empty:
        details["_group_key"] = details.apply(lambda r: _norm_for_block(_side_parent(r, detail_side)) or _norm_for_block(_side_top_parent(r, detail_side)) or _norm_for_block(section), axis=1)
        for group_key, group in details.groupby("_group_key", dropna=False):
            detail_sum = float(group["_side_value"].sum())
            diff = detail_sum - float(main_value)
            if abs(diff) <= max(BLOCK_SUM_TOLERANCE, ROUNDING_WARNING_LIMIT):
                suggestions.append({
                    "suggestion_type": "detail_sum_equals_main",
                    "main_label": main_label,
                    "main_value": float(main_value),
                    "subtotal_label": group_key,
                    "subtotal_value": None,
                    "detail_sum": detail_sum,
                    "detail_labels": group["_side_label"].astype(str).head(12).tolist(),
                    "difference": diff,
                    "confidence": "Medium" if abs(diff) <= TOLERANCE else "Low / rounding difference",
                    "reason": "The main item value matches the sum of account-level items in the same heading context.",
                })

    unique = []
    seen = set()
    for s in suggestions:
        key = (s["suggestion_type"], _compact_for_block(s.get("main_label")), _compact_for_block(s.get("subtotal_label")), round(float(s.get("difference") or 0), 2))
        if key not in seen:
            seen.add(key)
            unique.append(s)
    return unique


def build_hierarchical_one_to_many_suggestions(df: pd.DataFrame, tolerance: float = TOLERANCE) -> pd.DataFrame:
    """Build review suggestions for possible one-to-many presentation changes.

    The function checks whether a main item in one document matches a subtotal
    or the sum of detail rows in the other document. These findings are reported
    as suggestions, not automatic matches, because they require manual review.
    """
    columns = [
        "Osio", "Main item", "Main item side", "Main item value", "Vastinpuolen subtotal",
        "Subtotalin arvo", "Alaerien summa", "Difference", "Subitems", "Suggestion confidence", "Huomio"
    ]
    if df is None or df.empty:
        return pd.DataFrame(columns=columns)

    suggestions = []
    candidate_statuses = {"Missing row", "Could not verify", "Value differs"}
    main_rows = df[df["status"].isin(candidate_statuses)].copy() if "status" in df.columns else df.copy()
    if main_rows.empty:
        return pd.DataFrame(columns=columns)

    for _, row in main_rows.iterrows():
        for main_side, detail_side in [("older", "newer"), ("newer", "older")]:
            main_label = _side_label(row, main_side)
            main_value = _side_value(row, main_side)
            if not main_label or main_value is None:
                continue
            main_norm = _norm_for_block(main_label)
            is_main_like = (
                bool(row.get(f"forced_main_item_{main_side}"))
                or bool(row.get("summary_row"))
                or "yhteensa" in main_norm
                or main_norm in {"siirtovelat", "saamiset", "muut velat", "rahat ja pankkisaamiset"}
            )
            if not is_main_like:
                continue
            for s in _extract_block_candidates_for_main(df, row, main_side, detail_side):
                suggestions.append({
                    "Section": row.get("section"),
                    "Main item": s["main_label"],
                    "Main item side": "vanhempi" if main_side == "older" else "uudempi",
                    "Main item value": s["main_value"],
                    "Vastinpuolen subtotal": s.get("subtotal_label"),
                    "Subtotalin arvo": s.get("subtotal_value"),
                    "Alaerien summa": s.get("detail_sum"),
                    "Difference": s.get("difference"),
                    "Subitems": "; ".join(s.get("detail_labels") or []),
                    "Suggestion confidence": s.get("confidence"),
                    "Huomio": s.get("reason") + " Tarkastajan tulee varmistaa esitystavan muutos.",
                })
    if not suggestions:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(suggestions).drop_duplicates().reset_index(drop=True)


def build_one_to_many_suggestions(df: pd.DataFrame, tolerance: float = TOLERANCE) -> pd.DataFrame:

    return build_hierarchical_one_to_many_suggestions(df, tolerance=tolerance)


# =========================================================
# PRESENTATION CHANGE DETECTION AND REVIEW OUTPUT
# =========================================================
# This section identifies possible changes in presentation between two financial
# statements.
#
# It detects cases where:
# - The same item appears at a different structural level.
# - A main item may have been split into subitems.
# - Labels differ but the parent or context remains similar.
# - A total or main-level row is missing as a direct one-to-one match.
#
# The purpose is not to automatically accept these cases as reconciled. Instead,
# they are flagged for review because presentation changes require professional
# judgement.

PRESENTATION_CHANGE_NONE = "Ei muutosta"
PRESENTATION_CHANGE_POSSIBLE = "Possible presentation change"
PRESENTATION_CHANGE_CLEAR = "Clear presentation change"


def _same_contextish(row) -> bool:
   
    old_path = _safe_norm(row.get("category_path_older"))
    new_path = _safe_norm(row.get("category_path_newer"))
    old_parent = _safe_norm(row.get("parent_label_older"))
    new_parent = _safe_norm(row.get("parent_label_newer"))
    old_total = str(row.get("total_context_key_older") or "")
    new_total = str(row.get("total_context_key_newer") or "")
    return bool(
        (old_path and new_path and old_path == new_path)
        or (old_parent and new_parent and old_parent == new_parent)
        or (old_total and new_total and old_total == new_total)
    )


def detect_presentation_change_row(row) -> tuple[str, str]:
    """Detect whether a reconciliation row indicates a possible presentation change.

    The function compares labels, structural levels, parent paths and match
    reasons to identify cases where the same financial statement item may have
    been renamed, split, combined or moved within the statement structure."""
    method = str(row.get("match_method") or "")
    reason = str(row.get("match_reason") or "")
    status = str(row.get("status") or "")

    old_label = _safe_norm(row.get("label_older"))
    new_label = _safe_norm(row.get("label_newer"))
    old_scope = str(row.get("block_type_older") or "")
    new_scope = str(row.get("block_type_newer") or "")
    old_forced = bool(row.get("forced_main_item_older")) or bool(row.get("balance_total_root_item_older"))
    new_forced = bool(row.get("forced_main_item_newer")) or bool(row.get("balance_total_root_item_newer"))

    if "one_to_many" in method or "one-to-many" in reason or "one_to_many" in reason:
        return PRESENTATION_CHANGE_CLEAR, "The main item appears to be presented as multiple subitems in the other file."

    if old_label and new_label and status in {"Match", "Could not verify", "Value differs"}:
        if old_scope and new_scope and old_scope != new_scope:
            if old_forced or new_forced or _same_contextish(row):
                return PRESENTATION_CHANGE_POSSIBLE, "The row appears to represent the same context, but the structural level differs."

    if old_label and new_label and old_label != new_label and _same_contextish(row):
        if method in {"synonym", "fuzzy", "fuzzy_strong", "compact", "exact"} or "matched" in reason:
            return PRESENTATION_CHANGE_POSSIBLE, "Same context, but the item label or grouping differs."

    if status == "Missing row" and (old_forced or new_forced or "yhteensa" in old_label or "yhteensa" in new_label):
        return PRESENTATION_CHANGE_POSSIBLE, "A main-level or total row is missing as a 1:1 pair; check whether it has been split or combined."

    return PRESENTATION_CHANGE_NONE, ""


def apply_presentation_change_detection(df: pd.DataFrame) -> pd.DataFrame:
    """ Add presentation-change flags and reasons to reconciliation results.

    Rows that indicate possible or clear presentation changes are not accepted
    automatically. They are marked so that the user can review whether the
    difference is caused by presentation rather than a true accounting mismatch.
    """
    if df is None or df.empty:
        return df
    out = df.copy()
    flags = out.apply(detect_presentation_change_row, axis=1)
    out["presentation_change_flag"] = [x[0] for x in flags]
    out["presentation_change_reason"] = [x[1] for x in flags]
    if "exception_category" in out.columns:
        mask = out["presentation_change_flag"].isin([PRESENTATION_CHANGE_POSSIBLE, PRESENTATION_CHANGE_CLEAR])
        neutral = out["exception_category"].fillna("").isin(["", "no exception", "ei luokiteltu"])
        out.loc[mask & neutral, "exception_category"] = "presentation change"
    return out


def build_presentation_change_df(df: pd.DataFrame) -> pd.DataFrame:
    """Build a review table containing only presentation-change findings.

    The output preserves labels, values, parent context, structure paths and
    matching method so that the user can assess the presentation change manually.
    """
    columns = ["Section", "Older label", "Newer label", "Finding", "Peruste", "Old value", "New value", "Older parent", "Newer parent", "Older structure path", "Newer structure path", "Match-tapa", "Status"]
    if df is None or df.empty or "presentation_change_flag" not in df.columns:
        return pd.DataFrame(columns=columns)
    flagged = df[df["presentation_change_flag"].isin([PRESENTATION_CHANGE_POSSIBLE, PRESENTATION_CHANGE_CLEAR])].copy()
    if flagged.empty:
        return pd.DataFrame(columns=columns)
    rows = []
    for _, r in flagged.iterrows():
        rows.append({
            "Section": r.get("section"),
            "Older label": r.get("label_older"),
            "Newer label": r.get("label_newer"),
            "Finding": r.get("presentation_change_flag"),
            "Peruste": r.get("presentation_change_reason"),
            "Old value": r.get("display_older_value"),
            "New value": r.get("display_newer_value"),
            "Older parent": r.get("parent_label_older"),
            "Newer parent": r.get("parent_label_newer"),
            "Older structure path": r.get("category_path_older"),
            "Newer structure path": r.get("category_path_newer"),
            "Match-tapa": r.get("match_method"),
            "Status": r.get("status"),
        })
    return pd.DataFrame(rows, columns=columns)

# =========================================================
# MATCH LEVEL CLASSIFICATION AND CANDIDATE REVIEW OUTPUT
# =========================================================
# This section classifies each reconciliation outcome into a match level and
# builds a candidate-level candidate review table for review.
#
# Match levels separate:
# - verified matches that can be accepted automatically,
# - suggested matches that require manual review,
# - rejected or unresolved candidates.
#
# The purpose is to make the decision boundary transparent: the tool does not
# treat all matches equally, but distinguishes between review-safe reconciliation
# and merely plausible candidate relationships.


def classify_match_level_from_row(row) -> str:
    """Classify whether a match is verified, suggested, or rejected.

    Verified means accepted without manual review. Suggested means the tool found
    a plausible relationship, but the row must still be inspected by a human.
    """
    status = str(row.get("status", ""))
    method = str(row.get("match_method", "") or "")

    if status == "Match" and is_clear_verified_match_decision(row):
        return MATCH_VERIFIED

    if status == "Match" and method in {"canonical", "exact", "compact", "exact_value_match", "compact_value_match", "synonym", "fuzzy", "fuzzy_strong"}:
        return MATCH_SUGGESTED

    if status in {"Could not verify", "Value differs"} and method not in {"unmatched_newer", "unmatched_older", "None", ""}:
        return MATCH_SUGGESTED

    return MATCH_REJECTED


def build_candidate_review_df(df: pd.DataFrame) -> pd.DataFrame:
    """Build an candidate review table for suggested, rejected and review-required candidates.

    The output preserves the candidate label, values, match method, match reason,
    confidence level, structural context and source locations so that uncertain
    reconciliation decisions can be reviewed manually."""
    columns = [
        "Osio",
        "Queried item",
        "Kandidaatti",
        "Old value",
        "New value",
        "Difference",
        "Match-taso",
        "Match method",
        "Match-syy",
        "Confidence",
        "Older parent",
        "Newer parent",
        "Older category path",
        "Newer category path",
        "Older total context",
        "Newer total context",
        "Older page",
        "Newer page",
        "Older source row",
        "Newer source row",
    ]
    if df is None or df.empty:
        return pd.DataFrame(columns=columns)

    review_rows = []
    review_df = df[df["status"].isin(["Value differs", "Missing row", "Could not verify", "Manual review"])] if "status" in df.columns else df

    for _, row in review_df.iterrows():
        review_rows.append({
            "Section": row.get("section"),
            "Queried item": row.get("label_older") or row.get("label_newer"),
            "Kandidaatti": row.get("label_newer"),
            "Old value": row.get("value_older_current"),
            "New value": row.get("value_newer_comparison"),
            "Difference": row.get("difference"),
            "Match-taso": classify_match_level_from_row(row),
            "Match method": row.get("match_method"),
            "Match-syy": row.get("match_reason"),
            "Confidence": row.get("confidence_level"),
            "Older parent": row.get("parent_label_older"),
            "Newer parent": row.get("parent_label_newer"),
            "Older category path": row.get("category_path_older"),
            "Newer category path": row.get("category_path_newer"),
            "Older total context": row.get("total_context_key_older"),
            "Newer total context": row.get("total_context_key_newer"),
            "Older page": row.get("source_page_older"),
            "Newer page": row.get("source_page_newer"),
            "Older source row": row.get("source_line_older"),
            "Newer source row": row.get("source_line_newer"),
        })

    if not review_rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(review_rows, columns=columns)

# =========================================================
# EXCEL EXPORT, QUALITY REPORTS AND REVIEW WORKBOOK FORMATTING
# =========================================================
# This section builds the Excel workbook used as the main review output.
#
# It creates:
# - Main reconciliation results.
# - Exceptions and missing-row views.
# - Management summary and test metrics.
# - Decision trace and technical review sheets.
# - Vocabulary, parent-context and unknown-row quality reports.
# - Presentation-change and one-to-many review sheets.
#
# The purpose is to turn the reconciliation result into a traceable review
# workbook where the user can review both business-level findings and technical
# evidence behind each decision.


def _worksheet_header_map(ws) -> dict:
   
    return {str(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if cell.value is not None}


def _status_fill_for_value(value, fills: dict):
    """Return a row colour for user-facing status values in the Excel workbook."""
    value = str(value or "").strip()
    if value in {"Match", "Usable", "Täsmää", "Ei toimenpidettä."}:
        return fills["green"]
    if value in {"Could not verify", "Vaatii tarkistusta", "Possible presentation change", "Tarkista", "Tarkista manuaalisesti", "Manual review"}:
        return fills["yellow"]
    if value in {"Missing row", "Puuttuu", "Rivi puuttuu", "Row missing from one document", "Missing from comparative-period financial statement", "Missing from current-period financial statement"}:
        return fills["orange"]
    if value in {"Value differs", "Luku poikkeaa", "Not recommended without manual review", "Clear presentation change"}:
        return fills["red"]
    return None

def format_workbook_for_review(wb: Workbook, fills: dict, bold_font: Font):
    """ Apply review-oriented formatting to all workbook sheets.

    The function freezes headers, enables filters, formats header rows,
    applies status-based row colouring and adjusts column widths to make the
    workbook easier to review.
    """
    header_fill = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")
    for ws in wb.worksheets:
        if ws.max_row < 1:
            continue
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = bold_font
            cell.fill = header_fill
        headers = _worksheet_header_map(ws)
        status_cols = [headers.get(h) for h in ["Päätös", "Status", "Paatos", "Decision", "Finding", "Use classification"] if headers.get(h)]
        color_code_col = headers.get("Värikoodi")
        exception_col = headers.get("Exception category")
        confidence_col = headers.get("Luottamustaso") or headers.get("Confidence level") or headers.get("Confidence")
        presentation_col = headers.get("Presentation change") or headers.get("Finding") or headers.get("Päätös")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row_fill = None
            if color_code_col:
                code = str(row[color_code_col - 1].value or "").strip().lower()
                if code == "vihreä":
                    row_fill = fills["green"]
                elif code == "keltainen":
                    row_fill = fills["yellow"]
                elif code == "punainen":
                    row_fill = fills["red"]
                elif code == "harmaa":
                    row_fill = PatternFill(fill_type="solid", start_color="E7E6E6", end_color="E7E6E6")
            if row_fill is None:
                for col_idx in status_cols:
                    fill = _status_fill_for_value(row[col_idx - 1].value, fills)
                    if fill:
                        row_fill = fill
                        break
            if row_fill is None and presentation_col:
                row_fill = _status_fill_for_value(row[presentation_col - 1].value, fills)
            if row_fill is None and exception_col:
                exc_original = row[exception_col - 1].value
                exc_norm = normalize_exception_category(exc_original)
                exc = str(exc_norm or "").lower()
                if exc_norm == "No exception":
                    row_fill = fills["green"]
                elif "esitystavan" in exc:
                    row_fill = fills["yellow"]
                elif "puuttuva" in exc or "puuttuu" in exc:
                    row_fill = fills["orange"]
                elif "value differs" in exc or "numero" in exc or "poikkeama" in exc:
                    row_fill = fills["red"]
            if row_fill is None and confidence_col:
                conf = str(row[confidence_col - 1].value or "")
                if conf in {"Low", "Uncertain", "Matala"}:
                    row_fill = fills["yellow"]
            if row_fill:
                for cell in row:
                    cell.fill = row_fill
        for column_cells in ws.columns:
            max_length = 0
            col_letter = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, min(len(str(cell.value)), 100))
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 80)


def _strip_total_suffix(label: str) -> str:
    
    norm = canonicalize_common_labels(label)
    norm = re.sub(r"\b(yhteensa|total|totalt|summa)\b", " ", norm)
    norm = re.sub(r"\s+", " ", norm).strip()
    return norm


def _row_value(row, value_col: str):
    """Row value.
    
    Purpose: This function belongs to the number parsing and validation stage.
    Why: It reduces the risk that formatting differences or unreadable values are treated as reliable evidence.
    """
    try:
        value = row.get(value_col)
        if value is None or pd.isna(value):
            return None
        return float(value)
    except Exception:
        return None


def build_sum_recalculation_df(parsed_result: dict) -> pd.DataFrame:
    
    df = (parsed_result or {}).get("df_all", pd.DataFrame())
    document = (parsed_result or {}).get("document")
    columns = [
        "Document", "Section", "Total row", "Kausi", "PDF:n summa",
        "Uudelleen laskettu summa", "Difference", "Status", "Child row count", "Child rows", "Huomio"
    ]
    if df is None or df.empty:
        return pd.DataFrame(columns=columns)

    rows = []
    work = df.copy()
    work["_label_norm"] = work.get("label", pd.Series(index=work.index, dtype=str)).fillna("").astype(str).map(canonicalize_common_labels)
    work["_parent_norm"] = work.get("parent_label", pd.Series(index=work.index, dtype=str)).fillna("").astype(str).map(canonicalize_common_labels)
    work["_path_norm"] = work.get("category_path", pd.Series(index=work.index, dtype=str)).fillna("").astype(str).map(canonicalize_common_labels)

    summary_mask = work.get("summary_row", pd.Series(False, index=work.index)).fillna(False).astype(bool)
    summary_mask &= ~(work.get("total_level", pd.Series("", index=work.index)).fillna("").astype(str) == "balance_grand_total")

    for idx, summary_row in work[summary_mask].iterrows():
        base = _strip_total_suffix(summary_row.get("label", ""))
        if not base or len(base) < 3:
            continue
        section = summary_row.get("section")
        section_mask = work.get("section", pd.Series(index=work.index, dtype=str)).fillna("").astype(str) == str(section)
        child_mask = section_mask & (work.index != idx)
        child_mask &= ~work.get("summary_row", pd.Series(False, index=work.index)).fillna(False).astype(bool)
        child_mask &= (
            (work["_parent_norm"] == base)
            | work["_path_norm"].str.contains(re.escape(base), na=False)
        )
        children = work[child_mask].copy()
        direct_children = children[children["_parent_norm"] == base].copy()
        if len(direct_children) >= 2:
            children = direct_children
        if len(children) < 2:
            continue

        for value_col, period_label in [
            ("current_value", "Kuluva kausi"),
            ("comparison_value", "Vertailukausi"),
        ]:
            pdf_sum = _row_value(summary_row, value_col)
            if value_col not in children.columns:
                continue
            numeric_children = children[children[value_col].notna()].copy()
            if pdf_sum is None or len(numeric_children) < 2:
                continue
            recalculated = float(numeric_children[value_col].sum())
            diff = pdf_sum - recalculated
            if abs(diff) <= TOLERANCE:
                status = "Match"
                note = "The PDF total row matches the sum of detected child rows."
            elif abs(diff) < ROUNDING_WARNING_LIMIT:
                status = "Check rounding"
                note = "A small difference may be caused by rounding or PDF extraction."
            else:
                status = "Tarkista"
                note = "Total-rivi ei vastaa tunnistettujen lapsirivien summaa."
            child_labels = "; ".join(numeric_children.get("label", pd.Series(dtype=str)).fillna("").astype(str).head(12).tolist())
            rows.append({
                "Document": document,
                "Section": section,
                "Total row": summary_row.get("label"),
                "Kausi": period_label,
                "PDF:n summa": pdf_sum,
                "Uudelleen laskettu summa": recalculated,
                "Difference": diff,
                "Status": status,
                "Child row count": int(len(numeric_children)),
                "Child rows": child_labels,
                "Huomio": note,
            })

    return pd.DataFrame(rows, columns=columns)


def build_sum_recalculation_report(parsed_older: Optional[dict], parsed_newer: Optional[dict]) -> pd.DataFrame:
    """Build sum recalculation report.
    
    Purpose: This function belongs to the reporting and user-interface stage.
    Why: It turns technical reconciliation results into reviewable output for the user.
    """
    frames = []
    if parsed_older:
        frames.append(build_sum_recalculation_df(parsed_older))
    if parsed_newer:
        frames.append(build_sum_recalculation_df(parsed_newer))
    frames = [f for f in frames if f is not None and not f.empty]
    if not frames:
        return pd.DataFrame(columns=[
            "Document", "Section", "Total row", "Kausi", "PDF:n summa",
            "Uudelleen laskettu summa", "Difference", "Status", "Child row count", "Child rows", "Huomio"
        ])
    return pd.concat(frames, ignore_index=True)


def build_unknown_rows_df(parsed_result: dict) -> pd.DataFrame:
    
    df = (parsed_result or {}).get("df_all", pd.DataFrame())
    document = (parsed_result or {}).get("document")
    columns = [
        "Document", "Section", "Row", "Interpreted item", "Arvo kuluva", "Arvo vertailu",
        "Sivu", "Syy", "Label-osuman pisteet", "Rakennepolku"
    ]
    if df is None or df.empty:
        return pd.DataFrame(columns=columns)

    work = df.copy()
    legal_unknown = work.get("legal_basis_level", pd.Series("", index=work.index)).fillna("").astype(str).isin(["unknown_item", ""])
    low_label_score = work.get("label_match_score", pd.Series(1.0, index=work.index)).fillna(1.0).astype(float) < 0.70
    uncertain_hierarchy = work.get("hierarchy_status", pd.Series("", index=work.index)).fillna("").astype(str).isin(["Uncertain", "Ristiriitainen"])
    medium_parse = work.get("row_parse_quality", pd.Series("", index=work.index)).fillna("").astype(str) == "medium"
    mask = legal_unknown | low_label_score | uncertain_hierarchy | medium_parse

    rows = []
    for idx, row in work[mask].iterrows():
        reasons = []
        if bool(legal_unknown.loc[idx]):
            reasons.append("ei varmaa lakipohjaista/sanastollista luokitusta")
        if bool(low_label_score.loc[idx]):
            reasons.append("matala label-osuman pisteytys")
        if bool(uncertain_hierarchy.loc[idx]):
            reasons.append("uncertain or inconsistent structure path")
        if bool(medium_parse.loc[idx]):
            reasons.append("arvon lukulaatu keskitasoinen")
        rows.append({
            "Document": document,
            "Section": row.get("section"),
            "Row": row.get("source_line"),
            "Interpreted item": row.get("label"),
            "Arvo kuluva": row.get("current_value_raw_original"),
            "Arvo vertailu": row.get("comparison_value_raw_original"),
            "Sivu": row.get("source_page"),
            "Syy": "; ".join(reasons),
            "Label-osuman pisteet": row.get("label_match_score"),
            "Rakennepolku": row.get("category_path"),
        })
    return pd.DataFrame(rows, columns=columns)


def build_unknown_rows_report(parsed_older: Optional[dict], parsed_newer: Optional[dict]) -> pd.DataFrame:
    """Build unknown rows report.
    
    Purpose: This function belongs to the reporting and user-interface stage.
    Why: It turns technical reconciliation results into reviewable output for the user.
    """
    frames = []
    if parsed_older:
        frames.append(build_unknown_rows_df(parsed_older))
    if parsed_newer:
        frames.append(build_unknown_rows_df(parsed_newer))
    frames = [f for f in frames if f is not None and not f.empty]
    if not frames:
        return pd.DataFrame(columns=[
            "Document", "Section", "Row", "Interpreted item", "Arvo kuluva", "Arvo vertailu",
            "Sivu", "Syy", "Label-osuman pisteet", "Rakennepolku"
        ])
    return pd.concat(frames, ignore_index=True)


def build_dictionary_quality_report() -> pd.DataFrame:
    
    if not EXTERNAL_TERM_ROWS:
        return pd.DataFrame([{
            "Tarkistus": "Ulkoinen sanasto",
            "Tulos": "Varoitus",
            "Havainto": "terms.xlsx ei ole käytössä tai aktiivisia rivejä ei löytynyt.",
            "Määrä": 0,
            "Esimerkit": "",
        }])

    df = pd.DataFrame(EXTERNAL_TERM_ROWS).copy()
    for col in ["canonical", "synonym", "section", "parent", "language", "category", "type", "expected_side"]:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].fillna("").astype(str).str.strip()

    valid_sections = {"", "tuloslaskelma", "tase_vastaavaa", "tase_vastattavaa"}
    valid_levels = {
        "", "main", "detail", "total", "summary", "main_item", "key_item",
        "top_level", "grand_total", "section_total", "subgroup_total",
        "balance_detail", "income_detail", "allowed_item",
    }
    allowed_canonicals = set()
    for values in SECTION_ALLOWED_CANONICALS.values():
        allowed_canonicals.update(normalize_label(v) for v in values)
    allowed_canonicals.update(normalize_label(v) for v in FORCED_MAIN_ITEM_CANONICALS)

    rows = []

    def add_check(name: str, count: int, ok_status: str, warn_status: str, note: str, examples=None):
        """Append one Finnish vocabulary quality check row."""
        if examples is None:
            example_list = []
        else:
            try:
                example_list = list(examples)
            except TypeError:
                example_list = [examples]
        rows.append({
            "Tarkistus": name,
            "Tulos": ok_status if count == 0 else warn_status,
            "Havainto": note,
            "Määrä": int(count),
            "Esimerkit": ", ".join(map(str, example_list[:10])),
        })

    duplicate_mask = df.duplicated(subset=["language", "canonical", "synonym"], keep=False)
    add_check(
        "Duplikaattitermit",
        int(duplicate_mask.sum()),
        "OK",
        "Tarkista",
        "Sama language + canonical + synonym esiintyy useammin kuin kerran.",
        df.loc[duplicate_mask, "synonym"].dropna().unique(),
    )

    missing_canonical = df["canonical"].eq("")
    add_check(
        "Empty canonical values",
        int(missing_canonical.sum()),
        "OK",
        "Tarkista",
        "Rows missing a canonical value.",
        df.loc[missing_canonical, "synonym"].dropna().unique(),
    )

    missing_synonym = df["synonym"].eq("")
    add_check(
        "Empty synonym values",
        int(missing_synonym.sum()),
        "OK",
        "Tarkista",
        "Rows missing a synonym value.",
        df.loc[missing_synonym, "canonical"].dropna().unique(),
    )

    invalid_sections = ~df["section"].isin(valid_sections)
    add_check(
        "Tuntematon section",
        int(invalid_sections.sum()),
        "OK",
        "Tarkista",
        "The section value must be tuloslaskelma, tase_vastaavaa or tase_vastattavaa.",
        df.loc[invalid_sections, "section"].dropna().unique(),
    )

    invalid_levels = ~df["type"].isin(valid_levels)
    add_check(
        "Tuntematon item_type/type",
        int(invalid_levels.sum()),
        "OK",
        "Tarkista",
        "The row type must be a known value such as main, detail or total.",
        df.loc[invalid_levels, "type"].dropna().unique(),
    )

    synonym_conflicts = (
        df[df["synonym"].ne("")]
        .groupby(["language", "synonym"])["canonical"]
        .nunique()
    )
    synonym_conflicts = synonym_conflicts[synonym_conflicts > 1]
    add_check(
        "Synonyymi usealle canonicalille",
        len(synonym_conflicts),
        "OK",
        "Tarkista",
        "The same synonym points to multiple canonical concepts. This may cause an incorrect match.",
        [f"{lang}:{syn}" for (lang, syn) in synonym_conflicts.index[:10]],
    )

    section_conflicts = (
        df[df["section"].ne("")]
        .groupby("canonical")["section"]
        .nunique()
    )
    section_conflicts = section_conflicts[section_conflicts > 1]
    known_cross_section = {"tilikauden voitto", "tilikauden tappio", "tilikauden tulos"}
    suspicious_section_conflicts = [c for c in section_conflicts.index if c not in known_cross_section]
    add_check(
        "Canonical useassa osiossa",
        len(suspicious_section_conflicts),
        "OK",
        "Tarkista",
        "The canonical concept appears in multiple sections. Some cases may be justified, but check that this is not a vocabulary error.",
        suspicious_section_conflicts,
    )

    detail_mask = df["type"].isin(["detail", "balance_detail", "income_detail"])
    missing_parent = detail_mask & df["parent"].eq("")
    add_check(
        "Detail rows missing parent",
        int(missing_parent.sum()),
        "OK",
        "Info",
        "For detail rows, parent information improves context and reduces incorrect matches.",
        df.loc[missing_parent, "canonical"].dropna().unique(),
    )

    unknown_canonical = df["canonical"].apply(lambda x: normalize_label(x) not in allowed_canonicals if x else False)
    add_check(
        "Canonical PMA-rakenteen ulkopuolella",
        int(unknown_canonical.sum()),
        "OK",
        "Info",
        "The canonical concept is not found in the internal PMA structure. This may be an allowed company-specific detail row.",
        df.loc[unknown_canonical, "canonical"].dropna().unique(),
    )

    parent_unknown = df["parent"].apply(lambda x: bool(x) and normalize_label(x) not in allowed_canonicals)
    add_check(
        "Parent PMA-rakenteen ulkopuolella",
        int(parent_unknown.sum()),
        "OK",
        "Info",
        "The parent is not found in the PMA structure. Check whether it is a company-specific parent.",
        df.loc[parent_unknown, "parent"].dropna().unique(),
    )

    rows.insert(0, {
        "Tarkistus": "Aktiiviset termirivit",
        "Tulos": "OK",
        "Finding": "Excel-sanastosta luetut aktiiviset termirivit.",
        "Count": len(df),
        "Esimerkit": "",
    })

    return pd.DataFrame(rows)


def _language_match_counts_for_parsed(parsed_result: Optional[dict]) -> Counter:
    """Language match counts for parsed.
    
    Purpose: This function belongs to the number parsing and validation stage.
    Why: It reduces the risk that formatting differences or unreadable values are treated as reliable evidence.
    """
    counts = Counter()
    if not parsed_result or not EXTERNAL_TERM_ROWS:
        return counts
    df = parsed_result.get("df_all", pd.DataFrame())
    if df is None or df.empty:
        return counts
    labels = [normalize_label(x) for x in df.get("label", pd.Series(dtype=str)).fillna("").astype(str).tolist()]
    term_by_lang = {}
    for row in EXTERNAL_TERM_ROWS:
        lang = (row.get("language") or "?").lower()
        syn = normalize_label(row.get("synonym") or "")
        if syn:
            term_by_lang.setdefault(lang, set()).add(syn)
    for label in labels:
        if not label:
            continue
        for lang, terms in term_by_lang.items():
            if label in terms or any((t and t in label) for t in terms if len(t) >= 4):
                counts[lang] += 1
    return counts


def build_language_reliability_report(parsed_older: Optional[dict], parsed_newer: Optional[dict], comparison_df: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    
    rows = []
    for label, parsed in [("Older document", parsed_older), ("Newer document", parsed_newer)]:
        counts = _language_match_counts_for_parsed(parsed)
        total_hits = sum(counts.values())
        dominant = counts.most_common(1)[0][0].upper() if counts else "Ei tunnistettu"
        df_all = parsed.get("df_all", pd.DataFrame()) if parsed else pd.DataFrame()
        row_count = len(df_all) if df_all is not None else 0
        rows.append({"Document": parsed.get("document") if parsed else label, "Rows parsed": row_count, "Vocabulary matches total": total_hits, "Detected main language": dominant, "FI matches": counts.get("fi", 0), "EN matches": counts.get("en", 0), "SV matches": counts.get("sv", 0), "Vocabulary matches per row %": round(total_hits / max(1, row_count) * 100, 1)})
    if comparison_df is not None and not comparison_df.empty:
        canon_cols = [c for c in ["canonical_label_older", "canonical_label_newer"] if c in comparison_df.columns]
        canonical_covered = int(comparison_df[canon_cols].notna().any(axis=1).sum()) if canon_cols else 0
        rows.append({"Document": "Reconciliation total", "Rows parsed": len(comparison_df), "Vocabulary matches total": canonical_covered, "Detected main language": "-", "FI matches": "-", "EN matches": "-", "SV matches": "-", "Vocabulary matches per row %": round(canonical_covered / max(1, len(comparison_df)) * 100, 1)})
    return pd.DataFrame(rows)


def build_parent_validation_report(df: pd.DataFrame) -> pd.DataFrame:
    """Highlight parent and context risks based on canonical metadata."""
    columns = ["Section", "Older label", "Newer label", "Status", "Older context", "Newer context", "Expected older parent", "Expected newer parent", "Older reason", "Newer reason"]
    if df is None or df.empty:
        return pd.DataFrame(columns=columns)
    old_status = df.get("canonical_context_status_older", pd.Series("", index=df.index)).fillna("").astype(str)
    new_status = df.get("canonical_context_status_newer", pd.Series("", index=df.index)).fillna("").astype(str)
    mask = old_status.isin(["Uncertain", "Ristiriitainen"]) | new_status.isin(["Uncertain", "Ristiriitainen"])
    subset = df[mask].copy()
    if subset.empty:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame({"Section": subset.get("section", pd.Series("", index=subset.index)), "Older label": subset.get("label_older", pd.Series("", index=subset.index)), "Newer label": subset.get("label_newer", pd.Series("", index=subset.index)), "Status": subset.get("status", pd.Series("", index=subset.index)), "Older context": subset.get("canonical_context_status_older", pd.Series("", index=subset.index)), "Newer context": subset.get("canonical_context_status_newer", pd.Series("", index=subset.index)), "Expected older parent": subset.get("canonical_expected_parent_older", pd.Series("", index=subset.index)), "Expected newer parent": subset.get("canonical_expected_parent_newer", pd.Series("", index=subset.index)), "Older reason": subset.get("canonical_context_reason_older", pd.Series("", index=subset.index)), "Newer reason": subset.get("canonical_context_reason_newer", pd.Series("", index=subset.index))})

def build_extracted_rows_report(parsed_result: Optional[dict], file_role: str) -> pd.DataFrame:
    """Return a derived value used by the reconciliation workflow."""
    if not parsed_result:
        return pd.DataFrame([{"Information": "Ei aineistoa"}])

    df = parsed_result.get("df_all", pd.DataFrame())
    if df is None or df.empty:
        return pd.DataFrame([{"Information": "No extracted rows"}])

    cols = [
        "section", "label", "label_match_canonical", "extraction_status",
        "label_match_accepted", "row_type", "block_type", "current_value",
        "comparison_value", "current_value_raw_original",
        "comparison_value_raw_original", "parent_label", "top_parent_label",
        "category_path", "source_page", "source_line_idx", "source_line",
    ]
    existing = [c for c in cols if c in df.columns]
    out = df[existing].copy()
    out.insert(0, "file_role", file_role)
    return out




def _format_excel_number(value):
    """Return a stable Excel display value without hiding missing values."""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return value


def _finnish_row_type(row_type: str) -> str:
    """Map internal row types to Finnish labels for Excel and UI."""
    return {
        "main": "pääerä",
        "detail": "erittelyrivi",
        "section_total": "välisumma",
        "grand_total": "loppusumma",
        "result": "tulosrivi",
        "heading": "otsikko",
        "noise": "ei-tietorivi",
    }.get(str(row_type or ""), str(row_type or ""))


def _finnish_extraction_status(status: str) -> str:
    """Map internal extraction status values to Finnish labels."""
    return {
        "taxonomy_recognized": "tunnistettu sanastolla",
        "extracted_unmapped": "poimittu, ei sanastotunnistusta",
    }.get(str(status or ""), str(status or ""))


def _finnish_column_status(status: str) -> str:
    """Map internal column parsing status values to Finnish labels."""
    return {
        "two_values_by_x_order": "kaksi arvoa sarakejärjestyksessä",
        "single_value_in_current_column": "yksi arvo kuluvan kauden sarakkeessa",
        "single_value_in_comparative_column": "yksi arvo vertailusarakkeessa",
        "single_value_in_ambiguous_column": "yhden arvon sarake epäselvä",
        "text_order_fallback": "tekstijärjestyksen varatulkinta",
        "single_value_text_order_fallback": "yhden arvon tekstijärjestystulkinta",
        "no_value_found": "ei arvoa",
    }.get(str(status or ""), str(status or ""))


def _file_role_label(file_role: str) -> str:
    """Return the Finnish file role label used in Excel."""
    return {"older": "Vanhempi tilinpäätös", "newer": "Uudempi tilinpäätös"}.get(file_role, file_role)


def build_all_extracted_rows_report(parsed_result: Optional[dict], file_role: str) -> pd.DataFrame:
    """Build a Finnish all-rows transparency report for one parsed document.

    This sheet is intentionally broader than the reconciliation input. Its purpose
    is to show what the program extracted from the income statement and balance
    sheet areas, even when a row was not used in automatic reconciliation.
    """
    columns = [
        "Tiedosto", "Tiedoston rooli", "Osio", "Sivu", "Rivinumero", "Alkuperäinen rivi",
        "Poimittu erän nimi", "Normalisoitu erän nimi", "Sanastokäsite", "Poiminnan tila",
        "Rivityyppi", "Rakennetaso", "Kuluvan kauden arvo", "Vertailukauden arvo",
        "Kuluvan kauden alkuperäinen arvo", "Vertailukauden alkuperäinen arvo",
        "Saraketulkinta", "Rakennelogiikka", "Parent-rivi", "Ylätason parent", "Rakennepolku",
        "Käytettävissä täsmäytykseen", "Tarkistuksen syy",
    ]
    if not parsed_result:
        return pd.DataFrame(columns=columns)
    df = parsed_result.get("df_all", pd.DataFrame())
    if df is None or df.empty:
        return pd.DataFrame(columns=columns)

    out = pd.DataFrame(index=df.index)
    out["Tiedosto"] = parsed_result.get("document", "")
    out["Tiedoston rooli"] = _file_role_label(file_role)
    out["Osio"] = df.get("section", pd.Series("", index=df.index))
    out["Sivu"] = df.get("source_page", pd.Series("", index=df.index))
    out["Rivinumero"] = df.get("source_line_idx", pd.Series("", index=df.index))
    out["Alkuperäinen rivi"] = df.get("source_line", pd.Series("", index=df.index)).apply(strip_column_value_marker)
    out["Poimittu erän nimi"] = df.get("label", pd.Series("", index=df.index)).apply(_clean_display_label)
    out["Normalisoitu erän nimi"] = df.get("normalized_label", pd.Series("", index=df.index))
    out["Sanastokäsite"] = df.get("label_match_canonical", pd.Series("", index=df.index))
    out["Poiminnan tila"] = df.get("extraction_status", pd.Series("", index=df.index)).apply(_finnish_extraction_status)
    out["Rivityyppi"] = df.get("row_type", pd.Series("", index=df.index)).apply(_finnish_row_type)
    out["Rakennetaso"] = df.get("total_level", pd.Series("", index=df.index))
    out["Kuluvan kauden arvo"] = df.get("current_value", pd.Series("", index=df.index)).apply(_format_excel_number)
    out["Vertailukauden arvo"] = df.get("comparison_value", pd.Series("", index=df.index)).apply(_format_excel_number)
    out["Kuluvan kauden alkuperäinen arvo"] = df.get("current_value_raw_original", pd.Series("", index=df.index))
    out["Vertailukauden alkuperäinen arvo"] = df.get("comparison_value_raw_original", pd.Series("", index=df.index))
    out["Saraketulkinta"] = df.get("column_parse_status", pd.Series("", index=df.index)).apply(_finnish_column_status)
    out["Rakennelogiikka"] = df.get("hierarchy_status", pd.Series("", index=df.index))
    out["Parent-rivi"] = df.get("parent_label", pd.Series("", index=df.index))
    out["Ylätason parent"] = df.get("top_parent_label", pd.Series("", index=df.index))
    out["Rakennepolku"] = df.get("category_path", pd.Series("", index=df.index))

    match_usable = (
        df.get("label_match_accepted", pd.Series(False, index=df.index)).fillna(False).astype(bool)
        & df.get("current_value", pd.Series(index=df.index)).notna()
    )
    out["Käytettävissä täsmäytykseen"] = match_usable.map({True: "Kyllä", False: "Ei"})

    reasons = []
    for _, row in df.iterrows():
        row_reasons = []
        if not bool(row.get("label_match_accepted", False)):
            row_reasons.append("sanastotunnistus puuttuu tai on epävarma")
        if row.get("current_value") is None or pd.isna(row.get("current_value")):
            row_reasons.append("kuluvan kauden arvo puuttuu")
        if str(row.get("column_parse_status", "")) in {"single_value_in_comparative_column", "single_value_in_ambiguous_column"}:
            row_reasons.append("saraketulkinta vaatii tarkistusta")
        if str(row.get("canonical_context_status", "")) in {"Uncertain", "Ristiriitainen"}:
            row_reasons.append("rakenteellinen konteksti epävarma")
        reasons.append("; ".join(dict.fromkeys(row_reasons)))
    out["Tarkistuksen syy"] = reasons
    return out.reset_index(drop=True)[columns]


def build_combined_extracted_rows_report(parsed_older: Optional[dict], parsed_newer: Optional[dict]) -> pd.DataFrame:
    """Combine all extracted rows from both documents into one transparency sheet."""
    frames = [
        build_all_extracted_rows_report(parsed_older, "older"),
        build_all_extracted_rows_report(parsed_newer, "newer"),
    ]
    frames = [f for f in frames if f is not None and not f.empty]
    if not frames:
        return pd.DataFrame([{"Tieto": "Poimittuja rivejä ei ole"}])
    return pd.concat(frames, ignore_index=True)


def _lookup_source_location_for_raw_line(parsed: dict, line: str, used_counter: dict) -> tuple[object, object]:
    """Return page and line index for a raw section line when available."""
    lookup = parsed.get("source_line_lookup", {}) if parsed else {}
    key = normalize_keyword_text(strip_column_value_marker(line))
    matches = lookup.get(key, []) if isinstance(lookup, dict) else []
    use_idx = used_counter.get(key, 0)
    if not matches:
        return "", ""
    chosen = matches[min(use_idx, len(matches) - 1)]
    used_counter[key] = use_idx + 1
    return chosen.get("source_page", ""), chosen.get("source_line_idx", "")


def _build_parsed_line_index(parsed: dict) -> dict:
    """Index parsed monetary rows by their original source line."""
    df = parsed.get("df_all", pd.DataFrame()) if parsed else pd.DataFrame()
    index = {}
    if df is None or df.empty or "source_line" not in df.columns:
        return index
    for _, row in df.iterrows():
        key = normalize_keyword_text(strip_column_value_marker(str(row.get("source_line", ""))))
        if key:
            index.setdefault(key, []).append(row)
    return index


def _classify_raw_structure_line(line: str, section_key: str, parsed_row=None) -> tuple[str, str, str, str, str, str]:
    """Classify a raw statement line for transparent Excel reporting."""
    clean_line = strip_column_value_marker(line)
    if parsed_row is not None:
        label = _clean_display_label(parsed_row.get("label", ""))
        canonical = parsed_row.get("label_match_canonical", "")
        row_type = _finnish_row_type(parsed_row.get("row_type", ""))
        current_value = _format_excel_number(parsed_row.get("current_value", ""))
        comparison_value = _format_excel_number(parsed_row.get("comparison_value", ""))
        status = "poimittu täsmäytysdatan käsittelyyn"
        return label, canonical, row_type, current_value, comparison_value, status

    heading = _heading_candidate_from_line(clean_line, section_key)
    if heading:
        return clean_line, heading, "otsikko", "", "", "rakenteellinen rivi, ei rahamääräinen täsmäytysrivi"

    if NUMBER_PATTERN.search(clean_line):
        return clean_line, "", "poimittu raakarivi", "", "", "rivi näkyy läpinäkyvyysnäkymässä, mutta sitä ei voitu jäsentää täsmäytysriviksi"

    if looks_like_non_data_line(clean_line):
        return clean_line, "", "ohitettu rivi", "", "", "rivi ei vaikuta täsmäytykseen"

    return clean_line, "", "rakenteellinen tai tunnistamaton rivi", "", "", "näytetään tarkistettavuuden vuoksi"




def _normalize_source_line_key(value) -> str:
    """Build a stable key for linking reconciliation rows back to structure rows."""
    return normalize_keyword_text(strip_column_value_marker(str(value or "")))


def _safe_display_status(value) -> str:
    """Return a Finnish status label used in structure sheets."""
    return map_user_status(str(value or "")) if value is not None else ""


def _structure_match_lookup_key(section, source_line, label, canonical) -> tuple[str, str, str, str]:
    """Build a fallback key for structure-sheet reconciliation status lookup."""
    return (
        normalize_section_for_matching(section),
        _normalize_source_line_key(source_line),
        canonicalize_common_labels(label or ""),
        canonicalize_common_labels(canonical or ""),
    )


def build_structure_status_lookup(df: pd.DataFrame) -> dict:
    """Create lookup tables for colouring income statement and balance sheet rows.

    The lookup links raw structure rows back to the reconciliation result. Source
    line keys are preferred because they are the most transparent evidence. A
    canonical fallback is also stored for cases where the PDF line was cleaned
    before reconciliation.
    """
    lookup = {"source": {}, "canonical": {}}
    if df is None or df.empty:
        return lookup

    for _, row in df.iterrows():
        status = str(row.get("status", "") or "")
        confidence = str(row.get("confidence_level", "") or "")
        difference = row.get("difference", "")
        decision = _safe_display_status(status)

        if status == "Match":
            colour = "vihreä"
            review_status = "Täsmää"
        elif status in {"Value differs"}:
            colour = "punainen"
            review_status = "Poikkeama"
        elif status in {"Missing row", "Could not verify", "Manual review"}:
            colour = "keltainen" if status != "Missing row" else "punainen"
            review_status = "Tarkistettava" if status != "Missing row" else "Puuttuva vastinrivi"
        else:
            colour = "keltainen" if confidence in {"Low", "Uncertain", "Matala"} else ""
            review_status = "Tarkistettava" if colour else ""

        if not colour:
            continue

        payload = {
            "Täsmäytyksen tila": review_status,
            "Täsmäytyksen päätös": decision,
            "Värikoodi": colour,
            "Täsmäytyksen peruste": row.get("match_reason", ""),
            "Ero": difference,
        }

        for side in ["older", "newer"]:
            section = row.get("section", "")
            source_line = row.get(f"source_line_{side}", "")
            label = row.get(f"label_{side}", "")
            canonical = row.get(f"label_match_canonical_{side}", "")
            source_key = (normalize_section_for_matching(section), _normalize_source_line_key(source_line))
            if source_key[1]:
                lookup["source"][source_key] = payload
            canon_key = (normalize_section_for_matching(section), canonicalize_common_labels(canonical or label))
            if canon_key[1]:
                # Do not overwrite a stronger source-based or earlier exact decision.
                lookup["canonical"].setdefault(canon_key, payload)
    return lookup


def _status_for_structure_row(parsed_row, clean_line: str, section_key: str, status_lookup: dict) -> dict:
    """Return reconciliation status fields for one row in a structure sheet."""
    label = parsed_row.get("label", "") if parsed_row is not None else clean_line
    canonical = parsed_row.get("label_match_canonical", "") if parsed_row is not None else canonicalize_common_labels(clean_line)
    source_key = (normalize_section_for_matching(section_key), _normalize_source_line_key(clean_line))
    payload = status_lookup.get("source", {}).get(source_key)
    if payload:
        return payload

    canon_key = (normalize_section_for_matching(section_key), canonicalize_common_labels(canonical or label))
    payload = status_lookup.get("canonical", {}).get(canon_key)
    if payload:
        return payload

    if parsed_row is None:
        if NUMBER_PATTERN.search(clean_line):
            return {
                "Täsmäytyksen tila": "Tarkistettava",
                "Täsmäytyksen päätös": "Ei jäsennetty täsmäytysriviksi",
                "Värikoodi": "keltainen",
                "Täsmäytyksen peruste": "Rivi sisältää luvun, mutta sitä ei voitu jäsentää turvallisesti täsmäytykseen.",
                "Ero": "",
            }
        return {
            "Täsmäytyksen tila": "Ei täsmäytettävä",
            "Täsmäytyksen päätös": "Rakenteellinen rivi",
            "Värikoodi": "harmaa",
            "Täsmäytyksen peruste": "Rivi näytetään rakenteen läpinäkyvyyden vuoksi.",
            "Ero": "",
        }

    # Parsed monetary row exists but it was not found in the final reconciliation result.
    if parsed_row.get("current_value") is not None or parsed_row.get("comparison_value") is not None:
        return {
            "Täsmäytyksen tila": "Tarkistettava",
            "Täsmäytyksen päätös": "Ei mukana varsinaisessa täsmäytyksessä",
            "Värikoodi": "keltainen",
            "Täsmäytyksen peruste": "Rivi poimittiin rahariviksi, mutta sille ei löytynyt turvallista täsmäytyspäätöstä.",
            "Ero": "",
        }

    return {
        "Täsmäytyksen tila": "Ei täsmäytettävä",
        "Täsmäytyksen päätös": "Rakenteellinen rivi",
        "Värikoodi": "harmaa",
        "Täsmäytyksen peruste": "Rivi näytetään rakenteen läpinäkyvyyden vuoksi.",
        "Ero": "",
    }



def _loose_display_split_label_and_values(line: str) -> tuple[str, str, str]:
    """Split a raw PDF-like row into label, current value and comparative value for display only.

    This helper is intentionally presentation-only. It makes the Excel structure
    easier to read when a line contains values but the reconciliation parser did
    not accept it as a safe match row. It does not force the row into automatic
    reconciliation.
    """
    clean_line = strip_column_value_marker(line)
    matches = list(NUMBER_PATTERN.finditer(clean_line))
    if not matches:
        return clean_line, "", ""
    value_matches = matches
    if value_matches and value_matches[0].start() <= 1 and re.fullmatch(r"\d{3,6}", value_matches[0].group(0).strip()):
        value_matches = value_matches[1:]
    if not value_matches:
        return clean_line, "", ""
    values = [m.group(0).strip() for m in value_matches]
    first_value_start = value_matches[0].start()
    label = clean_line[:first_value_start].strip(" -–,;:") or clean_line
    if len(values) >= 2:
        return label, values[-2], values[-1]
    return label, values[-1], ""


def _pdf_look_display_label(label: str, parsed_row=None) -> str:
    """Return an indented label for the PDF-like Excel presentation."""
    label = _clean_display_label(label)
    level = 0
    if parsed_row is not None:
        try:
            raw_level = parsed_row.get("hierarchy_level", 0)
            if raw_level is not None and not pd.isna(raw_level):
                level = max(0, min(int(raw_level), 4))
        except Exception:
            level = 0
        row_type = str(parsed_row.get("row_type", "") or "")
        if row_type == "detail":
            level = max(level, 1)
    return f"{'    ' * level}{label}" if label else label


def build_pdf_like_financial_statement_report(
    parsed_result: Optional[dict],
    file_role: str,
    comparison_df: Optional[pd.DataFrame] = None,
    colour_rows: bool = True,
) -> pd.DataFrame:
    """Build a PDF-like Excel view of one financial statement.

    The sheet keeps the income statement and balance sheet in one document-like
    view: item name, current-period value and comparative-period value. Colouring
    is added as a review layer and does not change reconciliation decisions.
    """
    columns = [
        "Tiedosto", "Tiedoston rooli", "Osio", "Erä", "Kuluva kausi", "Vertailukausi",
        "Täsmäytyksen tila", "Täsmäytyksen päätös", "Selite", "Värikoodi",
        "Sivu", "Rivinumero", "Sanastokäsite", "Rivityyppi", "Saraketulkinta", "Alkuperäinen rivi",
    ]
    if not parsed_result:
        return pd.DataFrame(columns=columns)

    status_lookup = build_structure_status_lookup(comparison_df) if colour_rows else {"source": {}, "canonical": {}}
    parsed_index = _build_parsed_line_index(parsed_result)
    section_lines = parsed_result.get("section_lines", {}) or {}
    section_order = [
        ("tuloslaskelma", "TULOSLASKELMA"),
        ("tase_vastaavaa", "TASE VASTAAVAA"),
        ("tase_vastattavaa", "TASE VASTATTAVAA"),
    ]
    used_source_counter: dict[str, int] = {}
    used_parsed_counter: dict[str, int] = {}
    rows: list[dict] = []

    for section_key, section_title in section_order:
        raw_lines = section_lines.get(section_key, []) or []
        if not raw_lines:
            continue
        rows.append({
            "Tiedosto": parsed_result.get("document", ""),
            "Tiedoston rooli": _file_role_label(file_role),
            "Osio": section_title.title(),
            "Erä": section_title,
            "Kuluva kausi": "",
            "Vertailukausi": "",
            "Täsmäytyksen tila": "Ei täsmäytettävä",
            "Täsmäytyksen päätös": "Rakenteellinen otsikko",
            "Selite": "Tilinpäätöksen osion otsikko.",
            "Värikoodi": "harmaa" if colour_rows else "",
            "Sivu": "",
            "Rivinumero": "",
            "Sanastokäsite": "",
            "Rivityyppi": "otsikko",
            "Saraketulkinta": "",
            "Alkuperäinen rivi": section_title,
        })

        for raw_line in raw_lines:
            clean_line = strip_column_value_marker(raw_line)
            line_key = normalize_keyword_text(clean_line)
            parsed_matches = parsed_index.get(line_key, [])
            parsed_idx = used_parsed_counter.get(line_key, 0)
            parsed_row = parsed_matches[min(parsed_idx, len(parsed_matches) - 1)] if parsed_matches else None
            if parsed_matches:
                used_parsed_counter[line_key] = parsed_idx + 1

            page, line_idx = _lookup_source_location_for_raw_line(parsed_result, clean_line, used_source_counter)
            label, canonical, row_type, current_value, comparison_value, extraction_status = _classify_raw_structure_line(clean_line, section_key, parsed_row)

            # Presentation-only fallback: show values in the PDF-like sheet even
            # when a raw value line was not accepted for reconciliation.
            if parsed_row is None and NUMBER_PATTERN.search(clean_line):
                loose_label, loose_current, loose_comparison = _loose_display_split_label_and_values(clean_line)
                label = loose_label or label
                current_value = loose_current
                comparison_value = loose_comparison

            status_payload = _status_for_structure_row(parsed_row, clean_line, section_key, status_lookup) if colour_rows else {
                "Täsmäytyksen tila": "",
                "Täsmäytyksen päätös": "",
                "Värikoodi": "",
                "Täsmäytyksen peruste": "",
                "Ero": "",
            }
            if not colour_rows:
                # In the older comparative-period file, colour only rows that are
                # directly marked as missing or uncertain through reconciliation.
                missing_or_uncertain = str(status_payload.get("Täsmäytyksen tila", "")) in {"Puuttuva vastinrivi", "Tarkistettava"}
                if not missing_or_uncertain:
                    status_payload = {"Täsmäytyksen tila": "", "Täsmäytyksen päätös": "", "Värikoodi": "", "Täsmäytyksen peruste": "", "Ero": ""}

            display_label = _pdf_look_display_label(label, parsed_row)
            rows.append({
                "Tiedosto": parsed_result.get("document", ""),
                "Tiedoston rooli": _file_role_label(file_role),
                "Osio": {
                    "tuloslaskelma": "Tuloslaskelma",
                    "tase_vastaavaa": "Tase vastaavaa",
                    "tase_vastattavaa": "Tase vastattavaa",
                }.get(section_key, section_key),
                "Erä": display_label,
                "Kuluva kausi": current_value,
                "Vertailukausi": comparison_value,
                "Täsmäytyksen tila": status_payload.get("Täsmäytyksen tila", ""),
                "Täsmäytyksen päätös": status_payload.get("Täsmäytyksen päätös", ""),
                "Selite": status_payload.get("Täsmäytyksen peruste", extraction_status),
                "Värikoodi": status_payload.get("Värikoodi", ""),
                "Sivu": page,
                "Rivinumero": line_idx,
                "Sanastokäsite": canonical,
                "Rivityyppi": row_type,
                "Saraketulkinta": _finnish_column_status(parsed_row.get("column_parse_status", "")) if parsed_row is not None else "",
                "Alkuperäinen rivi": clean_line,
            })

    if not rows:
        return pd.DataFrame([{"Tieto": "Tilinpäätösrivejä ei ole"}])
    return pd.DataFrame(rows, columns=columns)

def build_statement_structure_report(parsed_older: Optional[dict], parsed_newer: Optional[dict], statement: str, comparison_df: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    """Build a full Finnish structure sheet from raw statement section lines.

    The report intentionally shows more than the reconciliation engine uses:
    all raw income statement and balance sheet lines are listed, and parsed
    monetary rows are annotated with their values and canonical recognition.
    """
    statement_key = normalize_section_name(statement)
    section_keys = ["tuloslaskelma"] if statement_key == "tuloslaskelma" else ["tase_vastaavaa", "tase_vastattavaa"]
    status_lookup = build_structure_status_lookup(comparison_df)
    frames = []

    for parsed, role in [(parsed_older, "older"), (parsed_newer, "newer")]:
        if not parsed:
            continue
        parsed_index = _build_parsed_line_index(parsed)
        used_source_counter = {}
        used_parsed_counter = {}
        section_lines = parsed.get("section_lines", {}) or {}
        for section_key in section_keys:
            raw_lines = section_lines.get(section_key, []) or []
            if not raw_lines:
                continue
            rows = []
            for order, raw_line in enumerate(raw_lines, start=1):
                clean_line = strip_column_value_marker(raw_line)
                line_key = normalize_keyword_text(clean_line)
                parsed_matches = parsed_index.get(line_key, [])
                parsed_idx = used_parsed_counter.get(line_key, 0)
                parsed_row = parsed_matches[min(parsed_idx, len(parsed_matches) - 1)] if parsed_matches else None
                if parsed_matches:
                    used_parsed_counter[line_key] = parsed_idx + 1

                page, line_idx = _lookup_source_location_for_raw_line(parsed, clean_line, used_source_counter)
                label, canonical, row_type, current_value, comparison_value, extraction_status = _classify_raw_structure_line(clean_line, section_key, parsed_row)
                section_display = {
                    "tuloslaskelma": "Tuloslaskelma",
                    "tase_vastaavaa": "Tase vastaavaa",
                    "tase_vastattavaa": "Tase vastattavaa",
                }.get(section_key, section_key)

                rows.append({
                    "Tiedosto": parsed.get("document", ""),
                    "Tiedoston rooli": _file_role_label(role),
                    "Sivu": page,
                    "Rivinumero": line_idx,
                    "Järjestys": order,
                    "Osio": section_display,
                    "Puoli": "Vastaavaa" if section_key == "tase_vastaavaa" else ("Vastattavaa" if section_key == "tase_vastattavaa" else ""),
                    "Alkuperäinen rivi": clean_line,
                    "Erän nimi": label,
                    "Sanastokäsite": canonical,
                    "Kuluvan kauden arvo": current_value,
                    "Vertailukauden arvo": comparison_value,
                    "Rivityyppi": row_type,
                    "Poiminnan tila": extraction_status,
                    "Rakennepolku": parsed_row.get("category_path", "") if parsed_row is not None else "",
                    **_status_for_structure_row(parsed_row, clean_line, section_key, status_lookup),
                    "Saraketulkinta": _finnish_column_status(parsed_row.get("column_parse_status", "")) if parsed_row is not None else "",
                })
            section_df = pd.DataFrame(rows)
            if statement_key == "tuloslaskelma" and "Puoli" in section_df.columns:
                section_df = section_df.drop(columns=["Puoli"])
            frames.append(section_df)

    if not frames:
        return pd.DataFrame([{"Tieto": "Rakennerivejä ei ole"}])
    return pd.concat(frames, ignore_index=True)

def build_term_recognition_report(parsed_older: Optional[dict], parsed_newer: Optional[dict]) -> pd.DataFrame:
    """Show how extracted labels were mapped to vocabulary terms."""
    rows = build_combined_extracted_rows_report(parsed_older, parsed_newer)
    if rows is None or rows.empty or "Poimittu erän nimi" not in rows.columns:
        return pd.DataFrame([{"Tieto": "Termien tunnistusta ei ole"}])
    cols = [
        "Tiedosto", "Tiedoston rooli", "Osio", "Poimittu erän nimi", "Normalisoitu erän nimi",
        "Sanastokäsite", "Poiminnan tila", "Rivityyppi", "Tarkistuksen syy",
    ]
    existing = [c for c in cols if c in rows.columns]
    return rows[existing].copy()



def build_pretax_result_control_report(df: pd.DataFrame) -> pd.DataFrame:
    """Highlight how the pre-tax result row was handled in reconciliation."""
    columns = [
        "Osio", "Erä vanhemmassa", "Erä uudemmassa", "Päätös", "Varmuustaso",
        "Vanhemman arvo", "Uudemman vertailuarvo", "Ero", "Täsmäytyksen peruste",
        "Vanhemman lähderivi", "Uudemman lähderivi",
    ]
    if df is None or df.empty:
        return pd.DataFrame(columns=columns)
    target = "voitto tappio ennen tilinpaatossiirtoja ja veroja"
    old_can = df.get("label_match_canonical_older", pd.Series("", index=df.index)).fillna("").astype(str).map(canonicalize_common_labels)
    new_can = df.get("label_match_canonical_newer", pd.Series("", index=df.index)).fillna("").astype(str).map(canonicalize_common_labels)
    old_label = df.get("label_older", pd.Series("", index=df.index)).fillna("").astype(str).map(canonicalize_common_labels)
    new_label = df.get("label_newer", pd.Series("", index=df.index)).fillna("").astype(str).map(canonicalize_common_labels)
    mask = old_can.eq(target) | new_can.eq(target) | old_label.eq(target) | new_label.eq(target)
    subset = df[mask].copy()
    if subset.empty:
        return pd.DataFrame([{
            "Osio": "Tuloslaskelma",
            "Erä vanhemmassa": "voitto/tulos ennen veroja",
            "Erä uudemmassa": "voitto/tulos ennen veroja",
            "Päätös": "Tarkista",
            "Varmuustaso": "Matala",
            "Vanhemman arvo": "",
            "Uudemman vertailuarvo": "",
            "Ero": "",
            "Täsmäytyksen peruste": "Riviä ei löytynyt täsmäytysdatasta. Tarkista Tuloslaskelma-välilehdeltä, näkyykö rivi raakarivinä.",
            "Vanhemman lähderivi": "",
            "Uudemman lähderivi": "",
        }])
    out = pd.DataFrame(index=subset.index)
    out["Osio"] = subset.get("section", pd.Series("", index=subset.index))
    out["Erä vanhemmassa"] = subset.get("label_older", pd.Series("", index=subset.index))
    out["Erä uudemmassa"] = subset.get("label_newer", pd.Series("", index=subset.index))
    out["Päätös"] = subset.get("status", pd.Series("", index=subset.index)).map(map_user_status)
    out["Varmuustaso"] = subset.get("confidence_level", pd.Series("", index=subset.index)).map(map_user_confidence)
    out["Vanhemman arvo"] = subset.get("display_older_value", pd.Series("", index=subset.index))
    out["Uudemman vertailuarvo"] = subset.get("display_newer_value", pd.Series("", index=subset.index))
    out["Ero"] = subset.get("difference", pd.Series("", index=subset.index))
    out["Täsmäytyksen peruste"] = subset.get("match_reason", pd.Series("", index=subset.index))
    out["Vanhemman lähderivi"] = subset.get("source_line_older", pd.Series("", index=subset.index))
    out["Uudemman lähderivi"] = subset.get("source_line_newer", pd.Series("", index=subset.index))
    return out.reset_index(drop=True)

def build_excel_usage_guide_df() -> pd.DataFrame:
    """Describe the Finnish workbook structure for the user."""
    rows = [
        {"Järjestys": 1, "Välilehti": "Ohje", "Tarkoitus": "Kertoo, mitä työkalu tekee ja miten raporttia luetaan.", "Käyttäjän toiminta": "Aloita tästä. Huomaa, että työkalu tukee täsmäytystä eikä tee lopullisia johtopäätöksiä käyttäjän puolesta."},
        {"Järjestys": 2, "Välilehti": "Yhteenveto", "Tarkoitus": "Näyttää täsmäytyksen keskeiset määrät, poikkeamat ja tarkistusta vaativat havainnot.", "Käyttäjän toiminta": "Katso ensin kokonaiskuva ja tarkistettavien rivien määrä."},
        {"Järjestys": 3, "Välilehti": "Kuluvan kauden tilinpäätös", "Tarkoitus": "Näyttää uudemman PDF-tilinpäätöksen Excel-muodossa: tuloslaskelma, tase vastaavaa ja tase vastattavaa PDF:n kaltaisessa järjestyksessä.", "Käyttäjän toiminta": "Tarkista erityisesti vertailukausi-sarake. Vihreä = täsmää, keltainen = tarkistettava, punainen = poikkeama tai puuttuva vastinrivi, harmaa = rakenteellinen rivi."},
        {"Järjestys": 4, "Välilehti": "Vertailukauden tilinpäätös", "Tarkoitus": "Näyttää vanhemman PDF-tilinpäätöksen Excel-muodossa vertailun lähteeksi. Värejä käytetään vain, jos rivi vaatii huomiota kuluvan kauden tilinpäätökseen nähden.", "Käyttäjän toiminta": "Käytä tätä lähdenäkymänä, jos kuluvan kauden tilinpäätöksen vertailuluku ei täsmää tai vastinrivi puuttuu."},
        {"Järjestys": 5, "Välilehti": "Tarkistusta vaativat", "Tarkoitus": "Nostaa yhteen näkymään rivit, joissa on poikkeama, puuttuva vastinrivi tai epävarma tulkinta.", "Käyttäjän toiminta": "Käy nämä rivit läpi tuloslaskelma- ja tasevälilehtien jälkeen."},
        {"Järjestys": 6, "Välilehti": "Täsmäytys", "Tarkoitus": "Näyttää varsinaiset vertailukauden lukujen täsmäytystulokset.", "Käyttäjän toiminta": "Käytä tätä varsinaisten täsmäytysten läpikäyntiin."},
        {"Järjestys": 8, "Välilehti": "Kaikki poimitut rivit", "Tarkoitus": "Näyttää kaikki tuloslaskelman ja taseen alueilta poimitut rivit, myös ne joita ei käytetty täsmäytyksessä.", "Käyttäjän toiminta": "Käytä tätä läpinäkyvyyden ja kattavuuden tarkistamiseen."},
        {"Järjestys": 9, "Välilehti": "Päätöksen lisätiedot", "Tarkoitus": "Näyttää täsmäytyspäätösten tarkemmat perusteet, lähdesivut ja lähderivit.", "Käyttäjän toiminta": "Käytä tätä, jos haluat ymmärtää miksi rivi hyväksyttiin tai nostettiin tarkistukseen."},
        {"Järjestys": 10, "Välilehti": "Termien tunnistus", "Tarkoitus": "Näyttää, miten poimitut eränimet yhdistyivät sanastokäsitteisiin.", "Käyttäjän toiminta": "Käytä tätä sanaston kehittämiseen ja epäselvien termien tarkistamiseen."},
        {"Järjestys": 11, "Välilehti": "Sanaston laatu", "Tarkoitus": "Näyttää terms.xlsx-sanaston tekniset laatuhuomiot.", "Käyttäjän toiminta": "Käytä tätä sanaston ylläpidossa."},
    ]
    return pd.DataFrame(rows)


def build_exceptions_display_df(df: pd.DataFrame) -> pd.DataFrame:
    """Return a derived value used by the reconciliation workflow."""
    if df is None or df.empty:
        return pd.DataFrame()
    mask = df["status"].isin(["Value differs", "Missing row"])
    return make_display_df(df[mask].copy())


def build_missing_rows_display_df(df: pd.DataFrame) -> pd.DataFrame:
    """Return a derived value used by the reconciliation workflow."""
    if df is None or df.empty:
        return pd.DataFrame()
    mask = df["status"].isin(["Missing row"])
    return make_display_df(df[mask].copy())


def export_comparison_to_excel_bytes(df: pd.DataFrame, parsed_older: Optional[dict] = None, parsed_newer: Optional[dict] = None):
    """Build and export a Finnish reconciliation workbook as Excel bytes.

    The workbook separates the core reconciliation result from transparency
    views. This keeps the thesis scope focused on comparative-period
    reconciliation while still showing all extracted income statement and
    balance sheet rows for user review.
    """
    wb = Workbook()
    green_fill = PatternFill(fill_type="solid", start_color="D9F2D9", end_color="D9F2D9")
    yellow_fill = PatternFill(fill_type="solid", start_color="FFF3CD", end_color="FFF3CD")
    orange_fill = PatternFill(fill_type="solid", start_color="FFE5B4", end_color="FFE5B4")
    red_fill = PatternFill(fill_type="solid", start_color="F8D7DA", end_color="F8D7DA")
    blue_fill = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")
    fills = {"green": green_fill, "yellow": yellow_fill, "orange": orange_fill, "red": red_fill, "blue": blue_fill}
    bold_font = Font(bold=True)

    def write_sheet(ws, data: pd.DataFrame):
        if data is None or data.empty:
            data = pd.DataFrame([{"Tieto": "Ei rivejä"}])
        for row in dataframe_to_rows(data, index=False, header=True):
            ws.append(row)
        return ws

    def add_sheet(name: str, data: pd.DataFrame):
        ws = wb.create_sheet(name[:31])
        return write_sheet(ws, data)

    ws_default = wb.active
    ws_default.title = "Ohje"
    write_sheet(ws_default, build_excel_usage_guide_df())

    add_sheet("Yhteenveto", build_management_dashboard_df(df))
    add_sheet("Kuluvan kauden tilinpäätös", build_pdf_like_financial_statement_report(parsed_newer, "newer", df, colour_rows=True))
    add_sheet("Vertailukauden tilinpäätös", build_pdf_like_financial_statement_report(parsed_older, "older", df, colour_rows=True))

    review_df = filter_review_rows(df)
    review_display = make_display_df(review_df) if not review_df.empty else pd.DataFrame([{"Tieto": "Ei tarkistusta vaativia rivejä"}])
    add_sheet("Tarkistusta vaativat", review_display)

    add_sheet("Täsmäytys", make_display_df(df))
    add_sheet("Kaikki poimitut rivit", build_combined_extracted_rows_report(parsed_older, parsed_newer))
    add_sheet("Päätöksen lisätiedot", build_decision_trace_display_df(df))
    add_sheet("Termien tunnistus", build_term_recognition_report(parsed_older, parsed_newer))
    add_sheet("Sanaston laatu", build_dictionary_quality_report())

    o2m_df = build_one_to_many_suggestions(df)
    if not o2m_df.empty:
        add_sheet("Yksi-moneen ehdotukset", o2m_df)

    pres_df = build_presentation_change_df(df)
    if not pres_df.empty:
        add_sheet("Esitystavan muutokset", pres_df)

    unknown_rows_df = build_unknown_rows_report(parsed_older, parsed_newer)
    if not unknown_rows_df.empty:
        add_sheet("Tuntemattomat rivit", unknown_rows_df)

    parent_validation_df = build_parent_validation_report(df)
    if not parent_validation_df.empty:
        add_sheet("Rakennevalidointi", parent_validation_df)

    format_workbook_for_review(wb, fills, bold_font)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# =========================================================
# STREAMLIT USER INTERFACE
# =========================================================

def main() -> None:
    
    st.set_page_config(page_title="Tilinpäätöksen vertailulukujen täsmäytys", layout="wide")

    if "comparison_df" not in st.session_state:
        st.session_state["comparison_df"] = None
        st.session_state["parsed_older"] = None
        st.session_state["parsed_newer"] = None
        st.session_state["timestamp"] = None

    st.title("Tilinpäätöksen vertailulukujen täsmäytystyökalu")
    st.caption(EXTERNAL_TERMS_MESSAGE)

    st.markdown(
        """
        Tämä opinnäytetyön prototyyppi täsmäyttää vanhemman tilinpäätöksen kuluvan kauden luvut
        uudemman tilinpäätöksen vertailukauden lukuihin. Raportti näyttää päätöksen, perusteen,
        käytetyt lähderivit sekä tarkistusta vaativat havainnot.
        """
    )

    col1, col2 = st.columns(2)
    with col1:
        older_file = st.file_uploader("1. Vanhempi PDF-tilinpäätös", type=["pdf"])
    with col2:
        newer_file = st.file_uploader("2. Uudempi PDF-tilinpäätös, jossa on vertailukausi", type=["pdf"])

    if st.button("Täsmäytä tilinpäätökset", type="primary"):
        if older_file is None or newer_file is None:
            st.warning("Lataa molemmat PDF-tiedostot.")
        else:
            try:
                older_bytes = older_file.read()
                newer_bytes = newer_file.read()
                parsed_older = parse_sme_financial_statement_from_bytes(older_bytes, older_file.name)
                parsed_newer = parse_sme_financial_statement_from_bytes(newer_bytes, newer_file.name)
                comparison_df = compare_all_rows(parsed_older, parsed_newer)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                st.session_state["comparison_df"] = comparison_df
                st.session_state["parsed_older"] = parsed_older
                st.session_state["parsed_newer"] = parsed_newer
                st.session_state["timestamp"] = timestamp
                st.success("Täsmäytys valmistui.")
            except Exception as e:
                st.error("Täsmäytystä ei voitu suorittaa luotettavasti.")
                st.warning(str(e))
                st.info("Tarkista, että PDF on tekstipohjainen ja sisältää tuloslaskelman sekä taseen. Tekninen virheloki näkyy alla.")
                with st.expander("Tekninen virheloki"):
                    st.code(traceback.format_exc())

    if st.session_state["comparison_df"] is not None:
        comparison_df = st.session_state["comparison_df"]
        parsed_older = st.session_state["parsed_older"]
        parsed_newer = st.session_state["parsed_newer"]
        timestamp = st.session_state["timestamp"]

        total_rows = len(comparison_df)
        ok_rows = int((comparison_df["status"] == "Match").sum())
        diff_rows = int((comparison_df["status"] == "Value differs").sum())
        missing_rows = int((comparison_df["status"] == "Missing row").sum())
        review_df = filter_review_rows(comparison_df)
        review_rows = len(review_df)
        older_found_rows = _count_present_reconciliation_rows(comparison_df, "label_older", "value_older_current")
        newer_found_rows = _count_present_reconciliation_rows(comparison_df, "label_newer", "value_newer_comparison")
        success_rate = round(ok_rows / max(1, total_rows) * 100, 1)

        rel_old = parsed_older.get("document_reliability", {}) or {}
        rel_new = parsed_newer.get("document_reliability", {}) or {}
        combined_score = None
        if rel_old.get("score") is not None and rel_new.get("score") is not None:
            combined_score = round((float(rel_old.get("score")) + float(rel_new.get("score"))) / 2, 1)

        # =====================================================
        # 1. Summary
        # =====================================================
        st.subheader("Yhteenveto")
        kpi1, kpi2, kpi3, kpi4, kpi5, kpi6 = st.columns(6)
        kpi1.metric("Kuluvan kauden rivit", older_found_rows)
        kpi2.metric("Vertailukauden rivit", newer_found_rows)
        kpi3.metric("Täsmää", ok_rows)
        kpi4.metric("Luku poikkeaa", diff_rows)
        kpi5.metric("Tarkistettavat", review_rows)
        kpi6.metric("Onnistumisaste", f"{success_rate} %")

        if combined_score is not None:
            if combined_score >= 80:
                st.success(f"Aineistojen keskimääräinen luotettavuuspisteytys: {combined_score}/100")
            elif combined_score >= 60:
                st.warning(f"Aineistojen keskimääräinen luotettavuuspisteytys: {combined_score}/100")
            else:
                st.error(f"Aineistojen keskimääräinen luotettavuuspisteytys: {combined_score}/100")

        fresh_excel_bytes = export_comparison_to_excel_bytes(comparison_df, parsed_older, parsed_newer)
        st.download_button(
            "Lataa Excel-raportti",
            data=fresh_excel_bytes,
            file_name=f"tilinpaatos_vertailu_{timestamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # =====================================================
        # =====================================================
        st.subheader("Poikkeamat")
        exceptions_df = comparison_df[comparison_df["status"].isin(["Value differs", "Missing row"])].copy()
        if exceptions_df.empty:
            st.success("Poikkeavia lukuja tai puuttuvia rivejä ei havaittu.")
        else:
            st.dataframe(make_display_df(exceptions_df).style.apply(status_row_style, axis=1), use_container_width=True, height=420)

        # =====================================================
        # 3. To be checked
        # =====================================================
        st.subheader("Tarkistusta vaativat rivit")
        if review_df.empty:
            st.success("Yksikään rivi ei vaadi manuaalista tarkistusta.")
        else:
            filter_options = ["Kaikki", "Manuaalinen tarkistus", "Ei voitu varmistaa", "Esitystavan muutos"]
            selected_filter = st.selectbox("Suodata tarkistettavia rivejä", filter_options)
            filtered_review = review_df.copy()
            if selected_filter == "Esitystavan muutos":
                filtered_review = filtered_review[
                    filtered_review.get("presentation_change_flag", pd.Series(index=filtered_review.index, dtype=str)).fillna("").isin([PRESENTATION_CHANGE_POSSIBLE, PRESENTATION_CHANGE_CLEAR])
                ]
            elif selected_filter != "Kaikki":
                status_map = {"Manuaalinen tarkistus": "Manual review", "Ei voitu varmistaa": "Could not verify"}
                filtered_review = filtered_review[filtered_review["status"] == status_map.get(selected_filter, selected_filter)]
            st.dataframe(make_display_df(filtered_review).style.apply(status_row_style, axis=1), use_container_width=True, height=420)

        # =====================================================
        # =====================================================
        st.subheader("Täsmäytys")
        st.dataframe(make_display_df(comparison_df).style.apply(status_row_style, axis=1), use_container_width=True, height=600)

        with st.expander("Tuloslaskelman rakenne"):
            st.dataframe(build_statement_structure_report(parsed_older, parsed_newer, "tuloslaskelma"), use_container_width=True, height=500)

        with st.expander("Taseen rakenne"):
            st.dataframe(build_statement_structure_report(parsed_older, parsed_newer, "tase"), use_container_width=True, height=500)

        with st.expander("Voitto ennen veroja -rivin tarkistus"):
            st.dataframe(build_pretax_result_control_report(comparison_df), use_container_width=True, height=260)

        with st.expander("Kaikki poimitut rivit"):
            st.dataframe(build_combined_extracted_rows_report(parsed_older, parsed_newer), use_container_width=True, height=600)

        # =====================================================
        # =====================================================
        with st.expander("Tärkeimmät havainnot"):
            top_findings = build_top_findings_df(comparison_df, limit=12)
            if top_findings.empty:
                st.success("Keskeisiä tarkistushavaintoja ei havaittu.")
            else:
                st.dataframe(top_findings.style.apply(status_row_style, axis=1), use_container_width=True, height=320)

        # =====================================================
        # =====================================================
        with st.expander("Yksi-moneen-ehdotukset ja esitystavan muutokset"):
            o2m_df = build_one_to_many_suggestions(comparison_df)
            presentation_df = build_presentation_change_df(comparison_df)

            st.markdown("**Yksi-moneen-ehdotukset**")
            if o2m_df.empty:
                st.info("Yksi-moneen-ehdotuksia ei havaittu.")
            else:
                st.dataframe(o2m_df, use_container_width=True, height=300)

            st.markdown("**Esitystavan muutokset**")
            if presentation_df.empty:
                st.info("Esitystavan muutokseen viittaavia havaintoja ei löytynyt.")
            else:
                st.dataframe(presentation_df, use_container_width=True, height=350)

        # =====================================================
        # =====================================================
        with st.expander("Päätöksen lisätiedot: miksi rivi täsmäytettiin"):
            st.dataframe(build_decision_trace_display_df(comparison_df), use_container_width=True, height=600)

        with st.expander("Luotettavuusraportit ja sanaston laadun tarkistus"):
            lang_report_df = build_language_reliability_report(parsed_older, parsed_newer, comparison_df)
            parent_validation_df = build_parent_validation_report(comparison_df)
            unknown_rows_df = build_unknown_rows_report(parsed_older, parsed_newer)

            st.markdown("**Kielikohtainen sanasto-osumaraportti**")
            if lang_report_df is not None and not lang_report_df.empty:
                lang_report_df = lang_report_df.rename(columns={
                    "Document": "Asiakirja",
                    "Rows parsed": "Poimitut rivit",
                    "Vocabulary matches total": "Sanasto-osumat yhteensä",
                    "Detected main language": "Tunnistettu pääkieli",
                    "Vocabulary matches per row %": "Sanasto-osumat / rivi %",
                })
            st.dataframe(lang_report_df, use_container_width=True)

            st.markdown("**Rakenne- ja kontekstivalidointi**")
            if parent_validation_df.empty:
                st.success("Parent- tai kontekstiristiriitoja ei havaittu.")
            else:
                st.dataframe(parent_validation_df, use_container_width=True, height=350)

            st.markdown("**Sanaston laadunvalvonta**")
            st.dataframe(build_dictionary_quality_report(), use_container_width=True, height=350)

            st.markdown("**Tuntemattomat tai epävarmat rivit**")
            if unknown_rows_df.empty:
                st.success("Tuntemattomia tai sanaston ulkopuolisia rivejä ei havaittu.")
            else:
                st.dataframe(unknown_rows_df, use_container_width=True, height=350)

        with st.expander("Tekninen tarkistus"):
            rel_df = pd.DataFrame([
                {
                    "Asiakirja": parsed_older["document"],
                    "Luotettavuuspisteet": rel_old.get("score"),
                    "Käyttöluokitus": rel_old.get("classification"),
                    "Syyperusteet": " | ".join(rel_old.get("reasons", [])),
                },
                {
                    "Asiakirja": parsed_newer["document"],
                    "Luotettavuuspisteet": rel_new.get("score"),
                    "Käyttöluokitus": rel_new.get("classification"),
                    "Syyperusteet": " | ".join(rel_new.get("reasons", [])),
                },
            ])
            st.dataframe(rel_df, use_container_width=True)
            st.markdown("**Tämän ajon testiraportti**")
            st.dataframe(build_test_case_metrics(parsed_older, parsed_newer, comparison_df), use_container_width=True)
            st.markdown("**Poikkeamaluokat**")
            st.dataframe(build_exception_category_summary(comparison_df), use_container_width=True)
            st.markdown("**Tekninen tarkistusdata**")
            st.dataframe(build_tech_debug_df(comparison_df), use_container_width=True, height=600)


# =========================================================
# EXTERNAL VOCABULARY LOADING
# =========================================================
# The terms.xlsx file may define these fields: canonical, synonym, language,
# category, section, parent, expected_side, item_type, active and notes.
#

EXTERNAL_TERMS_FILENAME = "terms.xlsx"
EXTERNAL_TERM_ROWS: list[dict] = []
EXTERNAL_TERMS_LOADED = False
EXTERNAL_TERMS_MESSAGE = "Käytössä sisäinen varasanasto."

SUPPORTED_TERM_SHEETS = ("FI", "EN", "SV")


def _external_terms_path() -> str:
    """Return a derived value used by the reconciliation workflow."""
    import os
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        base_dir = os.getcwd()

    env_path = os.environ.get("TERMS_XLSX")
    if env_path:
        return env_path

    pro_path = os.path.join(base_dir, "terms.xlsx")
    if os.path.exists(pro_path):
        return pro_path

    return os.path.join(base_dir, EXTERNAL_TERMS_FILENAME)


def _as_bool_active(value) -> bool:
   
    if value is None:
        return True
    text = str(value).strip().lower()
    return text not in {"0", "false", "ei", "no", "n", "inactive", "pois", "", "nan"}


def _append_unique_norm(target, value: str) -> None:
    
    norm = normalize_label(value or "")
    if not norm:
        return
    if isinstance(target, set):
        target.add(norm)
    elif norm not in target:
        target.append(norm)


def _append_unique_keyword(target, value: str) -> None:
    """Normalize input values for reliable comparison."""
    norm = normalize_keyword_text(value or "")
    if not norm:
        return
    if isinstance(target, set):
        target.add(norm)
    elif norm not in target:
        target.append(norm)


def _get_row_value(row: dict, *names: str, default=""):
    """Return a derived value used by the reconciliation workflow."""
    for name in names:
        key = name.lower()
        if key not in row:
            continue
        value = row[key]
        if value is None:
            continue
        try:
            if pd.isna(value):
                continue
        except Exception:
            pass
        return value
    return default


def _apply_external_term_to_runtime_lists(canonical: str, synonym: str, category: str, section: str, item_type: str) -> None:
    
    category = (category or "statement_item").strip().lower()

    if category in {"statement_item", "financial_statement_item", "line_item", "item"}:
        TERM_SYNONYM_MAP.setdefault(canonical, [])
        _append_unique_norm(TERM_SYNONYM_MAP[canonical], synonym)
        _append_unique_norm(TERM_SYNONYM_MAP[canonical], canonical)

        if section:
            SECTION_CANONICAL_TERMS.setdefault(section, [])
            _append_unique_norm(SECTION_CANONICAL_TERMS[section], canonical)
            if section in SECTION_ALLOWED_CANONICALS:
                _append_unique_norm(SECTION_ALLOWED_CANONICALS[section], canonical)

        if section == "tuloslaskelma":
            _append_unique_norm(INCOME_STATEMENT_LABEL_TERMS, canonical)
            _append_unique_norm(INCOME_STATEMENT_LABEL_TERMS, synonym)
        elif section == "tase_vastaavaa":
            _append_unique_norm(BALANCE_ASSET_LABEL_TERMS, canonical)
            _append_unique_norm(BALANCE_ASSET_LABEL_TERMS, synonym)
        elif section == "tase_vastattavaa":
            _append_unique_norm(BALANCE_LIABILITY_LABEL_TERMS, canonical)
            _append_unique_norm(BALANCE_LIABILITY_LABEL_TERMS, synonym)

        # Section-specific statement items are kept in section lists, not in generic.

    elif category in {"section_title", "section", "heading"}:
        _append_unique_keyword(TOC_RELATED_TERMS, synonym)
        _append_unique_keyword(TOC_RELATED_TERMS, canonical)
        _append_unique_norm(GENERIC_STATEMENT_LABEL_TERMS, synonym)
        _append_unique_norm(GENERIC_STATEMENT_LABEL_TERMS, canonical)
        if section:
            SECTION_CANONICAL_TERMS.setdefault(section, [])
            _append_unique_norm(SECTION_CANONICAL_TERMS[section], canonical)

    elif category in {"toc_keyword", "toc", "table_of_contents"}:
        _append_unique_keyword(TOC_KEYWORDS, synonym)
        _append_unique_keyword(TOC_KEYWORDS, canonical)

    elif category in {"skip_section", "skip", "ignore_section", "excluded_section"}:
        _append_unique_keyword(SKIP_SECTION_KEYWORDS, synonym)
        _append_unique_keyword(SKIP_SECTION_KEYWORDS, canonical)
        _append_unique_keyword(POST_BALANCE_STOP_TITLES, synonym)
        _append_unique_keyword(POST_BALANCE_STOP_TITLES, canonical)

    elif category in {"banned_label", "banned", "metadata", "document_metadata"}:
        _append_unique_norm(BANNED_LABEL_TERMS, synonym)
        _append_unique_norm(BANNED_LABEL_TERMS, canonical)

    elif category in {"total_word", "sum_word", "total", "summary"}:
        _append_unique_norm(SUM_KEYWORDS, synonym)
        _append_unique_norm(SUM_KEYWORDS, canonical)
        _append_unique_norm(SUBGROUP_TOTAL_ITEMS, canonical)
        if section == "tase_vastaavaa":
            _append_unique_norm(VASTAAVAA_TOTAL_TERMS, synonym)
            _append_unique_norm(VASTAAVAA_TOTAL_TERMS, canonical)
        elif section == "tase_vastattavaa":
            _append_unique_norm(VASTATTAVAA_TOTAL_TERMS, synonym)
            _append_unique_norm(VASTATTAVAA_TOTAL_TERMS, canonical)
    else:
        TERM_SYNONYM_MAP.setdefault(canonical, [])
        _append_unique_norm(TERM_SYNONYM_MAP[canonical], synonym)
        # Unknown categories are available through TERM_SYNONYM_MAP only.

    if item_type in {"summary", "grand_total", "total", "sub_total", "subtotal"}:
        _append_unique_norm(SUBGROUP_TOTAL_ITEMS, canonical)
        _append_unique_norm(SUM_KEYWORDS, synonym)


def _normalise_term_dataframe(frame: pd.DataFrame, sheet_name: str) -> Optional[pd.DataFrame]:
    """Return a derived value used by the reconciliation workflow."""
    if frame is None or frame.empty:
        return None

    lower_cols = {str(c).strip().lower(): c for c in frame.columns}
    if not {"canonical", "synonym"}.issubset(lower_cols):
        for idx in range(min(8, len(frame))):
            values = [str(v).strip().lower() for v in frame.iloc[idx].tolist()]
            if "canonical" in values and "synonym" in values:
                tmp = frame.iloc[idx + 1:].copy()
                tmp.columns = [str(v).strip() for v in frame.iloc[idx].tolist()]
                frame = tmp
                lower_cols = {str(c).strip().lower(): c for c in frame.columns}
                break

    if not {"canonical", "synonym"}.issubset(lower_cols):
        return None

    tmp = frame.copy()
    if "language" not in lower_cols:
        tmp["language"] = sheet_name.lower() if sheet_name in SUPPORTED_TERM_SHEETS else ""
    tmp["source_sheet"] = sheet_name
    return tmp


def _collect_external_term_frames(sheets: dict[str, pd.DataFrame]) -> list[pd.DataFrame]:
    
    frames: list[pd.DataFrame] = []
    for sheet_name in SUPPORTED_TERM_SHEETS:
        frame = _normalise_term_dataframe(sheets.get(sheet_name), sheet_name)
        if frame is not None:
            frames.append(frame)

    legacy = _normalise_term_dataframe(sheets.get("terms"), "terms")
    if legacy is not None and not frames:
        frames.append(legacy)
    elif legacy is not None and frames:
        frames.append(legacy)

    return frames


def _normalise_external_terms_table(frames: list[pd.DataFrame]) -> pd.DataFrame:
    """Normalize input values for reliable comparison."""
    if not frames:
        return pd.DataFrame()

    df_terms = pd.concat(frames, ignore_index=True)
    df_terms.columns = [str(c).strip() for c in df_terms.columns]
    lower_cols = {str(c).strip().lower(): c for c in df_terms.columns}

    defaults = {
        "section": "",
        "parent": "",
        "item_type": "detail",
        "expected_side": "",
        "language": "",
        "category": "statement_item",
        "match_policy": "",
        "active": 1,
        "notes": "",
        "source_sheet": "",
    }
    for col, default in defaults.items():
        if col not in lower_cols:
            df_terms[col] = default

    canonical_col = lower_cols.get("canonical", "canonical")
    synonym_col = lower_cols.get("synonym", "synonym")
    language_col = lower_cols.get("language", "language")

    df_terms["_dedupe_canonical"] = df_terms[canonical_col].astype(str).map(normalize_label)
    df_terms["_dedupe_synonym"] = df_terms[synonym_col].astype(str).map(normalize_label)
    df_terms["_dedupe_language"] = df_terms[language_col].astype(str).str.strip().str.lower()
    df_terms = df_terms.drop_duplicates(
        subset=["_dedupe_canonical", "_dedupe_synonym", "_dedupe_language"],
        keep="first",
    )
    return df_terms


def validate_external_terms_dataframe(df_terms: pd.DataFrame) -> list[dict]:
    """Validate the optional terms.xlsx vocabulary before it changes runtime matching rules."""
    issues: list[dict] = []
    if df_terms is None or df_terms.empty:
        return [{
            "severity": "high",
            "check": "empty_terms",
            "message": "No readable terms were found in the vocabulary.",
        }]

    allowed_sections = {"", "tuloslaskelma", "tase_vastaavaa", "tase_vastattavaa"}
    synonym_to_canonicals: dict[tuple[str, str], set[str]] = {}
    duplicate_keys: Counter = Counter()

    for _, raw in df_terms.iterrows():
        row = {str(k).strip().lower(): v for k, v in raw.items()}
        active = _as_bool_active(_get_row_value(row, "active", default=1))
        if not active:
            continue

        canonical = normalize_label(_get_row_value(row, "canonical"))
        synonym = normalize_label(str(_get_row_value(row, "synonym") or ""))
        section = normalize_section_name(_get_row_value(row, "section")) if str(_get_row_value(row, "section") or "").strip() else ""
        parent = normalize_label(_get_row_value(row, "parent")) if str(_get_row_value(row, "parent") or "").strip() else ""
        item_type = str(_get_row_value(row, "item_type", "type", default="detail") or "detail").strip().lower()
        language = str(_get_row_value(row, "language", default="") or "").strip().lower()

        if not canonical or not synonym:
            issues.append({
                "severity": "high",
                "check": "missing_required_value",
                "message": "An active vocabulary row is missing canonical or synonym.",
            })
            continue

        if section not in allowed_sections:
            issues.append({
                "severity": "high",
                "check": "unknown_section",
                "message": f"Canonical '{canonical}' uses unknown section '{section}'.",
            })

        if item_type == "detail" and section and not parent:
            issues.append({
                "severity": "info",
                "check": "detail_without_parent",
                "message": f"Detail item '{canonical}' has no parent. This may be allowed for detail rows.",
            })

        key = (language, canonical, synonym)
        duplicate_keys[key] += 1
        synonym_to_canonicals.setdefault((language, synonym), set()).add(canonical)

    for (language, canonical, synonym), count in duplicate_keys.items():
        if count > 1:
            issues.append({
                "severity": "warning",
                "check": "duplicate_term",
                "message": f"Sama termi esiintyy {count} kertaa: {language}/{canonical}/{synonym}.",
            })

    for (language, synonym), canonicals in synonym_to_canonicals.items():
        if synonym and len(canonicals) > 1:
            issues.append({
                "severity": "high",
                "check": "synonym_maps_to_multiple_canonicals",
                "message": f"Synonyymi '{synonym}' ({language or '?'}) osoittaa useaan canonicaliin: {', '.join(sorted(canonicals))}.",
            })

    return issues


TERM_VALIDATION_ISSUES: list[dict] = []


def load_external_terms_from_excel(path: Optional[str] = None) -> None:
    """Load optional vocabulary extensions from terms.xlsx if the file is available."""
    global EXTERNAL_TERMS_LOADED, EXTERNAL_TERMS_MESSAGE, EXTERNAL_TERM_ROWS, TERM_VALIDATION_ISSUES
    import os

    path = path or _external_terms_path()
    EXTERNAL_TERM_ROWS = []
    TERM_VALIDATION_ISSUES = []

    if not os.path.exists(path):
        EXTERNAL_TERMS_LOADED = False
        EXTERNAL_TERMS_MESSAGE = "terms.xlsx was not found in the same folder. Internal fallback vocabulary in use."
        return

    try:
        sheets = pd.read_excel(path, sheet_name=None)
    except Exception as exc:
        EXTERNAL_TERMS_LOADED = False
        EXTERNAL_TERMS_MESSAGE = f"terms.xlsx was found, but it could not be read: {exc}. Internal fallback vocabulary in use."
        return

    frames = _collect_external_term_frames(sheets)
    if not frames:
        EXTERNAL_TERMS_LOADED = False
        EXTERNAL_TERMS_MESSAGE = (
            "terms.xlsx did not contain a readable vocabulary sheet. "
            "Expected structure: FI/EN/SV or a legacy terms sheet containing canonical and synonym."
        )
        return

    df_terms = _normalise_external_terms_table(frames)
    TERM_VALIDATION_ISSUES = validate_external_terms_dataframe(df_terms)

    lower_cols = {str(c).strip().lower(): c for c in df_terms.columns}
    loaded = 0
    language_counter = Counter()
    source_sheet_counter = Counter()

    for _, raw in df_terms.iterrows():
        row = {str(k).strip().lower(): v for k, v in raw.items()}
        if not _as_bool_active(_get_row_value(row, "active", default=1)):
            continue

        canonical = normalize_label(_get_row_value(row, "canonical"))
        synonym = normalize_label(str(_get_row_value(row, "synonym") or ""))
        section_raw = _get_row_value(row, "section")
        section = normalize_section_name(section_raw) if str(section_raw or "").strip() else ""
        parent = normalize_label(_get_row_value(row, "parent")) if str(_get_row_value(row, "parent") or "").strip() else ""
        item_type = str(_get_row_value(row, "item_type", "type", default="detail") or "detail").strip().lower()
        expected_side = str(_get_row_value(row, "expected_side", default="") or "").strip().lower()
        language = str(_get_row_value(row, "language", default="") or "").strip().lower()
        category = str(_get_row_value(row, "category", default="statement_item") or "statement_item").strip().lower()
        match_policy = str(_get_row_value(row, "match_policy", default="") or "").strip().lower()
        source_sheet = str(_get_row_value(row, "source_sheet", default="") or "").strip()

        if not canonical or not synonym:
            continue

        _apply_external_term_to_runtime_lists(canonical, synonym, category, section, item_type)

        EXTERNAL_TERM_ROWS.append({
            "canonical": canonical,
            "synonym": synonym,
            "section": section,
            "parent": parent,
            "type": item_type,
            "expected_side": expected_side,
            "language": language,
            "category": category,
            "match_policy": match_policy,
            "source_sheet": source_sheet,
        })
        loaded += 1
        language_counter[language or "?"] += 1
        source_sheet_counter[source_sheet or "?"] += 1

    # terms.xlsx does not automatically expand the MAIN_ITEM_STARTERS list.

    EXTERNAL_TERMS_LOADED = loaded > 0
    issue_count = len(TERM_VALIDATION_ISSUES)
    high_issue_count = sum(1 for i in TERM_VALIDATION_ISSUES if i.get("severity") == "high")

    if loaded:
        lang_msg = ", ".join(f"{k.upper()}: {v}" for k, v in sorted(language_counter.items()))
        sheet_msg = ", ".join(f"{k}: {v}" for k, v in sorted(source_sheet_counter.items()))
        issue_msg = (
            f" Vocabulary validation: {issue_count} findings, of which {high_issue_count} are high-risk."
            if issue_count else " Vocabulary validation: no findings."
        )
        EXTERNAL_TERMS_MESSAGE = (
            f"External vocabulary loaded: {loaded} active term rows from terms.xlsx "
            f"({lang_msg}; sheets: {sheet_msg}).{issue_msg}"
        )
    else:
        EXTERNAL_TERMS_MESSAGE = "terms.xlsx was read, but no active terms were found. Internal fallback vocabulary in use."

def apply_external_metadata_from_terms() -> None:
    """Apply external metadata from terms.
    
    Purpose: This function belongs to the normalization and canonical recognition stage.
    Why: It converts inconsistent financial statement wording into comparable internal concepts.
    """
    for row in EXTERNAL_TERM_ROWS:
        if row.get("category") not in {"statement_item", "financial_statement_item", "line_item", "item", "total_word", "sum_word", "total", "summary"}:
            continue
        section = row.get("section")
        canonical = row.get("canonical")
        parent = row.get("parent") or None
        item_type = row.get("type") or "detail"
        expected_side = row.get("expected_side") or ""
        if section and canonical:
            _add_canonical_metadata(canonical, section, parent=parent, item_type=item_type, expected_side=expected_side)


# =========================================================
# RUNTIME TAXONOMY INITIALIZATION
# =========================================================
# The definitions above describe the internal vocabulary and all helper
# functions needed to validate an optional external vocabulary file. This
# initialization step is intentionally placed here: at this point the loader,
# normalizers and metadata builders are all available, so the active taxonomy
# can be built once in a controlled order before the Streamlit app is started.


def clear_taxonomy_caches() -> None:
    """Clear cached label lookups after the active vocabulary has changed."""
    try:
        _known_label_phrases_for_degluing.cache_clear()
        deglue_pdf_label_text.cache_clear()
        _all_allowed_canonical_keys.cache_clear()
        _safe_structural_canonical_from_label.cache_clear()
        forced_main_item_canonical.cache_clear()
    except Exception:
        # Cache clearing is a safety step. A failure here should not prevent
        # the prototype from starting, because the following metadata build
        # still uses the current in-memory taxonomy.
        pass


def initialize_runtime_taxonomy() -> None:
    """Build the active vocabulary, canonical indexes and metadata registry."""
    apply_pma_practical_extension_terms()
    load_external_terms_from_excel()
    rebuild_pma_taxonomy_indexes()
    clear_taxonomy_caches()
    _build_canonical_item_metadata()
    apply_external_metadata_from_terms()


initialize_runtime_taxonomy()


if __name__ == "__main__":
    main()
